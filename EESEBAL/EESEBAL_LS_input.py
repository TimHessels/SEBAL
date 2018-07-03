# -*- coding: utf-8 -*-

"""
pySEBAL_3.3.8

@author: Tim Hessels, Jonna van Opstal, Patricia Trambauer, Wim Bastiaanssen, Mohamed Faouzi Smiej, Yasir Mohamed, and Ahmed Er-Raji
         UNESCO-IHE
         September 2017
"""
import sys
import os
import re
import shutil
import numpy as np
import datetime
from osgeo import osr  
import gdal  
from math import sin, cos, pi, tan
import time
import subprocess
import numpy.polynomial.polynomial as poly	
from openpyxl import load_workbook
from pyproj import Proj, transform
import warnings

def main(number, Landsat_nr, inputExcel):
    
    print '---------------------------------------------------------'
    print '-------------------- Open Landsat -----------------------'
    print '---------------------------------------------------------'
   
    # Define bands used for each Landsat number
    if Landsat_nr == 5 or Landsat_nr == 7:
        Bands = np.array([1, 2, 3, 4, 5, 7, 6])
    elif Landsat_nr == 8:
       Bands = np.array([2, 3, 4, 5, 6, 7, 10, 11])  
    else:
        print 'Landsat image not supported, use Landsat 7 or 8'

    # Open MTL landsat and get the correction parameters   
    Landsat_meta_fileName = os.path.join(input_folder, '%s_MTL.txt' %Name_Landsat_Image)
    Lmin, Lmax, k1_c, k2_c = info_band_metadata(Landsat_meta_fileName, Bands)
    print 'Lmin= ', Lmin
    print 'Lmax= ', Lmax
    print 'k1= ', k1_c
    print 'k2= ', k2_c
    
    # Calibration parameters - comment
    #Lmin_L5 = [-2.8, -1.2, -1.5, -0.37, 1.2378, -0.150]    # REVISE LS 5 TM
    #Lmax_L5 = [296.8, 204.3, 206.2, 27.19, 15.303, 14.38]  # REVISE LS 5 TM

    # Mean solar exo-atmospheric irradiance for each band (W/m2/microm)
    # for the different Landsat images (L5, L7, or L8)
    ESUN_L5 = np.array([1983, 1796, 1536, 1031, 220, 83.44])
    ESUN_L7 = np.array([1997, 1812, 1533, 1039, 230.8, 84.9])
    ESUN_L8 = np.array([1973.28, 1842.68, 1565.17, 963.69, 245, 82.106])

    # Open one band - To get the metadata of the landsat images only once (to get the extend)
    src_FileName = os.path.join(input_folder, '%s_B2.TIF' %Name_Landsat_Image)  # before 10!
    ls, band_data, ulx, uly, lrx, lry, x_size_ls, y_size_ls = Get_Extend_Landsat(src_FileName)
    print '  Upper Left corner x, y: ', ulx, ', ', uly
    print '  Lower right corner x, y: ', lrx, ', ', lry
     
    # Crop  the Landsat images to the DEM extent -
    dst_FileName = os.path.join(output_folder, 'Output_temporary', '%s_cropped_LS_b2_%s_%s_%s.tif' %(sensor1, res1, year, DOY))  # Before 10 !!
    dir_name = os.path.dirname(dst_FileName)  # Directory of the file
 
    # If the directory does not exist, create it.
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)
         
    lsc, ulx, uly, lrx, lry, epsg_to = reproject_dataset_example(src_FileName, proyDEM_fileName)											
    
    #	Get the extend of the remaining landsat file	after clipping based on the DEM file	
    y_size_lsc = lsc.RasterYSize
    x_size_lsc = lsc.RasterXSize    
    shape_lsc = [x_size_lsc, y_size_lsc]
    
    print '--- '
    print 'Cropped LANDSAT Image - '
    print '  Size :', x_size_lsc, y_size_lsc
    print '  Upper Left corner x, y: ', ulx, ', ',  uly
    print '  Lower right corner x, y: ', lrx, ', ', lry
   
    # output names for resampling
    dst_LandsatMask = os.path.join(output_folder, 'Output_temporary', '%s_cropped_LANDSATMASK_%s_%s_%s.tif' %(sensor1, res1, year, DOY))

    # Open Landsat data only if all additional data is not defined.

    # Open the Additional input excel sheet
    ws = wb['Additional_Input']
							
    # If all additional fields are filled in than do not open the Landsat data
    if (ws['B%d' % number].value) is None or (ws['C%d' % number].value)  is None or (ws['D%d' % number].value)  is None or (ws['E%d' % number].value) is None:					

        # Collect the landsat Thermal and Spectral data
        # 1. Create mask for the landsat images
        # 2. Save the Thermal data in a 3D array
        # 3. Save the Spectral data in a 3D array
    
        #  1.)
        # find clipping extent for landsat images: NTIR is larger than VTIR
        # Open original Landsat image for the band number 11

        # if landsat 5 or 7 is used then first create a mask for removing the no data stripes       
        if Landsat_nr == 5 or Landsat_nr == 7:
            src_FileName = os.path.join(input_folder, '%s_B6.TIF' % (Name_Landsat_Image)) #open smallest band
            if not os.path.exists(src_FileName):
                src_FileName = os.path.join(input_folder, '%s_B6_VCID_2.TIF' % (Name_Landsat_Image))														
            src_FileName_2 = os.path.join(input_folder, '%s_B1.TIF' % (Name_Landsat_Image)) #open smallest band
            src_FileName_3 = os.path.join(input_folder, '%s_B3.TIF' % (Name_Landsat_Image)) #open smallest band
            src_FileName_4 = os.path.join(input_folder, '%s_B4.TIF' % (Name_Landsat_Image)) #open smallest band
            src_FileName_5 = os.path.join(input_folder, '%s_B7.TIF' % (Name_Landsat_Image)) #open smallest band
            src_FileName_6 = os.path.join(input_folder, '%s_B2.TIF' % (Name_Landsat_Image)) #open smallest band
            src_FileName_7 = os.path.join(input_folder, '%s_B5.TIF' % (Name_Landsat_Image)) #open smallest band
            ls_data=Open_landsat(src_FileName,proyDEM_fileName)
            ls_data_2=Open_landsat(src_FileName_2,proyDEM_fileName)
            ls_data_3=Open_landsat(src_FileName_3,proyDEM_fileName)
            ls_data_4=Open_landsat(src_FileName_4,proyDEM_fileName)
            ls_data_5=Open_landsat(src_FileName_5,proyDEM_fileName)
            ls_data_6=Open_landsat(src_FileName_6,proyDEM_fileName)
            ls_data_7=Open_landsat(src_FileName_7,proyDEM_fileName)
                    
            # create and save the landsat mask for all images based on band 11 (smallest map)
            ClipLandsat=np.ones((shape_lsc[1], shape_lsc[0]))
            ClipLandsat=np.where(np.logical_or(np.logical_or(np.logical_or(np.logical_or(np.logical_or(np.logical_or(ls_data==0,ls_data_2==0),ls_data_3==0),ls_data_4==0),ls_data_5==0),ls_data_6==0),ls_data_7==0),0,1)

        # If landsat 8 then use landsat band 10 and 11
        elif Landsat_nr == 8:
             src_FileName_11 = os.path.join(input_folder, '%s_B11.TIF' % (Name_Landsat_Image)) #open smallest band
             ls_data_11=Open_landsat(src_FileName_11,proyDEM_fileName)

             src_FileName_10 = os.path.join(input_folder, '%s_B10.TIF' % (Name_Landsat_Image)) #open smallest band
             ls_data_10=Open_landsat(src_FileName_10, proyDEM_fileName)
 
             # create and save the landsat mask for all images based on band 10 and 11
             ClipLandsat=np.ones((shape_lsc[1], shape_lsc[0]))
             ClipLandsat=np.where(np.logical_or(ls_data_11==0, ls_data_10==0),0,1)

        else:
            print 'Landsat image not supported, use Landsat 7 or 8'

        # Create Cloud mask is BQA map is available (newer version Landsat images)
        BQA_LS_Available = 0
        if os.path.exists(os.path.join(input_folder, '%s_BQA.TIF' %Name_Landsat_Image)):
            src_FileName_BQA = os.path.join(input_folder, '%s_BQA.TIF' %Name_Landsat_Image)
            ls_data_BQA = Open_landsat(src_FileName_BQA,proyDEM_fileName)
            if Landsat_nr == 8:
                Cloud_Treshold = 2721
            if Landsat_nr == 5 or Landsat_nr == 7:
                Cloud_Treshold = 700                    
            QC_mask_Cloud = np.copy(ls_data_BQA)
            QC_mask_Cloud[ls_data_BQA<Cloud_Treshold] = 0
            QC_mask_Cloud[ls_data_BQA>=Cloud_Treshold] = 1                             
            BQA_LS_Available = 1               
            

        # Open data of the landsat mask                            
        ls_data=Open_landsat(src_FileName, proyDEM_fileName)
   
        # Save Landsat mask as a tiff file							
        save_GeoTiff_proy(lsc, ClipLandsat, dst_LandsatMask, shape_lsc, nband=1)
   
        # 2.)          
        # Create 3D array to store the Termal band(s) (nr10(&11) for LS8 and n6 for LS7) 
        therm_data = Landsat_therm_data(Bands,input_folder,Name_Landsat_Image,output_folder,shape_lsc,ClipLandsat, proyDEM_fileName)          
   
        # 3.)
        # Create 3D array to store Spectral radiance and Reflectivity for each band
        Reflect, Spec_Rad = Landsat_Reflect(Bands, input_folder, Name_Landsat_Image, output_folder, shape_lsc, ClipLandsat, Lmax, Lmin, ESUN_L5, ESUN_L7, ESUN_L8, cos_zn, dr, Landsat_nr, proyDEM_fileName)
   
        # save spectral data 
        for i in range(0,6):							
            spec_ref_fileName = os.path.join(output_folder, 'Output_radiation_balance','%s_spectral_reflectance_B%s_%s_%s_%s.tif' %(Bands[i], sensor1, res3, year, DOY))
            save_GeoTiff_proy(lsc, Reflect[:, :, i], spec_ref_fileName, shape_lsc, nband=1)
      								
        # ------------------------------------------------------------------------
        # ------------------------------------------------------------------------
        # ----   MODULE 3 - Vegetation properties

    print '---------------------------------------------------------'
    print '-------------------- Module 3 ---------------------------'
    print '---------------------------------------------------------'
 						
    # Check NDVI							
    try:
        if (ws['B%d' % number].value) is not None:					
            # Output folder NDVI								
            ndvi_fileName2 = os.path.join(output_folder, 'Output_vegetation', 'User_NDVI_%s_%s_%s.tif' %(res3, year, DOY))
            NDVI=Reshape_Reproject_Input_data(r'%s' %str(ws['B%d' % number].value),ndvi_fileName2,proyDEM_fileName)		 														 

            water_mask_temp = np.zeros((shape_lsc[1], shape_lsc[0]))
            water_mask_temp[NDVI < 0.0] = 1.0
			  								
        else:
            # use the Landsat reflectance to calculate the surface albede, NDVI
            NDVI = Calc_NDVI(Reflect)
      
            # save landsat NDVI	
            ndvi_fileName2 = os.path.join(output_folder, 'Output_vegetation', '%s_NDVI_%s_%s_%s.tif' %(sensor3, res3, year, DOY))			
            save_GeoTiff_proy(lsc, NDVI, ndvi_fileName2, shape_lsc, nband=1)	
            
            # Calculate temporal water mask
            water_mask_temp=Water_Mask(shape_lsc,Reflect)
            
    except:
        assert "Please check the NDVI input path"
        
    # Check Water Mask	            
    try:
        if (ws['C%d' % number].value) is not None:	
				
            # Overwrite the Water mask and change the output name					
            water_mask_temp_fileName = os.path.join(output_folder, 'Output_soil_moisture', 'User_Water_mask_temporary_%s_%s_%s.tif' %(res2, year, DOY))
            water_mask_temp = Reshape_Reproject_Input_data(r'%s' %str(ws['C%d' % number].value), water_mask_temp_fileName, proyDEM_fileName)		 														 
            
    except:
        assert "Please check the Water Mask input path"            
												
    # Check Surface albedo	
    try:
        if (ws['D%d' % number].value) is not None:					
            # Output folder surface albedo						
            surface_albedo_fileName = os.path.join(output_folder, 'Output_vegetation','User_surface_albedo_%s_%s_%s.tif' %(res2, year, DOY))
            Surf_albedo=Reshape_Reproject_Input_data(r'%s' %str(ws['D%d' % number].value),surface_albedo_fileName,proyDEM_fileName)		 					
		 									
        else:
            surface_albedo_fileName = os.path.join(output_folder, 'Output_vegetation','%s_surface_albedo_%s_%s_%s.tif' %(sensor1, res2, year, DOY))

            # use the Landsat reflectance to calculate the surface albede, NDVI
            Surf_albedo = Calc_albedo(Reflect,path_radiance,Apparent_atmosf_transm)
 
            # save landsat surface albedo
            save_GeoTiff_proy(lsc, Surf_albedo, surface_albedo_fileName, shape_lsc, nband=1)																  
    except:
          assert "Please check the Albedo input path"        
	 						
    # calculate vegetation properties
    FPAR,tir_emis,Nitrogen,vegt_cover,LAI,b10_emissivity=Calc_vegt_para(NDVI, water_mask_temp,shape_lsc)

    # Save output maps that will be used in SEBAL
    save_GeoTiff_proy(lsc, water_mask_temp, water_mask_temp_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, FPAR, fpar_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, tir_emis, tir_emissivity_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, Nitrogen, nitrogen_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, vegt_cover, veg_cover_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, LAI, lai_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, b10_emissivity, b10_emissivity_fileName, shape_lsc, nband=1)
   
    # ------------------------------------------------------------------------
    # ------------------------------------------------------------------------
    # ----   MODULE 4  - Surface temperature, Cloud, Water, and Snow mask
 
    print '---------------------------------------------------------'
    print '-------------------- Module 4 ---------------------------'
    print '---------------------------------------------------------'
 
    # Check if a surface temperature dataset is defined. If so use this one instead of the Landsat, otherwise Landsat
 
    # Check Surface temperature	
    try:
        if (ws['E%d' % number].value) is not None:				
            # Output folder surface temperature						
            surf_temp_fileName = os.path.join(output_folder, 'Output_vegetation','User_surface_temp_%s_%s_%s.tif' %(res2, year, DOY))
            Surface_temp=Reshape_Reproject_Input_data(r'%s' %str(ws['E%d' % number].value),surf_temp_fileName,proyDEM_fileName)		 					
            cloud_mask = np.zeros([int(np.shape(Surface_temp)[1]),int(np.shape(Surface_temp)[0])])
            
            snow_mask, water_mask, ts_moist_veg_min, NDVI_max, NDVI_std = CalculateSnowWaterMask(NDVI,shape_lsc,water_mask_temp,Surface_temp)
            QC_Map = np.zeros((shape_lsc[1], shape_lsc[0]))
            QC_Map[np.isnan(Surface_temp)] = 1	 									
        else:
            
            # Calculate surface temperature and create a cloud mask
            Surface_temp,cloud_mask = Calc_surface_water_temp(Temp_inst, Landsat_nr, Lmax, Lmin, therm_data, b10_emissivity, k1_c, k2_c, eact_inst, shape_lsc, water_mask_temp, Bands_thermal, Rp, tau_sky, surf_temp_offset, Image_Type)    

            
            # Replace clouds mask is a better one is already created
            if BQA_LS_Available == 1:
                cloud_mask = QC_mask_Cloud
            
            Surface_temp[cloud_mask == 1] = np.nan
            
    except:
        assert "Please check the surface temperature input path"																

    # Perform Thermal sharpening for the thermal LANDSAT image
    if Thermal_Sharpening_not_needed is 1:  
        temp_surface_sharpened = Surface_temp
        
    if Thermal_Sharpening_not_needed is 0:

		    # Upscale DEM to 90m
        pixel_spacing_upscale=90

        dest_90, ulx_dem_90, lry_dem_90, lrx_dem_90, uly_dem_90, epsg_to = reproject_dataset(
            DEM_fileName, pixel_spacing_upscale, UTM_Zone = UTM_Zone)

        DEM_90 = dest_90.GetRasterBand(1).ReadAsArray()
        Y_raster_size_90 = dest_90.RasterYSize				
        X_raster_size_90 = dest_90.RasterXSize
        shape_90=([X_raster_size_90, Y_raster_size_90])
						
        save_GeoTiff_proy(dest_90, DEM_90, proyDEM_fileName_90, shape_90, nband=1)
	  	
		
        # save landsat surface temperature
        surf_temp_fileName = os.path.join(output_folder, 'Output_vegetation','%s_%s_surface_temp_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
        save_GeoTiff_proy(lsc, Surface_temp, surf_temp_fileName, shape_lsc, nband=1)							

        # Upscale NDVI data																																	
        dest_up, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
                                   ndvi_fileName2, proyDEM_fileName_90)
  
        NDVI_Landsat_up = dest_up.GetRasterBand(1).ReadAsArray()

        # Upscale Thermal data
        dest_up, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
                                   surf_temp_fileName, proyDEM_fileName_90)
        surface_temp_up = dest_up.GetRasterBand(1).ReadAsArray()
 
        # Define the width of the moving window box
        Box=7	
  
        # Apply thermal sharpening           
        temp_surface_sharpened = Thermal_Sharpening(surface_temp_up, NDVI_Landsat_up, NDVI, Box, dest_up, output_folder, ndvi_fileName2, shape_lsc, lsc, temp_surface_sharpened_fileName)	
			
    # Divide temporal watermask in snow and water mask by using surface temperature
    snow_mask, water_mask, ts_moist_veg_min, NDVI_max, NDVI_std = CalculateSnowWaterMask(NDVI,shape_lsc,water_mask_temp,temp_surface_sharpened)

    # Save the water mask
    save_GeoTiff_proy(lsc, water_mask, water_mask_fileName, shape_lsc, nband=1)
								
    if Thermal_Sharpening_not_needed is 0:
        temp_surface_sharpened[water_mask == 1] = Surface_temp[water_mask == 1]
        
		   # save landsat surface temperature
        surf_temp_fileName = os.path.join(output_folder, 'Output_vegetation','%s_%s_surface_temp_sharpened_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
        save_GeoTiff_proy(lsc, temp_surface_sharpened, surf_temp_fileName, shape_lsc, nband=1)	        
						
    # remove low temperature values
    temp_surface_sharpened[temp_surface_sharpened<=253.0]=np.nan
		 					
    # Calculate the tempearture of the water       							
    Temperature_water_std=np.nanstd(temp_surface_sharpened[water_mask != 0])
    Temperature_water_mean=np.nanmean(temp_surface_sharpened[water_mask != 0])
  
    print 'Mean water Temperature = %0.3f (K)' % Temperature_water_mean
    print 'Standard deviation water temperature = %0.3f (K)' % Temperature_water_std

    # Check if Quality dataset is defined. If so use this one instead of using Landsat otherwise landsat

    # Check Quality
    try:
        if (ws['F%d' % number].value) is not None:					
            # Output folder QC defined by the user							
            QC_Map_fileName = os.path.join(output_folder, 'Output_cloud_masked', 'User_quality_mask_%s_%s_%s.tif' %(res2, year, DOY))
 
            # Reproject and reshape users NDVI  
            QC_Map = Reshape_Reproject_Input_data(r'%s' %str(ws['F%d' % number].value),QC_Map_fileName,proyDEM_fileName)		 														 

		    # if the users QC data cannot be reprojected than use the original Landsat data as imported into SEBAL		
        else:
											
            # if there are no cold water pixels than use cold vegetation pixels       
            if np.isnan(Temperature_water_mean) == True or Temperature_water_mean < 0.0: 
                ts_cold_land=ts_moist_veg_min
            else:
                ts_cold_land=Temperature_water_mean

            # Make shadow mask
            mask=np.zeros((shape_lsc[1], shape_lsc[0]))
            mask[np.logical_and.reduce((temp_surface_sharpened < (ts_cold_land+Temperature_offset_shadow),Surf_albedo < Maximum_shadow_albedo,water_mask!=1))]=1
            shadow_mask=np.copy(mask)
            shadow_mask = Create_Buffer(shadow_mask)
   
            # Make cloud mask
            if BQA_LS_Available != 1:
                mask=np.zeros((shape_lsc[1], shape_lsc[0]))
                mask[np.logical_and.reduce((temp_surface_sharpened < (ts_cold_land+Temperature_offset_clouds),Surf_albedo > Minimum_cloud_albedo,NDVI<0.7,snow_mask!=1))]=1
                cloud_mask=np.copy(mask)
                cloud_mask = Create_Buffer(cloud_mask)                
       
            # Save output maps
            save_GeoTiff_proy(lsc, cloud_mask, cloud_mask_fileName, shape_lsc, nband=1)
            save_GeoTiff_proy(lsc, snow_mask, snow_mask_fileName, shape_lsc, nband=1)                  
            save_GeoTiff_proy(lsc, shadow_mask, shadow_mask_fileName, shape_lsc, nband=1)

            # Total Quality Mask
            QC_Map = np.empty(cloud_mask.shape)
            ClipLandsat_reverse = np.where(ClipLandsat==1,0,1)
            Landsat_Mask = cloud_mask + snow_mask + shadow_mask + ClipLandsat_reverse
            QC_Map[Landsat_Mask>0] = 1								
											
            # Output folder QC defined by the user							
            QC_Map_fileName = os.path.join(output_folder, 'Output_cloud_masked', '%s_quality_mask_%s_%s_%s.tif.tif' %(sensor1, res2, year, DOY))
											
            # Save the PROBA-V NDVI as tif file												
            save_GeoTiff_proy(lsc, QC_Map, QC_Map_fileName, shape, nband=1)
            save_GeoTiff_proy(dest, QC_Map, QC_Map_fileName, shape, nband=1)								  

    except:                
        assert "Please check the quality path"									


    print '...................................................................'
    print '............................DONE!..................................'
    print '...................................................................'

    # ------------------------------------------------------------------------
    # ------------------------------------------------------------------------
    # FUNCTIONS
    #-------------------------------------------------------------------------

def Create_Buffer(Data_In):
    
   '''
   This function creates a 3D array which is used to apply the moving window
   '''
   Buffer_area = 2 # A block of 2 times Buffer_area + 1 will be 1 if there is the pixel in the middle is 1
   Data_Out=np.empty((len(Data_In),len(Data_In[1])))   
   Data_Out[:,:] = Data_In
   for ypixel in range(1,Buffer_area + 1):
        Data_Out[:,0:-ypixel] += Data_In[:,ypixel:]
        Data_Out[:,ypixel:] += Data_In[:,:-ypixel]
        
        for xpixel in range(1,Buffer_area + 1):
           Data_Out[0:-xpixel,ypixel:] += Data_In[xpixel:,:-ypixel]
           Data_Out[xpixel:,ypixel:] += Data_In[:-xpixel,:-ypixel]
           Data_Out[0:-xpixel,0:-ypixel] += Data_In[xpixel:,ypixel:]
           Data_Out[xpixel:,0:-ypixel] += Data_In[:-xpixel,ypixel:]

           if ypixel==1:
                Data_Out[xpixel:,:] += Data_In[:-xpixel,:]
                Data_Out[:,0:-xpixel] += Data_In[:,xpixel:]


   Data_Out[Data_Out>0.1] = 1
   Data_Out[Data_Out<=0.1] = 0  
           
   return(Data_Out)

def Calc_Biomass_production(LAI,ETP_24,moisture_stress_biomass,ETA_24,Ra_mountain_24,Transm_24,FPAR,esat_24,eact_24,Th,Kt,Tl,Temp_24,LUEmax):
    """
    Function to calculate the biomass production and water productivity
    """ 
  
    Ksolar = Ra_mountain_24 * Transm_24
    
    # Incident Photosynthetically active radiation (PAR, MJ/m2) per time period
    PAR = 0.48 * Ksolar
    
    # Aborbed Photosynthetical Active Radiation (APAR) by the vegetation:
    APAR = FPAR * PAR

    vapor_stress = 0.88 - 0.183 * np.log(esat_24 - eact_24)
    vapor_stress_biomass = vapor_stress.clip(0.0, 1.0)
    Jarvis_coeff = (Th - Kt) / (Kt - Tl)
    heat_stress_biomass = ((Temp_24 - Tl) * np.power(Th - Temp_24, Jarvis_coeff) /
                           ((Kt - Tl) * np.power(Th - Kt, Jarvis_coeff)))
    print 'vapor stress biomass =', '%0.3f' % np.nanmean(vapor_stress_biomass)
    print 'heat stress biomass =', '%0.3f' % np.nanmean(heat_stress_biomass)
				
    # Light use efficiency, reduced below its potential value by low
    # temperature or water shortage:
    LUE = (LUEmax * heat_stress_biomass * vapor_stress_biomass * moisture_stress_biomass)
        
    # Dry matter production (kg/ha/d):
    Biomass_prod = APAR * LUE * 0.864             # C3 vegetation

    # Water productivity
    Biomass_wp = Biomass_prod/ (ETA_24 * 10)  # C3 vegetation
    Biomass_wp[ETA_24 == 0.0] = 0.0
    
    # Water deficit
    Biomass_deficit = (Biomass_prod / moisture_stress_biomass -
                          Biomass_prod)
                          
                          
    return(LUE,Biomass_prod,Biomass_wp,Biomass_deficit)

#------------------------------------------------------------------------------    
def Classify_Irrigation(moisture_stress_biomass, vegt_cover):
   '''
   This function makes a classification with 4 categories which show the irrigation needs
   '''
   for_irrigation = np.copy(moisture_stress_biomass)
   
   # make a discreed irrigation needs map with the following categories  
      # Irrigation needs:
       # 0: No need for irrigation
       # 1: Perhaps irrigate
       # 2: Irrigate
       # 3: Irrigate immediately
   irrigation_needs = np.copy(for_irrigation)
   irrigation_needs[np.where(irrigation_needs >= 1.0)] == 0.0
   irrigation_needs[np.logical_and(irrigation_needs >= 0.9, irrigation_needs < 1.0)] = 1.0
   irrigation_needs[np.where((irrigation_needs >= 0.8) & (irrigation_needs < 0.9))] = 2.0
   irrigation_needs[np.where(irrigation_needs < 0.8)] = 3.0   
   irrigation_needs[vegt_cover <= 0.3] = 0.0                                 
   return(irrigation_needs)   

#------------------------------------------------------------------------------   
def Separate_E_T(Light_use_extinction_factor,LAI,ETP_24,Theta_res_top,Theta_res_sub, Theta_sat_top, Theta_sat_sub, top_soil_moisture,sl_es_24, Psychro_c,moisture_stress_biomass_first,vegt_cover,ETA_24,SM_stress_trigger,root_zone_moisture_first,total_soil_moisture):
   '''
   Separate the Evapotranspiration into evaporation and Transpiration
   '''
   # constants
      
   Tpot_24_estimate=(1-np.exp(-Light_use_extinction_factor*LAI))*ETP_24
   SE_top = (top_soil_moisture-Theta_res_top)/(Theta_sat_top-Theta_res_top)
   Eact_24_estimate=np.minimum(1,1 / np.power(SE_top + 0.1,-2.0))*(ETP_24-Tpot_24_estimate)
   #RS_soil = RS_soil_min * np.power(SE_top,-2.0)
   #Eact_24_estimate=(sl_es_24+Psychro_c*(1+RS_soil_min/Rah_PM))/(sl_es_24+Psychro_c*(1+RS_soil/Rah_PM))*(ETP_24-Tpot_24_estimate)
   n66_memory = moisture_stress_biomass_first * Tpot_24_estimate
   
   # calulate the first estimation of actual daily tranpiration
   Tact_24_estimate = np.copy(n66_memory)
   Tact_24_estimate[n66_memory > 0.99*ETA_24]=ETA_24[n66_memory > 0.99*ETA_24]
   Tact_24_estimate[vegt_cover == 0.0] = 0.0
   
   # calculate the second estimation and end estimation of the actual daily tranpiration
   Tact_24 = np.abs((Tact_24_estimate/(Tact_24_estimate + Eact_24_estimate))*ETA_24)
  
   # calculate the actual daily potential transpiration
   Tpot_24 = np.copy(Tpot_24_estimate)
   Tpot_24[Tpot_24_estimate < Tact_24] = Tact_24[Tpot_24_estimate < Tact_24]
   
   # calculate moisture stress biomass
   moisture_stress_biomass = Tact_24 / Tpot_24
   
   # Calculate root zone moisture final
   Se_Poly=2.23*np.power(moisture_stress_biomass,3)-3.35*np.power(moisture_stress_biomass,2)+1.98*moisture_stress_biomass+0.07
   root_zone_moisture1=Se_Poly*(SM_stress_trigger+0.02-Theta_res_sub)+Theta_res_sub 
   root_zone_moisture_final=np.where(root_zone_moisture1>root_zone_moisture_first,root_zone_moisture1,root_zone_moisture_first)
   
   # Calculate top zone moisture final
   top_zone_moisture1=(total_soil_moisture-root_zone_moisture_final*vegt_cover)/(1-vegt_cover)
   top_zone_moisture_final=top_zone_moisture1.clip(Theta_res_top,Theta_sat_top)
    
   # calculate the actual daily evaporation
   Eact_24 = ETA_24 - Tact_24
   
   # calculate the Transpiration deficit
   T24_deficit = Tpot_24 - Tact_24
   
   # calculate the beneficial fraction
   beneficial_fraction=Tact_24 / ETA_24
   beneficial_fraction[ETA_24 == 0.0] = 0.0
   
   return(Eact_24,Tpot_24,Tact_24,moisture_stress_biomass,T24_deficit,beneficial_fraction,root_zone_moisture_final,top_zone_moisture_final)
  
#------------------------------------------------------------------------------    
def Calc_Soil_Moisture(ETA_24,EF_inst,QC_Map, water_mask,vegt_cover,Theta_sat_top, Theta_sat_sub,Theta_res_top, Theta_res_sub,depl_factor,Field_Capacity,FPAR, Soil_moisture_wilting_point):
    """
    Function to calculate soil characteristics
    """
    # constants:
    Veg_Cover_Threshold_RZ = 0.9 # Threshold vegetation cover for root zone moisture
       
    # Average fraction of TAW that can be depleted from the root zone
    # before stress:
    p_factor = depl_factor + 0.04 * (5.0 - ETA_24)  # page 163 of FAO 56
    # The factor p differs from one crop to another. It normally varies from
    # 0.30 for shallow rooted plants at high rates of ETc (> 8 mm d-1)
    # to 0.70 for deep rooted plants at low rates of ETc (< 3 mm d-1)

    # Critical value under which plants get stressed:
    SM_stress_trigger = Field_Capacity - p_factor * (Field_Capacity - Soil_moisture_wilting_point) 
    EF_inst[EF_inst >= 1.0] = 0.999    

    # Total soil water content (cm3/cm3):
    total_soil_moisture = Theta_sat_sub * np.exp((EF_inst - 1.0) / 0.421)   # asce paper Scott et al. 2003
    total_soil_moisture[np.logical_or(water_mask == 1.0,QC_Map == 1.0)] = 1.0  # In water and snow is 1                               
    total_soil_moisture[QC_Map == 1.0] = np.nan  # Where clouds no data                              
    
    # Root zone soil moisture:
    RZ_SM = np.copy(total_soil_moisture)
    RZ_SM[vegt_cover <= Veg_Cover_Threshold_RZ] = np.nan
    if np.isnan(np.nanmean(RZ_SM)) == True:
        Veg_Cover_Threshold_RZ = np.nanpercentile(vegt_cover, 80)   
        RZ_SM = np.copy(total_soil_moisture)
        RZ_SM[vegt_cover <= Veg_Cover_Threshold_RZ] = np.nan
        print 'No RZ_SM so the vegetation Threshold for RZ is adjusted from 0,9 to =', '%0.3f' % Veg_Cover_Threshold_RZ
        
    #RZ_SM = RZ_SM.clip(Theta_res, (0.85 * Theta_sat))
    #RZ_SM[np.logical_or(water_mask == 1.0, water_mask == 2.0)] = 1.0                            
    RZ_SM_NAN = np.copy(RZ_SM)
    RZ_SM_NAN[RZ_SM==0] = np.nan
    RZ_SM_min = np.nanmin(RZ_SM_NAN)
    RZ_SM_max = np.nanmax(RZ_SM_NAN)
    RZ_SM_mean = np.nanmean(RZ_SM_NAN)
    print 'Root Zone Soil moisture mean =', '%0.3f (cm3/cm3)' % RZ_SM_mean
    print 'Root Zone Soil moisture min =', '%0.3f (cm3/cm3)' % RZ_SM_min
    print 'Root Zone Soil moisture max =', '%0.3f (cm3/cm3)' % RZ_SM_max
    
    Max_moisture_RZ = vegt_cover * (RZ_SM_max - RZ_SM_mean) + RZ_SM_mean
    
    # Soil moisture in the top (temporary)
    top_soil_moisture_temp = np.copy(total_soil_moisture)
    top_soil_moisture_temp[np.logical_or(vegt_cover <= 0.02, vegt_cover >= 0.1)] = 0
    top_soil_moisture_temp[top_soil_moisture_temp == 0] = np.nan 
    top_soil_moisture_std = np.nanstd(top_soil_moisture_temp)
    top_soil_moisture_mean = np.nanmean(top_soil_moisture_temp)
    print 'Top Soil moisture mean =', '%0.3f (cm3/cm3)' % top_soil_moisture_mean
    print 'Top Soil moisture Standard Deviation', '%0.3f (cm3/cm3)' % top_soil_moisture_std
    
    # calculate root zone moisture
    root_zone_moisture_temp = (total_soil_moisture - (top_soil_moisture_mean + top_soil_moisture_std) * (1-vegt_cover))/vegt_cover # total soil moisture = soil moisture no vegtatation *(1-vegt_cover)+soil moisture root zone * vegt_cover
    try:
        root_zone_moisture_temp[root_zone_moisture_temp <= Theta_res_sub] = Theta_res_sub[root_zone_moisture_temp <= Theta_res_sub]
    except:
        root_zone_moisture_temp[root_zone_moisture_temp <= Theta_res_sub] = Theta_res_sub
 	
		
    root_zone_moisture_temp[root_zone_moisture_temp >= Max_moisture_RZ] = Max_moisture_RZ[root_zone_moisture_temp >= Max_moisture_RZ]
    
    root_zone_moisture_first = np.copy(root_zone_moisture_temp)
    root_zone_moisture_first[np.logical_or(QC_Map ==1.0 ,np.logical_or(water_mask == 1.0, vegt_cover < 0.0))] = 0

    # Normalized stress trigger:
    norm_trigger = (root_zone_moisture_first - Soil_moisture_wilting_point)/ (SM_stress_trigger + 0.02 - Soil_moisture_wilting_point)
    norm_trigger[norm_trigger > 1.0] = 1.0
    
    # moisture stress biomass:
    moisture_stress_biomass_first = norm_trigger - (np.sin(2 * np.pi * norm_trigger)) / (2 * np.pi)
    moisture_stress_biomass_first=np.where(moisture_stress_biomass_first<0.5*FPAR,0.5*FPAR,moisture_stress_biomass_first)
    moisture_stress_biomass_first[moisture_stress_biomass_first <= 0.0] = 0
    moisture_stress_biomass_first[moisture_stress_biomass_first > 1.0] = 1.0
    
     # Soil moisture in the top layer - Recalculated ??
    top_soil_moisture = ((total_soil_moisture - root_zone_moisture_first * vegt_cover) / (1.0 - vegt_cover))
    
    try:
        top_soil_moisture[top_soil_moisture > Theta_sat_top] = Theta_sat_top [top_soil_moisture > Theta_sat_top]
    except:
        top_soil_moisture[top_soil_moisture > Theta_sat_top] = Theta_sat_top
 	
    top_soil_moisture[np.logical_or(water_mask == 1.0, QC_Map == 1.0)] = 1.0

    return(SM_stress_trigger, total_soil_moisture, root_zone_moisture_first, moisture_stress_biomass_first,top_soil_moisture,RZ_SM_NAN)

#------------------------------------------------------------------------------
def Calc_Bulk_surface_resistance(sl_es_24,Rn_24,Refl_rad_water,air_dens,esat_24,eact_24,rah_pm_act,ETA_24,Lhv,Psychro_c):
    """
    Function to calculate the bulk surface resistance
    """
    # Bulk surface resistance (s/m):
    bulk_surf_resis_24 = ((((sl_es_24 * (Rn_24 - Refl_rad_water) + air_dens *
                          1004 * (esat_24 - eact_24) / rah_pm_act) / (ETA_24 * Lhv / 86400) -
                          sl_es_24) / Psychro_c - 1.0) * rah_pm_act)
    bulk_surf_resis_24[ETA_24 <= 0.0] = 100000.0
    bulk_surf_resis_24 = bulk_surf_resis_24.clip(0.0, 100000.0)
    return(bulk_surf_resis_24)

#------------------------------------------------------------------------------
def Calc_ETact(esat_24, eact_24, EF_inst, Rn_24, Refl_rad_water, Lhv, Image_Type):
    """
    Function to calculate the daily evaporation
    """
    # Advection factor
    if Image_Type == 2:
        AF = np.ones(Rn_24.shape)
    else:
        AF = 1 + 0.985 * (np.exp((esat_24 - eact_24) * 0.08) - 1.0) * EF_inst
    # Daily evapotranspiration:
    ETA_24 = EF_inst * AF * (Rn_24 - Refl_rad_water) / (Lhv * 1000) * 86400000
    ETA_24=ETA_24.clip(0,15.0)
    return(ETA_24, AF)
    
#------------------------------------------------------------------------------   
def Calc_instantaneous_ET_fraction(LE_inst,rn_inst,g_inst):   
    """
    Function to calculate the evaporative fraction
    """ 
    EF_inst = LE_inst / (rn_inst - g_inst)    # Evaporative fraction
    EF_inst = EF_inst.clip(0.0, 1.8)
    EF_inst[LE_inst<0] = 0

    return(EF_inst)

#------------------------------------------------------------------------------
def Calc_Ref_Pot_ET(LAI,Surface_temp,sl_es_24,Rn_ref,air_dens,esat_24,eact_24,rah_grass,Psychro_c,Rn_24,Refl_rad_water,rah_pm_pot,rl):
    """
    Function to calculate the reference potential evapotransporation and potential evaporation
    """ 

    # Effective leaf area index involved, see Allen et al. (2006):
    LAI_eff = LAI / (0.3 * LAI + 1.2)
    rs_min = rl / LAI_eff  # Min (Bulk) surface resistance (s/m)
    # Latent heat of vaporization (J/kg):
    Lhv = (2.501 - 2.361e-3 * (Surface_temp - 273.15)) * 1E6

    # Reference evapotranspiration- grass
    # Penman-Monteith of the combination equation (eq 3 FAO 56) (J/s/m2)
    LET_ref_24 = ((sl_es_24 * Rn_ref + air_dens * 1004 * (esat_24 - eact_24) /
                  rah_grass) / (sl_es_24 + Psychro_c * (1 + 70.0/rah_grass)))
    # Reference evaportranspiration (mm/d):
    ETref_24 = LET_ref_24 / (Lhv * 1000) * 86400000

    # Potential evapotranspiration
    # Penman-Monteith of the combination equation (eq 3 FAO 56) (J/s/m2)
    LETpot_24 = ((sl_es_24 * (Rn_24 - Refl_rad_water) + air_dens * 1004 *
               (esat_24 - eact_24)/rah_pm_pot) / (sl_es_24 + Psychro_c * (1 + rs_min/rah_pm_pot)))
    # Potential evaportranspiration (mm/d)
    ETpot_24 = LETpot_24 / (Lhv * 1000) * 86400000
    ETpot_24[ETpot_24 > 15.0] = 15.0
    return(ETpot_24,ETref_24,Lhv,rs_min)
     
#------------------------------------------------------------------------------    
def Calc_Rn_Ref(shape_lsc,water_mask,Rn_24,Ra_mountain_24,Transm_24,Rnl_24_FAO,Wind_24):
    """
    Function to calculate the net solar radiation
    """  
    # constants:
    G24_water = 0.1  # G24 ratio for water - reflectivity?

    # Reflected radiation at water surface: ??
    Refl_rad_water = np.zeros((shape_lsc[1], shape_lsc[0]))
    Refl_rad_water = np.where(water_mask != 0.0, G24_water * Rn_24, 0.0)

    # Aerodynamic resistance (s/m) for grass surface:
    rah_grass = 208.0 / Wind_24
    print  'rah_grass=', '%0.3f (s/m)' % np.nanmean(rah_grass)
    # Net radiation for grass Rn_ref, eq 40, FAO56:
    Rn_ref = Ra_mountain_24 * Transm_24 * (1 - 0.23) - Rnl_24_FAO  # Rnl avg(fao-slob)?
    return(Rn_ref, Refl_rad_water,rah_grass)
    

#------------------------------------------------------------------------------
def Iterate_Friction_Velocity(k_vk,u_200,Surf_roughness,g_inst,rn_inst, ts_dem, ts_dem_hot, ts_dem_cold,air_dens, Surface_temp,L,psi,psi_m200,psi_m200_stable,QC_Map, hot_pixels, slope):
    """
    Function to correct the windspeed and aerodynamic resistance for the iterative process the output can be used as the new input for this model
    """    
    # Sensible heat 2 (Step 6)
    # Corrected value for the friction velocity, unstable
    ustar_corr_unstable = (k_vk * u_200 / (np.log(200.0 / Surf_roughness) -
                        psi_m200))
    # Corrected value for the friction velocity, stable
    ustar_corr_stable = (k_vk * u_200 / (np.log(200.0 / Surf_roughness) -
                      psi_m200_stable))
    ustar_corr = np.where(L > 0.0, ustar_corr_stable, ustar_corr_unstable)
    ustar_corr[ustar_corr < 0.02] = 0.02

    rah_corr_unstable = (np.log(2.0/0.01) - psi) / (k_vk * ustar_corr)     # unstable
    rah_corr_stable = (np.log(2.0/0.01) - 0.0) / (k_vk * ustar_corr)       # stable
    rah_corr = np.where(L > 0.0, rah_corr_stable, rah_corr_unstable)
    L_corr, psi_m200_corr_stable, psi_corr, psi_m200_corr,h,dT, slope_dt, offset_dt = sensible_heat(
            rah_corr, ustar_corr, rn_inst, g_inst, ts_dem, ts_dem_hot, ts_dem_cold,
            air_dens, Surface_temp, k_vk,QC_Map, hot_pixels, slope)
    return(L_corr,psi_corr,psi_m200_corr,psi_m200_corr_stable,h,ustar_corr,rah_corr,dT,slope_dt, offset_dt)

#------------------------------------------------------------------------------
def Calc_Wind_Speed_Friction(h_obst,Wind_inst,zx,LAI,NDVI,Surf_albedo,water_mask,surf_roughness_equation_used):    
    """
    Function to calculate the windspeed and friction by using the Raupach or NDVI model
    """    
    
    # constants
    k_vk = 0.41      # Von Karman constant
    h_grass = 0.12   # Grass height (m)
    cd = 53          # Free parameter for displacement height, default = 20.6
    # 1) Raupach model
    zom_Raupach=Raupach_Model(h_obst,cd,LAI)

    # 2) NDVI model
    zom_NDVI=NDVI_Model(NDVI,Surf_albedo,water_mask)

    if surf_roughness_equation_used == 1:
        Surf_roughness = zom_NDVI
    else:
        Surf_roughness = zom_Raupach

    zom_grass = 0.123 * h_grass
    # Friction velocity for grass (m/s):
    ustar_grass = k_vk * Wind_inst / np.log(zx / zom_grass)
    print 'u*_grass = ', '%0.3f (m/s)' % np.mean(ustar_grass)
    # Wind speed (m/s) at the "blending height" (200m):
    u_200 = ustar_grass * np.log(200 / zom_grass) / k_vk
    print 'Wind speed at the blending height, u200 =', '%0.3f (m/s)' % np.mean(u_200)
    # Friction velocity (m/s):
    ustar_1 = k_vk * u_200 / np.log(200 / Surf_roughness)
    return(Surf_roughness,u_200,ustar_1)
    
#------------------------------------------------------------------------------
def Raupach_Model(h_obst,cd,LAI):    
    """
    Function for the Raupach model to calculate the surface roughness (based on Raupach 1994)  
    """   
    # constants
    cw = 2.0
    LAIshelter = 2.5
    
    # calculate psi
    psi = np.log(cw) - 1 + np.power(2.0, -1)  # Vegetation influence function
    
    # Calculate Ustar divided by U
    ustar_u = np.power((0.003+0.3*LAI/2), 0.5)
    ustar_u[LAI<LAIshelter] = 0.3
           
    # calculate: 1 - d/hv      
    inv_d_hv =np.power((1-np.exp(np.power(-1*(cd*LAI),0.5)))/(cd * LAI),0.5)
              
    # Calculate: surface roughness/hv           
    zom_hv = inv_d_hv * np.exp(-0.41/ustar_u-psi)
    
    # Calculate: surface roughness
    zom_Raupach = zom_hv * h_obst 
    
    return(zom_Raupach)
 
#------------------------------------------------------------------------------
def NDVI_Model(NDVI,Surf_albedo,water_mask):
    """
    Function for the NDVI model to calculate the surface roughness
    """    
    zom_NDVI = np.exp(1.096 * NDVI / Surf_albedo - 5.307)
    zom_NDVI[water_mask == 1.0] = 0.001
    zom_NDVI[zom_NDVI > 10.0] = 10.0
    return(zom_NDVI)   
       
#------------------------------------------------------------------------------
def Correct_Surface_Temp(Surface_temp,Temp_lapse_rate,DEM_resh,Pair,dr,Transm_corr,cos_zn,Sun_elevation,deg2rad,ClipLandsat):
    """
    Function to correct the surface temperature based on the DEM map
    """     
    #constants:
    Gsc = 1367        # Solar constant (W / m2)   
    
    cos_zenith_flat = np.cos((90 - Sun_elevation) * deg2rad)
    Temp_corr = Surface_temp + Temp_lapse_rate * DEM_resh  # rescale everything to sea level
    Temp_corr[Surface_temp == 350.0] = 0.0
    air_dens = 1000 * Pair / (1.01 * Surface_temp * 287)
    #
    ts_dem = (Temp_corr + (Gsc * dr * Transm_corr * cos_zn -
              Gsc * dr * Transm_corr * cos_zenith_flat) / (air_dens * 1004 * 0.050))
    #(Temp_corr - (Gsc * dr * Transm_corr * cos_zn -
    #          Gsc * dr * Transm_corr * cos_zenith_flat) / (air_dens * 1004 * 0.050))
    ts_dem[ClipLandsat==1]=np.nan
    ts_dem[ts_dem==0]=np.nan
    ts_dem[ts_dem<273]=np.nan   
    ts_dem[ts_dem>350]=np.nan                
  
    return(ts_dem,air_dens,Temp_corr)
 
#------------------------------------------------------------------------------
def Calc_Hot_Pixels(ts_dem,QC_Map, water_mask, NDVI,NDVIhot_low,NDVIhot_high,Hot_Pixel_Constant, ts_dem_cold):
    """
    Function to calculates the hot pixels based on the surface temperature and NDVI
    """ 
    for_hot = np.copy(ts_dem)
    for_hot[NDVI <= NDVIhot_low] = 0.0
    for_hot[NDVI >= NDVIhot_high] = 0.0
    for_hot[np.logical_or(water_mask != 0.0, QC_Map != 0.0)] = 0.0
    hot_pixels = np.copy(for_hot)
    hot_pixels[for_hot < ts_dem_cold] = np.nan
    ts_dem_hot_max = np.nanmax(hot_pixels)    # Max
    ts_dem_hot_mean = np.nanmean(hot_pixels)  # Mean
    ts_dem_hot_std = np.nanstd(hot_pixels)    # Standard deviation
    #ts_dem_hot = ts_dem_hot_max - 0.25 * ts_dem_hot_std
    #ts_dem_hot = (ts_dem_hot_max + ts_dem_hot_mean)/2
    
   
    ts_dem_hot=ts_dem_hot_mean + Hot_Pixel_Constant * ts_dem_hot_std

    
    print 'hot : max= %0.3f (Kelvin)' % ts_dem_hot_max, ', sd= %0.3f (Kelvin)' % ts_dem_hot_std, \
           ', mean= %0.3f (Kelvin)' % ts_dem_hot_mean, ', value= %0.3f (Kelvin)' % ts_dem_hot
    return(ts_dem_hot,hot_pixels)    
    
#------------------------------------------------------------------------------    
def Calc_Cold_Pixels(ts_dem,water_mask,QC_Map,ts_dem_cold_veg,Cold_Pixel_Constant):
    """
    Function to calculates the the cold pixels based on the surface temperature
    """ 
    for_cold = np.copy(ts_dem)
    for_cold[water_mask != 1.0] = 0.0
    for_cold[QC_Map != 0] = 0.0
    cold_pixels = np.copy(for_cold)
    cold_pixels[for_cold < 278.0] = np.nan
    cold_pixels[for_cold > 320.0] = np.nan
    # cold_pixels[for_cold < 285.0] = 285.0
    ts_dem_cold_std = np.nanstd(cold_pixels)     # Standard deviation
    ts_dem_cold_min = np.nanmin(cold_pixels)     # Min
    ts_dem_cold_mean = np.nanmean(cold_pixels)   # Mean

    # If average temperature is below zero or nan than use the vegetation cold pixel
    if ts_dem_cold_mean <= 0.0:
        ts_dem_cold = ts_dem_cold_veg + Cold_Pixel_Constant * ts_dem_cold_std
    if np.isnan(ts_dem_cold_mean) == True:
        ts_dem_cold = ts_dem_cold_veg + Cold_Pixel_Constant * ts_dem_cold_std 							
    else:
        ts_dem_cold = ts_dem_cold_mean + Cold_Pixel_Constant * ts_dem_cold_std

    if ts_dem_cold > ts_dem_cold_veg:
        ts_dem_cold = ts_dem_cold_veg 
            
    print 'cold water: min=%0.3f (Kelvin)' %ts_dem_cold_min , ', sd= %0.3f (Kelvin)' % ts_dem_cold_std, \
           ', mean= %0.3f (Kelvin)' % ts_dem_cold_mean, ', value= %0.3f (Kelvin)' % ts_dem_cold 
    return(ts_dem_cold,cold_pixels,ts_dem_cold_mean)
     
#------------------------------------------------------------------------------
def Calc_Cold_Pixels_Veg(NDVI,NDVI_max,NDVI_std,QC_Map,ts_dem,Image_Type, Cold_Pixel_Constant):
    """
    Function to calculates the the cold pixels based on vegetation
    """   
    cold_pixels_vegetation = np.copy(ts_dem)
    cold_pixels_vegetation[np.logical_and(NDVI <= (NDVI_max-0.1*NDVI_std),QC_Map != 0.0)] = 0.0
    cold_pixels_vegetation[cold_pixels_vegetation==0.0] = np.nan
    ts_dem_cold_std_veg = np.nanstd(cold_pixels_vegetation)
    ts_dem_cold_min_veg = np.nanmin(cold_pixels_vegetation)
    ts_dem_cold_mean_veg = np.nanmean(cold_pixels_vegetation)
    
    if Image_Type == 1:    
            ts_dem_cold_veg = ts_dem_cold_mean_veg + Cold_Pixel_Constant * ts_dem_cold_std_veg
    if Image_Type == 2:    
            ts_dem_cold_veg = ts_dem_cold_mean_veg + Cold_Pixel_Constant * ts_dem_cold_std_veg
    if Image_Type == 3:    
            ts_dem_cold_veg = ts_dem_cold_mean_veg + Cold_Pixel_Constant * ts_dem_cold_std_veg    
            
    print 'cold vegetation: min=%0.3f (Kelvin)' %ts_dem_cold_min_veg , ', sd= %0.3f (Kelvin)' % ts_dem_cold_std_veg, \
				', mean= %0.3f (Kelvin)' % ts_dem_cold_mean_veg, ', value= %0.3f (Kelvin)' % ts_dem_cold_veg 
    return(ts_dem_cold_veg)
   
#------------------------------------------------------------------------------   
def Calc_Meteo(Rs_24,eact_24,Temp_24,Surf_albedo,dr,tir_emis,Surface_temp,water_mask,NDVI,Transm_24,SB_const,lw_in_inst,Rs_in_inst):
    """       
    Calculates the instantaneous Ground heat flux and solar radiation.
    """ 
    
    # Net shortwave radiation (W/m2):
    Rns_24 = Rs_24 * (1 - Surf_albedo)


    # Net outgoing longwave radiation (W/m2):
    Rnl_24_FAO = (SB_const * np.power(Temp_24 + 273.15, 4) * (0.34-0.14 *
                  np.power(eact_24, 0.5)) * (1.35 * Transm_24 / 0.8 - 0.35))
                  
    Rnl_24_Slob = 110 * Transm_24            
    print 'Mean Daily Net Radiation (Slob) = %0.3f (W/m2)' % np.nanmean(Rnl_24_Slob)    
				
    # Net 24 hrs radiation (W/m2):
    Rn_24_FAO = Rns_24 - Rnl_24_FAO          # FAO equation
    Rn_24_Slob = Rns_24 - Rnl_24_Slob       # Slob equation
    Rn_24 = (Rn_24_FAO + Rn_24_Slob) / 2  # Average
   

    # Instantaneous outgoing longwave radiation:
    lw_out_inst = tir_emis * SB_const * np.power(Surface_temp, 4)
    
    # Instantaneous net radiation
    rn_inst = (Rs_in_inst * (1 - Surf_albedo) + lw_in_inst - lw_out_inst -
               (1 - tir_emis) * lw_in_inst)
    # Instantaneous Soil heat flux
    g_inst = np.where(water_mask != 0.0, 0.4 * rn_inst,
                      ((Surface_temp - 273.15) * (0.0038 + 0.0074 * Surf_albedo) *
                       (1 - 0.978 * np.power(NDVI, 4))) * rn_inst)
    return(Rn_24,rn_inst,g_inst,Rnl_24_FAO)
    
#------------------------------------------------------------------------------    
def Calc_surface_water_temp(Temp_inst,Landsat_nr,Lmax,Lmin,therm_data,b10_emissivity,k1_c,k2_c,eact,shape_lsc,water_mask_temp,Bands_thermal,Rp,tau_sky,surf_temp_offset,Image_Type):    
    """
    Calculates the surface temperature and create a water mask
    """ 
    
    # Spectral radiance for termal
    if Landsat_nr == 8:
        if Bands_thermal == 1:
            k1 = k1_c[0]
            k2 = k2_c[0]
            L_lambda_b10 = (Lmax[-1] - Lmin[-1]) / (65535-1) * therm_data[:, :, 0] + Lmin[-1]
            
            # Get Temperature
            Temp_TOA = Get_Thermal(L_lambda_b10,Rp,Temp_inst,tau_sky,b10_emissivity,k1,k2) 
                      
        elif Bands_thermal == 2:
            L_lambda_b10 = (Lmax[-2] - Lmin[-2]) / (65535-1) * therm_data[:, :, 0] + Lmin[-2]
            L_lambda_b11 = (Lmax[-1] - Lmin[-1]) / (65535-1) * therm_data[:, :, 1] + Lmin[-1]
    
            # Brightness temperature
            # From Band 10:
            Temp_TOA_10 = (k2_c[0] / np.log(k1_c[0] / L_lambda_b10 + 1.0))
            # From Band 11:
            Temp_TOA_11 = (k2_c[1] / np.log(k1_c[1] / L_lambda_b11 + 1.0))
            # Combined:
            Temp_TOA = (Temp_TOA_10 + 1.378 * (Temp_TOA_10 - Temp_TOA_11) +
                           0.183 * np.power(Temp_TOA_10 - Temp_TOA_11, 2) - 0.268 +
                           (54.30 - 2.238 * eact) * (1 - b10_emissivity))
       
    elif Landsat_nr == 7:
        k1=666.09
        k2=1282.71
        L_lambda_b6 = (Lmax[-1] - Lmin[-1]) / (256-1) * therm_data[:, :, 0] + Lmin[-1]

        # Brightness temperature - From Band 6:
        Temp_TOA = Get_Thermal(L_lambda_b6,Rp,Temp_inst,tau_sky,b10_emissivity,k1,k2) 
        
    elif Landsat_nr == 5:
        k1=607.76
        k2=1260.56
        L_lambda_b6 = ((Lmax[-1] - Lmin[-1]) / (256-1) * therm_data[:, :, 0] +
                       Lmin[-1])
   
       # Brightness temperature - From Band 6:
        Temp_TOA = Get_Thermal(L_lambda_b6,Rp,Temp_inst,tau_sky,b10_emissivity,k1,k2) 
            
    # Surface temperature
    Surface_temp = Temp_TOA
    Surface_temp = Surface_temp.clip(230.0, 360.0)

    # Cloud mask:
    temp_water = np.zeros((shape_lsc[1], shape_lsc[0]))
    temp_water = np.copy(Surface_temp)
    temp_water[water_mask_temp == 0.0] = np.nan
    temp_water_sd = np.nanstd(temp_water)     # Standard deviation
    temp_water_mean = np.nanmean(temp_water)  # Mean
    print 'Mean water temperature = ', '%0.3f (Kelvin)' % temp_water_mean
    print 'SD water temperature = ', '%0.3f (Kelvin)' % temp_water_sd
    cloud_mask = np.zeros((shape_lsc[1], shape_lsc[0]))
    cloud_mask[Surface_temp < np.minimum((temp_water_mean - 1.0 * temp_water_sd -
               surf_temp_offset),290)] = 1.0
                          
    return(Surface_temp,cloud_mask)           


#------------------------------------------------------------------------------
def Get_Thermal(lambda_b10,Rp,Temp_inst,tau_sky,TIR_Emissivity,k1,k2):
    
    # Narrow band downward thermal radiation from clear sky, rsky (W/m2/sr/µm)
    rsky = (1.807E-10 * np.power(Temp_inst + 273.15, 4) * (1 - 0.26 *
            np.exp(-7.77E-4 * np.power((-Temp_inst), -2))))
    print 'Rsky = ', '%0.3f (W/m2/sr/µm)' % np.nanmean(rsky)
    
    # Corrected thermal radiance from the surface, Wukelikc et al. (1989):
    correc_lambda_b10 = ((lambda_b10 - Rp) / tau_sky -
                               (1.0 - TIR_Emissivity) * rsky)
    # Brightness temperature - From Band 10:
    Temp_TOA = (k2 / np.log(TIR_Emissivity * k1 /
                       correc_lambda_b10 + 1.0))  
                       
    return(Temp_TOA)                   
#------------------------------------------------------------------------------
def Calc_vegt_para(NDVI,water_mask_temp,shape_lsc):
    """
    Calculates the Fraction of PAR, Thermal infrared emissivity, Nitrogen, Vegetation Cover, LAI, b10_emissivity
    """
    # Fraction of PAR absorbed by the vegetation canopy (FPAR):
    FPAR = -0.161 + 1.257 * NDVI
    FPAR[NDVI < 0.125] = 0.0

    # Termal infrared emissivity
    tir_emis = 1.009 + 0.047 * np.log(NDVI)
    tir_emis[np.logical_or(water_mask_temp == 1.0, water_mask_temp == 2.0)] = 1.0
    tir_emis[np.logical_and(NDVI < 0.125, water_mask_temp == 0.0)] = 0.92

    # Vegetation Index - Regression model from Bagheri et al. (2013)
    VI = 38.764 * np.square(NDVI) - 24.605 * NDVI + 5.8103

    # Nitrogen computation
    Nitrogen = np.copy(VI)
    Nitrogen[VI <= 0.0] = 0.0
    Nitrogen[NDVI <= 0.0] = 0.0
                  
    # Vegetation cover:
    vegt_cover = 1 - np.power((0.8 - NDVI)/(0.8 - 0.125), 0.7)
    vegt_cover[NDVI < 0.125] = 0.0
    vegt_cover[NDVI > 0.8] = 0.99

    # Leaf Area Index (LAI)
    LAI_1 = np.log(-(vegt_cover - 1)) / -0.45
    LAI_1[LAI_1 > 8] = 8.0
    LAI_2 = (9.519 * np.power(NDVI, 3) + 0.104 * np.power(NDVI, 2) +
             1.236 * NDVI - 0.257)

    LAI = (LAI_1 + LAI_2) / 2.0  # Average LAI
    LAI[LAI < 0.001] = 0.001

    b10_emissivity = np.zeros((shape_lsc[1], shape_lsc[0]))
    b10_emissivity = np.where(LAI <= 3.0, 0.95 + 0.01 * LAI, 0.98)
    b10_emissivity[water_mask_temp != 0.0] = 1.0
                  
    return(FPAR,tir_emis,Nitrogen,vegt_cover,LAI,b10_emissivity)
    
#------------------------------------------------------------------------------    
def Water_Mask(shape_lsc,Reflect):
    """
    Calculates the water and cloud mask
    """
    mask = np.zeros((shape_lsc[1], shape_lsc[0]))
    mask[np.logical_and(Reflect[:, :, 3] < Reflect[:, :, 2],
                        Reflect[:, :, 4] < Reflect[:, :, 1])] = 1.0
    water_mask_temp = np.copy(mask)
    
    return(water_mask_temp)

#------------------------------------------------------------------------------   
def Calc_albedo(Reflect,path_radiance,Apparent_atmosf_transm):
    """
    This function calculates and returns the Surface albedo, NDVI by using the refectance from the landsat image.
    """ 
    # Surface albedo:
    Surf_albedo = (0.254 * Reflect[:, :, 0] + 0.149 * Reflect[:, :, 1] +
                   0.147 * Reflect[:, :, 2] + 0.311 * Reflect[:, :, 3] +
                   0.103 * Reflect[:, :, 4] + 0.036 * Reflect[:, :, 5] -
                   path_radiance) / np.power(Apparent_atmosf_transm, 2)

    # Better tsw instead of Apparent_atmosf_transm ??
    Surf_albedo = Surf_albedo.clip(0.0, 0.6)

    return(Surf_albedo)
#------------------------------------------------------------------------------  
def Calc_NDVI(Reflect):
    """
    This function calculates and returns the Surface albedo, NDVI by using the refectance from the landsat image.
    """ 
    # Computation of Normalized Difference Vegetation Index (NDVI)
    NDVI = ((Reflect[:, :, 3] - Reflect[:, :, 2]) /
            (Reflect[:, :, 3] + Reflect[:, :, 2]))

    return(NDVI)
							
#------------------------------------------------------------------------------  
def CalculateSnowWaterMask(NDVI,shape_lsc,water_mask_temp,Surface_temp):
   '''
   Devides the temporaly water mask into a snow and water mask by using the surface temperature
   '''
   NDVI_nan=np.copy(NDVI)
   NDVI_nan[NDVI==0]=np.nan
   NDVI_nan=np.float32(NDVI_nan)
   NDVI_std=np.nanstd(NDVI_nan)
   NDVI_max=np.nanmax(NDVI_nan)
   NDVI_treshold_cold_pixels=NDVI_max-0.1*NDVI_std
   print 'NDVI treshold for cold pixels = ', '%0.3f' % NDVI_treshold_cold_pixels
   ts_moist_veg_min=np.nanmin(Surface_temp[NDVI>NDVI_treshold_cold_pixels])
   
   # calculate new water mask 
   mask=np.zeros((shape_lsc[1], shape_lsc[0]))
   mask[np.logical_and(np.logical_and(water_mask_temp==1, Surface_temp <= 275),NDVI>=0.3)]=1   
   snow_mask=np.copy(mask)
   
   # calculate new water mask 
   mask=np.zeros((shape_lsc[1], shape_lsc[0]))
   mask[np.logical_and(water_mask_temp==1, Surface_temp > 273)]=1   
   water_mask=np.copy(mask)
   
   return(snow_mask,water_mask,ts_moist_veg_min, NDVI_max, NDVI_std)
    
#------------------------------------------------------------------------------    
def Landsat_Reflect(Bands,input_folder,Name_Landsat_Image,output_folder,shape_lsc,ClipLandsat,Lmax,Lmin,ESUN_L5,ESUN_L7,ESUN_L8,cos_zn,dr,Landsat_nr, proyDEM_fileName):
    """
    This function calculates and returns the reflectance and spectral radiation from the landsat image.
    """ 
    
    Spec_Rad = np.zeros((shape_lsc[1], shape_lsc[0], 7))
    Reflect = np.zeros((shape_lsc[1], shape_lsc[0], 7))
    for band in Bands[:-(len(Bands)-6)]:
        # Open original Landsat image for the band number
        src_FileName = os.path.join(input_folder, '%s_B%1d.TIF'
                                    % (Name_Landsat_Image, band))

        ls_data=Open_landsat(src_FileName, proyDEM_fileName)
        ls_data = ls_data*ClipLandsat
        # stats = band_data.GetStatistics(0, 1)

        index = np.where(Bands[:-(len(Bands)-6)] == band)[0][0]
        if Landsat_nr == 8:
            # Spectral radiance for each band:
            L_lambda = Landsat_L_lambda(Lmin, Lmax, ls_data, index, Landsat_nr)
            # Reflectivity for each band:
            rho_lambda = Landsat_rho_lambda(L_lambda, ESUN_L8, index, cos_zn, dr)
        elif Landsat_nr == 7:
            # Spectral radiance for each band:
            L_lambda=Landsat_L_lambda(Lmin, Lmax, ls_data, index, Landsat_nr)
            # Reflectivity for each band:
            rho_lambda = Landsat_rho_lambda(L_lambda, ESUN_L7, index, cos_zn, dr)
        elif Landsat_nr == 5:
            # Spectral radiance for each band:
            L_lambda=Landsat_L_lambda(Lmin, Lmax, ls_data, index, Landsat_nr)
            # Reflectivity for each band:
            rho_lambda =Landsat_rho_lambda(L_lambda, ESUN_L5, index, cos_zn, dr)
        else:
            print 'Landsat image not supported, use Landsat 5, 7 or 8'

        Spec_Rad[:, :, index] = L_lambda
        Reflect[:, :, index] = rho_lambda
    Reflect = Reflect.clip(0.0, 1.0)
    return(Reflect,Spec_Rad)
     

#------------------------------------------------------------------------------    
def Landsat_L_lambda(Lmin,Lmax,ls_data,index,Landsat_nr):
    """
    Calculates the lambda from landsat
    """
    if Landsat_nr==8:
        L_lambda = ((Lmax[index] - Lmin[index]) / (65535 - 1) * ls_data + Lmin[index]) 
    elif Landsat_nr == 5 or Landsat_nr ==7:
        L_lambda = (Lmax[index] - Lmin[index]) / 255 * ls_data + Lmin[index]
    return(L_lambda)
    
    
#------------------------------------------------------------------------------   
def Landsat_rho_lambda(L_lambda,ESUN,index,cos_zn,dr):
    """
    Calculates the rho from landsat
    """
    rho_lambda = np.pi * L_lambda / (ESUN[index] * cos_zn * dr)
    return(rho_lambda)
    
    
#------------------------------------------------------------------------------
def Landsat_therm_data(Bands,input_folder,Name_Landsat_Image,output_folder, shape_lsc,ClipLandsat, proyDEM_fileName):          
    """
    This function calculates and returns the thermal data from the landsat image.
    """                             
    therm_data = np.zeros((shape_lsc[1], shape_lsc[0], len(Bands)-6))
    for band in Bands[-(len(Bands)-6):]:
        # Open original Landsat image for the band number
        src_FileName = os.path.join(input_folder, '%s_B%1d.TIF'
                                    % (Name_Landsat_Image, band))
        if not os.path.exists(src_FileName):
             src_FileName = os.path.join(input_folder, '%s_B%1d_VCID_2.TIF'
                                    % (Name_Landsat_Image, band))		    																																
																																				
        ls_data=Open_landsat(src_FileName, proyDEM_fileName) 																																	
        ls_data = ls_data*ClipLandsat
        index = np.where(Bands[:] == band)[0][0] - 6
        therm_data[:, :, index] = ls_data
								
    return(therm_data)
   
#------------------------------------------------------------------------------   
def Open_landsat(src_FileName, proyDEM_fileName):
    """
    This function opens a landsat image and returns the data array of a specific landsat band.
    """                           
    # crop band to the DEM extent
    ls, ulx, uly, lrx, lry, epsg_to = reproject_dataset_example(src_FileName, proyDEM_fileName)											
    
    # Open the cropped Landsat image for the band number
    ls_data = ls.GetRasterBand(1).ReadAsArray()
    return(ls_data) 
  
#------------------------------------------------------------------------------
def Get_Extend_Landsat(src_FileName):
    """
    This function gets the extend of the landsat image
    """
    ls = gdal.Open(src_FileName)       # Open Landsat image
    print 'Original LANDSAT Image - '
    geo_t_ls = ls.GetGeoTransform()    # Get the Geotransform vector
    x_size_ls = ls.RasterXSize         # Raster xsize - Columns
    y_size_ls = ls.RasterYSize         # Raster ysize - Rows
    print '  Size :', x_size_ls, y_size_ls
    (ulx, uly) = geo_t_ls[0], geo_t_ls[3]
    (lrx, lry) = (geo_t_ls[0] + geo_t_ls[1] * x_size_ls,
                  geo_t_ls[3] + geo_t_ls[5] * y_size_ls)
    band_data = ls.GetRasterBand(1)
    
    return(ls,band_data,ulx,uly,lrx,lry,x_size_ls,y_size_ls)
    
#------------------------------------------------------------------------------
def Calc_Ra_Mountain(lon,DOY,hour,minutes,lon_proy,lat_proy,slope,aspect):    
    """
    Calculates the extraterrestiral solar radiation by using the date, slope and aspect.
    """
    
    # Constants
    deg2rad = np.pi / 180.0  # Factor to transform from degree to rad
    Min_cos_zn = 0.1  # Min value for cos zenith angle
    Max_cos_zn = 1.0  # Max value for cos zenith angle
    Gsc = 1367        # Solar constant (W / m2)
    try:
        Loc_time = float(hour) + float(minutes)/60  # Local time (hours)
    except:
        Loc_time = np.float_(hour) + np.float_(minutes)/60  # Local time (hours)
    # Rounded difference of the local time from Greenwich (GMT) (hours):
    offset_GTM = round(np.sign(lon[int(lon.shape[0])/2, int(lon.shape[1])/2]) * lon[int(lon.shape[0])/2,int(lon.shape[1])/2] * 24 / 360)
                       
    print '  Local time: ', '%0.3f' % np.nanmean(Loc_time)
    print '  Difference of local time (LT) from Greenwich (GMT): ', offset_GTM

    # 1. Calculation of extraterrestrial solar radiation for slope and aspect
    # Computation of Hour Angle (HRA = w)
    B = 360./365 * (DOY-81)           # (degrees)
    # Computation of cos(theta), where theta is the solar incidence angle
    # relative to the normal to the land surface
    delta=np.arcsin(np.sin(23.45*deg2rad)*np.sin(np.deg2rad(B))) # Declination angle (radians)
    phi = lat_proy * deg2rad                                     # latitude of the pixel (radians)
    s = slope * deg2rad                                          # Surface slope (radians)
    gamma = (aspect-180) * deg2rad                               # Surface aspect angle (radians)
    w=w_time(Loc_time, lon_proy, DOY)                            # Hour angle (radians)
    a,b,c = Constants(delta,s,gamma,phi)
    cos_zn= AngleSlope(a,b,c,w)
    cos_zn = cos_zn.clip(Min_cos_zn, Max_cos_zn)
    
    print 'Average Cos Zenith Angle: ', '%0.3f (Radians)' % np.nanmean(cos_zn)

    dr = 1 + 0.033 * cos(DOY*2*pi/365)  # Inverse relative distance Earth-Sun
    # Instant. extraterrestrial solar radiation (W/m2), Allen et al.(2006):
    Ra_inst = Gsc * cos_zn * dr              

    # 24-hours extraterrestrial radiation
    # 1.) determine if there are one or two periods of sun
    # 2.) calculate the 24-hours extraterrestrial radiation if there are two periods of sun
    # 3.) calculate the 24-hours extraterrestrial radiation if there is one period of sun
    
    #1.) determine amount of sun periods
    Ra_24 = np.zeros(np.shape(lat_proy))*np.nan
    constant=Gsc*dr/(2*np.pi)
    TwoPeriod= TwoPeriods(delta,s,phi)  # all input in radians
    
    #2.) calculate the 24-hours extraterrestrial radiation (2 periods)
    ID = np.where(np.ravel(TwoPeriod==True))
    Ra_24.flat[ID]=TwoPeriodSun(constant,delta,s.flat[ID],gamma.flat[ID],phi.flat[ID])
    
    #3.) calculate the 24-hours extraterrestrial radiation (1 period)
    ID = np.where(np.ravel(TwoPeriod==False))
    Ra_24.flat[ID]=OnePeriodSun(constant,delta,s.flat[ID],gamma.flat[ID],phi.flat[ID])
        
    # Horizontal surface
    ws = np.arccos(-np.tan(delta) * np.tan(phi))  # Sunrise/sunset time angle
    
    # Extraterrestial radiation for a horizontal surface for 24-h period:
    Ra_hor_24 = (Gsc * dr / np.pi * (np.sin(delta) * np.sin(phi) * ws + np.cos(delta) * np.cos(phi) * np.sin(ws)))
    # cos_theta_flat = (np.sin(delta) * np.sin(phi) + np.cos(delta) * np.cos(phi) * np.cos(w))

    # Mountain radiation
    Ra_mountain_24 = np.where(Ra_24 > Min_cos_zn * Ra_hor_24, Ra_24 / np.cos(s),
                           Ra_hor_24)
    Ra_mountain_24[Ra_mountain_24 > 600.0] = 600.0
                  
    return(Ra_mountain_24,Ra_inst,cos_zn,dr,phi,delta)
    
#------------------------------------------------------------------------------    
def OnePeriodSun(constant,delta,s,gamma,phi):
    '''
    Based on Richard G. Allen 2006
    Calculate the 24-hours extraterrestrial radiation when there is one sun period
    '''
    sunrise,sunset = SunHours(delta,s,gamma,phi)    
    Vals=IntegrateSlope(constant,sunrise,sunset,delta,s,gamma,phi)
    
    return(Vals)
    
#------------------------------------------------------------------------------   
def TwoPeriodSun(constant,delta,s,gamma,phi):  
    '''
    Based on Richard G. Allen 2006
    Calculate the 24-hours extraterrestrial radiation when there are two sun period
    '''
    A1, A2 = SunHours(delta,s,gamma,phi)
    a,b,c = Constants(delta,s,gamma,phi)
    riseSlope, setSlope = BoundsSlope(a,b,c)
    B1 = np.maximum(riseSlope,setSlope)
    B2 = np.minimum(riseSlope,setSlope)
    Angle_B1 = AngleSlope(a,b,c,B1)
    Angle_B2 = AngleSlope(a,b,c,B2)
   
    B1[abs(Angle_B1) > 0.001] = np.pi - B1[abs(Angle_B1) > 0.001]  
    B2[abs(Angle_B2) > 0.001] = -np.pi - B2[abs(Angle_B2) > 0.001]    

    # Check if two periods really exist
    ID = np.ravel_multi_index(np.where(np.logical_and(B2 >= A1, B1 >= A2) == True),a.shape)
    Val = IntegrateSlope(constant,B2.flat[ID],B1.flat[ID],delta,s.flat[ID],gamma.flat[ID],phi.flat[ID])
    ID = ID[Val < 0]

    # Finally calculate resulting values
    Vals = np.zeros(B1.shape)

    Vals.flat[ID] = (IntegrateSlope(constant,A1.flat[ID],B2.flat[ID],delta,s.flat[ID],gamma.flat[ID],phi.flat[ID])  + 
                   IntegrateSlope(constant,B1.flat[ID],A2.flat[ID],delta,s.flat[ID],gamma.flat[ID],phi.flat[ID]))
    ID = np.ravel_multi_index(np.where(Vals == 0),a.shape)   
    Vals.flat[ID] = IntegrateSlope(constant,A1.flat[ID],A2.flat[ID],delta,s.flat[ID],gamma.flat[ID],phi.flat[ID])
    
    return(Vals)
    
#------------------------------------------------------------------------------       
def IntegrateSlope(constant,sunrise,sunset,delta,s,gamma,phi):
    '''
    Based on Richard G. Allen 2006 equation 5
    Calculate the 24 hours extraterrestrial radiation
    '''
    # correct the sunset and sunrise angels for days that have no sunset or no sunrise 
    SunOrNoSun = np.logical_or(((np.abs(delta + phi)) > (np.pi/2)),((np.abs(delta - phi)) > (np.pi/2)))
    integral=np.zeros(s.shape)
    ID = np.where(np.ravel(SunOrNoSun==True))
    
    # No sunset
    if abs(delta+phi.flat[ID])>(np.pi/2):
        sunset1=np.pi
        sunrise1=-np.pi
        integral.flat[ID] = constant * (np.sin(delta)*np.sin(phi)*np.cos(s)*(sunset1-sunrise1) 
            - np.sin(delta)*np.cos(phi)*np.sin(s)*np.cos(gamma)*(sunset1-sunrise1)
            + np.cos(delta)*np.cos(phi)*np.cos(s)*(np.sin(sunset1)-np.sin(sunrise1))
            + np.cos(delta)*np.sin(phi)*np.sin(s)*np.cos(gamma)*(np.sin(sunset1)-np.sin(sunrise1))
            - np.cos(delta)*np.sin(s)*np.sin(gamma)*(np.cos(sunset1)-np.cos(sunrise1)))
   
    # No sunrise
    elif np.abs(delta-phi.flat[ID])>(np.pi/2):
        integral.flat[ID]=constant * (np.sin(delta)*np.sin(phi)*np.cos(s)*(0) 
            - np.sin(delta)*np.cos(phi)*np.sin(s)*np.cos(gamma)*(0)
            + np.cos(delta)*np.cos(phi)*np.cos(s)*(np.sin(0)-np.sin(0))
            + np.cos(delta)*np.sin(phi)*np.sin(s)*np.cos(gamma)*(np.sin(0)-np.sin(0))
            - np.cos(delta)*np.sin(s)*np.sin(gamma)*(np.cos(0)-np.cos(0)))
        
    ID = np.where(np.ravel(SunOrNoSun==False))
    integral.flat[ID] = constant * (np.sin(delta)*np.sin(phi)*np.cos(s)*(sunset-sunrise) 
            - np.sin(delta)*np.cos(phi)*np.sin(s)*np.cos(gamma)*(sunset-sunrise)
            + np.cos(delta)*np.cos(phi)*np.cos(s)*(np.sin(sunset)-np.sin(sunrise))
            + np.cos(delta)*np.sin(phi)*np.sin(s)*np.cos(gamma)*(np.sin(sunset)-np.sin(sunrise))
            - np.cos(delta)*np.sin(s)*np.sin(gamma)*(np.cos(sunset)-np.cos(sunrise)))
    
    return(integral)     

#------------------------------------------------------------------------------    
def TwoPeriods(delta,s,phi):
    '''
    Based on Richard G. Allen 2006
    Create a boolean map with True values for places with two sunsets
    '''
    TwoPeriods = (np.sin(s) > np.ones(s.shape)*np.sin(phi)*np.sin(delta)+np.cos(phi)*np.cos(delta))
    
    return(TwoPeriods)

#------------------------------------------------------------------------------
def SunHours(delta,slope,slopedir,lat):
    # Define sun hours in case of one sunlight period
    
    a,b,c = Constants(delta,slope,slopedir,lat)
    riseSlope, setSlope = BoundsSlope(a,b,c)  
    bound = BoundsHorizontal(delta,lat)  

    Calculated = np.zeros(slope.shape, dtype = bool)    
    RiseFinal = np.zeros(slope.shape)    
    SetFinal = np.zeros(slope.shape)
    
    # First check sunrise is not nan
    # This means that their is either no sunrise (whole day night) or no sunset (whole day light)
    # For whole day light, use the horizontal sunrise and whole day night a zero..
    Angle4 = AngleSlope(a,b,c,-bound)    
    RiseFinal[np.logical_and(np.isnan(riseSlope),Angle4 >= 0)] = -bound[np.logical_and(np.isnan(riseSlope),Angle4 >= 0)]
    Calculated[np.isnan(riseSlope)] = True 
    
    # Step 1 > 4    
    Angle1 = AngleSlope(a,b,c,riseSlope)
    Angle2 = AngleSlope(a,b,c,-bound)    
    
    ID = np.ravel_multi_index(np.where(np.logical_and(np.logical_and(Angle2 < Angle1+0.001 ,Angle1 < 0.001),Calculated == False) == True),a.shape)
    RiseFinal.flat[ID] = riseSlope.flat[ID]
    Calculated.flat[ID] = True
    # step 5 > 7
    Angle3 = AngleSlope(a,b,c,-np.pi - riseSlope)
    
    ID = np.ravel_multi_index(np.where(np.logical_and(np.logical_and(-bound<(-np.pi-riseSlope),Angle3 <= 0.001),Calculated == False) == True),a.shape) 
    RiseFinal.flat[ID] = -np.pi -riseSlope.flat[ID]
    Calculated.flat[ID] = True
    
    # For all other values we use the horizontal sunset if it is positive, otherwise keep a zero   
    RiseFinal[Calculated == False] = -bound[Calculated == False] 
    
    # Then check sunset is not nan or < 0 
    Calculated = np.zeros(slope.shape, dtype = bool)     
    
    Angle4 = AngleSlope(a,b,c,bound)    
    SetFinal[np.logical_and(np.isnan(setSlope),Angle4 >= 0)] = bound[np.logical_and(np.isnan(setSlope),Angle4 >= 0)]
    Calculated[np.isnan(setSlope)] = True     
    
    # Step 1 > 4    
    Angle1 = AngleSlope(a,b,c,setSlope)
    Angle2 = AngleSlope(a,b,c,bound)    
    
    ID = np.ravel_multi_index(np.where(np.logical_and(np.logical_and(Angle2 < Angle1+0.001,Angle1 < 0.001),Calculated == False) == True),a.shape)
    SetFinal.flat[ID] = setSlope.flat[ID]
    Calculated.flat[ID] = True
    # step 5 > 7
    Angle3 = AngleSlope(a,b,c,np.pi - setSlope)
    
    ID = np.ravel_multi_index(np.where(np.logical_and(np.logical_and(bound>(np.pi-setSlope),Angle3 <= 0.001),Calculated == False) == True),a.shape) 
    SetFinal.flat[ID] = np.pi - setSlope.flat[ID]
    Calculated.flat[ID] = True
    
    # For all other values we use the horizontal sunset if it is positive, otherwise keep a zero   
    SetFinal[Calculated == False] = bound[Calculated == False]   
    
    #    Angle4 = AngleSlope(a,b,c,bound)    
    #    SetFinal[np.logical_and(Calculated == False,Angle4 >= 0)] = bound[np.logical_and(Calculated == False,Angle4 >= 0)]

    # If Sunrise is after Sunset there is no sunlight during the day
    SetFinal[SetFinal <= RiseFinal] = 0
    RiseFinal[SetFinal <= RiseFinal] = 0
    
    return(RiseFinal,SetFinal)

#------------------------------------------------------------------------------    
def Constants(delta,s,gamma,phi):
    '''
    Based on Richard G. Allen 2006 equation 11
    determines constants for calculating the exterrestial solar radiation
    '''
    a = np.sin(delta)*np.cos(phi)*np.sin(s)*np.cos(gamma) - np.sin(delta)*np.sin(phi)*np.cos(s)
    b = np.cos(delta)*np.cos(phi)*np.cos(s) + np.cos(delta)*np.sin(phi)*np.sin(s)*np.cos(gamma)
    c = np.cos(delta)*np.sin(s)*np.sin(gamma)
    
    return(a,b,c)

#------------------------------------------------------------------------------
def BoundsSlope(a,b,c):
    '''
    Based on Richard G. Allen 2006 equation 13
    This function calculates candidate values for sunrise and sunset hour angles
    '''
    Div = (b**2+c**2)   
    Div[Div <= 0] = 0.00001
    sinB = (a*c + b*np.sqrt(b**2+c**2-a**2)) / Div
    sinA = (a*c - b*np.sqrt(b**2+c**2-a**2)) / Div

    sinB[sinB < -1] = -1; sinB[sinB > 1] = 1    # Limits see appendix A.2.i
    sinA[sinA < -1] = -1; sinA[sinA > 1] = 1    # Limits see appendix A.2.i

    sunrise = np.arcsin(sinA)
    sunset = np.arcsin(sinB)
    
    return(sunrise,sunset)
    
#------------------------------------------------------------------------------   
def BoundsHorizontal(delta,phi):
    ''''
    Based on Richard G. Allen 2006
    This function calculates sunrise hours based on earth inclination and latitude
    If there is no sunset or sunrise hours the values are either set to 0 (polar night) or pi (polar day)    
    '''
    bound = np.arccos(-np.tan(delta)*np.tan(phi))
    bound[abs(delta+phi) > np.pi/2] = np.pi
    bound[abs(delta-phi) > np.pi/2] = 0
          
    return(bound)

#------------------------------------------------------------------------------
def AngleSlope(a,b,c,w):
    '''
    Based on Richard G. Allen 2006
    Calculate the cos zenith angle by using the hour angle and constants
    '''
    angle = -a + b*np.cos(w) + c*np.sin(w)
    
    return(angle)    

#------------------------------------------------------------------------------
def Calc_Gradient(dataset,pixel_spacing):
    """
    This function calculates the slope and aspect of a DEM map.
    """
    # constants
    deg2rad = np.pi / 180.0  # Factor to transform from degree to rad
    rad2deg = 180.0 / np.pi  # Factor to transform from rad to degree
    
    # Calculate slope		
    x, y = np.gradient(dataset)
    slope = np.arctan(np.sqrt(np.square(x/pixel_spacing) + np.square(y/pixel_spacing))) * rad2deg
    
    # calculate aspect                  
    aspect = np.arctan2(y/pixel_spacing, -x/pixel_spacing) * rad2deg
    aspect = 180 + aspect

    return(deg2rad,rad2deg,slope,aspect)

#------------------------------------------------------------------------------
def DEM_lat_lon(DEM_fileName,output_folder):
    """
    This function retrieves information about the latitude and longitude of the
    DEM map. 
    
    """
    # name for output
    lat_fileName = os.path.join(output_folder, 'Output_radiation_balance','latitude.tif')
    lon_fileName = os.path.join(output_folder, 'Output_radiation_balance','longitude.tif')
                                
    g = gdal.Open(DEM_fileName)     # Open DEM
    geo_t = g.GetGeoTransform()     # Get the Geotransform vector:
    x_size = g.RasterXSize          # Raster xsize - Columns
    y_size = g.RasterYSize          # Raster ysize - Rows
    
    # create a longitude and a latitude array 
    lon = np.zeros((y_size, x_size))
    lat = np.zeros((y_size, x_size))
    for col in np.arange(x_size):
        lon[:, col] = geo_t[0] + col * geo_t[1] + geo_t[1]/2
        # ULx + col*(E-W pixel spacing) + E-W pixel spacing
    for row in np.arange(y_size):
        lat[row, :] = geo_t[3] + row * geo_t[5] + geo_t[5]/2
        # ULy + row*(N-S pixel spacing) + N-S pixel spacing,
        # negative as we will be counting from the UL corner
    
    # Define shape of the raster    
    shape = [x_size, y_size]
    
    # Save lat and lon files in geo- coordinates
    save_GeoTiff_proy(g, lat, lat_fileName, shape, nband=1)
    save_GeoTiff_proy(g, lon, lon_fileName, shape, nband=1)
    
    return(lat,lon,lat_fileName,lon_fileName)

#------------------------------------------------------------------------------
def reproject_dataset(dataset, pixel_spacing, UTM_Zone):
    """
    A sample function to reproject and resample a GDAL dataset from within
    Python. The idea here is to reproject from one system to another, as well
    as to change the pixel size. The procedure is slightly long-winded, but
    goes like this:

    1. Set up the two Spatial Reference systems.
    2. Open the original dataset, and get the geotransform
    3. Calculate bounds of new geotransform by projecting the UL corners
    4. Calculate the number of pixels with the new projection & spacing
    5. Create an in-memory raster dataset
    6. Perform the projection
    """

    # 1) Open the dataset
    g = gdal.Open(dataset)
    if g is None:
        print 'input folder does not exist'

     # Define the EPSG code...
    EPSG_code = '326%02d' % UTM_Zone
    epsg_to = int(EPSG_code)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    try:
        proj = g.GetProjection()
        Proj_in=proj.split('EPSG","')
        epsg_from=int((str(Proj_in[-1]).split(']')[0])[0:-1])		  
    except:
        epsg_from = int(4326)    # Get the Geotransform vector:
    geo_t = g.GetGeoTransform()
    
    # Vector components:
    # 0- The Upper Left easting coordinate (i.e., horizontal)
    # 1- The E-W pixel spacing
    # 2- The rotation (0 degrees if image is "North Up")
    # 3- The Upper left northing coordinate (i.e., vertical)
    # 4- The rotation (0 degrees)
    # 5- The N-S pixel spacing, negative as it is counted from the UL corner
    x_size = g.RasterXSize  # Raster xsize
    y_size = g.RasterYSize  # Raster ysize

    epsg_to = int(epsg_to)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    osng = osr.SpatialReference()
    osng.ImportFromEPSG(epsg_to)
    wgs84 = osr.SpatialReference()
    wgs84.ImportFromEPSG(epsg_from)

    inProj = Proj(init='epsg:%d' %epsg_from)
    outProj = Proj(init='epsg:%d' %epsg_to)

    nrow_skip = round((0.07*y_size)/2)
    ncol_skip = round((0.04*x_size)/2)
				
    # Up to here, all  the projection have been defined, as well as a
    # transformation from the from to the to
    ulx, uly = transform(inProj,outProj,geo_t[0], geo_t[3] + nrow_skip * geo_t[5])
    lrx, lry = transform(inProj,outProj,geo_t[0] + geo_t[1] * (x_size-ncol_skip),
                                        geo_t[3] + geo_t[5] * (y_size-nrow_skip))

    # See how using 27700 and WGS84 introduces a z-value!
    # Now, we create an in-memory raster
    mem_drv = gdal.GetDriverByName('MEM')

    # The size of the raster is given the new projection and pixel spacing
    # Using the values we calculated above. Also, setting it to store one band
    # and to use Float32 data type.
    col = int((lrx - ulx)/pixel_spacing)
    rows = int((uly - lry)/pixel_spacing)

    # Re-define lr coordinates based on whole number or rows and columns
    (lrx, lry) = (ulx + col * pixel_spacing, uly -
                  rows * pixel_spacing)
																		
    dest = mem_drv.Create('', col, rows, 1, gdal.GDT_Float32)
    
    if dest is None:
        print 'input folder to large for memory, clip input map'
     
   # Calculate the new geotransform
    new_geo = (ulx, pixel_spacing, geo_t[2], uly,
               geo_t[4], - pixel_spacing)
    
    # Set the geotransform
    dest.SetGeoTransform(new_geo)
    dest.SetProjection(osng.ExportToWkt())
      
    # Perform the projection/resampling
    gdal.ReprojectImage(g, dest, wgs84.ExportToWkt(), osng.ExportToWkt(),gdal.GRA_Bilinear)						

    return dest, ulx, lry, lrx, uly, epsg_to
#------------------------------------------------------------------------------
def reproject_dataset_example(dataset, dataset_example, method = 1):
   
    # open example dataset 
    g_ex = gdal.Open(dataset_example)
    try:
        proj = g_ex.GetProjection()
        Proj=proj.split('EPSG","')
        epsg_to=int((str(Proj[-1]).split(']')[0])[0:-1])		
    except:
        epsg_to = int(4326)       
      
    Y_raster_size = g_ex.RasterYSize				
    X_raster_size = g_ex.RasterXSize
				
    Geo = g_ex.GetGeoTransform()	
    ulx = Geo[0]
    uly = Geo[3]
    lrx = ulx + X_raster_size * Geo[1]				
    lry = uly + Y_raster_size * Geo[5]	
 
    # open dataset that must be transformed    
    g_in = gdal.Open(dataset)
    try:
        proj = g_in.GetProjection()
        Proj=proj.split('EPSG","')
        epsg_from=int((str(Proj[-1]).split(']')[0])[0:-1])		   
    except:
        epsg_from = int(4326)

    # Set the EPSG codes
    osng = osr.SpatialReference()
    osng.ImportFromEPSG(epsg_to)
    wgs84 = osr.SpatialReference()
    wgs84.ImportFromEPSG(epsg_from)

    # Create new raster			
    mem_drv = gdal.GetDriverByName('MEM')
    dest1 = mem_drv.Create('', X_raster_size, Y_raster_size, 1, gdal.GDT_Float32)
    dest1.SetGeoTransform(Geo)
    dest1.SetProjection(osng.ExportToWkt())
    
    # Perform the projection/resampling
    if method == 1:
        gdal.ReprojectImage(g_in, dest1, wgs84.ExportToWkt(), osng.ExportToWkt(), gdal.GRA_NearestNeighbour)
    if method == 2:
        gdal.ReprojectImage(g_in, dest1, wgs84.ExportToWkt(), osng.ExportToWkt(), gdal.GRA_Average)
    if method == 3:
        gdal.ReprojectImage(g_in, dest1, wgs84.ExportToWkt(), osng.ExportToWkt(), gdal.GRA_Cubic)
        
    return(dest1, ulx, lry, lrx, uly, epsg_to)			

#------------------------------------------------------------------------------
def save_GeoTiff_proy(src_dataset, dst_dataset_array, dst_fileName, shape_lsc, nband):
    """
    This function saves an array dataset in GeoTiff, using the parameters
    from the source dataset, in projected coordinates

    """
    dst_dataset_array	= np.float_(dst_dataset_array)		
    dst_dataset_array[dst_dataset_array<-9999] = np.nan					
    geotransform = src_dataset.GetGeoTransform()
    spatialreference = src_dataset.GetProjection()
    
    # create dataset for output
    fmt = 'GTiff'
    driver = gdal.GetDriverByName(fmt)
    dir_name = os.path.dirname(dst_fileName)
    
    # If the directory does not exist, make it.
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)
    dst_dataset = driver.Create(dst_fileName, shape_lsc[0], shape_lsc[1], nband,gdal.GDT_Float32)
    dst_dataset.SetGeoTransform(geotransform)
    dst_dataset.SetProjection(spatialreference)
    dst_dataset.GetRasterBand(1).SetNoDataValue(-9999)				
    dst_dataset.GetRasterBand(1).WriteArray(dst_dataset_array)
    dst_dataset = None

#------------------------------------------------------------------------------
def w_time(LT,lon_proy, DOY):
    """
    This function computes the hour angle (radians) of an image given the
    local time, longitude, and day of the year.

    """
    nrow, ncol = lon_proy.shape
   
    # Difference of the local time (LT) from Greenwich Mean Time (GMT) (hours):
    delta_GTM = np.sign(lon_proy[nrow/2, ncol/2]) * lon_proy[nrow/2, ncol/2] * 24 / 360
    if np.isnan(delta_GTM) == True:
         delta_GTM = np.nanmean(lon_proy) * np.nanmean(lon_proy)  * 24 / 360
    
    
    # Local Standard Time Meridian (degrees):
    LSTM = 15 * delta_GTM
    
    # Ecuation of time (EoT, minutes):
    B = 360./365 * (DOY-81)  # (degrees)
    EoT = 9.87*sin(np.deg2rad(2*B))-7.53*cos(np.deg2rad(B))-1.5*sin(np.deg2rad(B))
   
    # Net Time Correction Factor (minutes) at the center of the image:
    TC = 4 * (lon_proy - LSTM) + EoT     # Difference in time over the longitude
    LST = LT + delta_GTM + TC/60         # Local solar time (hours)
    HRA = 15 * (LST-12)                  # Hour angle HRA (degrees)
    deg2rad = np.pi / 180.0              # Factor to transform from degree to rad
    w = HRA * deg2rad                    # Hour angle HRA (radians)
    return w

#------------------------------------------------------------------------------
def info_general_metadata(filename):
    """
    This function retrieves general information of the Landsat image
    (date and time aquired, UTM zone, sun elevation) from the
    metadata file.

    """
    Landsat_meta = open(filename, "r")  # Open metadata file
    for line in Landsat_meta:
        if re.match("(.*)SCENE_CENTER_TIME(.*)", line): # search in metadata for line SCENE_CENTER_TIME
            words = line.split()# make groups of words which are divided by an open space
            time_list = words[2].split(':', 2) # Take the second word of words and split the word which are divided by :
            if len(time_list[0])== 3: 
                time_list[0]=time_list[0][1:3]
                time_list[2]=time_list[2][0:-1]
            hour = float(time_list[0]) # take the first word of time_list
            minutes = float(time_list[1]) + float(time_list[2][:-1]) / 60 # Take the second and third word of time_list and place :-1 to remove Z behined minutes
    Landsat_meta = open(filename, "r")  # Open metadata file    
    for line in Landsat_meta:
        if re.match("(.*)DATE_ACQUIRED(.*)", line):
            words = line.split()
            DOY = time.strptime(words[2], "%Y-%m-%d").tm_yday
            year = time.strptime(words[2], "%Y-%m-%d").tm_year
    Landsat_meta = open(filename, "r")  # Open metadata file
    for line in Landsat_meta:
        if re.match("(.*)UTM_ZONE(.*)", line):
            words = line.split()
            UTM_Zone = int(words[2])
    Landsat_meta = open(filename, "r")  # Open metadata file
    for line in Landsat_meta:
        if re.match("(.*)SUN_ELEVATION(.*)", line):
            words = line.split()
            Sun_elevation = float(words[2])

    return year, DOY, hour, minutes, UTM_Zone, Sun_elevation

#------------------------------------------------------------------------------
def info_band_metadata(filename, Bands):
    """
    This function retrieves Landsat band information (minimum and maximum
    radiance) from the metadata file.

    """
    Lmin = np.zeros(len(Bands))  # Minimum band radiance, for each band
    Lmax = np.zeros(len(Bands))  # Maximum band radiance, for each band
    k1_const = np.zeros(len(Bands)-6)  # TIRS_Thermal constant k1 ######
    k2_const = np.zeros(len(Bands)-6)  # TIRS_Thermal constant k2 ######
    for band in Bands:
        Landsat_meta = open(filename, "r")  # Open metadata file
        for line in Landsat_meta:
            if re.match("(.*)RADIANCE_MINIMUM_BAND_%1d(.*)" % band, line):
                words = line.split()
                value = float(words[2])
                Lmin[np.where(Bands == band)[0][0]] = value
            if re.match("(.*)RADIANCE_MAXIMUM_BAND_%1d(.*)" % band, line):
                words = line.split()
                value = float(words[2])
                Lmax[np.where(Bands == band)[0][0]] = value
            if re.match("(.*)K1_CONSTANT_BAND_%1d(.*)" % band, line):  # #####
                words = line.split()
                value = float(words[2])
                k1_const[np.where(Bands == band)[0][0]-6] = value
            if re.match("(.*)K2_CONSTANT_BAND_%1d(.*)" % band, line):  # #####
                words = line.split()
                value = float(words[2])
                k2_const[np.where(Bands == band)[0][0]-6] = value
    return Lmin, Lmax, k1_const, k2_const


#------------------------------------------------------------------------------
def sensible_heat(rah, ustar, rn_inst, g_inst, ts_dem, ts_dem_hot, ts_dem_cold,
                  air_dens, Surf_temp, k_vk, QC_Map, hot_pixels = None, slope = None, offset_dt = None, slope_dt = None):
    """
    This function computes the instantaneous sensible heat given the
    instantaneous net radiation, ground heat flux, and other parameters.

    """
     # Near surface temperature difference (dT):
    dT_ini = (rn_inst - g_inst) * rah / (air_dens * 1004)
    dT_hot = np.copy(dT_ini)
    
    #dT_hot_fileName = os.path.join(output_folder, 'Output_cloud_masked','test.tif')
    #save_GeoTiff_proy(dest, dT_hot, dT_hot_fileName,shape, nband=1)
    
    if (hot_pixels is not None and slope is not None): 
        # dT for hot pixels - hot, (dry) agricultural fields with no green veget.:
        dT_hot[ts_dem <= (ts_dem_hot - 0.5)] = np.nan
        dT_hot[QC_Map == 1] = np.nan
        dT_hot[dT_hot == 0] = np.nan
        if np.all(np.isnan(dT_hot)) == True:
           dT_hot = np.copy(dT_ini)       
           ts_dem_hot = np.nanpercentile(hot_pixels, 99.5)
           dT_hot[ts_dem <= (ts_dem_hot - 0.5)] = np.nan
           dT_hot[dT_hot == 0] = np.nan 
           
        dT_hot=np.float32(dT_hot)
        dT_hot[slope > 10]=np.nan  
              
        dT_hot_mean = np.nanmean(dT_hot)
    
        # Compute slope and offset of linear relationship dT = b + a * Ts
        slope_dt = (dT_hot_mean - 0.0) / (ts_dem_hot - ts_dem_cold)  # EThot = 0.0
        offset_dt = dT_hot_mean - slope_dt * ts_dem_hot
    elif (slope_dt is not None and offset_dt is not None):
        pass
    
    dT = offset_dt + slope_dt * ts_dem
    
    # Sensible heat flux:
    h = air_dens * 1004 * dT / rah
    h[QC_Map == 1] = np.nan
    h[h==0]=np.nan
    h[QC_Map != 0] = np.nan
     
    # Monin-Obukhov length (m):
    L_MO = ((-1.0 * air_dens * 1004 * np.power(ustar, 3) * Surf_temp) /
            (k_vk * 9.81 * h))
    L_MO[L_MO < -1000] = -1000
        
    # Stability correction for momentum, stable conditions (L_MO >= 0):
    psi_200_stable = -0.05 * 200 / L_MO
    
    # Stability correction for momentum and heat transport, unstable
    # conditions (L_MO < 0):
    x2 = np.power((1.0 - 16.0 * (2.0/L_MO)), 0.25)  # x at 2m
    x200 = np.power(1.0 - 16.0 * (200/L_MO), 0.25)  # x at 200m
    psi_h = 2 * np.log((1 + np.power(x2, 2))/2)
    psi_m200 = (2 * np.log((1 + x200) / 2) + np.log((1 + np.power(x200, 2)) /
                2) - 2 * np.arctan(x200) + 0.5*np.pi)


    return L_MO, psi_200_stable, psi_h, psi_m200, h, dT, slope_dt, offset_dt

#------------------------------------------------------------------------------   
    
def Reshape_Reproject_Input_data(input_File_Name, output_File_Name, Example_extend_fileName):

   # Reproject the dataset based on the example       
   data_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
       input_File_Name, Example_extend_fileName)
   
   # Get the array information from the new created map
   band_data = data_rep.GetRasterBand(1) # Get the reprojected dem band
   ncol_data = data_rep.RasterXSize
   nrow_data = data_rep.RasterYSize
   shape_data=[ncol_data, nrow_data]
 
   # Save new dataset 
   #stats = band.GetStatistics(0, 1)
   data = band_data.ReadAsArray(0, 0, ncol_data, nrow_data)
   save_GeoTiff_proy(data_rep, data, output_File_Name, shape_data, nband=1)
   
   return(data)

#------------------------------------------------------------------------------   
def Thermal_Sharpening(surface_temp_up, NDVI_up, NDVI, Box, dest_up, output_folder, ndvi_fileName, shape_down, dest_down, temp_surface_sharpened_fileName):

    # Creating arrays to store the coefficients						
    CoefA=np.zeros((len(surface_temp_up),len(surface_temp_up[1])))
    CoefB=np.zeros((len(surface_temp_up),len(surface_temp_up[1])))
    CoefC=np.zeros((len(surface_temp_up),len(surface_temp_up[1])))
  
    # Fit a second polynominal fit to the NDVI and Thermal data and save the coefficients for each pixel  
    # NOW USING FOR LOOPS PROBABLY NOT THE FASTEST METHOD     
    for i in range(0,len(surface_temp_up)):
        for j in range(0,len(surface_temp_up[1])):
            if np.isnan(np.sum(surface_temp_up[i,j]))==False and np.isnan(np.sum(NDVI_up[i,j]))==False:
                x_data=NDVI_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)][np.logical_and(np.logical_not(np.isnan(NDVI_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)])),np.logical_not(np.isnan(surface_temp_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)])))]
                y_data=surface_temp_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)][np.logical_and(np.logical_not(np.isnan(NDVI_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)])),np.logical_not(np.isnan(surface_temp_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)])))]
                if len(x_data)>6:
                    coefs = poly.polyfit(x_data, y_data, 2)
                    CoefA[i,j] = coefs[2]
                    CoefB[i,j] = coefs[1]
                    CoefC[i,j] = coefs[0]
                else:
                    CoefA[i,j] = np.nan
                    CoefB[i,j] = np.nan
                    CoefC[i,j] = np.nan
            else:
                CoefA[i,j] = np.nan
                CoefB[i,j] = np.nan
                CoefC[i,j] = np.nan
            
    # Define the shape of the surface temperature with the resolution of 400m      
    shape_up=[len(surface_temp_up[1]),len(surface_temp_up)]
           
    # Save the coefficients
    CoefA_fileName_Optie2 = os.path.join(output_folder, 'Output_temporary','coef_A.tif')
    save_GeoTiff_proy(dest_up,CoefA, CoefA_fileName_Optie2,shape_up, nband=1)
        
    CoefB_fileName_Optie2 = os.path.join(output_folder, 'Output_temporary','coef_B.tif')
    save_GeoTiff_proy(dest_up,CoefB, CoefB_fileName_Optie2,shape_up, nband=1)
 
    CoefC_fileName_Optie2 = os.path.join(output_folder, 'Output_temporary','coef_C.tif')
    save_GeoTiff_proy(dest_up,CoefC, CoefC_fileName_Optie2,shape_up, nband=1)
       
    # Downscale the fitted coefficients
    CoefA_Downscale, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
                                  CoefA_fileName_Optie2, ndvi_fileName)
    CoefA = CoefA_Downscale.GetRasterBand(1).ReadAsArray()
       
    CoefB_Downscale, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
                                  CoefB_fileName_Optie2, ndvi_fileName)
    CoefB = CoefB_Downscale.GetRasterBand(1).ReadAsArray()
        
    CoefC_downscale, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
                                CoefC_fileName_Optie2, ndvi_fileName)
    CoefC = CoefC_downscale.GetRasterBand(1).ReadAsArray()

    # Calculate the surface temperature based on the fitted coefficents and NDVI
    temp_surface_sharpened=CoefA*NDVI**2+CoefB*NDVI+CoefC
    temp_surface_sharpened[temp_surface_sharpened < 250] = np.nan					
    temp_surface_sharpened[temp_surface_sharpened > 400] = np.nan	
				
    save_GeoTiff_proy(dest_down,temp_surface_sharpened, temp_surface_sharpened_fileName,shape_down, nband=1)	
    return(temp_surface_sharpened)				

#------------------------------------------------------------------------------   

def Check_dT(rn_inst, g_inst, LAI, vegt_cover, z0m, u_200, Temp_inst, RH_inst, air_dens, Psychro_c, rl, rs_min, rsoil_min):    

    esat_inst = 0.6108*np.exp((17.27*Temp_inst)/(237.3+Temp_inst))
    eact_inst = RH_inst*0.01*esat_inst
    u_starr = (0.41*u_200)/np.log(200/z0m)	
    rah = 0.8 * np.log(2/(0.1*z0m))/(0.41*u_starr)		 # 0.8 to skip the iteration part and just assume 80% lower rah end of iteration			
    rs = 1/(vegt_cover/rs_min+(1-vegt_cover)/rsoil_min)
    vpd = esat_inst - eact_inst
    sa_slope = (4098*esat_inst)/np.power(Temp_inst + 237.3, 2)    
    LE = (sa_slope*(rn_inst-g_inst)+air_dens*1004*vpd/rah)/(sa_slope+Psychro_c*(1+rs/rah))
    Aearodynamic_res_SEBAL =np.log(2/0.01)/(0.41*u_starr)								
    H = rn_inst-g_inst-LE		 					
    dT = (H*Aearodynamic_res_SEBAL)/(air_dens*1004)
    T0_DEM = dT + Temp_inst		
    T0_DEM += 273.15
									
    return(dT, T0_DEM)	

#------------------------------------------------------------------------------   
def reproject_MODIS(input_name, output_name, epsg_to):   
    
    '''
    Reproject the merged data file
	
    Keywords arguments:
    output_folder -- 'C:/file/to/path/'
    '''                    
    
    # Get environmental variable
    SEBAL_env_paths = os.environ["SEBAL"].split(';')
    GDAL_env_path = SEBAL_env_paths[0]
    GDALWARP_PATH = os.path.join(GDAL_env_path, 'gdalwarp.exe')

    split_input = input_name.split('hdf":')
    inputname = '%shdf":"%s"' %(split_input[0],split_input[1])

    # find path to the executable
    fullCmd = ' '.join(["%s" %(GDALWARP_PATH), '-overwrite -s_srs "+proj=sinu +lon_0=0 +x_0=0 +y_0=0 +a=6371007.181 +b=6371007.181 +units=m +no_defs"', '-t_srs EPSG:%s -of GTiff' %(epsg_to), inputname, output_name])   
    Run_command_window(fullCmd)
				
    return()  

#------------------------------------------------------------------------------   				
def Open_reprojected_hdf(input_name, Band, epsg_to, scale_factor, proyDEM_fileName):

    g=gdal.Open(input_name, gdal.GA_ReadOnly)
    
    folder_out = os.path.dirname(input_name)
    output_name_temp = os.path.join(folder_out, "temporary.tif")
    
    # Open and reproject           
    name_in = g.GetSubDatasets()[Band][0]

    reproject_MODIS(name_in, output_name_temp, epsg_to)

    dest, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(output_name_temp, proyDEM_fileName)
    Array = dest.GetRasterBand(1).ReadAsArray() * scale_factor

    os.remove(output_name_temp)
                              
    return(Array)  

#------------------------------------------------------------------------------   	

def Modis_Time(src_FileName_LST, epsg_to, proyDEM_fileName):
      
    Time = Open_reprojected_hdf(src_FileName_LST, 2, epsg_to, 0.1, proyDEM_fileName)

    hour = np.floor(Time)
    hour[hour == 0] = np.nan
    minutes = (Time - hour) * 60
    
    return(hour, minutes) 

#------------------------------------------------------------------------------  

def Run_command_window(argument):
    """
    This function runs the argument in the command window without showing cmd window

    Keyword Arguments:
    argument -- string, name of the adf file
    """  
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW		
    
    process = subprocess.Popen(argument, startupinfo=startupinfo, stderr=subprocess.PIPE, stdout=subprocess.PIPE)
    process.wait()  
    
    return()
