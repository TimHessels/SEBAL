# -*- coding: utf-8 -*-
"""
Created on Thu Sep 08 15:09:49 2016

@author: tih
"""
import numpy as np
import os
import scipy.interpolate
import gdal
from openpyxl import load_workbook
import osr
from datetime import datetime, timedelta
import pandas as pd
import shutil
import glob
from netCDF4 import Dataset
import warnings

import SEBAL.pySEBAL.pySEBAL_code as SBF

def main():
       
####################################################################################################################   
############################################# CREATE INPUT FOR SEBAL RUN ###########################################		
####################################################################################################################
    print 'preSEBAL version 1.0 Github'
####################################################################################################################   
##################################################### PreHANTS ####################################################		
####################################################################################################################

# PreHANTS
# Part 1: Define input by user 
# Part 2: Set parameters and output folder 
# Part 3: RUN SEBAL
# Part 4: HANTS
# Part 5: post HANTS
# Part 6: Write output

####################################################################################################################   
################################################# PreHANTS part 1 ##################################################		
####################################################################################################################

    VegetationExcel =r"$HOME\SEBAL\Excel_PreSEBAL_v1_0.xlsx"  # This excel defines the p and c factor and vegetation height.

####################################################################################################################   
################################################# PreHANTS part 2 ##################################################		
####################################################################################################################

    # Open Excel workbook used for Vegetation c and p factor conversions				
    wb_veg = load_workbook(VegetationExcel, data_only=True)		
    ws_veg = wb_veg['General_Input']	
    
    # Input for preSEBAL.py
    start_date = "%s" %str(ws_veg['B2'].value)  
    end_date = "%s" %str(ws_veg['B3'].value)
    inputExcel= r"%s" %str(ws_veg['B4'].value)               # The excel with all the SEBAL input data
    LU_data_FileName = r"%s" %str(ws_veg['B5'].value)       # Path to Land Use map
    output_folder = r"%s" %str(ws_veg['B7'].value)  
    
    # optional paramater 
    DSSF_Folder= r"%s" %str(ws_veg['B6'].value)
    
    ######################## Load Excels ##########################################	
    # Open Excel workbook for SEBAL inputs
    wb = load_workbook(inputExcel)
				
    # Get length of EXCEL sheet
    ws = wb['General_Input']	
    ws2 = wb['VIIRS_PROBAV_Input']	
    endExcel=int(ws.max_row)
    
    # Create Dict
    SEBAL_RUNS = dict()
    
    for number in range(2,endExcel+1):
        input_folder_SEBAL = str(ws['B%d' % number].value)  
        output_folder_SEBAL = str(ws['C%d' % number].value)                    
        Image_Type = int(ws['D%d' % number].value)
        PROBA_V_name  = str(ws2['D%d' % number].value)
        VIIRS_name  = str(ws2['B%d' % number].value) 
        SEBAL_RUNS[number] = {'input_folder': input_folder_SEBAL, 'output_folder': output_folder_SEBAL, 'image_type': Image_Type,'PROBA_V_name': PROBA_V_name,'VIIRS_name': VIIRS_name}
    
    Kind_Of_Runs_Dict = {}
    for k, v in SEBAL_RUNS.iteritems():
        Kind_Of_Runs_Dict.setdefault(v['image_type'], []).append(k)

    ######################## Create output folders  ##########################################	
  
    # Define output
    output_folder_PreSEBAL_SEBAL = os.path.join(output_folder,'PreSEBAL_SEBAL_out') 
    input_folder_HANTS = os.path.join(output_folder,'HANTS_in')   
    output_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_out') 
    temp_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_temp') 
    temp_folder_PreSEBAL_LST = os.path.join(temp_folder_PreSEBAL,'LST')         
    NDVI_outfolder = os.path.join(output_folder_PreSEBAL_SEBAL,'NDVI') 				
    Albedo_outfolder = os.path.join(output_folder_PreSEBAL_SEBAL,'Albedo') 											
    WaterMask_outfolder = os.path.join(output_folder_PreSEBAL_SEBAL,'Water_Mask') 								
    LAI_outfolder = os.path.join(output_folder_PreSEBAL,'LAI') 
    ALBEDO_outfolder_end = os.path.join(output_folder_PreSEBAL,'ALBEDO')        
    NDVI_outfolder_end = os.path.join(output_folder_PreSEBAL,'NDVI')
    WaterMask_outfolder_end = os.path.join(output_folder_PreSEBAL,'Water_Mask') 	     
    TRANS_outfolder = os.path.join(output_folder_PreSEBAL,'Transmissivity') 
    Surface_Temperature_outfolder = os.path.join(output_folder_PreSEBAL_SEBAL,'Surface_Temperature')  
    output_folder_HANTS_end_sharp = os.path.join(output_folder_PreSEBAL, 'LST_Sharpened')
    output_folder_HANTS_end_Veg = os.path.join(output_folder_PreSEBAL, 'Vegetation_Height')
    output_folder_p_factor =  os.path.join(output_folder_PreSEBAL, 'p_factor')
    output_folder_LUE =  os.path.join(output_folder_PreSEBAL, 'LUE')
    
    if not os.path.exists(output_folder_PreSEBAL_SEBAL):
        os.makedirs(output_folder_PreSEBAL_SEBAL)
    if not os.path.exists(output_folder_PreSEBAL):
        os.mkdir(output_folder_PreSEBAL)
    if not os.path.exists(temp_folder_PreSEBAL):
        os.mkdir(temp_folder_PreSEBAL)        
    if not os.path.exists(NDVI_outfolder):
        os.makedirs(NDVI_outfolder)				
    if not os.path.exists(Albedo_outfolder):
        os.makedirs(Albedo_outfolder)     
    if not os.path.exists(WaterMask_outfolder):
        os.makedirs(WaterMask_outfolder)        
    if not os.path.exists(LAI_outfolder):
        os.makedirs(LAI_outfolder)
    if not os.path.exists(ALBEDO_outfolder_end):
        os.makedirs(ALBEDO_outfolder_end)	        
    if not os.path.exists(NDVI_outfolder_end):
        os.makedirs(NDVI_outfolder_end)
    if not os.path.exists(WaterMask_outfolder_end):
        os.makedirs(WaterMask_outfolder_end)		        
    if not os.path.exists(temp_folder_PreSEBAL_LST):
        os.makedirs(temp_folder_PreSEBAL_LST)        
    if not os.path.exists(Surface_Temperature_outfolder):
        os.makedirs(Surface_Temperature_outfolder)
    if not os.path.exists(TRANS_outfolder):
        os.makedirs(TRANS_outfolder)	
    if not os.path.exists(output_folder_HANTS_end_sharp):
        os.mkdir(output_folder_HANTS_end_sharp)
    if not os.path.exists(output_folder_HANTS_end_Veg):
        os.mkdir(output_folder_HANTS_end_Veg)           
    if not os.path.exists(output_folder_p_factor):
        os.mkdir(output_folder_p_factor)     
    if not os.path.exists(output_folder_LUE):
        os.mkdir(output_folder_LUE)       
        
    # Do not show warnings
    warnings.filterwarnings('ignore')  

####################################################################################################################   
################################################### RUN SEBAL part 3 ###############################################	
####################################################################################################################    
    
    ############################## Define General info ############################ 
    for number in Kind_Of_Runs_Dict[2]: # Number defines the column of the inputExcel
        print number
        if not (SEBAL_RUNS[number]['PROBA_V_name'] == 'None' and SEBAL_RUNS[number]['VIIRS_name'] == 'None'):
            Rp = 0.91                        # Path radiance in the 10.4-12.5 µm band (W/m2/sr/µm)
            tau_sky = 0.866                  # Narrow band transmissivity of air, range: [10.4-12.5 µm]
            surf_temp_offset = 3             # Surface temperature offset for water 
            
        ######################## Open General info from SEBAL Excel ################### 
    
            # Open the General_Input sheet			
            ws = wb['General_Input']
        
            # Extract the input and output folder, and Image type from the excel file			
            input_folder = str(ws['B%d' % number].value)                             
            Image_Type = int(2)                              # Type of Image (1=Landsat & 2 = VIIRS & GLOBA-V)     
      
            # Extract the Path to the DEM map from the excel file
            DEM_fileName = '%s' %str(ws['E%d' % number].value) #'DEM_HydroShed_m'  
    
            # Open DEM and create Latitude and longitude files
            lat,lon,lat_fileName,lon_fileName= SBF.DEM_lat_lon(DEM_fileName, temp_folder_PreSEBAL)
    
        ######################## Extract general data for Landsat ##########################################	
            if Image_Type == 1:	
           
                # Open the Landsat_Input sheet				
                ws = wb['Landsat_Input']		
          
                # Extract Landsat name, number and amount of thermal bands from excel file 
                Name_Landsat_Image = str(ws['B%d' % number].value)    # From glovis.usgs.gov
                Landsat_nr = int(ws['C%d' % number].value)            # Type of Landsat (LS) image used (LS5, LS7, or LS8)
                Bands_thermal = int(ws['D%d' %number].value)         # Number of LS bands to use to retrieve land surface 
     
                # Pixel size of the model
                pixel_spacing=int(30) 
                
                # the path to the MTL file of landsat				
                Landsat_meta_fileName = os.path.join(input_folder, '%s_MTL.txt' % Name_Landsat_Image)
                
                # read out the general info out of the MTL file in Greenwich Time
                year, DOY, hour, minutes, UTM_Zone, Sun_elevation = SBF.info_general_metadata(Landsat_meta_fileName) # call definition info_general_metadata
                date=datetime.strptime('%s %s'%(year,DOY), '%Y %j')
                month = date.month
                day = date.day
                
                # define the kind of sensor and resolution of the sensor
                sensor1 = 'L%d' % Landsat_nr
                sensor2 = 'L%d' % Landsat_nr
                sensor3 = 'L%d' % Landsat_nr
                res1 = '30m'
                res2 = '%sm' %int(pixel_spacing)
                res3 = '30m'
                
                # Set the start parameter for determining transmissivity at 0
                Determine_transmissivity = 0
                
        ######################## Extract general data for VIIRS-PROBAV ##########################################	
            if Image_Type == 2:	
                
                # Open the VIIRS_PROBAV_Input sheet					
                ws = wb['VIIRS_PROBAV_Input']
                
                # Extract the name of the thermal and quality VIIRS image from the excel file	
                Name_VIIRS_Image_TB = '%s' %str(ws['B%d' % number].value)
                
                # Extract the name to the PROBA-V image from the excel file	
                Name_PROBAV_Image = '%s' %str(ws['D%d' % number].value)    # Must be a tiff file 
                     
                # Pixel size of the model
                pixel_spacing=int(100) 	
     		
                # UTM Zone of the end results					
                UTM_Zone = float(ws['G%d' % number].value)
    
                if not Name_VIIRS_Image_TB == 'None':
    
                    #Get time from the VIIRS dataset name (IMPORTANT TO KEEP THE TEMPLATE OF THE VIIRS NAME CORRECT example: npp_viirs_i05_20150701_124752_wgs84_fit.tif)
                    Total_Day_VIIRS = Name_VIIRS_Image_TB.split('_')[3]
                    Total_Time_VIIRS = Name_VIIRS_Image_TB.split('_')[4]
        				
                    # Get the information out of the VIIRS name in GMT (Greenwich time)
                    year = int(Total_Day_VIIRS[0:4])
                    month = int(Total_Day_VIIRS[4:6])
                    day = int(Total_Day_VIIRS[6:8])
                    Startdate = '%d-%02d-%02d' % (year,month,day)
                    DOY=datetime.strptime(Startdate,'%Y-%m-%d').timetuple().tm_yday
                    hour = int(Total_Time_VIIRS[0:2])
                    minutes = int(Total_Time_VIIRS[2:4])
                    
                    # If this is runned correctly, we can determine transmissivity
                    ws = wb['Meteo_Input']
                    Field_Radiation_24 = '%s' %str(ws['J%d' % number].value)  
                    Field_Trans_24 = '%s' %str(ws['K%d' % number].value)  
    
                    Determine_transmissivity = 1
    
                    
                # else use PROBA-V day but than no transmissivity can be determined for now
                else:
                   
                    # Get the day and time from the PROBA-V                   
                    Band_PROBAVhdf_fileName = os.path.join(input_folder, '%s.HDF5' % (Name_PROBAV_Image))   
                    g=gdal.Open(Band_PROBAVhdf_fileName, gdal.GA_ReadOnly)
                        
                    Meta_data = g.GetMetadata()
                    Date_PROBAV = str(Meta_data['LEVEL3_RADIOMETRY_BLUE_OBSERVATION_START_DATE'])
                    year = int(Date_PROBAV.split("-")[0])
                    month = int(Date_PROBAV.split("-")[1])
                    day = int(Date_PROBAV.split("-")[2])                        
                    Var_name = '%d%02d%02d' %(year, month, day)  
                    DOY=datetime.strptime(Var_name,'%Y%m%d').timetuple().tm_yday
    
                    # We cannot determine transmissivity                                     
                    Determine_transmissivity = 0
                  
                # Determine the transmissivity if possible (Determine_transmissivity = 1)    
                if Determine_transmissivity == 1:                
                    
                    # Rounded difference of the local time from Greenwich (GMT) (hours):
                    delta_GTM = round(np.sign(lon[int(np.shape(lon)[0]/2), int(np.shape(lon)[1]/2)]) * lon[int(np.shape(lon)[0]/2), int(np.shape(lon)[1]/2)] * 24 / 360)
                    if np.isnan(delta_GTM) == True:
                        delta_GTM = round(np.nanmean(lon) * np.nanmean(lon)  * 24 / 360)
        
                    # Calculate local time 
                    hour += delta_GTM
                    if hour < 0.0:
                        day -= 1
                        hour += 24
                    if hour >= 24:
                        day += 1
                        hour -= 24         
                
                    # define the kind of sensor and resolution of the sensor	
                    sensor1 = 'PROBAV'
                    sensor2 = 'VIIRS'
                    res1 = '375m'
                    res2 = '%sm' %int(pixel_spacing)
                    res3 = '30m'
     
        ######################## Extract general data from DEM file and create Slope map ##########################################	
                
            # Variable date name
            Var_name = '%d%02d%02d' %(year, month, day)
        								
            # Reproject from Geog Coord Syst to UTM -
            # 1) DEM - Original DEM coordinates is Geographic: lat, lon
            dest, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SBF.reproject_dataset(
                           DEM_fileName, pixel_spacing, UTM_Zone=UTM_Zone)
            band = dest.GetRasterBand(1)   # Get the reprojected dem band
            ncol = dest.RasterXSize        # Get the reprojected dem column size
            nrow = dest.RasterYSize        # Get the reprojected dem row size
            shape=[ncol, nrow]
           
            # Read out the DEM band and print the DEM properties
            data_DEM = band.ReadAsArray(0, 0, ncol, nrow)
    
            # 2) Latitude file - reprojection    
            # reproject latitude to the landsat projection and save as tiff file																																
            lat_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SBF.reproject_dataset(
                            lat_fileName, pixel_spacing, UTM_Zone=UTM_Zone)
                        
            # Get the reprojected latitude data															
            lat_proy = lat_rep.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)
         
            # 3) Longitude file - reprojection
            # reproject longitude to the landsat projection	 and save as tiff file	
            lon_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SBF.reproject_dataset(lon_fileName, pixel_spacing, UTM_Zone=UTM_Zone)
    
            # Get the reprojected longitude data	
            lon_proy = lon_rep.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)
            
            lon_fileName = os.path.join(temp_folder_PreSEBAL,'lon_resh.tif')
            SBF.save_GeoTiff_proy(dest, lon_proy, lon_fileName, shape, nband=1)	
    				
            # Calculate slope and aspect from the reprojected DEM
            deg2rad,rad2deg,slope,aspect=SBF.Calc_Gradient(data_DEM, pixel_spacing)
 
    
            if Determine_transmissivity == 1:
                
                # calculate the coz zenith angle
                Ra_mountain_24, Ra_inst, cos_zn_resh, dr, phi, delta = SBF.Calc_Ra_Mountain(lon,DOY,hour,minutes,lon_proy,lat_proy,slope,aspect) 
                cos_zn_fileName = os.path.join(temp_folder_PreSEBAL,'cos_zn.tif')
                SBF.save_GeoTiff_proy(dest, cos_zn_resh, cos_zn_fileName, shape, nband=1)
            
                # Save the Ra
                Ra_inst_fileName = os.path.join(temp_folder_PreSEBAL,'Ra_inst.tif')
                SBF.save_GeoTiff_proy(dest, Ra_inst, Ra_inst_fileName, shape, nband=1)	
                Ra_mountain_24_fileName = os.path.join(temp_folder_PreSEBAL,'Ra_mountain_24.tif')
                SBF.save_GeoTiff_proy(dest, Ra_mountain_24, Ra_mountain_24_fileName, shape, nband=1)	
                
                #################### Calculate Transmissivity ##########################################	           
                # Open the General_Input sheet			
                ws = wb['Meteo_Input']
        
                # Extract the method radiation value	
                Value_Method_Radiation_inst = '%s' %str(ws['L%d' % number].value)
        
                # Values to check if data is created
                Check_Trans_inst = 0
                Check_Trans_24 = 0
                
                '''  This is now turned of, so you need to fill in the instantanious transmissivity or Radiation
                # Extract the data to the method of radiation
                if int(Value_Method_Radiation_inst) == 2:
                    Field_Radiation_inst = '%s' %str(ws['N%d' % number].value)        
                        
                    if Field_Radiation_inst == 'None':
        
                        # Instantanious Transmissivity files must be created
                        Check_Trans_inst = 1
                        
                        # Calculate Transmissivity
                        quarters_hours = np.ceil(minutes/30.) * 30
                        hours_GMT = hour - delta_GTM
                        if quarters_hours >= 60:
                            hours_GMT += 1
                            quarters_hours = 0
                        
                        # Define the instantanious LANDSAF file
                        name_Landsaf_inst = 'HDF5_LSASAF_MSG_DSSF_MSG-Disk_%d%02d%02d%02d%02d.tif' %(year, month,day, hours_GMT, quarters_hours)
                        file_Landsaf_inst = os.path.join(DSSF_Folder,name_Landsaf_inst)  
                                     
                        # Reproject the Ra_inst data to match the LANDSAF data
                        Ra_inst_3Km_dest, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(Ra_inst_fileName, file_Landsaf_inst, method = 1)   
                        Ra_inst_3Km = Ra_inst_3Km_dest.GetRasterBand(1).ReadAsArray()
                        Ra_inst_3Km[Ra_inst_3Km==0] = np.nan
                        
                        # Open the Rs LANDSAF data           
                        dest_Rs_inst_3Km = gdal.Open(file_Landsaf_inst)
                        Rs_inst_3Km = dest_Rs_inst_3Km.GetRasterBand(1).ReadAsArray()
                        Rs_inst_3Km = np.float_(Rs_inst_3Km)/10
                        Rs_inst_3Km[Rs_inst_3Km<0]=np.nan
                        
                        # Get shape LANDSAF data
                        shape_trans=[dest_Rs_inst_3Km.RasterXSize , dest_Rs_inst_3Km.RasterYSize ]
                        
                        # Calculate Transmissivity 3Km
                        Transmissivity_3Km = Rs_inst_3Km/Ra_inst_3Km
                        Transmissivity_3Km_fileName = os.path.join(output_folder_temp,'Transmissivity_3Km.tif')
                        SEBAL.save_GeoTiff_proy(Ra_inst_3Km_dest, Transmissivity_3Km, Transmissivity_3Km_fileName, shape_trans, nband=1)	
                        
                        # Reproject Transmissivity to match DEM (now this is done by using the nearest neighbour method)
                        Transmissivity_inst_dest, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(Transmissivity_3Km_fileName, cos_zn_fileName, method = 3)  
                        Transmissivity_inst = Transmissivity_inst_dest.GetRasterBand(1).ReadAsArray()
                        Transmissivity_inst[Transmissivity_inst>0.98] = 0.98                
                        Transmissivity_inst_fileName = os.path.join(TRANS_outfolder,'Transmissivity_inst_%s.tif' %Var_name)                
                        SEBAL.save_GeoTiff_proy(Transmissivity_inst_dest, Transmissivity_inst, Transmissivity_inst_fileName, shape, nband=1)	
    
                '''                    
                # Extract the method radiation value	
                Value_Method_Radiation_24 = '%s' %str(ws['I%d' % number].value)
        
                # Extract the data to the method of radiation
                if int(Value_Method_Radiation_24) == 2:
                    Field_Radiation_24 = '%s' %str(ws['K%d' % number].value)        
                        
                    if Field_Radiation_24 == 'None':
                        
                        # Daily Transmissivity files must be created
                        Check_Trans_24 = 1
        
                        # Create times that are needed to calculate daily Rs (LANDSAF)
                        Starttime_GMT = datetime.strptime(Startdate,'%Y-%m-%d') + timedelta(hours=-delta_GTM)
                        Endtime_GMT = Starttime_GMT + timedelta(days=1)
                        Times = pd.date_range(Starttime_GMT, Endtime_GMT,freq = '30min')
         
                        for Time in Times[:-1]:
                            year_LANDSAF = Time.year
                            month_LANDSAF = Time.month
                            day_LANDSAF = Time.day
                            hour_LANDSAF = Time.hour
                            min_LANDSAF = Time.minute
         
                            # Define the instantanious LANDSAF file
                            #re = glob.glob('')
                            name_Landsaf_inst = 'HDF5_LSASAF_MSG_DSSF_MSG-Disk_%d%02d%02d%02d%02d.tif' %(year_LANDSAF, month_LANDSAF,day_LANDSAF, hour_LANDSAF, min_LANDSAF)
                            file_Landsaf_inst = os.path.join(DSSF_Folder,name_Landsaf_inst)  
        
                            # Open the Rs LANDSAF data           
                            dest_Rs_inst_3Km = gdal.Open(file_Landsaf_inst)
                            Rs_one_3Km = dest_Rs_inst_3Km.GetRasterBand(1).ReadAsArray()
                            Rs_one_3Km = np.float_(Rs_one_3Km)/10
                            Rs_one_3Km[Rs_one_3Km < 0]=np.nan
                          
                            if Time == Times[0]:
                                Rs_24_3Km_tot = Rs_one_3Km
                            else:
                                Rs_24_3Km_tot += Rs_one_3Km
                        
                        Rs_24_3Km = Rs_24_3Km_tot / len(Times[:-1])  
                         
                        # Reproject the Ra_inst data to match the LANDSAF data
                        Ra_24_3Km_dest, ulx, lry, lrx, uly, epsg_to = SBF.reproject_dataset_example(Ra_mountain_24_fileName, file_Landsaf_inst, method = 3)   
                        Ra_24_3Km = Ra_24_3Km_dest.GetRasterBand(1).ReadAsArray()
                        Ra_24_3Km[Ra_24_3Km==0] = np.nan  
                         
                        # Do gapfilling
                        Ra_24_3Km = gap_filling(Ra_24_3Km,np.nan)
                                 
                        # Get shape LANDSAF data
                        shape_trans=[dest_Rs_inst_3Km.RasterXSize , dest_Rs_inst_3Km.RasterYSize ]
                        
                        # Calculate Transmissivity 3Km
                        Transmissivity_24_3Km = Rs_24_3Km/Ra_24_3Km
                    
                        Transmissivity_24_3Km_fileName = os.path.join(temp_folder_PreSEBAL,'Transmissivity_24_3Km.tif')
                        SBF.save_GeoTiff_proy(Ra_24_3Km_dest, Transmissivity_24_3Km, Transmissivity_24_3Km_fileName, shape_trans, nband=1)	
                        
                        # Reproject Transmissivity to match DEM (now this is done by using the nearest neighbour method)
                        Transmissivity_24_dest, ulx, lry, lrx, uly, epsg_to = SBF.reproject_dataset_example(Transmissivity_24_3Km_fileName, lon_fileName, method = 3)  
                        Transmissivity_24 = Transmissivity_24_dest.GetRasterBand(1).ReadAsArray()
                        Transmissivity_24[Transmissivity_24>0.98] = 0.98
                        Transmissivity_24_fileName = os.path.join(TRANS_outfolder,'Transmissivity_24_%s.tif' %Var_name)                
                        SBF.save_GeoTiff_proy(Transmissivity_24_dest, Transmissivity_24, Transmissivity_24_fileName, shape, nband=1)	
    
        #################### Calculate NDVI for LANDSAT ##########################################	
    
            if Image_Type == 1:	
    
                # Define bands used for each Landsat number
                if Landsat_nr == 5 or Landsat_nr == 7:
                    Bands = np.array([1, 2, 3, 4, 5, 7, 6])
                elif Landsat_nr == 8:
                    Bands = np.array([2, 3, 4, 5, 6, 7, 10, 11])  
                else:
                    print 'Landsat image not supported, use Landsat 7 or 8'
    
                # Open MTL landsat and get the correction parameters   
                Landsat_meta_fileName = os.path.join(input_folder, '%s_MTL.txt' % Name_Landsat_Image)
                Lmin, Lmax, k1_c, k2_c = SBF.info_band_metadata(Landsat_meta_fileName, Bands)
               
                # Mean solar exo-atmospheric irradiance for each band (W/m2/microm)
                # for the different Landsat images (L5, L7, or L8)
                ESUN_L5 = np.array([1983, 1796, 1536, 1031, 220, 83.44])
                ESUN_L7 = np.array([1997, 1812, 1533, 1039, 230.8, 84.9])
                ESUN_L8 = np.array([1973.28, 1842.68, 1565.17, 963.69, 245, 82.106])
        
                # Open one band - To get the metadata of the landsat images only once (to get the extend)
                src_FileName = os.path.join(input_folder, '%s_B2.TIF' % Name_Landsat_Image)  # before 10!
                ls,band_data,ulx,uly,lrx,lry,x_size_ls,y_size_ls = SBF.Get_Extend_Landsat(src_FileName)
             
                # Crop the Landsat images to the DEM extent -
                dst_FileName = os.path.join(temp_folder_PreSEBAL,'cropped_LS_b2.tif')  # Before 10 !!
         							
                # Clip the landsat image to match the DEM map											
                lsc, ulx, lry, lrx, uly, epsg_to = SBF.reproject_dataset_example(src_FileName, lon_fileName)	
                data_LS = lsc.GetRasterBand(1).ReadAsArray()
                SBF.save_GeoTiff_proy(dest, data_LS, dst_FileName, shape, nband=1)	
        
                # Get the extend of the remaining landsat file after clipping based on the DEM file	
                lsc,band_data,ulx,uly,lrx,lry,x_size_lsc,y_size_lsc = SBF.Get_Extend_Landsat(dst_FileName)
    
                # Create the corrected signals of Landsat in 1 array
                Reflect = SBF.Landsat_Reflect(Bands,input_folder,Name_Landsat_Image,output_folder,shape,Lmax,Lmin,ESUN_L5,ESUN_L7,ESUN_L8,cos_zn_resh,dr,Landsat_nr, cos_zn_fileName)
    
                # Calculate temporal water mask
                water_mask_temp=SBF.Water_Mask(shape,Reflect)       
    
                # Calculate NDVI
                NDVI = SBF.Calc_NDVI(Reflect)
    
                # Calculate albedo
                albedo = SBF.Calc_albedo(Reflect)
    
                # Save NDVI
                NDVI_FileName = os.path.join(NDVI_outfolder,'NDVI_LS_%s.tif'%Var_name)
                SBF.save_GeoTiff_proy(dest, NDVI, NDVI_FileName, shape, nband=1)
    
                # Save albedo								
                albedo_FileName = os.path.join(Albedo_outfolder,'Albedo_LS_%s.tif'%Var_name)
                SBF.save_GeoTiff_proy(dest, albedo, albedo_FileName, shape, nband=1)
    
        ################### Extract Meteo data for Landsat days from SEBAL Excel ##################
    
            # Open the Meteo_Input sheet	
            ws = wb['Meteo_Input']	
            # ---------------------------- Instantaneous Air Temperature ------------
            # Open meteo data, first try to open as value, otherwise as string (path)	  
            try:
               Temp_inst = float(ws['B%d' %number].value)                # Instantaneous Air Temperature (°C)
    
            # if the data is not a value, than open as a string	
            except:
                Temp_inst_name = '%s' %str(ws['B%d' %number].value) 
                Temp_inst_fileName = os.path.join(output_folder, 'Temp', 'Temp_inst_input.tif')
                Temp_inst = SBF.Reshape_Reproject_Input_data(Temp_inst_name, Temp_inst_fileName, lon_fileName)
    
            try:
                RH_inst = float(ws['D%d' %number].value)                # Instantaneous Relative humidity (%)
     
            # if the data is not a value, than open as a string							
            except:
                RH_inst_name = '%s' %str(ws['D%d' %number].value) 
                RH_inst_fileName = os.path.join(output_folder, 'Temp', 'RH_inst_input.tif')
                RH_inst = SBF.Reshape_Reproject_Input_data(RH_inst_name, RH_inst_fileName, lon_fileName)			
             
            esat_inst = 0.6108 * np.exp(17.27 * Temp_inst / (Temp_inst + 237.3)) 													
            eact_inst = RH_inst * esat_inst / 100
            
         #################### Calculate NDVI for VIIRS-PROBAV ##########################################	
         
            if Image_Type == 2:	
                
                if Name_PROBAV_Image == 'None':
                    
                    offset_all = [-1, 1, -2, 2, -3, 3,-4, 4,-5 ,5 ,-6 , 6, -7, 7, -8, 8]
                    found_Name_PROBAV_Image = 0
                    for offset in offset_all:
                        
                        if found_Name_PROBAV_Image == 1:
                            continue
                        else:
                            try:
                                Name_PROBAV_Image = SEBAL_RUNS[number + offset]['PROBA_V_name']
                                if not Name_PROBAV_Image == 'None':
                                    found_Name_PROBAV_Image = 1  
                            except:
                                pass
 
                    # Get the day and time from the PROBA-V                   
                    Band_PROBAVhdf_fileName = os.path.join(input_folder, '%s.HDF5' % (Name_PROBAV_Image))   
                    g=gdal.Open(Band_PROBAVhdf_fileName, gdal.GA_ReadOnly)
                        
                    Meta_data = g.GetMetadata()
                    Date_PROBAV = str(Meta_data['LEVEL3_RADIOMETRY_BLUE_OBSERVATION_START_DATE'])
                    year = int(Date_PROBAV.split("-")[0])
                    month = int(Date_PROBAV.split("-")[1])
                    day = int(Date_PROBAV.split("-")[2])                        
                    Var_name_2 = '%d%02d%02d' %(year, month, day)  
                     
                    # Define the output name
                    NDVI_FileName = os.path.join(NDVI_outfolder,'NDVI_PROBAV_%s.tif' %Var_name_2)              
                    Albedo_FileName = os.path.join(Albedo_outfolder, 'Albedo_PROBAV_%s.tif' %Var_name_2) 	
                    water_mask_temp_FileName = os.path.join(WaterMask_outfolder, 'Water_Mask_PROBAV_%s.tif' %Var_name_2) 	
    
                else: 
                    NDVI_FileName = os.path.join(NDVI_outfolder,'NDVI_PROBAV_%s.tif' %Var_name)
                    Albedo_FileName = os.path.join(Albedo_outfolder, 'Albedo_PROBAV_%s.tif' %Var_name) 	
                    water_mask_temp_FileName = os.path.join(WaterMask_outfolder, 'Water_Mask_PROBAV_%s.tif' %Var_name) 	
    
                # vegetation maps that will be generated
               

                if not os.path.exists(NDVI_FileName):
             
                    # Define the bands that will be used
                    bands=['SM', 'B1', 'B2', 'B3', 'B4']  #'SM', 'BLUE', 'RED', 'NIR', 'SWIR'
        
                    # Set the index number at 0
                    index=0
        							
                    # create a zero array with the shape of the reprojected DEM file						
                    data_PROBAV=np.zeros((shape[1], shape[0]))
                    spectral_reflectance_PROBAV=np.zeros([shape[1], shape[0], 5])
               
                    # constants
                    n188_float=248       # Now it is 248, but we do not exactly know what this really means and if this is for constant for all images.
           
                    # write the data one by one to the spectral_reflectance_PROBAV       
                    for bandnmr in bands:
        
                        # Translate the PROBA-V names to the Landsat band names                
                        Band_number = {'SM':7,'B1':8,'B2':10,'B3':9,'B4':11}
        
                        # Open the dataset 
                        Band_PROBAVhdf_fileName = os.path.join(input_folder, '%s.HDF5' % (Name_PROBAV_Image))   
                        g=gdal.Open(Band_PROBAVhdf_fileName, gdal.GA_ReadOnly)
                        
                        # define data if it is not there yet
                        if not 'Var_name' in locals():
                            Meta_data = g.GetMetadata()
                            Date_PROBAV = str(Meta_data['LEVEL3_RADIOMETRY_BLUE_OBSERVATION_START_DATE'])
                            year = int(Date_PROBAV.split("-")[0])
                            month = int(Date_PROBAV.split("-")[0])
                            day = int(Date_PROBAV.split("-")[0])                        
                            Var_name = '%d%02d%02d' %(year, month, day)
        
                        
                        # Open the .hdf file              
                        name_out = os.path.join(input_folder, '%s_test.tif' % (Name_PROBAV_Image))   
                        name_in = g.GetSubDatasets()[Band_number[bandnmr]][0]
                        
                        # Get environmental variable
                        SEBAL_env_paths = os.environ["SEBAL"].split(';')
                        GDAL_env_path = SEBAL_env_paths[0]
                        GDAL_TRANSLATE = os.path.join(GDAL_env_path, 'gdal_translate.exe')
         
                        # run gdal translate command               
                        FullCmd = '%s -of GTiff %s %s' %(GDAL_TRANSLATE, name_in, name_out)            
                        SBF.Run_command_window(FullCmd)
                        
                        # Open data
                        dest_PV = gdal.Open(name_out)
                        Data = dest_PV.GetRasterBand(1).ReadAsArray()               
                        dest_PV = None
                        
                        # Remove temporary file
                        os.remove(name_out)
         
                        # Define the x and y spacing
                        Meta_data = g.GetMetadata()
                        Lat_Bottom = float(Meta_data['LEVEL3_GEOMETRY_BOTTOM_LEFT_LATITUDE'])
                        Lat_Top = float(Meta_data['LEVEL3_GEOMETRY_TOP_RIGHT_LATITUDE'])
                        Lon_Left = float(Meta_data['LEVEL3_GEOMETRY_BOTTOM_LEFT_LONGITUDE'])
                        Lon_Right = float(Meta_data['LEVEL3_GEOMETRY_TOP_RIGHT_LONGITUDE'])           
                        Pixel_size = float((Meta_data['LEVEL3_GEOMETRY_VNIR_VAA_MAPPING']).split(' ')[-3])
                        
                        # Define the georeference of the PROBA-V data
                        geo_PROBAV=[Lon_Left-0.5*Pixel_size, Pixel_size, 0, Lat_Top+0.5*Pixel_size, 0, -Pixel_size] #0.000992063492063
        		 									
                        # Define the name of the output file	
                        PROBAV_data_name=os.path.join(input_folder, '%s_%s.tif' % (Name_PROBAV_Image,bandnmr)) 	 																		
                        dst_fileName=os.path.join(input_folder, PROBAV_data_name)
                        
                        # create gtiff output with the PROBA-V band
                        fmt = 'GTiff'
                        driver = gdal.GetDriverByName(fmt)
        
                        dst_dataset = driver.Create(dst_fileName, int(Data.shape[1]), int(Data.shape[0]), 1,gdal.GDT_Float32)
                        dst_dataset.SetGeoTransform(geo_PROBAV)
                    
                        # set the reference info
                        srs = osr.SpatialReference()
                        srs.SetWellKnownGeogCS("WGS84")
                        dst_dataset.SetProjection(srs.ExportToWkt())
                   
                        # write the array in the geotiff band
                        dst_dataset.GetRasterBand(1).WriteArray(Data)
                        dst_dataset = None
         
                        # Open the PROBA-V band in SEBAL											
                        g=gdal.Open(PROBAV_data_name.replace("\\","/")) 
        	 										
                        # If the data cannot be opened, change the extension											
                        if g is None:
                            PROBAV_data_name=os.path.join(input_folder, '%s_%s.tiff' % (Name_PROBAV_Image,bandnmr))  
          
                        # Reproject the PROBA-V band  to match DEM's resolution          
                        PROBAV, ulx, lry, lrx, uly, epsg_to = SBF.reproject_dataset_example(
                                          PROBAV_data_name, lon_fileName)
        
                        # Open the reprojected PROBA-V band data                         
                        data_PROBAV_DN = PROBAV.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)
        	 										
                        # Define the filename to store the cropped Landsat image
                        dst_FileName = os.path.join(output_folder, 'Output_PROBAV','proy_PROBAV_%s.tif' % bandnmr)
        		 									
                        # close the PROBA-V 											
                        g=None
                                           
                        # If the band data is not SM change the DN values into PROBA-V values and write into the spectral_reflectance_PROBAV                      
                        if bandnmr is not 'SM':  
                            data_PROBAV[:, :]=data_PROBAV_DN/2000                           
                            spectral_reflectance_PROBAV[:, :, index]=data_PROBAV[:, :]
                        
                        # If the band data is the SM band than write the data into the spectral_reflectance_PROBAV  and create cloud mask            
                        else:
                            data_PROBAV[:, :]=data_PROBAV_DN
                            Cloud_Mask_PROBAV=np.zeros((shape[1], shape[0]))
                            Cloud_Mask_PROBAV[data_PROBAV[:,:]!=n188_float]=1
                            spectral_reflectance_PROBAV[:, :, index]=Cloud_Mask_PROBAV
          
                        # Change the spectral reflectance to meet certain limits                               
                        spectral_reflectance_PROBAV[:, :, index]=np.where(spectral_reflectance_PROBAV[:, :, index]<=0,np.nan,spectral_reflectance_PROBAV[:, :, index])   
                        spectral_reflectance_PROBAV[:, :, index]=np.where(spectral_reflectance_PROBAV[:, :, index]>=150,np.nan,spectral_reflectance_PROBAV[:, :, index])   
          										
                        # Go to the next index 									
                        index=index+1
         
                    # Bands in PROBAV spectral reflectance
                    # 0 = MS
                    # 1 = BLUE
                    # 2 = NIR
                    # 3 = RED
                    # 4 = SWIR
            
                    # Calculate surface albedo based on PROBA-V
                    Surface_Albedo_PROBAV = 0.219 * spectral_reflectance_PROBAV[:, :, 1] + 0.361 * spectral_reflectance_PROBAV[:, :, 2] + 0.379 * spectral_reflectance_PROBAV[:, :, 3] + 0.041 * spectral_reflectance_PROBAV[:, :, 4]
                        
                    # Create Water mask based on PROBA-V             
                    water_mask_temp = np.zeros((shape[1], shape[0])) 
                    water_mask_temp[np.logical_and(spectral_reflectance_PROBAV[:, :, 2] >= spectral_reflectance_PROBAV[:, :, 3],data_DEM>0)]=1
       
                    # Calculate the NDVI based on PROBA-V     
                    n218_memory = spectral_reflectance_PROBAV[:, :, 2] + spectral_reflectance_PROBAV[:, :, 3]
                    NDVI = np.zeros((shape[1], shape[0]))
                    NDVI[n218_memory != 0] =  ( spectral_reflectance_PROBAV[:, :, 3][n218_memory != 0] - spectral_reflectance_PROBAV[:, :, 2][n218_memory != 0] )/ ( spectral_reflectance_PROBAV[:, :, 2][n218_memory != 0] + spectral_reflectance_PROBAV[:, :, 3][n218_memory != 0] )
    
                    # Save Albedo for PROBA-V		
                    SBF.save_GeoTiff_proy(dest, Surface_Albedo_PROBAV, Albedo_FileName, shape, nband=1)	  
        
                    # Save NDVI for PROBA-V
                    SBF.save_GeoTiff_proy(dest, NDVI, NDVI_FileName, shape, nband=1)														

                    # Save Water Mask for PROBA-V	        
                    SBF.save_GeoTiff_proy(dest, water_mask_temp, water_mask_temp_FileName, shape, nband=1)	  
  
                else:
                    
                    dest_NDVI = gdal.Open(NDVI_FileName)
                    dest_water_mask_temp = gdal.Open(water_mask_temp_FileName)
                    NDVI = dest_NDVI.GetRasterBand(1).ReadAsArray()
                    water_mask_temp = dest_water_mask_temp.GetRasterBand(1).ReadAsArray()

             ############################ Calculate LAI ##########################################	
        
            # Calculate the LAI
            FPAR,tir_emis,Nitrogen,vegt_cover,LAI,b10_emissivity = SBF.Calc_vegt_para(NDVI,water_mask_temp,shape)
    
            # Create LAI name
            if Image_Type == 1:
                LAI_FileName = os.path.join(LAI_outfolder,'LAI_LS_%s.tif' %Var_name)
                SBF.save_GeoTiff_proy(dest, LAI, LAI_FileName, shape, nband=1)	
            
         #################### Calculate thermal for Landsat ##########################################	
    
            if Image_Type == 1:
    				
                # Calculate thermal				
                therm_data = SBF.Landsat_therm_data(Bands,input_folder,Name_Landsat_Image,output_folder,ulx_dem,lry_dem,lrx_dem,uly_dem,shape)          
    
                # Calculate surface temperature
                Surface_temp=SBF.Calc_surface_water_temp(Temp_inst,Landsat_nr,Lmax,Lmin,therm_data,b10_emissivity,k1_c,k2_c,eact_inst,shape,water_mask_temp,Bands_thermal,Rp,tau_sky,surf_temp_offset,Image_Type)
    
                # Save surface temperature
                therm_data_FileName = os.path.join(Surface_Temperature_outfolder,'Surface_Temperature_LS_%s.tif' %Var_name)
                SBF.save_GeoTiff_proy(dest, Surface_temp, therm_data_FileName, shape, nband=1)	
    
    
        ################################## Calculate VIIRS surface temperature ########################
    
            if Image_Type == 2:
     
               # If there is VIIRS data
               if not Name_VIIRS_Image_TB == 'None':
    			
                    # Define the VIIRS thermal data name
                    VIIRS_data_name=os.path.join(input_folder, '%s' % (Name_VIIRS_Image_TB))
    			
                    # Reproject VIIRS thermal data								
                    VIIRS, ulx, lry, lrx, uly, epsg_to = SBF.reproject_dataset_example(VIIRS_data_name, lon_fileName)
        																
                    # Open VIIRS thermal data																		
                    data_VIIRS = VIIRS.GetRasterBand(1).ReadAsArray()    
        				
                    # Define the thermal VIIRS output name
                    proyVIIRS_fileName = os.path.join(temp_folder_PreSEBAL, 'Surface_Temp_VIIRS_%s.tif' %Var_name)
        	 											
                    # Save the thermal VIIRS data 												
                    SBF.save_GeoTiff_proy(dest, data_VIIRS, proyVIIRS_fileName, shape, nband=1)	
        
                    # Set the conditions for the brightness temperature (100m)
                    brightness_temp=np.where(data_VIIRS>=250, data_VIIRS, np.nan)
         
                    # Constants
                    k1=606.399172
                    k2=1258.78
                    L_lambda_b10_100=((2*6.63e-34*(3.0e8)**2)/((11.45e-6)**5*(np.exp((6.63e-34*3e8)/(1.38e-23*(11.45e-6)*brightness_temp))-1)))*1e-6
                 
                    # Get Temperature for 100 and 375m resolution
                    Temp_TOA_100 = SBF.Get_Thermal(L_lambda_b10_100,Rp,Temp_inst,tau_sky,tir_emis,k1,k2) 
               
                    # Conditions for surface temperature (100m)
                    n120_surface_temp=Temp_TOA_100.clip(250, 450)
        
                    # Save the surface temperature of the VIIRS in 100m resolution
                    temp_surface_100_fileName_beforeTS = os.path.join(Surface_Temperature_outfolder,'Surface_Temperature_VIIRS_%s.tif' %Var_name)
                    SBF.save_GeoTiff_proy(dest, n120_surface_temp, temp_surface_100_fileName_beforeTS, shape, nband=1)     
     

###################################################################################################################
################################################### HANTS part 4 ##################################################
###################################################################################################################    

    # Select files for PROBA-V that needs to be used (sometimes a composite product is used)
    PROBA_V_Dict = {}
    for k, v in SEBAL_RUNS.iteritems():
        if str(v['PROBA_V_name']) != 'None':
            PROBA_V_Dict.setdefault(v['PROBA_V_name'], []).append(k)
        
    
    Amount_Unique_PROBA_V_images = len(PROBA_V_Dict.keys())
    Back_names = []
    
    # Define HANTS PROBA-V variables
    VARS = ["NDVI", "Albedo"]

    for VAR in VARS:
        
        output_folder_preprocessing_VAR = os.path.join(output_folder_PreSEBAL_SEBAL, VAR)
        os.chdir(output_folder_preprocessing_VAR)
        
        for PROBA_V_image in PROBA_V_Dict.keys():
            Band_PROBAVhdf_fileName = os.path.join(input_folder_SEBAL, '%s.HDF5' % (PROBA_V_image))   
            g=gdal.Open(Band_PROBAVhdf_fileName, gdal.GA_ReadOnly)
                
            Meta_data = g.GetMetadata()
            Date_PROBAV = str(Meta_data['LEVEL3_RADIOMETRY_BLUE_OBSERVATION_START_DATE'])
            year = int(Date_PROBAV.split("-")[0])
            month = int(Date_PROBAV.split("-")[1])
            day = int(Date_PROBAV.split("-")[2])         
            
            Back_name = '%s_PROBAV_%d%02d%02d.tif' %(VAR, year, month, day)

            # Create HANTS input NDVI
            input_folder_HANTS_VAR = os.path.join(temp_folder_PreSEBAL, VAR)
            if not os.path.exists(input_folder_HANTS_VAR):
                os.mkdir(input_folder_HANTS_VAR)
                   
            shutil.copy(os.path.join(output_folder_preprocessing_VAR,Back_name),os.path.join(input_folder_HANTS_VAR,Back_name))

    # VIIRS parameter copy
    VIIRS_Dict = {}
    for k, v in SEBAL_RUNS.iteritems():
        if str(v['VIIRS_name']) != 'None':
            VIIRS_Dict.setdefault(v['VIIRS_name'], []).append(k)

    THERM = 'Surface_Temperature'
    output_folder_preprocessing_THERM = os.path.join(output_folder_PreSEBAL_SEBAL, THERM)
    
    for VIIRS_image in VIIRS_Dict.keys():
        
        try:
            Date_VIIRS = (VIIRS_image.split("d")[1])
            year = int(Date_VIIRS.split("-")[0][0:4])
            month = int(Date_VIIRS.split("-")[0][4:6])
            day = int(Date_VIIRS.split("-")[0][6:8])         
        
        except:
            Date_VIIRS = (VIIRS_image.split("_")[3])
            year = int(Date_VIIRS.split("-")[0][0:4])
            month = int(Date_VIIRS.split("-")[0][4:6])
            day = int(Date_VIIRS.split("-")[0][6:8])               

        Back_name_TB = '%s_VIIRS_%d%02d%02d.tif' %(THERM, year, month, day)

        # Create HANTS input NDVI
        input_folder_HANTS_THERM = os.path.join(temp_folder_PreSEBAL, THERM)
        if not os.path.exists(input_folder_HANTS_THERM):
            os.mkdir(input_folder_HANTS_THERM)
               
        shutil.copy(os.path.join(output_folder_preprocessing_THERM,Back_name_TB),os.path.join(input_folder_HANTS_THERM,Back_name_TB))

    
    ############################################ Solve shift in PROBA=V ##############################################    

    VAR = 'Albedo'
    os.chdir(os.path.join(temp_folder_PreSEBAL, VAR))
    re = glob.glob('%s*.tif' %(VAR))
    i = 0
    
    while i < int(len(re)-1):

        filename1 = re[0]                  # maak hier misschien later van dat alleen 0 word genomen als de hoeveelheid pixels minder dan 40% van totaal is
        filename2 = re[i + 1]
        
        dest1 = gdal.Open(filename1)
        dest2 = gdal.Open(filename2)
    
        Array1 = dest1.GetRasterBand(1).ReadAsArray().flatten()
        Array2 = dest2.GetRasterBand(1).ReadAsArray().flatten()
        Array3 = dest1.GetRasterBand(1).ReadAsArray()[1:,:].flatten()
        Array4 = dest2.GetRasterBand(1).ReadAsArray()[:-1,:].flatten() 
        
        Array1_flat = Array1[np.logical_and(~np.isnan(Array1),~np.isnan(Array2))]
        Array2_flat = Array2[np.logical_and(~np.isnan(Array1),~np.isnan(Array2))]
        Array3_flat = Array3[np.logical_and(~np.isnan(Array3),~np.isnan(Array4))]
        Array4_flat = Array4[np.logical_and(~np.isnan(Array3),~np.isnan(Array4))]
        
        Corr = np.corrcoef(Array1_flat,Array2_flat)[0,1]
        Corr2 = np.corrcoef(Array3_flat,Array4_flat)[0,1] 
        
        if Corr2 > Corr:
            x,y = dest1.GetRasterBand(1).ReadAsArray().shape
                                     
            for VAR_check in VARS:
                os.chdir(os.path.join(temp_folder_PreSEBAL, VAR_check))
                endname = filename2.split('_')[-1]
                re_vars = glob.glob('%s*_%s' %(VAR_check,endname))
                filename3 = re_vars[0]
                dest3 = gdal.Open(filename3)
                New_Array = np.ones(dest1.GetRasterBand(1).ReadAsArray().shape) * np.nan
                New_Array[1:,:] = dest3.GetRasterBand(1).ReadAsArray()[:-1,:]
                filename_out = os.path.join(temp_folder_PreSEBAL, VAR_check, filename2)
                SBF.save_GeoTiff_proy(dest3, New_Array, filename_out, [int(y),int(x)], nband=1)	         
            
        i += 1    
        
    ################################################### General HANTS ###############################################

    # Open one image
    PROBA_V_IMAGE = os.path.join(input_folder_HANTS_VAR,Back_name)
    destPROBAV = gdal.Open(PROBA_V_IMAGE)
    VIIRS_IMAGE = os.path.join(input_folder_HANTS_THERM,Back_name_TB)
    destVIIRS = gdal.Open(VIIRS_IMAGE)    
    
    # Get Geotransform
    Geo_PROBAV = destPROBAV.GetGeoTransform()
    x_size_PROBAV = destPROBAV.RasterXSize
    y_size_PROBAV = destPROBAV.RasterYSize
    Geo_VIIRS = destVIIRS.GetGeoTransform()
    x_size_VIIRS = destVIIRS.RasterXSize
    y_size_VIIRS = destVIIRS.RasterYSize
    
    # Get projection
    proj = Get_epsg(destPROBAV)
    projVIIRS = Get_epsg(destVIIRS)
            
    # Data parameters
    latlim = [Geo_PROBAV[3] + y_size_PROBAV * Geo_PROBAV[5],Geo_PROBAV[3]]
    lonlim = [Geo_PROBAV[0], Geo_PROBAV[0] + x_size_PROBAV * Geo_PROBAV[1]]
    cellsize = Geo_PROBAV[1]
    latlimVIIRS = [Geo_VIIRS [3] + y_size_VIIRS  * Geo_VIIRS [5],Geo_VIIRS [3]]
    lonlimVIIRS  = [Geo_VIIRS [0], Geo_VIIRS [0] + x_size_VIIRS  * Geo_VIIRS [1]]
    cellsizeVIIRS  = Geo_VIIRS [1]    

    # Get the HANTS parameters
    ws_para = wb_veg['HANTS_Input']	

    ###################################################### HANTS Thermal ###############################################
    
    # Define parameters for the NDVI
    THERM = 'Surface_Temperature'
    
    # Define paths for NDVI
    input_folder_HANTS_THERM = os.path.join(temp_folder_PreSEBAL, THERM)
    name_format = '%s_VIIRS_{0}.tif' %THERM
    nc_path_TB = os.path.join(input_folder_HANTS_THERM,'%s_NC.nc' %THERM)
    
    # Create Output folder
    rasters_path_out = os.path.join(temp_folder_PreSEBAL, THERM)
    if not os.path.exists(rasters_path_out):
        os.mkdir(rasters_path_out)
        
    # HANTS parameters for NDVI
    Dates = pd.date_range(start_date, end_date, freq = 'D')
    nb = int(len(Dates)) # nr of images
    nf = int(ws_para['D2'].value)   # number of frequencies to be considered above the zero frequency
    low = float(ws_para['D3'].value) # valid range minimum
    high = float(ws_para['D4'].value) # valid range maximum
    HiLo = str(ws_para['D5'].value) # 2-character string indicating rejection of high or low outliers
    fet = float(ws_para['D6'].value)  # fit error tolerance (point eviating more than fet from curve fit are rejected)
    delta = float(ws_para['D7'].value) # small positive number e.g. 0.1 to supress high amplitudes
    dod = float(ws_para['D8'].value) # degree of overdeterminedness (iteration stops if number of points reaches the minimum required for curve fitting, plus dod). This is a safety measure

    from hants import wa_gdal 
    # Run
    wa_gdal.run_HANTS(input_folder_HANTS_THERM, name_format,
                      start_date, end_date, latlimVIIRS, lonlimVIIRS, cellsizeVIIRS, nc_path_TB,
                      nb, nf, HiLo, low, high, fet, dod, delta,
                      projVIIRS, -9999.0, rasters_path_out, export_hants_only=True)
    
    ###################################################### HANTS NDVI ###############################################
    
    # Define parameters for the NDVI
    VAR = 'NDVI'
    
    # Define paths for NDVI
    input_folder_HANTS_VAR = os.path.join(temp_folder_PreSEBAL, VAR)
    name_format = '%s_PROBAV_{0}.tif' %VAR
    nc_path_ndvi = os.path.join(input_folder_HANTS_VAR,'%s_NC.nc' %VAR)
    
    # Create Output folder
    rasters_path_out = os.path.join(temp_folder_PreSEBAL, VAR + "_HANTS")
    if not os.path.exists(rasters_path_out):
        os.mkdir(rasters_path_out)
        
    # HANTS parameters for NDVI
    Dates = pd.date_range(start_date, end_date, freq = 'D')
    nb = int(len(Dates)) # nr of images
    nf = int(ws_para['C2'].value)   # number of frequencies to be considered above the zero frequency
    low = float(ws_para['C3'].value) # valid range minimum
    high = float(ws_para['C4'].value) # valid range maximum
    HiLo = str(ws_para['C5'].value) # 2-character string indicating rejection of high or low outliers
    fet = float(ws_para['C6'].value)  # fit error tolerance (point eviating more than fet from curve fit are rejected)
    delta = float(ws_para['C7'].value) # small positive number e.g. 0.1 to supress high amplitudes
    dod = float(ws_para['C8'].value) # degree of overdeterminedness (iteration stops if number of points reaches the minimum required for curve fitting, plus dod). This is a safety measure

    from hants import wa_gdal 
    # Run
    wa_gdal.run_HANTS(input_folder_HANTS_VAR, name_format,
                      start_date, end_date, latlim, lonlim, cellsize, nc_path_ndvi,
                      nb, nf, HiLo, low, high, fet, dod, delta,
                      proj, -9999.0, rasters_path_out, export_hants_only=True)

    ###################################################### HANTS Albedo ##############################################
    
    # Define parameters for the albedo
    VAR = 'Albedo'
    
    # Define paths for NDVI
    input_folder_HANTS_VAR = os.path.join(temp_folder_PreSEBAL, VAR + "_HANTS")
    name_format = '%s_PROBAV_{0}.tif' %VAR
    nc_path_albedo = os.path.join(temp_folder_PreSEBAL,'%s_NC.nc' %VAR)
    
    # Create Output folder
    rasters_path_out = os.path.join(output_folder_PreSEBAL, VAR)
    if not os.path.exists(rasters_path_out):
        os.mkdir(rasters_path_out)
        
    # HANTS parameters for NDVI
    Dates = pd.date_range(start_date, end_date, freq = 'D')
    nb = int(len(Dates)) # nr of images
    nf = int(ws_para['B2'].value)   # number of frequencies to be considered above the zero frequency
    low = float(ws_para['B3'].value) # valid range minimum
    high = float(ws_para['B4'].value) # valid range maximum
    HiLo = str(ws_para['B5'].value) # 2-character string indicating rejection of high or low outliers
    fet = float(ws_para['B6'].value)  # fit error tolerance (point eviating more than fet from curve fit are rejected)
    delta = float(ws_para['B7'].value) # small positive number e.g. 0.1 to supress high amplitudes
    dod = float(ws_para['B8'].value) # degree of overdeterminedness (iteration stops if number of points reaches the minimum required for curve fitting, plus dod). This is a safety measure

    from hants import wa_gdal 
    # Run
    wa_gdal.run_HANTS(input_folder_HANTS_VAR, name_format,
                      start_date, end_date, latlim, lonlim, cellsize, nc_path_albedo,
                      nb, nf, HiLo, low, high, fet, dod, delta,
                      proj, -9999.0, rasters_path_out, export_hants_only=True)

###################################################################################################################
################################################### post HANTS part 5 #############################################
###################################################################################################################    


    ############################################# Create Outlier maps for PROBA-V #######################################
     # Create output folder if not exists
    output_folder_HANTS_outliers_PROBAV = os.path.join(temp_folder_PreSEBAL, 'Outliers_PROBAV')
    if not os.path.exists(output_folder_HANTS_outliers_PROBAV):
        os.mkdir(output_folder_HANTS_outliers_PROBAV)

    fh = Dataset(nc_path_albedo, mode='r')
    Var = fh.variables.keys()[-1]
    
    lat = fh.variables[fh.variables.keys()[1]][:]
    lon = fh.variables[fh.variables.keys()[2]][:] 
    time = fh.variables[fh.variables.keys()[3]][:]     
    minimum_lon = np.min(lon)
    maximum_lat = np.max(lat)
    diff_lon = lon[1] - lon[0]
    diff_lat = lat[1] - lat[0]
        
    if not ('shape' in locals() or 'dest' in locals()):
        Example_file = os.path.join(output_folder_preprocessing_VAR, Back_name)                   
        dest = gdal.Open(Example_file)
        ncol = dest.RasterXSize        # Get the reprojected dem column size
        nrow = dest.RasterYSize        # Get the reprojected dem row size
        shape=[ncol, nrow]
        
    for i in range(0,int(np.shape(time)[0])):
        time_now = time[i]
        data = fh.variables['outliers'][i,:,:] 
        geo = tuple([minimum_lon, diff_lon, 0, maximum_lat, 0, diff_lat]) 
        name_out = os.path.join(output_folder_HANTS_outliers_PROBAV, 'Outliers_PROBAV_%s.tif' %time_now)      
        SBF.save_GeoTiff_proy(dest, data, name_out, shape, nband=1)	        
 
    ############################################# Create ALBEDO and NDVI #########################################
        
    # Create the end thermal files date by date        
    for date in Dates:
        
        # Define date
        year = date.year
        month = date.month
        day = date.day
    
        # input filenames needed for creating end thermal file
        filename_outliers = os.path.join(output_folder_HANTS_outliers_PROBAV,"Outliers_PROBAV_%d%02d%02d.tif" %(year,month,day))
        VAR = 'Albedo'
        input_folder_PreSEBAL_ALBEDO = os.path.join(temp_folder_PreSEBAL, VAR + "_HANTS")
        filename_Albedo_original =  os.path.join(Albedo_outfolder, "%s_PROBAV_%d%02d%02d.tif" %(VAR,year,month,day))
        filename_Albedo_HANTS = os.path.join(input_folder_PreSEBAL_ALBEDO, "%s_PROBAV_%d%02d%02d.tif" %(VAR,year,month,day))
        VAR = 'NDVI'
        input_folder_PreSEBAL_NDVI = os.path.join(temp_folder_PreSEBAL, VAR + "_HANTS")
        filename_NDVI_original =  os.path.join(NDVI_outfolder, "%s_PROBAV_%d%02d%02d.tif" %(VAR,year,month,day))
        filename_NDVI_HANTS = os.path.join(input_folder_PreSEBAL_NDVI, "%s_PROBAV_%d%02d%02d.tif" %(VAR,year,month,day))
  
        # Open the input filenames
        dest_outliers = gdal.Open(filename_outliers)
        dest_PROBAV_ALBEDO = gdal.Open(filename_Albedo_original)
        dest_PROBAV_NDVI = gdal.Open(filename_NDVI_original)
        dest_HANTS_ALBEDO = gdal.Open(filename_Albedo_HANTS)
        dest_HANTS_NDVI = gdal.Open(filename_NDVI_HANTS)
        
        # If original exists, this will be the basis for the end thermal map
        if not dest_PROBAV_ALBEDO == None:
                    
            # Open arrays of the input files
            Array_outliers = dest_outliers.GetRasterBand(1).ReadAsArray()[:,:]
            Array_ALBEDO_original = dest_PROBAV_ALBEDO.GetRasterBand(1).ReadAsArray()
            Array_ALBEDO_HANTS = dest_HANTS_ALBEDO.GetRasterBand(1).ReadAsArray()[:,:]
            Array_NDVI_original = dest_PROBAV_NDVI.GetRasterBand(1).ReadAsArray()
            Array_NDVI_HANTS = dest_HANTS_NDVI.GetRasterBand(1).ReadAsArray()[:,:]
            
            # Create outlier Mask
            Array_outliers[Array_outliers==-9999.] = 0
            Array_outliers_mask = np.zeros(np.shape(Array_outliers))
            Array_outliers_mask[Array_outliers==1.]=0   
            Array_outliers_mask[Array_outliers==0.]=1                                         
            Array_outliers_mask[Array_outliers_mask==0]=2
            Array_outliers_mask[Array_outliers_mask==1]=0 
            Array_outliers_mask[Array_outliers_mask==2]=1
                               
            # Create a buffer zone arround the bad pixels                   
            Array_outliers_mask = Create_Buffer(Array_outliers_mask)
            Array_outliers_mask[Array_outliers_mask==1] = 2
            Array_outliers_mask[Array_outliers_mask==0] = 1
            Array_outliers_mask[Array_outliers_mask==2] = 0  
            
            # If there are more than 300 Good pixels                   
            if np.nansum(Array_outliers_mask) > 300:                  
                 
                # Use the mask to find the good original pixels and HANTS pixels  
                Array_ALBEDO_original_mask_nan = Array_ALBEDO_original * Array_outliers_mask                         
                Array_ALBEDO_HANTS_mask_nan = Array_ALBEDO_HANTS * Array_outliers_mask  
                Array_NDVI_original_mask_nan = Array_NDVI_original * Array_outliers_mask                       
                Array_NDVI_HANTS_mask_nan = Array_NDVI_HANTS * Array_outliers_mask 
                
                # Create a 1D array of those pixels
                Array_ALBEDO_original_mask_nan_flatten = Array_ALBEDO_original_mask_nan.flatten()
                Array_ALBEDO_HANTS_mask_nan_flatten = Array_ALBEDO_HANTS_mask_nan.flatten()
                Array_NDVI_original_mask_nan_flatten = Array_NDVI_original_mask_nan.flatten()
                Array_NDVI_HANTS_mask_nan_flatten = Array_NDVI_HANTS_mask_nan.flatten()
                
                # Remove pixels with high and low values
                Array_ALBEDO_HANTS_mask_nan_flatten[Array_ALBEDO_HANTS_mask_nan_flatten<-0.2] = np.nan
                Array_ALBEDO_HANTS_mask_nan_flatten[Array_ALBEDO_HANTS_mask_nan_flatten>0.6] = np.nan
                Array_ALBEDO_original_mask_nan_flatten[Array_ALBEDO_original_mask_nan_flatten<-0.2] = np.nan
                Array_ALBEDO_original_mask_nan_flatten[Array_ALBEDO_original_mask_nan_flatten>0.6] = np.nan
                Array_NDVI_HANTS_mask_nan_flatten[Array_NDVI_HANTS_mask_nan_flatten<-0.2] = np.nan
                Array_NDVI_HANTS_mask_nan_flatten[Array_NDVI_HANTS_mask_nan_flatten>0.6] = np.nan
                Array_NDVI_original_mask_nan_flatten[Array_NDVI_original_mask_nan_flatten<-0.2] = np.nan
                Array_NDVI_original_mask_nan_flatten[Array_NDVI_original_mask_nan_flatten>0.6] = np.nan
                                                      
                # Remove the nan values (if there is a nan in one of the arrays remove also the same value in the other array)                                     
                Array_ALBEDO_original_mask_nan_flatten2 = Array_ALBEDO_original_mask_nan_flatten[np.logical_or(~np.isnan(Array_ALBEDO_original_mask_nan_flatten),~np.isnan(Array_ALBEDO_HANTS_mask_nan_flatten))]
                Array_ALBEDO_HANTS_mask_nan_flatten2 = Array_ALBEDO_HANTS_mask_nan_flatten[np.logical_or(~np.isnan(Array_ALBEDO_original_mask_nan_flatten),~np.isnan(Array_ALBEDO_HANTS_mask_nan_flatten))]
                Array_NDVI_original_mask_nan_flatten2 = Array_NDVI_original_mask_nan_flatten[np.logical_or(~np.isnan(Array_NDVI_original_mask_nan_flatten),~np.isnan(Array_NDVI_HANTS_mask_nan_flatten))]
                Array_NDVI_HANTS_mask_nan_flatten2 = Array_NDVI_HANTS_mask_nan_flatten[np.logical_or(~np.isnan(Array_NDVI_HANTS_mask_nan_flatten),~np.isnan(Array_NDVI_original_mask_nan_flatten))]
                Array_ALBEDO_original_mask_nan_flatten = Array_ALBEDO_original_mask_nan_flatten2
                Array_ALBEDO_HANTS_mask_nan_flatten = Array_ALBEDO_HANTS_mask_nan_flatten2
                Array_NDVI_original_mask_nan_flatten = Array_NDVI_original_mask_nan_flatten2
                Array_NDVI_HANTS_mask_nan_flatten = Array_NDVI_HANTS_mask_nan_flatten2
                
    
                # Remove all zero values
                Array_ALBEDO_original_mask_nan_flatten_without_zero =Array_ALBEDO_original_mask_nan_flatten[Array_ALBEDO_original_mask_nan_flatten != 0.0]
                Array_NDVI_original_mask_nan_flatten_without_zero =Array_NDVI_original_mask_nan_flatten[Array_NDVI_original_mask_nan_flatten != 0.0]

                # Caluculate the value of the 40 and 90 percent percentiles of the original arrays good pixels
                Array_ALBEDO_original_mask_value_cold = np.nanpercentile(Array_ALBEDO_original_mask_nan_flatten_without_zero,40)
                Array_ALBEDO_original_mask_value_hot = np.nanpercentile(Array_ALBEDO_original_mask_nan_flatten_without_zero,90)
                Array_NDVI_original_mask_value_cold = np.nanpercentile(Array_NDVI_original_mask_nan_flatten_without_zero,40)
                Array_NDVI_original_mask_value_hot = np.nanpercentile(Array_NDVI_original_mask_nan_flatten_without_zero,90)

                # Delete the colder and hotter pixel values in both 1D arrays (this is to exclude large areas of seas)
                Array_ALBEDO_HANTS_mask_nan_flatten_exc_coldest = Array_ALBEDO_HANTS_mask_nan_flatten[np.logical_and(Array_ALBEDO_original_mask_nan_flatten > Array_ALBEDO_original_mask_value_cold,Array_ALBEDO_original_mask_nan_flatten < Array_ALBEDO_original_mask_value_hot)]
                Array_ALBEDO_original_mask_nan_flatten_exc_coldest = Array_ALBEDO_original_mask_nan_flatten[np.logical_and(Array_ALBEDO_original_mask_nan_flatten > Array_ALBEDO_original_mask_value_cold,Array_ALBEDO_original_mask_nan_flatten < Array_ALBEDO_original_mask_value_hot)]
                Array_NDVI_HANTS_mask_nan_flatten_exc_coldest = Array_NDVI_HANTS_mask_nan_flatten[np.logical_and(Array_NDVI_original_mask_nan_flatten > Array_NDVI_original_mask_value_cold,Array_NDVI_original_mask_nan_flatten < Array_NDVI_original_mask_value_hot)]
                Array_NDVI_original_mask_nan_flatten_exc_coldest = Array_NDVI_original_mask_nan_flatten[np.logical_and(Array_NDVI_original_mask_nan_flatten > Array_NDVI_original_mask_value_cold,Array_NDVI_original_mask_nan_flatten < Array_NDVI_original_mask_value_hot)]
                
                #Calculate the mean of those arrays
                Ave_ALBEDO_HANTS = np.nanmean(Array_ALBEDO_HANTS_mask_nan_flatten_exc_coldest)
                Ave_ALBEDO_original = np.nanmean(Array_ALBEDO_original_mask_nan_flatten_exc_coldest)
                Ave_NDVI_HANTS = np.nanmean(Array_NDVI_HANTS_mask_nan_flatten_exc_coldest)
                Ave_NDVI_original = np.nanmean(Array_NDVI_original_mask_nan_flatten_exc_coldest)
                
                # Calculate the correction factor for the simulated image
                Factor_Albedo = Ave_ALBEDO_original/Ave_ALBEDO_HANTS                
                Factor_NDVI = Ave_NDVI_original/Ave_NDVI_HANTS

                # Apply this factor over the simulated HANTS image
                Array_ALBEDO_HANTS_Corrected = Array_ALBEDO_HANTS * Factor_Albedo
                Array_NDVI_HANTS_Corrected = Array_NDVI_HANTS * Factor_NDVI
                
                # Create the end array by replacing the bad pixels of the original array by the corrected simulated HANTS values
                End_array_Albedo = np.ones(np.shape(Array_outliers_mask)) * np.nan
                End_array_Albedo[Array_outliers_mask==0] =Array_ALBEDO_HANTS_Corrected[Array_outliers_mask==0]
                End_array_Albedo[Array_outliers_mask==1] =Array_ALBEDO_original[Array_outliers_mask==1]
                End_array_NDVI = np.ones(np.shape(Array_outliers_mask)) * np.nan
                End_array_NDVI[Array_outliers_mask==0] =Array_NDVI_HANTS_Corrected[Array_outliers_mask==0]
                End_array_NDVI[Array_outliers_mask==1] =Array_NDVI_original[Array_outliers_mask==1]
            
            # If the original images is to bad than replace the whole image by the simulated HANTS image
            else:
                End_array_Albedo = Array_ALBEDO_HANTS
                End_array_NDVI = Array_NDVI_HANTS
                
            # Get the geolocation information of the image
            geo = dest_PROBAV_ALBEDO.GetGeoTransform()
            proj = dest_outliers.GetProjection()

        # If there is no original image, use the simulated HANTS image     
        else:
            Array_ALBEDO_HANTS = dest_HANTS_ALBEDO.GetRasterBand(1).ReadAsArray()
            End_array_Albedo = Array_ALBEDO_HANTS
            Array_NDVI_HANTS = dest_HANTS_NDVI.GetRasterBand(1).ReadAsArray()
            End_array_NDVI = Array_NDVI_HANTS
            dest_test = None
            i = 0
            
            while dest_test == None:
                
                # Get the date of the first image that exists to get the geolocation information
                date2 = Dates[i]
                year2 = date2.year
                month2= date2.month
                day2 = date2.day
                
                try:
                    filename_ALBEDO_original2 = os.path.join(input_folder_PreSEBAL_ALBEDO, "Albedo_PROBAV_%d%02d%02d.tif" %(year2,month2,day2))
                    dest_test = gdal.Open(filename_ALBEDO_original2)
                    geo = dest_test.GetGeoTransform()
                    proj = dest_test.GetProjection()
                except:    
                    i+=1
        
        # Save the end array            
        output_name_end_ALBEDO = os.path.join(ALBEDO_outfolder_end, "Albedo_PROBAV_%d%02d%02d.tif"%(year,month,day))
        SBF.save_GeoTiff_proy(dest, End_array_Albedo, output_name_end_ALBEDO, shape, nband=1)
        output_name_end_NDVI = os.path.join(NDVI_outfolder_end, "NDVI_PROBAV_%d%02d%02d.tif"%(year,month,day))
        SBF.save_GeoTiff_proy(dest, End_array_NDVI, output_name_end_NDVI, shape, nband=1)	
          

    ############################################# Create Outlier maps for VIIRS #########################################

    # Create output folder if not exists
    output_folder_HANTS_outliers_VIIRS = os.path.join(temp_folder_PreSEBAL, 'Outliers_VIIRS')
    if not os.path.exists(output_folder_HANTS_outliers_VIIRS):
        os.mkdir(output_folder_HANTS_outliers_VIIRS)

    fh = Dataset(nc_path_TB, mode='r')
    Var = fh.variables.keys()[-1]
    
    lat = fh.variables[fh.variables.keys()[1]][:]
    lon = fh.variables[fh.variables.keys()[2]][:] 
    time = fh.variables[fh.variables.keys()[3]][:]     
    minimum_lon = np.min(lon)
    maximum_lat = np.max(lat)
    diff_lon = lon[1] - lon[0]
    diff_lat = lat[1] - lat[0]
        
    if not ('shape' in locals() or 'dest' in locals()):
        Example_file = os.path.join(output_folder_preprocessing_THERM,Back_name_TB)                   
        dest = gdal.Open(Example_file)
        ncol = dest.RasterXSize        # Get the reprojected dem column size
        nrow = dest.RasterYSize        # Get the reprojected dem row size
        shape=[ncol, nrow]
        
    for i in range(0,int(np.shape(time)[0])):
        time_now = time[i]
        data = fh.variables['outliers'][i,:,:] 
        geo = tuple([minimum_lon, diff_lon, 0, maximum_lat, 0, diff_lat]) 
        name_out = os.path.join(output_folder_HANTS_outliers_VIIRS, 'Outliers_VIIRS_%s.tif' %time_now)      
        SBF.save_GeoTiff_proy(dest, data[:,:,i], name_out, shape, nband=1)	        
 
    ############################################# Create end thermal #########################################
        
    # Create the end thermal files date by date        
    for date in Dates:
        
        # Define date
        year = date.year
        month = date.month
        day = date.day
    
        # input filenames needed for creating end thermal file
        filename_outliers = os.path.join(output_folder_HANTS_outliers_VIIRS,"Outliers_VIIRS_%d%02d%02d.tif" %(year,month,day))
        filename_VIIRS_original =  os.path.join(input_folder_HANTS_THERM, "Surface_Temperature_VIIRS_%d%02d%02d.tif" %(year,month,day))
        filename_VIIRS_HANTS = os.path.join(temp_folder_PreSEBAL, THERM, "Surface_Temperature_VIIRS_%d%02d%02d.tif" %(year,month,day))
        
        # Open the input filenames
        dest_outliers = gdal.Open(filename_outliers)
        dest_VIIRS_original = gdal.Open(filename_VIIRS_original)
        dest_VIIRS_HANTS = gdal.Open(filename_VIIRS_HANTS)
        
        # If original exists, this will be the basis for the end thermal map
        if not dest_VIIRS_original == None:
        
            # Open arrays of the input files
            Array_outliers = dest_outliers.GetRasterBand(1).ReadAsArray()[:,:]
            Array_VIIRS_original = dest_VIIRS_original.GetRasterBand(1).ReadAsArray()
            Array_VIIRS_HANTS = dest_VIIRS_HANTS.GetRasterBand(1).ReadAsArray()[:,:]
            
            # Create outlier Mask
            Array_outliers[Array_outliers==-9999.] = 0
            Array_outliers_mask = np.zeros(np.shape(Array_outliers))
            Array_outliers_mask[Array_outliers==1.]=0   
            Array_outliers_mask[Array_outliers==0.]=1                                         
            Array_outliers_mask[Array_outliers_mask==0]=2
            Array_outliers_mask[Array_outliers_mask==1]=0 
            Array_outliers_mask[Array_outliers_mask==2]=1
                               
            # Create a buffer zone arround the bad pixels                   
            Array_outliers_mask = Create_Buffer(Array_outliers_mask)
            Array_outliers_mask[Array_outliers_mask==1] = 2
            Array_outliers_mask[Array_outliers_mask==0] = 1
            Array_outliers_mask[Array_outliers_mask==2] = 0  
            
            # If there are more than 300 Good pixels                   
            if np.nansum(Array_outliers_mask) > 300:                  
                 
                # Use the mask to find the good original pixels and HANTS pixels  
                Array_VIIRS_original_mask_nan = Array_VIIRS_original * Array_outliers_mask                 
                Array_VIIRS_HANTS_mask_nan = Array_VIIRS_HANTS * Array_outliers_mask                     
                
                # Create a 1D array of those pixels
                Array_VIIRS_original_mask_nan_flatten = Array_VIIRS_original_mask_nan.flatten()
                Array_VIIRS_HANTS_mask_nan_flatten = Array_VIIRS_HANTS_mask_nan.flatten()
                
                # Remove pixels with high and low values
                Array_VIIRS_HANTS_mask_nan_flatten[Array_VIIRS_HANTS_mask_nan_flatten<250] = np.nan
                Array_VIIRS_HANTS_mask_nan_flatten[Array_VIIRS_HANTS_mask_nan_flatten>350] = np.nan
                Array_VIIRS_original_mask_nan_flatten[Array_VIIRS_original_mask_nan_flatten<250] = np.nan
                Array_VIIRS_original_mask_nan_flatten[Array_VIIRS_original_mask_nan_flatten>350] = np.nan
                      
                # Remove the nan values (if there is a nan in one of the arrays remove also the same value in the other array)                                     
                Array_VIIRS_original_mask_nan_flatten = Array_VIIRS_original_mask_nan_flatten[~np.isnan(Array_VIIRS_original_mask_nan_flatten)]
                Array_VIIRS_HANTS_mask_nan_flatten = Array_VIIRS_HANTS_mask_nan_flatten[~np.isnan(Array_VIIRS_HANTS_mask_nan_flatten)]
 
                # Remove all zero values
                Array_VIIRS_original_mask_nan_flatten_without_zero =Array_VIIRS_original_mask_nan_flatten[Array_VIIRS_original_mask_nan_flatten>0]

                # Caluculate the value of the 40 and 90 percent percentiles of the original arrays good pixels
                Array_VIIRS_original_mask_value_cold = np.nanpercentile(Array_VIIRS_original_mask_nan_flatten_without_zero,40)
                Array_VIIRS_original_mask_value_hot = np.nanpercentile(Array_VIIRS_original_mask_nan_flatten_without_zero,90)

                # Delete the colder and hotter pixel values in both 1D arrays (this is to exclude large areas of seas)
                Array_VIIRS_HANTS_mask_nan_flatten_exc_coldest = Array_VIIRS_HANTS_mask_nan_flatten[np.logical_and(Array_VIIRS_original_mask_nan_flatten > Array_VIIRS_original_mask_value_cold,Array_VIIRS_original_mask_nan_flatten < Array_VIIRS_original_mask_value_hot)]
                Array_VIIRS_original_mask_nan_flatten_exc_coldest = Array_VIIRS_original_mask_nan_flatten[np.logical_and(Array_VIIRS_original_mask_nan_flatten > Array_VIIRS_original_mask_value_cold,Array_VIIRS_original_mask_nan_flatten < Array_VIIRS_original_mask_value_hot)]
                
                #Calculate the mean of those arrays
                Ave_VIIRS_HANTS = np.nanmean(Array_VIIRS_HANTS_mask_nan_flatten_exc_coldest)
                Ave_VIIRS_original = np.nanmean(Array_VIIRS_original_mask_nan_flatten_exc_coldest)
                
                # Calculate the correction factor for the simulated image
                Factor = Ave_VIIRS_original/Ave_VIIRS_HANTS
                
                # Apply this factor over the simulated HANTS image
                Array_VIIRS_HANTS_Corrected = Array_VIIRS_HANTS * Factor
                
                # Create the end array by replacing the bad pixels of the original array by the corrected simulated HANTS values
                End_array = np.ones(np.shape(Array_outliers_mask)) * np.nan
                End_array[Array_outliers_mask==0] =Array_VIIRS_HANTS_Corrected[Array_outliers_mask==0]
                End_array[Array_outliers_mask==1] =Array_VIIRS_original[Array_outliers_mask==1]
            
            # If the original images is to bad than replace the whole image by the simulated HANTS image
            else:
                End_array = Array_VIIRS_HANTS
            
            # Get the geolocation information of the image
            geo = dest_VIIRS_original.GetGeoTransform()
            proj = dest_outliers.GetProjection()

        # If there is no original image, use the simulated HANTS image     
        else:
            Array_VIIRS_HANTS = dest_VIIRS_HANTS.GetRasterBand(1).ReadAsArray()
            End_array = Array_VIIRS_HANTS
            dest_test = None
            i = 0
            
            while dest_test == None:
                
                # Get the date of the first image that exists to get the geolocation information
                date2 = Dates[i]
                year2 = date2.year
                month2= date2.month
                day2 = date2.day
                
                try:
                    filename_VIIRS_original2 = os.path.join(input_folder_HANTS_THERM, "Surface_Temperature_VIIRS_%d%02d%02d.tif" %(year2,month2,day2))
                    dest_test = gdal.Open(filename_VIIRS_original2)
                    geo = dest_test.GetGeoTransform()
                    proj = dest_test.GetProjection()
                except:    
                    i+=1
        
        # Save the end array            
        output_name_end_LST = os.path.join(temp_folder_PreSEBAL_LST, "VIIRS_LST_%d%02d%02d.tif"%(year,month,day))
        SBF.save_GeoTiff_proy(dest, End_array, output_name_end_LST, shape, nband=1)	

###################################################################################################################
###################################################### preSEBAL continue ##########################################
###################################################################################################################    
        
    ############################################### Apply thermal sharpening ##########################################

    print '---------------------------------------------------------'
    print '-------------------- Downscale VIIRS --------------------'
    print '---------------------------------------------------------'

    # Upscale VIIRS and PROBA-V to 400m
    pixel_spacing_upscale = 400    

    # Open the General_Input sheet			
    ws = wb['General_Input']

    # Extract the input and output folder, and Image type from the excel file			
    DEM_fileName = str(ws['E2'].value)

    ws = wb['VIIRS_PROBAV_Input']
    UTM_Zone = int(str(ws['G2'].value))    

    # Reproject from Geog Coord Syst to UTM -
    # 1) DEM - Original DEM coordinates is Geographic: lat, lon
    proyDEM_fileName_100 = os.path.join(temp_folder_PreSEBAL,'DEM_100.tif')
    dest, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SBF.reproject_dataset(
                   DEM_fileName, pixel_spacing = 100, UTM_Zone=UTM_Zone)
    band = dest.GetRasterBand(1)   # Get the reprojected dem band
    ncol = dest.RasterXSize        # Get the reprojected dem column size
    nrow = dest.RasterYSize        # Get the reprojected dem row size
    shape=[ncol, nrow]
    DEM = band.ReadAsArray()

    # Save DEM file with the 100 meter resolution					
    SBF.save_GeoTiff_proy(dest, DEM, proyDEM_fileName_100, shape, nband=1)

    # Create upscaled DEM
    proyDEM_fileName_400 = os.path.join(temp_folder_PreSEBAL,'DEM_400.tif')
    dest_400, ulx_dem_400, lry_dem_400, lrx_dem_400, uly_dem_400, epsg_to = SBF.reproject_dataset(
                DEM_fileName, pixel_spacing_upscale, UTM_Zone = UTM_Zone)

    # find spatial parameters array
    DEM_400 = dest_400.GetRasterBand(1).ReadAsArray()
    Y_raster_size_400 = dest_400.RasterYSize				
    X_raster_size_400 = dest_400.RasterXSize
    shape_400=([X_raster_size_400, Y_raster_size_400])
	
    # Save DEM file with the 400 meter resolution					
    SBF.save_GeoTiff_proy(dest_400, DEM_400, proyDEM_fileName_400, shape_400, nband=1)

    for date in Dates:
        
        surf_temp_fileName = os.path.join(temp_folder_PreSEBAL, 'Surf_temp_After_TS_%d%02d%02d.tif' %(date.year, date.month, date.day))
        temp_surface_100_fileName_beforeTS = os.path.join(temp_folder_PreSEBAL_LST,'VIIRS_LST_%d%02d%02d.tif' %(date.year, date.month, date.day))
        
        ################################ Thermal Sharpening #####################################################
       
        # Define filename
        file_NDVI_after_HANTS = os.path.join(NDVI_outfolder_end, 'NDVI_PROBAV_%d%02d%02d.tif' %(date.year, date.month, date.day))

        # Open NDVI/LST destination folder
        dest_NDVI = gdal.Open(file_NDVI_after_HANTS)
        dest_LST = gdal.Open(temp_surface_100_fileName_beforeTS)
        
        # Open NDVI array
        NDVI = dest_NDVI.GetRasterBand(1).ReadAsArray()
 
        # Open LST array
        LST = dest_LST.GetRasterBand(1).ReadAsArray()
       										
        # Upscale thermal band VIIRS from 100m to 400m
        VIIRS_Upscale, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SBF.reproject_dataset_example(
                       temp_surface_100_fileName_beforeTS, proyDEM_fileName_400)
        data_Temp_Surf_400 = VIIRS_Upscale.GetRasterBand(1).ReadAsArray()
 
        # Upscale PROBA-V NDVI from 100m to 400m       
        NDVI_PROBAV_Upscale, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SBF.reproject_dataset_example(
                      file_NDVI_after_HANTS, proyDEM_fileName_400)
        data_NDVI_400 = NDVI_PROBAV_Upscale.GetRasterBand(1).ReadAsArray()

        # Define the width of the moving window box
        Box=9
   
        # Apply the surface temperature sharpening        
        temp_surface_sharpened = SBF.Thermal_Sharpening(data_Temp_Surf_400, data_NDVI_400, NDVI, Box, NDVI_PROBAV_Upscale, output_folder, proyDEM_fileName_100, shape, dest, surf_temp_fileName)	

        # Create Water mask based on HANTS NDVI output            
        water_mask = np.zeros((shape[1], shape[0])) 
        water_mask[NDVI<0.0]=1

        # Divide temporal watermask in snow and water mask by using surface temperature
        Snow_Mask_PROBAV, water_mask, ts_moist_veg_min, NDVI_max, NDVI_std = SBF.CalculateSnowWaterMask(NDVI,shape,water_mask,temp_surface_sharpened)

        # Replace water values
        temp_surface_sharpened[water_mask==1] = LST[water_mask == 1]
        temp_surface_sharpened = np.where(np.isnan(temp_surface_sharpened), LST, temp_surface_sharpened)          
        
        surf_temp_fileName = os.path.join(output_folder_HANTS_end_sharp, 'LST_surface_temp_sharpened_%d%02d%02d.tif' %(date.year, date.month, date.day))
        SBF.save_GeoTiff_proy(dest, temp_surface_sharpened, surf_temp_fileName, shape, nband=1)             
        
    ################################################## Calculate LAI ##################################################
    
        # Open NDVI destination folder
        dest_NDVI = gdal.Open(file_NDVI_after_HANTS)
        
        # Open NDVI array
        NDVI = dest_NDVI.GetRasterBand(1).ReadAsArray()
        
        LAI_FileName = os.path.join(LAI_outfolder,'LAI_%d%02d%02d.tif' %(date.year, date.month, date.day))

        # Calculate LAI
        FPAR, tir_emis, Nitrogen, vegt_cover, LAI, b10_emissivity = SBF.Calc_vegt_para(NDVI,water_mask, shape)
        SBF.save_GeoTiff_proy(dest, LAI, LAI_FileName, shape, nband=1) 

    ################################ Calculate the Vegetation height ########################
    
        # Open preprosessing excel the Vegetation_Height sheet				
        ws_veg = wb_veg['Vegetation_Height'] 
    
        # Define output name for the LandUse map 
        dst_FileName = os.path.join(output_folder,'LU.tif') 
    
        # Open LU data
        LU_dest = gdal.Open(LU_data_FileName)
        LU_data = LU_dest.GetRasterBand(1).ReadAsArray() 
     
        # Reproject the LAI to the same projection as LU
        dest1, ulx, lry, lrx, uly, epsg_to = SBF.reproject_dataset_example(LAI_FileName, LU_data_FileName)	 ## input after HANTS
        LAI_proj = dest1.GetRasterBand(1).ReadAsArray() 
    
        # Read out the excel file coefficient numbers			
        Array = np.zeros([ws_veg.max_row-1,4])
        for j in ['A','C','D','E']:
            j_number={'A' : 0, 'C' : 1, 'D' : 2, 'E' : 3}					
            for i in range(2,ws_veg.max_row+1):											
    	        Value = (ws_veg['%s%s' %(j,i)].value)  																
    	        Array[i-2, j_number[j]] = Value										
    
        # Create maps with the coefficient numbers for the right land cover
        coeff = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1]),3])
        for coeff_nmbr in range(0,3):				
            for Class in range(0,len(Array)):
    	        coeff[LU_data==Array[Class,0],coeff_nmbr] = Array[Class,coeff_nmbr+1]
    
        # Get some dimensions of the projected dataset 
        band_data = dest1.GetRasterBand(1) 
        ncol_data = dest1.RasterXSize
        nrow_data = dest1.RasterYSize
        shape_data=[ncol_data, nrow_data]
    
        # Calculate the vegetation height in the LU projection
        Veg_Height_proj = coeff[:,:,0] * np.power(LAI_proj,2) + coeff[:,:,1] * LAI_proj + coeff[:,:,2]
        Veg_Height_proj = np.clip(Veg_Height_proj, 0, 600)
    
        # Save the vegetation height in the lU projection in the temporary directory
        Veg_Height_proj_FileName = os.path.join(temp_folder_PreSEBAL,'Veg_Height_proj.tif') 				
        SBF.save_GeoTiff_proy(dest1, Veg_Height_proj, Veg_Height_proj_FileName, shape_data, nband=1)	
    				
        # Reproject the Veg_height to the LAI projection
        dest, ulx, lry, lrx, uly, epsg_to = SBF.reproject_dataset_example(Veg_Height_proj_FileName, LAI_FileName)			
    
        # Get some dimensions of the original dataset 
        band_data = dest.GetRasterBand(1)
        ncol_data = dest.RasterXSize
        nrow_data = dest.RasterYSize
    
        # Open the Veg_height with the same projection as LAI				
        Veg_Height = band_data.ReadAsArray(0, 0, ncol_data, nrow_data)
        Veg_Height[Veg_Height == 0] = 0.4			
    
        # Save Vegetation Height in the end folder				
        dst_FileName = os.path.join(output_folder_HANTS_end_Veg,'Vegetation_Height_%d%02d%02d.tif' %(date.year, date.month, date.day))
        SBF.save_GeoTiff_proy(dest, Veg_Height, dst_FileName, shape, nband=1)	
        
    ######################## calculate Water Mask #########################
    
    # Open all the water mask
    os.chdir(WaterMask_outfolder)
    re_water_mask = glob.glob('Water_Mask*.tif')

    #  Loop over all the files
    for water_mask_filename in re_water_mask:
        
        # Create the filepath to the water mask
        water_mask_filepath = os.path.join(WaterMask_outfolder,water_mask_filename)
        
        # Open Array
        water_mask_dest = gdal.Open(water_mask_filepath)
        
        # If the total water mask raster does not exists create this one
        if not 'water_mask_array' in locals():
            
            water_mask_array = np.zeros([water_mask_dest.RasterYSize, water_mask_dest.RasterXSize])
        
        # Add all the water masks      
        water_mask_array += water_mask_dest.GetRasterBand(1).ReadAsArray()
        
    # Calculate the end water mask if the area is more than 50 percent defined as water    
    water_mask_array_per = water_mask_array/len(re_water_mask)
    water_mask_array_end = np.zeros([water_mask_dest.RasterYSize, water_mask_dest.RasterXSize])
    water_mask_array_end[water_mask_array_per > 0.5] = 1
    
    # Save water mask
    WaterMask_outfolder_end_FileName = os.path.join(WaterMask_outfolder_end,'Water_Mask.tif') 
    SBF.save_GeoTiff_proy(dest, water_mask_array_end, WaterMask_outfolder_end_FileName, shape, nband=1)	

    ######################## calculate p-factor by using the Landuse map #########################
    ws_p = wb_veg['p-factor'] 
			
    Array_P = np.zeros([ws_p.max_row-1,2])
    for j in ['A','C']:
        j_number={'A' : 0, 'C' : 1}					
        for i in range(2,ws_p.max_row+1):											
            Value = (ws_p['%s%s' %(j,i)].value)  																
            Array_P[i-2, j_number[j]] = Value	

    p_factor = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1])])		
    for Class in range(0,len(Array_P)):
	    p_factor[LU_data==Array_P[Class,0]] = Array_P[Class,1]

    p_factor[p_factor == 0] = 0.5

    dst_FileName = os.path.join(temp_folder_PreSEBAL, 'p-factor_proj.tif') 	
    SBF.save_GeoTiff_proy(dest1, p_factor, dst_FileName, shape_data, nband=1)
    
    # Reproject the p factor map
    dest, ulx, lry, lrx, uly, epsg_to = SBF.reproject_dataset_example(dst_FileName, LAI_FileName)	

    # Open the new projected p factor map
    band_data = dest.GetRasterBand(1) # Get the reprojected p factor
    ncol_data = dest.RasterXSize
    nrow_data = dest.RasterYSize				
    p_factor = band_data.ReadAsArray(0, 0, ncol_data, nrow_data)
    p_factor[p_factor == 0] = 0.5

    dst_pfactor_FileName = os.path.join(output_folder_p_factor,'p_factor.tif') 
    SBF.save_GeoTiff_proy(dest, p_factor, dst_pfactor_FileName, shape, nband=1)	

######################## calculate c-factor by using the Landuse map #########################

    ws_c = wb_veg['C-factor'] 
			
    Array_C = np.zeros([ws_c.max_row-1,2])
    for j in ['A','C']:
        j_number={'A' : 0, 'C' : 1}					
        for i in range(2,ws_c.max_row+1):											
            Value = (ws_c['%s%s' %(j,i)].value)  																
            Array_C[i-2, j_number[j]] = Value	

    c_factor = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1])])		
    for Class in range(0,len(Array_C)):
	    c_factor[LU_data==Array_C[Class,0]] = Array_C[Class,1]

    c_factor[np.logical_and(c_factor != 3.0, c_factor != 4.0)] = np.nan

    LUE_max = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1])])	
    LUE_max[c_factor == 3] = 2.5
    LUE_max[c_factor == 4] = 4.5
    LUE_max[LUE_max == 0] = 2.5

    dst_FileName = os.path.join(temp_folder_PreSEBAL, 'LUE_max_proj.tif') 	
    SBF.save_GeoTiff_proy(dest1, LUE_max, dst_FileName, shape_data, nband=1)

    dest, ulx, lry, lrx, uly, epsg_to = SBF.reproject_dataset_example(dst_FileName, LAI_FileName)	

    band_data = dest.GetRasterBand(1) # Get the reprojected dem band	
    ncol_data = dest.RasterXSize
    nrow_data = dest.RasterYSize				
    LUE_max = band_data.ReadAsArray(0, 0, ncol_data, nrow_data)
    LUE_max[LUE_max == 0] = 2.5

    dst_LUEmax_FileName = os.path.join(output_folder_LUE,'LUE_max.tif') 
    SBF.save_GeoTiff_proy(dest, LUE_max, dst_LUEmax_FileName, shape, nband=1)	

####################################################################################################################   
################################################ Write output part 6 ###############################################		
####################################################################################################################

   ############################################# Fill in the additional input sheet ######################################### 
   # things to be filled in:
   # Transmissivity (optional)
   # NDVI (additional input)
   # Albedo (additional input)
   # LST (additional input)   
   # Water Mask (additional input)    
   # p-factor (soil input)
   # c-factor  (soil input)
   # Vegetation height (meteo input)
 
    # VIIRS parameter copy
    VIIRS_Dict = {}
    for k, v in SEBAL_RUNS.iteritems():
        VIIRS_Dict.setdefault(v['output_folder'], []).append(k)      

    # Define the excel sheets that will be filled in by PreSEBAL
    VARS = ["NDVI", "Albedo"]
    Letter_dict = {"NDVI":'B', "Albedo":'D'}
    xfile = load_workbook(inputExcel) 
    sheet_additional = xfile.get_sheet_by_name('Additional_Input')
    sheet_meteo = xfile.get_sheet_by_name('Meteo_Input')
    sheet_soil = xfile.get_sheet_by_name('Soil_Input')    
    sheet_out_name = ''.join([os.path.splitext(os.path.basename(inputExcel))[0],'_SEBAL.xlsx'])
    sheet_out_dir = os.path.dirname(inputExcel) 
    sheet_out_file_name = os.path.join(sheet_out_dir, sheet_out_name)
    
    for output_name_run in VIIRS_Dict.keys()[2:4]:
        
        # Get General parameters
        Row_number = VIIRS_Dict[output_name_run][0]
        Type_of_Run = SEBAL_RUNS.items()
        VIIRS_date = output_name_run.split('_')[-1]      
        VIIRS_datetime= datetime.strptime(VIIRS_date, '%d%m%Y')
        date_run = '%d%02d%02d' %(VIIRS_datetime.year,VIIRS_datetime.month,VIIRS_datetime.day)

        # import LST
        file_name_LST = os.path.join(output_folder_HANTS_end_sharp, 'LST_surface_temp_sharpened_%s.tif' %date_run )
        sheet_additional['E%d'%(Row_number)] = str(file_name_LST)

        # import NDVI and Albedo and water mask
        for VAR_SINGLE in VARS:
            Letter = Letter_dict[VAR_SINGLE]   
            file_name_VAR_single = os.path.join(output_folder_PreSEBAL, VAR_SINGLE, '%s_PROBAV_%s.tif' %(VAR_SINGLE, date_run))
            sheet_additional['%s%d'%(Letter, Row_number)] = str(file_name_VAR_single)

        # import Water Mask
        sheet_additional['C%d'%(Row_number)] = str(WaterMask_outfolder_end_FileName)

        # import p-factor
        file_name_p_factor = os.path.join(output_folder_p_factor,'p_factor.tif')
        sheet_soil['H%d'%(Row_number)] = str(file_name_p_factor)
        
        # import p-factor
        file_name_c_factor = os.path.join(output_folder_LUE, 'LUE_max.tif')
        sheet_soil['I%d'%(Row_number)] = str(file_name_c_factor)
        
        # import vegetation height        
        file_name_vegt_height = os.path.join(output_folder_HANTS_end_Veg,'Vegetation_Height_%s.tif' %date_run)
        sheet_meteo['O%d'%(Row_number)] = str(file_name_vegt_height)

    xfile.save(sheet_out_file_name) 

# Functions
#################################################################################   
def Create_Buffer(Data_In):
    
   '''
   This function creates a 3D array which is used to apply the moving window
   '''
   Buffer_area = 7 # A block of 2 times Buffer_area + 1 will be 1 if there is the pixel in the middle is 1
   Data_Out=np.empty((len(Data_In),len(Data_In[1])))   
   Data_Out[:,:] = Data_In
   for ypixel in range(0,Buffer_area + 1):
                   
        for xpixel in range(1,Buffer_area + 1):

           if ypixel==0:
                for xpixel in range(1,Buffer_area + 1):
                    Data_Out[:,0:-xpixel] += Data_In[:,xpixel:]
                    Data_Out[:,xpixel:] += Data_In[:,:-xpixel]              
                
                for ypixel in range(1,Buffer_area + 1):
                
                    Data_Out[ypixel:,:] += Data_In[:-ypixel,:]
                    Data_Out[0:-ypixel,:] += Data_In[ypixel:,:]

           else:
               Data_Out[0:-xpixel,ypixel:] += Data_In[xpixel:,:-ypixel]
               Data_Out[xpixel:,ypixel:] += Data_In[:-xpixel,:-ypixel]
               Data_Out[0:-xpixel,0:-ypixel] += Data_In[xpixel:,ypixel:]
               Data_Out[xpixel:,0:-ypixel] += Data_In[:-xpixel,ypixel:]

   Data_Out[Data_Out>0.1] = 1
   Data_Out[Data_Out<=0.1] = 0  
           
   return(Data_Out)

#------------------------------------------------------------------------------
def Get_epsg(g):				
			
    try:
        # Get info of the dataset that is used for transforming     
        gland_proj = g.GetProjection()
        Projection=gland_proj.split('EPSG","')
        epsg_to=int((str(Projection[-1]).split(']')[0])[0:-1])				      
    except:
        epsg_to=4326	
        print 'Was not able to get the projection, so WGS84 is assumed'							
    return(epsg_to)	

#------------------------------------------------------------------------------
def gap_filling(data,NoDataValue):
    """
    This function fills the no data gaps in a numpy array
				
    Keyword arguments:
    dataset -- Array
    NoDataValue -- Value that must be filled
    """
	        
    # fill the no data values
    if NoDataValue is np.nan:
        mask = ~(np.isnan(data))
    else:
        mask = ~(data==NoDataValue)
    xx, yy = np.meshgrid(np.arange(data.shape[1]), np.arange(data.shape[0]))
    xym = np.vstack( (np.ravel(xx[mask]), np.ravel(yy[mask])) ).T
    data0 = np.ravel( data[:,:][mask] )
    interp0 = scipy.interpolate.NearestNDInterpolator( xym, data0 )
    data_end = interp0(np.ravel(xx), np.ravel(yy)).reshape( xx.shape )
		
    return (data_end)


#------------------------------------------------------------------------------				
if __name__ == '__main__':
    main()