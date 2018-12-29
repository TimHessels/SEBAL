# -*- coding: utf-8 -*-
"""
Created on Mon May 28 08:39:23 2018

@author: tih
"""

# import standard modules
import os
import numpy as np
import gdal
import pandas as pd
import datetime
import glob
import scipy
import osr
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

###############################################################################
# Code
###############################################################################

def main(Excel_name, Startdate, Enddate ,output_folder, input_folder_VIIRS_PROBAV, Temp_format, PROBAV_format, WD_Temp_format, WD_RH_format, WD_Wind_format, WD_Rad_format, Time_loc_Temp_hour, Time_loc_Temp_minutes, Difference_time, Difference_day, DEM_map_filename, output_folder_SEBAL, UTM_Zone, wcpf2_topsoil_filename, wcpf42_topsoil_filename, wcsat_topsoil_filename, wcsat_subsoil_filename,dates_remove):
   
    # Create timesteps
    Dates_inst_Temp = pd.date_range(Startdate, Enddate, freq = WD_Temp_format[2])
    Dates_inst_RH = pd.date_range(Startdate, Enddate, freq = WD_RH_format[2])
    Dates_inst_Wind = pd.date_range(Startdate, Enddate, freq = WD_Wind_format[2])
    Dates_inst_Rad = pd.date_range(Startdate, Enddate, freq = WD_Rad_format[2])
    Dates_daily = pd.date_range(Startdate, Enddate, freq = "D")

    for date_remove in dates_remove:
        Dates_daily = Dates_daily.drop([datetime.datetime.strptime(date_remove, "%Y-%m-%d")])

    # Define paths
    output_folder_Temp_daily = os.path.join(output_folder, "Temperature", "Daily")
    output_folder_Wind_daily = os.path.join(output_folder, "Wind", "Daily")
    output_folder_RH_daily = os.path.join(output_folder, "Humidity", "Daily")
    output_folder_Rad_daily = os.path.join(output_folder, "Radiation", "Daily")

    output_folder_Temp_inst = os.path.join(output_folder, "Temperature", "Inst")
    output_folder_Wind_inst = os.path.join(output_folder, "Wind", "Inst")
    output_folder_RH_inst = os.path.join(output_folder, "Humidity", "Inst")
    output_folder_Rad_inst = os.path.join(output_folder, "Radiation", "Inst")

    # Define start Excel
    excel_number = 2

    # Open Excel workbook
    wb=load_workbook(Excel_name)
    
    # Calculate the average VIIRS time
    os.chdir(input_folder_VIIRS_PROBAV)
    VIIRS_files = glob.glob(Temp_format.format(yyyy = '*', mm = '*', dd = '*'))
    minutes_hours = []
    for VIIRS_file in VIIRS_files:
        minutes_hours = np.append(minutes_hours, float(VIIRS_file[Time_loc_Temp_hour[0]:Time_loc_Temp_hour[1]]) \
                        + float(VIIRS_file[Time_loc_Temp_minutes[0]:Time_loc_Temp_minutes[1]])/60)
    Average_VIIRS_time = np.nanmean(minutes_hours)
    hour_ave = "%02d" %int(np.floor(Average_VIIRS_time))
    minutes_ave = "%02d" %float((Average_VIIRS_time - int(hour_ave))*60)
    
    # Loop over days
    for date in Dates_daily:
        print(date)
        print(excel_number)
        # Define date
        year = "%d" %date.year
        month = "%02d" %date.month
        day = "%02d" %date.day

        day_viirs = "%02d" %(int(day) + Difference_day)
        
        #####################################################################
        # Calculate Meteo and save tiff file
        #####################################################################

        # Find instantanious time
        os.chdir(input_folder_VIIRS_PROBAV)
        file_Temp = glob.glob(Temp_format.format(yyyy = year, mm = month, dd = day_viirs))
        if len(file_Temp) > 0:
            file_Temp = file_Temp[0]
            
            #####################################################################
            # Calculations for Meteo Standard inst data
            #####################################################################

            # Define time for instantanious from VIIRS name in local time
            Hour = int(file_Temp[Time_loc_Temp_hour[0]:Time_loc_Temp_hour[1]]) + Difference_time
            Minutes = file_Temp[Time_loc_Temp_minutes[0]:Time_loc_Temp_minutes[1]]
            Date_inst = datetime.datetime(int(year),int(month),int(day),int(Hour), int(Minutes))
            Nearest_inst_Temp = Find_nearest_time(Dates_inst_Temp, Date_inst)
            Nearest_inst_RH = Find_nearest_time(Dates_inst_RH, Date_inst)
            Nearest_inst_Wind = Find_nearest_time(Dates_inst_Wind, Date_inst)
            Nearest_inst_Rad = Find_nearest_time(Dates_inst_Rad, Date_inst)
                                    
            # Define time for Weather Data
            hour_Temp = "%02d" %Nearest_inst_Temp.hour
            minute_Temp = "%02d" %Nearest_inst_Temp.minute
            hour_RH = "%02d" %Nearest_inst_RH.hour
            minute_RH = "%02d" %Nearest_inst_RH.minute
            hour_Wind = "%02d" %Nearest_inst_Wind.hour
            minute_Wind = "%02d" %Nearest_inst_Wind.minute
            hour_Rad = "%02d" %Nearest_inst_Rad.hour
            minute_Rad = "%02d" %Nearest_inst_Rad.minute

        else:

            # Average time of all the VIIRS images
            Date_inst = datetime.datetime(int(year),int(month),int(day),int(hour_ave), int(minutes_ave))
            Nearest_inst_Temp = Find_nearest_time(Dates_inst_Temp, Date_inst)
            Nearest_inst_RH = Find_nearest_time(Dates_inst_RH, Date_inst)
            Nearest_inst_Wind = Find_nearest_time(Dates_inst_Wind, Date_inst)
            Nearest_inst_Rad = Find_nearest_time(Dates_inst_Rad, Date_inst)
                                    
            # Define time for Weather Data
            hour_Temp = "%02d" %Nearest_inst_Temp.hour
            minute_Temp = "%02d" %Nearest_inst_Temp.minute
            hour_RH = "%02d" %Nearest_inst_RH.hour
            minute_RH = "%02d" %Nearest_inst_RH.minute
            hour_Wind = "%02d" %Nearest_inst_Wind.hour
            minute_Wind = "%02d" %Nearest_inst_Wind.minute
            hour_Rad = "%02d" %Nearest_inst_Rad.hour
            minute_Rad = "%02d" %Nearest_inst_Rad.minute
            file_Temp = ''

        # SEBAL output name
        SEBAL_output_folder = os.path.join(output_folder_SEBAL, "VIIRS_PROBAV_%s%s%s_H%s_M%s" %(year, month, day, Date_inst.hour, Date_inst.minute))

        #####################################################################
        # Fill in the excel file general
        #####################################################################

        ws = wb['General_Input']
        ws['B%d'%(excel_number)] = str(input_folder_VIIRS_PROBAV)
        ws['C%d'%(excel_number)] = str(SEBAL_output_folder)
        ws['D%d'%(excel_number)] = int(2)
        ws['E%d'%(excel_number)] = str(DEM_map_filename)

        #####################################################################
        # Fill in the excel file Soil Input
        #####################################################################
        ws = wb['Soil_Input']
        ws['B%d'%(excel_number)] = str(wcsat_topsoil_filename)
        ws['C%d'%(excel_number)] = str(wcsat_subsoil_filename)
        ws['D%d'%(excel_number)] = float(0.02)
        ws['E%d'%(excel_number)] = float(0.02)
        ws['F%d'%(excel_number)] = str(wcpf2_topsoil_filename)
        ws['G%d'%(excel_number)] = str(wcpf42_topsoil_filename)

        #####################################################################
        # Temperature
        #####################################################################
        try:
            if os.path.splitext(WD_Temp_format[1].split(":")[0])[1] == ".xlsx":
             
                output_name_Temp_inst, output_name_Temp_daily = Get_Meteo_From_Excel(WD_Temp_format, year, month, day, Date_inst)
    
            else:
                # Create output folder temperature
                if not os.path.exists(output_folder_Temp_daily):
                    os.makedirs(output_folder_Temp_daily)
            
                if not os.path.exists(output_folder_Temp_inst):
                    os.makedirs(output_folder_Temp_inst)
    
                # Create Daily file
                output_name_Temp_inst = os.path.join(output_folder_Temp_inst, "Temperature_inst_C_%d.%02d.%02d_H%02d.M%02d.tif" %(int(year),int(month),int(day),int(hour_Temp), int(minute_Temp)))
                output_name_Temp_daily = os.path.join(output_folder_Temp_daily, "Temperature_Daily_C_%d.%02d.%02d.tif" %(int(year), int(month), int(day)))
                if not (os.path.exists(output_name_Temp_daily) and os.path.exists(output_name_Temp_inst)):
    
                    # Create Daily file
                    os.chdir(WD_Temp_format[0])
                    variable_inst_day = WD_Temp_format[1].format(yyyy = year, mm = month, dd = day, HH = '*',  MM = '*')
                    files_whole_day = glob.glob(variable_inst_day)
                    Data_Temp_daily = Sum_files_avg(WD_Temp_format[0], files_whole_day)
                    Geo_WD, Proj_WD = get_proj(WD_Temp_format[0], files_whole_day[0])
                    Save_as_tiff(output_name_Temp_daily, Data_Temp_daily, Geo_WD, Proj_WD)
    
                    # Create Instantanious file
                    variable_inst_day = WD_Temp_format[1].format(yyyy = year, mm = month, dd = day, HH = hour_Temp,  MM = minute_Temp)
                    Data_Temp_inst = Return_file_inst(WD_Temp_format[0], variable_inst_day)
                    Save_as_tiff(output_name_Temp_inst, Data_Temp_inst, Geo_WD, Proj_WD)
        except:
            print("ERROR: Temperature data for %s-%02s-%02s not created, fill in by user is required!" %(year,month,day))
            output_name_Temp_inst = ""
            output_name_Temp_daily = ""
        #####################################################################
        # Wind
        #####################################################################
        try:
            if os.path.splitext(WD_Wind_format[1].split(":")[0])[1] == ".xlsx":
             
                output_name_Wind_inst, output_name_Wind_daily = Get_Meteo_From_Excel(WD_Wind_format, year, month, day, Date_inst)
    
            else:
    
                # Create output folder wind
                if not os.path.exists(output_folder_Wind_daily):
                    os.makedirs(output_folder_Wind_daily)
            
                if not os.path.exists(output_folder_Wind_inst):
                    os.makedirs(output_folder_Wind_inst)
    
                output_name_Wind_daily = os.path.join(output_folder_Wind_daily, "Wind_Daily_m_s-1_%d.%02d.%02d.tif" %(int(year), int(month), int(day)))
                output_name_Wind_inst = os.path.join(output_folder_Wind_inst, "Wind_inst_m_s-1_%d.%02d.%02d_H%02d.M%02d.tif" %(int(year),int(month),int(day),int(hour_Wind), int(minute_Wind)))
                if not (os.path.exists(output_name_Wind_daily) and os.path.exists(output_name_Wind_inst)):
    
                    # Create Daily file
                    os.chdir(WD_Wind_format[0])
                    variable_inst_day = WD_Wind_format[1].format(yyyy = year, mm = month, dd = day, HH = '*',  MM = '*')
                    files_whole_day = glob.glob(variable_inst_day)
                    Data_Wind_daily = Sum_files_avg(WD_Wind_format[0], files_whole_day)
                    Geo_WD, Proj_WD = get_proj(WD_Wind_format[0], variable_inst_day)
                    Save_as_tiff(output_name_Wind_daily, Data_Wind_daily, Geo_WD, Proj_WD)
    
                    # Create Instantanious file
                    variable_inst_day = WD_Wind_format[1].format(yyyy = year, mm = month, dd = day, HH = hour_Wind,  MM = minute_Wind)
                    Data_Wind_inst = Return_file_inst(WD_Wind_format[0], variable_inst_day)
                    Save_as_tiff(output_name_Wind_inst, Data_Wind_inst, Geo_WD, Proj_WD)
        except:
            print("ERROR: Wind data for %s-%02s-%02s not created, fill in by user is required!" %(year,month,day))
            output_name_Wind_daily = ""
            output_name_Wind_inst = ""
        #####################################################################
        # RH
        #####################################################################
        try:
            if os.path.splitext(WD_RH_format[1].split(":")[0])[1] == ".xlsx":
             
                output_name_RH_inst, output_name_RH_daily = Get_Meteo_From_Excel(WD_RH_format, year, month, day, Date_inst)
    
            else:
       
                # Create output folder relative humidity
                if not os.path.exists(output_folder_RH_daily):
                    os.makedirs(output_folder_RH_daily)
            
                if not os.path.exists(output_folder_RH_inst):
                    os.makedirs(output_folder_RH_inst)
    
                output_name_RH_daily = os.path.join(output_folder_RH_daily, "RH_Daily_percentage_%d.%02d.%02d.tif" %(int(year), int(month), int(day)))
                output_name_RH_inst = os.path.join(output_folder_RH_inst, "RH_inst_percentage_%d.%02d.%02d_H%02d.M%02d.tif" %(int(year),int(month),int(day),int(hour_RH), int(minute_RH)))
                if not (os.path.exists(output_name_RH_daily) and os.path.exists(output_name_RH_inst)):
    
                    # Create Daily file
                    os.chdir(WD_RH_format[0])
                    variable_inst_day = WD_RH_format[1].format(yyyy = year, mm = month, dd = day, HH = '*',  MM = '*')
                    files_whole_day = glob.glob(variable_inst_day)
                    Data_RH_daily = Sum_files_avg(WD_RH_format[0], files_whole_day)
                    Geo_WD, Proj_WD = get_proj(WD_RH_format[0], variable_inst_day)
                    Save_as_tiff(output_name_RH_daily, Data_RH_daily, Geo_WD, Proj_WD)
    
                    # Create Instantanious file
                    variable_inst_day = WD_RH_format[1].format(yyyy = year, mm = month, dd = day, HH = hour_RH,  MM = minute_RH)
                    Data_RH_inst = Return_file_inst(WD_RH_format[0], variable_inst_day)
                    Save_as_tiff(output_name_RH_inst, Data_RH_inst, Geo_WD, Proj_WD)
        except:
            print("ERROR: Humidity data for %s-%02s-%02s not created, fill in by user is required!" %(year,month,day))    
            output_name_RH_daily = ""
            output_name_RH_inst = ""
        #####################################################################
        # Radiation
        #####################################################################
        try:
            if os.path.splitext(WD_Rad_format[1].split(":")[0])[1] == ".xlsx":
             
                output_name_Rad_inst, output_name_Rad_daily = Get_Meteo_From_Excel(WD_Rad_format, year, month, day, Date_inst)
    
            else:
    
                # Create output folder radiation
                if not os.path.exists(output_folder_Rad_daily):
                    os.makedirs(output_folder_Rad_daily)
            
                if not os.path.exists(output_folder_Rad_inst):
                    os.makedirs(output_folder_Rad_inst)
    
                output_name_Rad_daily = os.path.join(output_folder_Rad_daily, "Rad_Daily_W_m-2_%d.%02d.%02d.tif" %(int(year), int(month), int(day)))
                output_name_Rad_inst = os.path.join(output_folder_Rad_inst, "Rad_inst_W_m-2_%d.%02d.%02d_H%02d.M%02d.tif" %(int(year),int(month),int(day),int(hour_Rad), int(minute_Rad)))
                if not (os.path.exists(output_name_Rad_daily) and os.path.exists(output_name_Rad_inst)):
    
                    # Create Daily file
                    os.chdir(WD_Rad_format[0])
                    variable_inst_day = WD_Rad_format[1].format(yyyy = year, mm = month, dd = day, HH = '*',  MM = '*')
                    files_whole_day = glob.glob(variable_inst_day)
                    Data_Rad_daily = Sum_files_avg(WD_Rad_format[0], files_whole_day)
                    Geo_WD, Proj_WD = get_proj(WD_Rad_format[0], variable_inst_day)
                    Save_as_tiff(output_name_Rad_daily, Data_Rad_daily, Geo_WD, Proj_WD)
    
                    # Create Instantanious file
                    variable_inst_day = WD_Rad_format[1].format(yyyy = year, mm = month, dd = day, HH = hour_Rad,  MM = minute_Rad)
                    Data_Rad_inst = Return_file_inst(WD_Rad_format[0], variable_inst_day)
                    Save_as_tiff(output_name_Rad_inst, Data_Rad_inst, Geo_WD, Proj_WD)
        except:
            print("ERROR: Radiation data for %s-%02s-%02s not created, fill in by user is required!" %(year,month,day))
            output_name_Rad_inst = ""
            output_name_Rad_daily = ""
        #####################################################################
        # Fill in Excel Meteo tab
        #####################################################################
        ws = wb['Meteo_Input']
        ws['B%d'%(excel_number)] = str(output_name_Temp_inst)
        ws['C%d'%(excel_number)] = str(output_name_Temp_daily)
        ws['D%d'%(excel_number)] = str(output_name_RH_inst)
        ws['E%d'%(excel_number)] = str(output_name_RH_daily)
        ws['F%d'%(excel_number)] = int(10)
        ws['G%d'%(excel_number)] = str(output_name_Wind_inst)
        ws['H%d'%(excel_number)] = str(output_name_Wind_daily)
        ws['I%d'%(excel_number)] = int(1)
        ws['J%d'%(excel_number)] = str(output_name_Rad_inst)
        ws['L%d'%(excel_number)] = int(1)
        ws['M%d'%(excel_number)] = str(output_name_Rad_daily)
        ws['O%d'%(excel_number)] = float(0.4)

        #####################################################################
        # Fill in Excel VIIRS_PROBAV_Input VIIRS part
        #####################################################################

        ws = wb['VIIRS_PROBAV_Input']
        ws['B%d'%(excel_number)] = str(file_Temp)
        ws['E%d'%(excel_number)] = float(2)
        ws['F%d'%(excel_number)] = float(1.5)
        ws['G%d'%(excel_number)] = int(UTM_Zone)

        #####################################################################
        # Fill in Excel VIIRS_PROBAV_Input PROBA-V part
        #####################################################################

        # change working directory
        os.chdir(input_folder_VIIRS_PROBAV)

        # Define PROBAV name
        PROBA_V_name = PROBAV_format.format(yyyy = year, mm = month, dd = day)

        # Find PROBA-V file
        re_PROBAV = glob.glob(PROBA_V_name)

        if len(re_PROBAV) > 0:

            # PROBA-V name without extension
            PROBA_V_name = os.path.splitext(os.path.basename(re_PROBAV[0]))[0]

            # Add PROBA-V band if exists
            ws = wb['VIIRS_PROBAV_Input']
            ws['D%d'%(excel_number)] = str(PROBA_V_name)

        # go to next excel row
        excel_number += 1

    wb.save(Excel_name)
    return()

###############################################################################
# Functions
###############################################################################

def get_proj(input_folder_WD, variable_inst_day):

    # get projection
    os.chdir(input_folder_WD)
    first_file = glob.glob(variable_inst_day)[0]
    dest = gdal.Open(first_file)
    Proj_WD = dest.GetProjection()
    Geo_WD  = dest.GetGeoTransform()
    return(Geo_WD, Proj_WD)


def Sum_files_avg(input_folder_WD, files_whole_day):

    os.chdir(input_folder_WD)
    i = 0
    for file_whole_day in files_whole_day:

        # create empty 3D array
        if file_whole_day == files_whole_day[0]:
            dest = gdal.Open(file_whole_day)
            data = dest.GetRasterBand(1).ReadAsArray()
            size_Y, size_X = np.int_(data.shape[-2:])
            DataTot = np.zeros([len(files_whole_day), size_Y, size_X])

        dest = gdal.Open(file_whole_day)
        data = dest.GetRasterBand(1).ReadAsArray()
        DataTot[i,:,:] = data
        
        i += 1

    DataTot_day = np.nanmean(DataTot, axis = 0)
    DataTot_day[np.isnan(DataTot_day)] = -9999
    DataTot_day = gap_filling(DataTot_day, -9999)
    
    return(DataTot_day)

def gap_filling(data,NoDataValue, method = 1):
    """
    This function fills the no data gaps in a numpy array

    Keyword arguments:
    dataset -- 'C:/'  path to the source data (dataset that must be filled)
    NoDataValue -- Value that must be filled
    """
    # fill the no data values
    if NoDataValue is np.nan:
        mask = ~(np.isnan(data))
    else:
        mask = ~(data==NoDataValue)
    xx, yy = np.meshgrid(np.arange(data.shape[1]), np.arange(data.shape[0]))
    xym = np.vstack( (np.ravel(xx[mask]), np.ravel(yy[mask])) ).T
    data0 = np.ravel( data[:,:][mask] )

    if method == 1:
        interp0 = scipy.interpolate.NearestNDInterpolator( xym, data0 )
        data_end = interp0(np.ravel(xx), np.ravel(yy)).reshape( xx.shape )

    if method == 2:
        interp0 = scipy.interpolate.LinearNDInterpolator( xym, data0 )
        data_end = interp0(np.ravel(xx), np.ravel(yy)).reshape( xx.shape )

    return (data_end)

def Return_file_inst(input_folder_WD, file_inst):

    os.chdir(input_folder_WD)
    dest = gdal.Open(file_inst)
    data = dest.GetRasterBand(1).ReadAsArray()
    data[np.isnan(data)] = -9999
    data = gap_filling(data, -9999)
    return(data)


def Save_as_tiff(name='', data='', geo='', projection=''):
    """
    This function save the array as a geotiff

    Keyword arguments:
    name -- string, directory name
    data -- [array], dataset of the geotiff
    geo -- [minimum lon, pixelsize, rotation, maximum lat, rotation,
            pixelsize], (geospatial dataset)
    projection -- integer, the EPSG code
    """
    # save as a geotiff
    driver = gdal.GetDriverByName("GTiff")
    dst_ds = driver.Create(name, int(data.shape[1]), int(data.shape[0]), 1,
                           gdal.GDT_Float32, ['COMPRESS=LZW'])
    srse = osr.SpatialReference()
    if projection == '':
        srse.SetWellKnownGeogCS("WGS84")

    else:
        try:
            if not srse.SetWellKnownGeogCS(projection) == 6:
                srse.SetWellKnownGeogCS(projection)
            else:
                try:
                    srse.ImportFromEPSG(int(projection))
                except:
                    srse.ImportFromWkt(projection)
        except:
            try:
                srse.ImportFromEPSG(int(projection))
            except:
                srse.ImportFromWkt(projection)

    dst_ds.SetProjection(srse.ExportToWkt())
    dst_ds.GetRasterBand(1).SetNoDataValue(-9999)
    dst_ds.SetGeoTransform(geo)
    dst_ds.GetRasterBand(1).WriteArray(data)
    dst_ds = None
    return()


def Find_nearest_time(Dates, Date):
    return min(Dates, key=lambda x: abs(x - Date))

def Get_Meteo_From_Excel(WD_format, year, month, day, Date_inst):                 
       
        
    # Open excel file
    name_excel_in_Temp = os.path.join(WD_format[0], WD_format[1].split(":")[0])
    wb_meteo=load_workbook(name_excel_in_Temp)
    #sheet_name = wb_meteo.get_sheet_names()
    ws_meteo = wb_meteo['Sheet1']
    length_excel = ws_meteo.max_row

    # read out all the content, including the date and the temperature
    Meteo_data = np.ones([length_excel]) * np.nan
    Date_data = list()
    Column_temp = WD_format[1].split(":")[1]
    for i in range(2, length_excel):
        Date_data.append(str(ws_meteo["A%s" %i].value))
        Meteo_data[i-2] = np.float_(ws_meteo["%s%s" %(Column_temp, i)].value)
        
    # Correct length    
    Meteo_data = Meteo_data[:len(Date_data)]
    
    # Find nearest date for meteo to instantanious VIIRS image
    dates_excel = pd.to_datetime(pd.Series(Date_data), format= "%Y-%m-%d %H:%M:%S")
    Nearest_inst = Find_nearest_time(dates_excel, Date_inst)
    row = np.where(Nearest_inst==dates_excel)[0][0]
    output_inst = Meteo_data[row]
    
    # Calculate daily average
    dates_day = Meteo_data[np.logical_and(dates_excel >= pd.datetime(int(year),int(month),int(day),0,0), dates_excel <= pd.datetime(int(year),int(month),int(day),23,59))]
    if len(dates_day) > 1:
        output_daily = np.nanmean(dates_day)
    else:
        output_daily= -9999    
        output_inst = -9999
        
    return(output_inst, output_daily)    
