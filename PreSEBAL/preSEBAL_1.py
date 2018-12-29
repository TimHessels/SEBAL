# -*- coding: utf-8 -*-
"""
Created on Thu May 31 13:23:47 2018

@author: tih
"""

import numpy as np
import os
import gdal
from openpyxl import load_workbook
import osr
from datetime import datetime, timedelta
import warnings
from pyproj import Proj, transform

def main(VegetationExcel):

####################################################################################################################
############################################# CREATE INPUT FOR preSEBAL RUN ########################################
####################################################################################################################

####################################################################################################################
##################################################### PreSEBAL_1 ###################################################
####################################################################################################################

# PreSEBAL_1
# Part 1: Define input by user
# Part 2: Set parameters and output folder
# Part 3: Calculate NDVI and Albedo based on PROBA-V

####################################################################################################################
#################################### part 1: Define input by user ##################################################
####################################################################################################################

    import SEBAL.pySEBAL.pySEBAL_code as SEBAL
    import SEBAL.pySEBAL.pySEBAL_input_PROBAV_VIIRS as input_PROBAV_VIIRS

####################################################################################################################
################################# part 2: Set parameters and output folder #########################################
####################################################################################################################

    # Open Excel workbook used for Vegetation c and p factor conversions
    wb_veg = load_workbook(VegetationExcel, data_only=True)
    ws_veg = wb_veg['General_Input']

    # Input for preSEBAL.py
    start_date = "%s" %str(ws_veg['B2'].value)
    end_date = "%s" %str(ws_veg['B3'].value)
    inputExcel= r"%s" %str(ws_veg['B4'].value)              # The excel with all the SEBAL input data
    output_folder = r"%s" %str(ws_veg['B7'].value)

    ######################## Load Excels ##########################################
    # Open Excel workbook for SEBAL inputs
    wb = load_workbook(inputExcel)

    # Get length of EXCEL sheet
    ws = wb['General_Input']
    ws2 = wb['VIIRS_PROBAV_Input']
    endExcel=int(ws.max_row)

    # Create Dict
    SEBAL_RUNS = dict()

    for number in range(2,endExcel+1):
        input_folder_SEBAL = str(ws['B%d' % number].value)
        output_folder_SEBAL = str(ws['C%d' % number].value)
        Image_Type = int(ws['D%d' % number].value)
        PROBA_V_name  = str(ws2['D%d' % number].value)
        VIIRS_name  = str(ws2['B%d' % number].value)
        SEBAL_RUNS[number] = {'input_folder': input_folder_SEBAL, 'output_folder': output_folder_SEBAL, 'image_type': Image_Type,'PROBA_V_name': PROBA_V_name,'VIIRS_name': VIIRS_name}

    Kind_Of_Runs_Dict = {}
    for k, v in iter(SEBAL_RUNS.items()):
        Kind_Of_Runs_Dict.setdefault(v['image_type'], []).append(k)

    ######################## Create output folders  ##########################################

    # Define main directories
    output_folder_PreSEBAL_sub = os.path.join(output_folder,'PreSEBAL_SEBAL_out')         # outputs before HANTS
    output_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_out')                     # End outputs of preSEBAL
    temp_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_temp')                      # Temp outputs of preSEBAL

    # Define sub results folders
    NDVI_outfolder = os.path.join(output_folder_PreSEBAL_sub,'NDVI')
    Albedo_outfolder = os.path.join(output_folder_PreSEBAL_sub,'Albedo')
    WaterMask_outfolder = os.path.join(output_folder_PreSEBAL_sub,'Water_Mask')
    Mask_outfolder = os.path.join(output_folder_PreSEBAL_sub,'Mask')
    
    # file name of the example dataset based on DEM
    Example_fileName = os.path.join(temp_folder_PreSEBAL, "DEM_Example_Projection.tif")

    if not os.path.exists(output_folder_PreSEBAL_sub):
        os.makedirs(output_folder_PreSEBAL_sub)
    if not os.path.exists(output_folder_PreSEBAL):
        os.mkdir(output_folder_PreSEBAL)
    if not os.path.exists(temp_folder_PreSEBAL):
        os.mkdir(temp_folder_PreSEBAL)
    if not os.path.exists(NDVI_outfolder):
        os.makedirs(NDVI_outfolder)
    if not os.path.exists(Albedo_outfolder):
        os.makedirs(Albedo_outfolder)
    if not os.path.exists(WaterMask_outfolder):
        os.makedirs(WaterMask_outfolder)
    if not os.path.exists(Mask_outfolder):
        os.makedirs(Mask_outfolder)
        
    # Do not show warnings
    warnings.filterwarnings('ignore')

####################################################################################################################
########################## Calculate NDVI and Albedo based on PROBA-V part 3 #######################################
####################################################################################################################
    
    PV_count = 0
    
    # Find pixels that are 90% of the time defined as clouds (this is wrong, so create an anti cloud mask)
    for number in Kind_Of_Runs_Dict[2][:]: 

      ############################ Create the example file based on DEM  ######################################

        # Open the General_Input sheet
        ws = wb['General_Input']

        # Extract the Path to the DEM map from the excel file
        DEM_fileName = '%s' %str(ws['E%d' % number].value) #'DEM_HydroShed_m'

        # Extract the name of the thermal and quality VIIRS image from the excel file
        Name_VIIRS_Image_TB = '%s' %str(ws['B%d' % number].value)

        if not os.path.exists(Example_fileName):

            # Pixel size of the model
            pixel_spacing= 0.001

            # Open DEM and create Latitude and longitude files
            lat,lon,lat_fileName,lon_fileName=SEBAL.DEM_lat_lon(DEM_fileName, temp_folder_PreSEBAL)

            # Reproject from Geog Coord Syst to UTM -
            # 1) DEM - Original DEM coordinates is Geographic: lat, lon
            dest, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_to_example(DEM_fileName, pixel_spacing, 4326)

            # Open data
            data_DEM = dest.GetRasterBand(1).ReadAsArray()
            band = dest.GetRasterBand(1)   # Get the reprojected dem band
            ncol = dest.RasterXSize        # Get the reprojected dem column size
            nrow = dest.RasterYSize        # Get the reprojected dem row size
            shape=[ncol, nrow]

            # Save file
            SEBAL.save_GeoTiff_proy(dest, data_DEM, Example_fileName, shape, nband = 1)

        else:
            dest = gdal.Open(Example_fileName)
            band = dest.GetRasterBand(1)   # Get the reprojected dem band
            ncol = dest.RasterXSize        # Get the reprojected dem column size
            nrow = dest.RasterYSize        # Get the reprojected dem row size
            shape=[ncol, nrow]

        
        ######################## Extract general data for VIIRS-PROBAV ##########################################

        # Open the VIIRS_PROBAV_Input sheet
        ws = wb['VIIRS_PROBAV_Input']

        # Extract the name to the PROBA-V image from the excel file
        Name_PROBAV_Image = '%s' %str(ws['D%d' % number].value)        
          
       ############################ Open General info from SEBAL Excel ###########################

        # Open the General_Input sheet
        ws = wb['General_Input']

        # Extract the input and output folder, and Image type from the excel file
        input_folder = str(ws['B%d' % number].value)    
        
        if not SEBAL_RUNS[number]['PROBA_V_name'] == 'None':
            
            PV_count += 1
            
            # Get all cloud mask
            spectral_reflectance_PROBAV, cloud_mask_temp = input_PROBAV_VIIRS.Open_PROBAV_Reflectance(Name_PROBAV_Image, input_folder, temp_folder_PreSEBAL, Example_fileName)
                    
            # If the total water mask raster does not exists create this one
            if not 'anti_cloud_mask_array' in locals():
    
                anti_cloud_mask_array = cloud_mask_temp
    
            else:
                # Add all the water masks
                anti_cloud_mask_array += cloud_mask_temp   
                
    # Define wrong pixels       
    Percentage_Cloud_Mask = anti_cloud_mask_array/PV_count
    Anti_Cloud_Mask = np.zeros(Percentage_Cloud_Mask.shape)
    Anti_Cloud_Mask[Percentage_Cloud_Mask>0.9] = 1        
      
    # save anti cloud mask
    WaterMask_outfolder_end_FileName = os.path.join(Mask_outfolder,'Anti_Mask.tif')
    SEBAL.save_GeoTiff_proy(dest, Anti_Cloud_Mask, WaterMask_outfolder_end_FileName, shape, nband=1)

    for number in Kind_Of_Runs_Dict[2][:]: # Number defines the column of the inputExcel

        print(number)

        ######################## Extract general data for VIIRS-PROBAV ##########################################

        # Open the VIIRS_PROBAV_Input sheet
        ws = wb['VIIRS_PROBAV_Input']

        # Extract the name to the PROBA-V image from the excel file
        Name_PROBAV_Image = '%s' %str(ws['D%d' % number].value)

        ######################## Create Example tiff file based on DEM ##########################################

        if not SEBAL_RUNS[number]['PROBA_V_name'] == 'None':

       ############################ Open General info from SEBAL Excel ###########################

            # Open the General_Input sheet
            ws = wb['General_Input']

            # Extract the input and output folder, and Image type from the excel file
            input_folder = str(ws['B%d' % number].value)

            Name_PROBAV, Name_PROBAV_exe = os.path.splitext(Name_PROBAV_Image)
            
            if (Name_PROBAV_exe == '.hdf5' or Name_PROBAV_exe == '.HDF5'):
                
                # Get the day and time from the PROBA-V
                Band_PROBAVhdf_fileName = os.path.join(input_folder, '%s' % (Name_PROBAV_Image))
                g=gdal.Open(Band_PROBAVhdf_fileName, gdal.GA_ReadOnly)
    
                Meta_data = g.GetMetadata()
                Date_PROBAV = str(Meta_data['LEVEL3_RADIOMETRY_BLUE_OBSERVATION_START_DATE'])
                year = int(Date_PROBAV.split("-")[0])
                month = int(Date_PROBAV.split("-")[1])
                day = int(Date_PROBAV.split("-")[2])
                
            else:
                year = int(Name_PROBAV.split("_")[-3][0:4])
                month = int(Name_PROBAV.split("_")[-3][4:6])
                day = int(Name_PROBAV.split("_")[-3][6:8])
                
                
            Var_name = '%d%02d%02d' %(year, month, day)
            DOY=datetime.strptime(Var_name,'%Y%m%d').timetuple().tm_yday

           ################################ Calculate NDVI from PROBAV ##########################################

            # Define output maps
            NDVI_FileName = os.path.join(NDVI_outfolder,'NDVI_PROBAV_%s.tif' %Var_name)
            Albedo_FileName = os.path.join(Albedo_outfolder, 'Albedo_PROBAV_%s.tif' %Var_name)
            water_mask_temp_FileName = os.path.join(WaterMask_outfolder, 'Water_Mask_PROBAV_%s.tif' %Var_name)

            # Run code if output map not exists
            if not (os.path.exists(NDVI_FileName) and os.path.exists(Albedo_FileName) and os.path.exists(water_mask_temp_FileName)):

                spectral_reflectance_PROBAV, cloud_mask_temp = input_PROBAV_VIIRS.Open_PROBAV_Reflectance(Name_PROBAV_Image, input_folder, temp_folder_PreSEBAL, Example_fileName)

                # remove wrong pixels
                spectral_reflectance_PROBAV[np.logical_and(cloud_mask_temp==1, Anti_Cloud_Mask == 0)] = np.nan
                
                ################################ Calculate Albedo from PROBAV ##########################################

                # Calculate surface albedo based on PROBA-V
                Surface_Albedo_PROBAV = 0.219 * spectral_reflectance_PROBAV[:, :, 1] + 0.361 * spectral_reflectance_PROBAV[:, :, 2] + 0.379 * spectral_reflectance_PROBAV[:, :, 3] + 0.041 * spectral_reflectance_PROBAV[:, :, 4]

                # Calculate the NDVI based on PROBA-V
                n218_memory = spectral_reflectance_PROBAV[:, :, 2] + spectral_reflectance_PROBAV[:, :, 3]
                NDVI = np.zeros((shape[1], shape[0]))
                NDVI[n218_memory != 0] =  ( spectral_reflectance_PROBAV[:, :, 3][n218_memory != 0] - spectral_reflectance_PROBAV[:, :, 2][n218_memory != 0] )/ ( spectral_reflectance_PROBAV[:, :, 2][n218_memory != 0] + spectral_reflectance_PROBAV[:, :, 3][n218_memory != 0] )

                ############################# Calculate Water Mask from PROBAV ##########################################

                # Create Water mask based on PROBA-V
                water_mask_temp = np.zeros((shape[1], shape[0]))
                water_mask_temp[np.logical_and(NDVI<0.0,Surface_Albedo_PROBAV<0.2)] = 1

                # Save Albedo for PROBA-V
                SEBAL.save_GeoTiff_proy(dest, Surface_Albedo_PROBAV, Albedo_FileName, shape, nband=1)

                # Save NDVI for PROBA-V
                SEBAL.save_GeoTiff_proy(dest, NDVI, NDVI_FileName, shape, nband=1)

                # Save Water Mask for PROBA-V
                SEBAL.save_GeoTiff_proy(dest, water_mask_temp, water_mask_temp_FileName, shape, nband=1)

    return()


#------------------------------------------------------------------------------
def reproject_dataset_to_example(dataset, pixel_spacing, proj_in):
    """
    A sample function to reproject and resample a GDAL dataset from within
    Python. The idea here is to reproject from one system to another, as well
    as to change the pixel size. The procedure is slightly long-winded, but
    goes like this:

    1. Set up the two Spatial Reference systems.
    2. Open the original dataset, and get the geotransform
    3. Calculate bounds of new geotransform by projecting the UL corners
    4. Calculate the number of pixels with the new projection & spacing
    5. Create an in-memory raster dataset
    6. Perform the projection
    """

    # 1) Open the dataset
    g = gdal.Open(dataset)
    if g is None:
        print('input folder does not exist')

     # Define the EPSG code...
    EPSG_code = '%d' % proj_in
    epsg_to = int(EPSG_code)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    try:
        proj = g.GetProjection()
        Proj_in=proj.split('EPSG","')
        epsg_from=int((str(Proj_in[-1]).split(']')[0])[0:-1])
    except:
        epsg_from = int(4326)    # Get the Geotransform vector:
    geo_t = g.GetGeoTransform()

    # Vector components:
    # 0- The Upper Left easting coordinate (i.e., horizontal)
    # 1- The E-W pixel spacing
    # 2- The rotation (0 degrees if image is "North Up")
    # 3- The Upper left northing coordinate (i.e., vertical)
    # 4- The rotation (0 degrees)
    # 5- The N-S pixel spacing, negative as it is counted from the UL corner
    x_size = g.RasterXSize  # Raster xsize
    y_size = g.RasterYSize  # Raster ysize

    epsg_to = int(epsg_to)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    osng = osr.SpatialReference()
    osng.ImportFromEPSG(epsg_to)
    wgs84 = osr.SpatialReference()
    wgs84.ImportFromEPSG(epsg_from)

    inProj = Proj(init='epsg:%d' %epsg_from)
    outProj = Proj(init='epsg:%d' %epsg_to)

    # Up to here, all  the projection have been defined, as well as a
    # transformation from the from to the to
    ulx, uly = transform(inProj,outProj,geo_t[0], geo_t[3])
    lrx, lry = transform(inProj,outProj,geo_t[0] + geo_t[1] * x_size,
                                        geo_t[3] + geo_t[5] * y_size)

    # See how using 27700 and WGS84 introduces a z-value!
    # Now, we create an in-memory raster
    mem_drv = gdal.GetDriverByName('MEM')

    # The size of the raster is given the new projection and pixel spacing
    # Using the values we calculated above. Also, setting it to store one band
    # and to use Float32 data type.
    col = int((lrx - ulx)/pixel_spacing)
    rows = int((uly - lry)/pixel_spacing)

    # Re-define lr coordinates based on whole number or rows and columns
    (lrx, lry) = (ulx + col * pixel_spacing, uly -
                  rows * pixel_spacing)

    dest = mem_drv.Create('', col, rows, 1, gdal.GDT_Float32)

    if dest is None:
        print('input folder to large for memory, clip input map')

   # Calculate the new geotransform
    new_geo = (ulx, pixel_spacing, geo_t[2], uly,
               geo_t[4], - pixel_spacing)

    # Set the geotransform
    dest.SetGeoTransform(new_geo)
    dest.SetProjection(osng.ExportToWkt())

    # Perform the projection/resampling
    gdal.ReprojectImage(g, dest, wgs84.ExportToWkt(), osng.ExportToWkt(),gdal.GRA_Bilinear)

    return dest, ulx, lry, lrx, uly, epsg_to
