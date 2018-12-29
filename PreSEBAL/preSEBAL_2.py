# -*- coding: utf-8 -*-
"""
Created on Thu May 31 17:21:06 2018

@author: tih
"""
import numpy as np
import os
import gdal
from openpyxl import load_workbook
import osr
from datetime import datetime, timedelta
import pandas as pd
import glob
from netCDF4 import Dataset
from scipy import interpolate


def main(VegetationExcel, Albedo_netcdf, NDVI_netcdf, Apply_Corrections_using_original):

####################################################################################################################
############################################# CREATE INPUT FOR preSEBAL RUN ########################################
####################################################################################################################

####################################################################################################################
##################################################### PreSEBAL_2 ###################################################
####################################################################################################################

# PreSEBAL_2
# Part 1: Define input by user
# Part 2: Set parameters and output folder
# Part 3: Calculate NDVI and Albedo based on PROBA-V

####################################################################################################################
#################################### part 1: Define input by user ##################################################
####################################################################################################################

    import SEBAL.pySEBAL.pySEBAL_code as SEBAL
    #VegetationExcel =r"G:\SEBAL_Tadla\Daily_SEBAL\Excel_PreSEBAL_v1_0.xlsx" # This excel defines the p and c factor and vegetation height.
    #Albedo_netcdf = r"G:\SEBAL_Tadla\Daily_SEBAL\Dataset_out\PreSEBAL_SEBAL_out\Albedo_HANTS5.nc"
    #NDVI_netcdf = r"G:\SEBAL_Tadla\Daily_SEBAL\Dataset_out\PreSEBAL_SEBAL_out\NDVI_HANTS5.nc"
    #Apply_Corrections_using_original = 1 # 1 = yes, 0 = no   #293 change amount of pixels

####################################################################################################################
################################# part 2: Set parameters and output folder #########################################
####################################################################################################################

    # Open Excel workbook used for Vegetation c and p factor conversions
    wb_veg = load_workbook(VegetationExcel, data_only=True)
    ws_veg = wb_veg['General_Input']

    # Input for preSEBAL.py
    start_date = "%s" %str(ws_veg['B2'].value)
    end_date = "%s" %str(ws_veg['B3'].value)
    inputExcel= r"%s" %str(ws_veg['B4'].value)              # The excel with all the SEBAL input data
    output_folder = r"%s" %str(ws_veg['B7'].value)

    ######################## Load Excels ##########################################
    # Open Excel workbook for SEBAL inputs
    wb = load_workbook(inputExcel)

    # Get length of EXCEL sheet
    ws = wb['General_Input']
    ws2 = wb['VIIRS_PROBAV_Input']
    ws3 = wb['Meteo_Input']
    endExcel=int(ws.max_row)

    # Create Dict
    SEBAL_RUNS = dict()

    for number in range(2,endExcel+1):
        input_folder_SEBAL = str(ws['B%d' % number].value)
        output_folder_SEBAL = str(ws['C%d' % number].value)
        Image_Type = int(ws['D%d' % number].value)
        PROBA_V_name  = str(ws2['D%d' % number].value)
        VIIRS_name  = str(ws2['B%d' % number].value)
        SEBAL_RUNS[number] = {'input_folder': input_folder_SEBAL, 'output_folder': output_folder_SEBAL, 'image_type': Image_Type,'PROBA_V_name': PROBA_V_name,'VIIRS_name': VIIRS_name}

    Kind_Of_Runs_Dict = {}
    for k, v in iter(SEBAL_RUNS.items()):
        Kind_Of_Runs_Dict.setdefault(v['image_type'], []).append(k)

    PROBA_V_dict = {}
    for k, v in iter(SEBAL_RUNS.items()):
        PROBA_V_dict.setdefault(v['PROBA_V_name'], []).append(k)

    ######################## Create output folders  ##########################################

    # output maps for main directories
    output_folder_PreSEBAL_sub = os.path.join(output_folder,'PreSEBAL_SEBAL_out')         # outputs before HANTS
    output_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_out')                     # End outputs of preSEBAL
    temp_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_temp')                      # Temp outputs of preSEBAL

    # output maps for the end products
    ALBEDO_outfolder_end = os.path.join(output_folder_PreSEBAL,'ALBEDO')
    NDVI_outfolder_end = os.path.join(output_folder_PreSEBAL,'NDVI')
    WaterMask_outfolder_end = os.path.join(output_folder_PreSEBAL,'Water_Mask')

    # preSEBAL side outputs
    Surface_Temperature_outfolder = os.path.join(output_folder_PreSEBAL_sub,'Surface_Temperature')
    WaterMask_outfolder = os.path.join(output_folder_PreSEBAL_sub,'Water_Mask')
    Cloud_free_PROBAV_Land_Pixels_outfolder = os.path.join(output_folder_PreSEBAL_sub,'Cloud_free_PROBAV_Land_Pixels')
    NCfractions_outfolder_NDVI = os.path.join(output_folder_PreSEBAL_sub,'Fractions_NDVI')
    NCfractions_outfolder_Albedo = os.path.join(output_folder_PreSEBAL_sub,'Fractions_Albedo')
    NDVI_outfolder = os.path.join(output_folder_PreSEBAL_sub,'NDVI')
    Albedo_outfolder = os.path.join(output_folder_PreSEBAL_sub,'Albedo')

    # preSEBAL temp outputs
    output_folder_tir_emis =  os.path.join(temp_folder_PreSEBAL, 'tir_emis')

    # file name of the example dataset based on DEM
    Example_fileName = os.path.join(temp_folder_PreSEBAL, "DEM_Example_Projection.tif")

    if not os.path.exists(WaterMask_outfolder_end):
        os.makedirs(WaterMask_outfolder_end)
    if not os.path.exists(Surface_Temperature_outfolder):
        os.makedirs(Surface_Temperature_outfolder)
    if not os.path.exists(output_folder_tir_emis):
        os.mkdir(output_folder_tir_emis)
    if not os.path.exists(Cloud_free_PROBAV_Land_Pixels_outfolder):
        os.makedirs(Cloud_free_PROBAV_Land_Pixels_outfolder)
    if not os.path.exists(NCfractions_outfolder_NDVI):
        os.makedirs(NCfractions_outfolder_NDVI)
    if not os.path.exists(NCfractions_outfolder_Albedo):
        os.makedirs(NCfractions_outfolder_Albedo)
    if not os.path.exists(NDVI_outfolder):
        os.makedirs(NDVI_outfolder)
    if not os.path.exists(Albedo_outfolder):
        os.makedirs(Albedo_outfolder)
    if not os.path.exists(ALBEDO_outfolder_end):
        os.makedirs(ALBEDO_outfolder_end)
    if not os.path.exists(NDVI_outfolder_end):
        os.makedirs(NDVI_outfolder_end)

####################################################################################################################
################################### part 3: Create outlier maps PROBA-V ############################################
####################################################################################################################

    # Get the geo information of the example filename
    Example_file = os.path.join(Example_fileName)
    dest = gdal.Open(Example_file)
    ncol = dest.RasterXSize        # Get the reprojected dem column size
    nrow = dest.RasterYSize        # Get the reprojected dem row size
    shape=[ncol, nrow]

    ####################################### Create Water Mask based on PROBA-V  ##################################

    WaterMask_outfolder_end_FileName = os.path.join(WaterMask_outfolder_end,'Water_Mask.tif')

    if not os.path.exists(WaterMask_outfolder_end_FileName):

        # Find all the water mask
        os.chdir(WaterMask_outfolder)
        re_water_mask = glob.glob('Water_Mask*.tif')

        #  Loop over all the water mask files
        for water_mask_filename in re_water_mask:

            # Create the filepath to one water mask
            water_mask_filepath = os.path.join(WaterMask_outfolder,water_mask_filename)

            # Open the array of the water mask
            water_mask_dest = gdal.Open(water_mask_filepath)

            # Get the area of the example dataset if data is a subset of total !!!!
            if (water_mask_dest.RasterXSize is not shape[0] or water_mask_dest.RasterYSize is not shape[1]):

                # reproject dataset
                water_mask_dest, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(water_mask_filepath, Example_file)

            # If the total water mask raster does not exists create this one
            if not 'water_mask_array' in locals():

                water_mask_array = water_mask_dest.GetRasterBand(1).ReadAsArray()

            else:
                # Add all the water masks
                water_mask_array += water_mask_dest.GetRasterBand(1).ReadAsArray()

        # Calculate the end water mask if the area is more than 50 percent defined as water
        water_mask_array_per = water_mask_array/len(re_water_mask)
        water_mask_array_end = np.zeros(water_mask_array.shape)
        water_mask_array_end[water_mask_array_per > 0.25] = 1

        # Save water mask
        SEBAL.save_GeoTiff_proy(dest, water_mask_array_end, WaterMask_outfolder_end_FileName, shape, nband=1)

    else:
        # Open water mask
        dest_Water_mask = gdal.Open(WaterMask_outfolder_end_FileName)
        water_mask_array_end = dest_Water_mask.GetRasterBand(1).ReadAsArray()

    #################################################################################
    # Get General values of nc file
    #################################################################################

    fh_ndvi = Dataset(NDVI_netcdf)
    fh_alb = Dataset(Albedo_netcdf)

    longitudes = fh_ndvi.variables['longitude'][:]
    latitudes = fh_ndvi.variables['latitude'][:]
    time_nc = fh_ndvi.variables['time'][:]

    ######################################## Create tiff outliers  ##################################

    # open netcdf outliers
    fh_alb = Dataset(Albedo_netcdf, mode='r')
    fh_ndvi = Dataset(NDVI_netcdf, mode='r')

    # open geo information of the netcdf file
    lat = fh_alb.variables['latitude'][:]
    lon = fh_alb.variables['longitude'][:]
    minimum_lat = np.min(lat)
    minimum_lon = np.min(lon)
    maximum_lat = np.max(lat)
    maximum_lon = np.max(lon)
    diff_lon = 0.001
    diff_lat = 0.001
    geo = tuple([minimum_lon, diff_lon, 0, maximum_lat, 0, diff_lat])

    for i in PROBA_V_dict.keys():
        if len(i)>5:                                            # This is still dirty, normally I need to remove the None string from the PROBA_V_dict.keys()!

            # Get right time dimension
            date = datetime.strptime(i.split('_')[3], "%Y%m%d")
            time_now = date.strftime("%Y%m%d")
            DOY = int(date.strftime('%j'))
            time = DOY - 1
            Date_or = date.toordinal()
            IDz_hants = np.argwhere(time_nc == Date_or)[0][0]
            
            Outliers_alb = os.path.join(output_folder_PreSEBAL_sub,'Outliers_alb')
            Outliers_ndvi = os.path.join(output_folder_PreSEBAL_sub,'Outliers_ndvi')
            #Outliers_com = os.path.join(output_folder_PreSEBAL_sub,'Outliers_com')
            if not os.path.exists(Outliers_alb):
                os.makedirs(Outliers_alb)
            if not os.path.exists(Outliers_ndvi):
                os.makedirs(Outliers_ndvi)
            #if not os.path.exists(Outliers_com):
            #    os.makedirs(Outliers_com)
            
            # define output name
            name_out_alb = os.path.join(Outliers_alb, 'Outliers_PROBAV_albedo_%s.tif' %time_now)
            name_out_ndvi = os.path.join(Outliers_ndvi, 'Outliers_PROBAV_ndvi_%s.tif' %time_now)
            #name_out_com = os.path.join(Outliers_com, 'Outliers_PROBAV_%s.tif' %time_now)
            #name_out_hants_alb = os.path.join(output_folder, 'Albedo_HANTS_%s.tif' %time_now)
            #name_out_hants_ndvi = os.path.join(output_folder, 'NDVI_HANTS_%s.tif' %time_now)
            name_out_good_PROBAV_pixels = os.path.join(Cloud_free_PROBAV_Land_Pixels_outfolder, 'Cloud_free_PROBAV_Land_Pixels_%s.tif' %time_now)

            if not os.path.exists(name_out_good_PROBAV_pixels):

                # Open arrays
                # Open outliers
                data_alb = fh_alb.variables['outliers'][IDz_hants,:,:]
                data_ndvi = fh_ndvi.variables['outliers'][IDz_hants,:,:]
                # Open Simulated HANTS values
                # data_alb_sim = fh_alb.variables['hants_values'][time,:,:]
                # data_ndvi_sim = fh_ndvi.variables['hants_values'][time,:,:]

                # Combine outliers map
                good_PROBAV_pixels_ndvi = Create_Buffer(data_ndvi, 10)
                good_PROBAV_pixels_albedo = Create_Buffer(data_alb, 10)
                good_PROBAV_pixels = good_PROBAV_pixels_ndvi + good_PROBAV_pixels_albedo
                good_PROBAV_pixels[good_PROBAV_pixels >= 1.5] = 1
                good_PROBAV_pixels[water_mask_array_end == 1] = np.nan
                good_PROBAV_pixels[good_PROBAV_pixels == 1] = np.nan
                good_PROBAV_pixels[good_PROBAV_pixels == 0] = 1

                # save data
                #SEBAL.save_GeoTiff_proy(dest, np.flipud(data_alb_sim), name_out_hants_alb, shape, nband=1)
                #SEBAL.save_GeoTiff_proy(dest, np.flipud(data_ndvi_sim), name_out_hants_ndvi, shape, nband=1)
                SEBAL.save_GeoTiff_proy(dest, data_alb, name_out_alb, shape, nband=1)
                SEBAL.save_GeoTiff_proy(dest, data_ndvi, name_out_ndvi, shape, nband=1)
                #SEBAL.save_GeoTiff_proy(dest, data_com, name_out_com, shape, nband=1)
                SEBAL.save_GeoTiff_proy(dest, good_PROBAV_pixels, name_out_good_PROBAV_pixels, shape, nband=1)

    #################################################################################
    # Get the highest NDVI values and determine which pixels needs to be corrected
    #################################################################################

    # Define the dates
    Dates = pd.date_range(start_date, end_date, freq = "D")
    monthly_Dates = pd.date_range(start_date, end_date, freq = "MS")

    if Apply_Corrections_using_original == 1:

        # Search for the high NDVI pixels by taking the first values of the month
        i = 0
        time_or_month = np.zeros(len(monthly_Dates))
        for Date_month in monthly_Dates:
            time_or_month[i] = int(Date_month.strftime('%j'))
            i += 1

        # Get the NDVI of all the dates
        NDVI_all_dates = np.zeros([len(monthly_Dates), shape[1], shape[0]])

        for i in range(0,len(time_or_month)):
            ndvi_one_date = fh_ndvi.variables['hants_values'][time_or_month[i],:,:]
            NDVI_all_dates[i,:,:] = ndvi_one_date

        NDVI_max = np.max(NDVI_all_dates, axis = 0)
        NDVI_max[NDVI_max==-9999] = np.nan
        NDVI_value = np.nanpercentile(NDVI_max, 0)

        NDVI_calc = np.zeros([shape[1], shape[0]])
        NDVI_calc[NDVI_max>NDVI_value] = 1
        NDVI_calc_flatten = NDVI_calc.flatten()
        pixels = np.argwhere(NDVI_calc == 1)

        # fractions layer
        fraction_layer_ndvi = np.zeros([len(Dates), len(pixels)])
        fraction_layer_albedo = np.zeros([len(Dates), len(pixels)])

        # delete some large parameters
        del NDVI_all_dates

        #################################### Calculate fractions #######################################

        for Date in Dates:

            # specify time units
            year = Date.year
            month = Date.month
            day = Date.day
            Date_or = Date.toordinal()

            # Define filenames
            fileName_mask = os.path.join(Cloud_free_PROBAV_Land_Pixels_outfolder, "Cloud_free_PROBAV_Land_Pixels_%d%02d%02d.tif" %(year, month, day))
            fileName_NDVI = os.path.join(NDVI_outfolder,"NDVI_PROBAV_%d%02d%02d.tif" %(year, month, day))
            fileName_Albedo = os.path.join(Albedo_outfolder,"Albedo_PROBAV_%d%02d%02d.tif" %(year, month, day))

            # Define time unit
            IDz = int(Date.strftime("%j")) - int(Dates[0].strftime("%j"))
            IDz_hants = np.argwhere(time_nc == Date_or)[0][0]

            if os.path.exists(fileName_NDVI):

                # Reproject dataset
                dest_NDVI, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(fileName_NDVI, Example_file)
                dest_Albedo, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(fileName_Albedo, Example_file)
                dest_mask = gdal.Open(fileName_mask)

                # Open required data
                NDVI_HANTS = fh_ndvi.variables['hants_values'][IDz_hants,:,:]
                Albedo_HANTS = fh_alb.variables['hants_values'][IDz_hants,:,:]
                NDVI_HANTS[NDVI_HANTS == -9999] = np.nan
                Albedo_HANTS[Albedo_HANTS == -9999] = np.nan
                NDVI_RAW = dest_NDVI.GetRasterBand(1).ReadAsArray()* NDVI_calc
                Albedo_RAW = dest_Albedo.GetRasterBand(1).ReadAsArray() * NDVI_calc
                MASK_PROBA_V = dest_mask.GetRasterBand(1).ReadAsArray() * NDVI_calc

                # Apply fraction calculations
                Fractions_NDVI = (NDVI_RAW * MASK_PROBA_V)/ NDVI_HANTS
                Fractions_NDVI[np.logical_or(Fractions_NDVI<1,water_mask_array_end == 1)] = np.nan  
                Fractions_Albedo = (Albedo_RAW * MASK_PROBA_V)/ Albedo_HANTS
                Fractions_Albedo[np.logical_or(Fractions_Albedo>1,water_mask_array_end == 1)] = np.nan               
                Fractions_NDVI[np.isnan(Fractions_NDVI)]=0
                Fractions_Albedo[np.isnan(Fractions_Albedo)]=0
                Fractions_NDVI_flatten = Fractions_NDVI.flatten()
                Fractions_Albedo_flatten = Fractions_Albedo.flatten()
                Fractions_NDVI_flatten = Fractions_NDVI_flatten[NDVI_calc_flatten==1]
                Fractions_Albedo_flatten = Fractions_Albedo_flatten[NDVI_calc_flatten==1]

                # write time in 2d array
                fraction_layer_ndvi[IDz,:] = Fractions_NDVI_flatten
                fraction_layer_albedo[IDz,:] = Fractions_Albedo_flatten

         #################################### interpolate fractinos #######################################

        # Set the end and begin of the fractions at 1 if there is no fraction
        fraction_layer_ndvi[0,:][fraction_layer_ndvi[0,:]==0] = 1
        fraction_layer_albedo[0,:][fraction_layer_albedo[0,:]==0] = 1
        fraction_layer_ndvi[-1,:][fraction_layer_ndvi[-1,:]==0] = 1
        fraction_layer_albedo[-1,:][fraction_layer_albedo[-1,:]==0] = 1

        # interpolate by looping over the pixels
        for j in range(0, fraction_layer_albedo.shape[1]):

            # Set general paramters
            y = np.array(range(0,len(Dates)))

            # Apply interpolation on NDVI
            x_ndvi = np.array(fraction_layer_ndvi[:,j])
            y_good_ndvi = np.array(y)[x_ndvi != 0.]
            x_ndvi_good = np.array(x_ndvi)[x_ndvi != 0.]
            f_ndvi = interpolate.interp1d(y_good_ndvi, x_ndvi_good)
            x_ndvi_inter = f_ndvi(y)
            fraction_layer_ndvi[:,j] = x_ndvi_inter

            # Apply interpolation on Albedo
            x_alb = np.array(fraction_layer_albedo[:,j])
            y_good_alb = np.array(y)[x_alb != 0.]
            x_alb_good = np.array(x_alb)[x_alb != 0.]
            f_alb = interpolate.interp1d(y_good_alb, x_alb_good)
            x_alb_inter = f_alb(y)
            fraction_layer_albedo[:,j] = x_alb_inter

    # Apply the fractions over the simulated HANTS albedo and NDVI values
    for Date in Dates:

        Var_name = Date.strftime("%Y%m%d")

        # filenames
        name_out_ndvi_end = os.path.join(NDVI_outfolder_end, 'NDVI_%s.tif' %Var_name)
        name_out_albedo_end = os.path.join(ALBEDO_outfolder_end, 'Albedo_%s.tif' %Var_name)
        tir_emis_FileName = os.path.join(output_folder_tir_emis,'tir_b10_emis_%s.tif' %Var_name)

        if not (os.path.exists(name_out_ndvi_end) and os.path.exists(name_out_albedo_end) and os.path.exists(tir_emis_FileName)):

            if Apply_Corrections_using_original == 1:
                # specify time units
                year = Date.year
                month = Date.month
                day = Date.day
                Date_or = Date.toordinal()

                # Define time unit
                IDz = int(Date.strftime("%j")) - int(Dates[0].strftime("%j"))
                IDz_hants = np.argwhere(time_nc == Date_or)[0][0]

                # Define output filenames
                fileName_fraction_NDVI = os.path.join(NCfractions_outfolder_NDVI, "Fractions_NDVI_%d%02d%02d.tif" %(year, month, day))
                fileName_fraction_Albedo = os.path.join(NCfractions_outfolder_Albedo, "Fractions_Albedo_%d%02d%02d.tif" %(year, month, day))

                # create the fraction map
                fraction_map_one_day_NDVI = np.ones(NDVI_calc.shape)
                fraction_map_one_day_Albedo = np.ones(NDVI_calc.shape)
                one_fraction_time_NDVI = fraction_layer_ndvi[IDz,:]
                one_fraction_time_Albedo = fraction_layer_albedo[IDz,:]
                fraction_map_one_day_NDVI[NDVI_calc==1] = one_fraction_time_NDVI
                fraction_map_one_day_Albedo[NDVI_calc==1] = one_fraction_time_Albedo

                # Open HANTS values
                NDVI_HANTS = fh_ndvi.variables['hants_values'][IDz,:,:]
                Albedo_HANTS = fh_alb.variables['hants_values'][IDz,:,:]

                # multiply hants values with fractions
                NDVI_new = NDVI_HANTS * fraction_map_one_day_NDVI
                Albedo_new = Albedo_HANTS * fraction_map_one_day_Albedo
                
                # Save results
                SEBAL.save_GeoTiff_proy(dest, fraction_map_one_day_NDVI, fileName_fraction_NDVI, shape, nband=1)
                SEBAL.save_GeoTiff_proy(dest, fraction_map_one_day_Albedo, fileName_fraction_Albedo, shape, nband=1)

            if Apply_Corrections_using_original == 0:

                ################################# Get HANTS albedo and NDVI #############################################
                # Variable date name
                DOY = int(Date.strftime('%j'))
                time = DOY - 1
                Date_or = Date.toordinal()

                # Define time unit
                IDz = int(Date.strftime("%j")) - int(Dates[0].strftime("%j"))
                IDz_hants = np.argwhere(time_nc == Date_or)[0][0]

                # Open Simulated HANTS values
                data_ndvi_sim = fh_alb.variables['hants_values'][IDz_hants,:,:]
                data_alb_sim = fh_ndvi.variables['hants_values'][IDz_hants,:,:]

                # flip NDVI and albedo
                NDVI_new = data_ndvi_sim
                Albedo_new = data_alb_sim
                                
    ################################## Save End products NDVI and Albedo #################################

        # Limit albedo and NDVI
        NDVI_new = NDVI_new.clip(-0.3, 1.0)
        Albedo_new = Albedo_new.clip(0.0, 0.9)
                
        NDVI_new[NDVI_new == -9999] = -0.1
        Albedo_new[Albedo_new == -9999] = 0.02

        # Save End NDVI and Albedo
        SEBAL.save_GeoTiff_proy(dest, NDVI_new, name_out_ndvi_end, shape, nband=1)
        SEBAL.save_GeoTiff_proy(dest, Albedo_new, name_out_albedo_end, shape, nband=1)

        ################################# Calculate Emissivity #######################################

        # Calculate the emissivity
        FPAR,tir_emis,Nitrogen,vegt_cover,LAI,b10_emissivity = SEBAL.Calc_vegt_para(NDVI_new, water_mask_array_end, shape)

        # Save Water Mask for PROBA-V
        SEBAL.save_GeoTiff_proy(dest, b10_emissivity, tir_emis_FileName, shape, nband=1)

    ################################## Calculate VIIRS surface temperature ########################

    for number in Kind_Of_Runs_Dict[2][:]: # Number defines the column of the inputExcel

        # Extract the name of the input folder for VIIRS data
        input_folder_SEBAL = '%s' %str(ws['B%d' % number].value)

        # Extract the name of the thermal and quality VIIRS image from the excel file
        Name_VIIRS_Image_TB = '%s' %str(ws2['B%d' % number].value)

        # If there is VIIRS data
        if not Name_VIIRS_Image_TB == 'None':


            ################################# General information ####################################

            # Get date of the image
            datum = Name_VIIRS_Image_TB.split('_')[3][:]

            # Define the VIIRS thermal data name
            VIIRS_data_name=os.path.join(input_folder_SEBAL, '%s' % (Name_VIIRS_Image_TB))

            # Get time
            hour = int(Name_VIIRS_Image_TB.split('_')[4][:2])
            minutes = int(Name_VIIRS_Image_TB.split('_')[4][2:4])

            # Get the variable name
            #Var_name_VIIRS = str(datum + "_H%02dM%02d" %(hour, minutes))
            Var_name_VIIRS = str(datum)
            Var_name = datum

            temp_surface_375_fileName_beforeTS = os.path.join(Surface_Temperature_outfolder,'Surface_Temperature_VIIRS_%s.tif' %Var_name_VIIRS)

            if (os.path.exists(temp_surface_375_fileName_beforeTS) or str(os.path.splitext(VIIRS_data_name)[-1]) == ".fake"):
                continue
            
            else:
                #################################### Contants ###########################################

                Rp = 0.91                        # Path radiance in the 10.4-12.5 µm band (W/m2/sr/µm)
                tau_sky = 0.866                  # Narrow band transmissivity of air, range: [10.4-12.5 µm]
                k1=606.399172
                k2=1258.78

                ################################## VIIRS Temperature ####################################

                data_VIIRS_clipped, Geo_out_VIIRS_clipped = clip_data(VIIRS_data_name, [minimum_lat - 0.02, maximum_lat+0.02], [minimum_lon - 0.02, maximum_lon + 0.02])
                if len(data_VIIRS_clipped) is not 0:
                    dest_VIIRS_clipped = Save_as_MEM(data_VIIRS_clipped, Geo_out_VIIRS_clipped, projection='WGS84')
                    shape_VIIRS_clipped = [dest_VIIRS_clipped.RasterXSize, dest_VIIRS_clipped.RasterYSize]

                ############################# brightness temperature ####################################
    
                    # Set the conditions for the brightness temperature (100m)
                    brightness_temp=np.where(data_VIIRS_clipped>=250, data_VIIRS_clipped, np.nan)
    
                    # Constants
                    L_lambda_b10_375=((2*6.63e-34*(3.0e8)**2)/((11.45e-6)**5*(np.exp((6.63e-34*3e8)/(1.38e-23*(11.45e-6)*brightness_temp))-1)))*1e-6

                ################################ tir emissivity #########################################

                    # Define the tir_emis_b10_filename
                    tir_emis_FileName = os.path.join(output_folder_tir_emis,'tir_b10_emis_%s.tif' %Var_name)
    
                    # reproject tir emis
                    tir_emis_375_dest, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(
                                          tir_emis_FileName, dest_VIIRS_clipped)
    
                    # Open tir emis data
                    tir_emis_375 = tir_emis_375_dest.GetRasterBand(1).ReadAsArray()
                    tir_emis_375 = gap_filling(tir_emis_375,0)

                ############################# instantanious temperature #################################

                    # Open path to Instantanious temperature
                    try:
                        Temp_inst_375 = float(ws3['B%d' % number].value)
                    except:
                        Temp_inst_name = '%s' %str(ws3['B%d' % number].value)
    
                        # reproject instantanious temperature
                        Temp_inst_375_dest, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(
                                              Temp_inst_name, dest_VIIRS_clipped)
    
                        # Open instantanious temperature
                        Temp_inst_375 = Temp_inst_375_dest.GetRasterBand(1).ReadAsArray()

                ################################ tir emissivity #########################################

                    # Get Temperature for 100 and 375m resolution
                    Temp_TOA_375 = SEBAL.Get_Thermal(L_lambda_b10_375,Rp,Temp_inst_375,tau_sky,tir_emis_375,k1,k2)
    
                    # Conditions for surface temperature (100m)
                    n120_surface_temp=Temp_TOA_375.clip(250, 450)
    
                    # Save the surface temperature of the VIIRS in 100m resolution
                    SEBAL.save_GeoTiff_proy(dest_VIIRS_clipped, n120_surface_temp, temp_surface_375_fileName_beforeTS, shape_VIIRS_clipped, nband=1)
                    del Temp_inst_375

# Functions
#################################################################################

def Create_Buffer(Data_In, Buffer_area):

   '''
   This function creates a 3D array which is used to apply the moving window
   '''

   # Buffer_area = 2 # A block of 2 times Buffer_area + 1 will be 1 if there is the pixel in the middle is 1
   Data_Out=np.empty((len(Data_In),len(Data_In[1])))
   Data_Out[:,:] = Data_In
   for ypixel in range(0,Buffer_area + 1):

        for xpixel in range(1,Buffer_area + 1):

           if ypixel==0:
                for xpixel in range(1,Buffer_area + 1):
                    Data_Out[:,0:-xpixel] += Data_In[:,xpixel:]
                    Data_Out[:,xpixel:] += Data_In[:,:-xpixel]

                for ypixel in range(1,Buffer_area + 1):

                    Data_Out[ypixel:,:] += Data_In[:-ypixel,:]
                    Data_Out[0:-ypixel,:] += Data_In[ypixel:,:]

           else:
               Data_Out[0:-xpixel,ypixel:] += Data_In[xpixel:,:-ypixel]
               Data_Out[xpixel:,ypixel:] += Data_In[:-xpixel,:-ypixel]
               Data_Out[0:-xpixel,0:-ypixel] += Data_In[xpixel:,ypixel:]
               Data_Out[xpixel:,0:-ypixel] += Data_In[:-xpixel,ypixel:]

   Data_Out[Data_Out>0.1] = 1
   Data_Out[Data_Out<=0.1] = 0

   return(Data_Out)

#__________________________________________________________________________________
def clip_data(input_file, latlim, lonlim):
    """
    Clip the data to the defined extend of the user (latlim, lonlim) or to the
    extend of the DEM tile

    Keyword Arguments:
    input_file -- output data, output of the clipped dataset
    latlim -- [ymin, ymax]
    lonlim -- [xmin, xmax]
    """
    try:
        if input_file.split('.')[-1] == 'tif':
            dest_in = gdal.Open(input_file)
        else:
            dest_in = input_file
    except:
        dest_in = input_file

    # Open Array
    data_in = dest_in.GetRasterBand(1).ReadAsArray()

    # Define the array that must remain
    Geo_in = dest_in.GetGeoTransform()
    Geo_in = list(Geo_in)
    Start_x = np.max([int(np.ceil(((lonlim[0]) - Geo_in[0])/ Geo_in[1])),0])
    End_x = np.min([int(np.floor(((lonlim[1]) - Geo_in[0])/ Geo_in[1])),int(dest_in.RasterXSize)])

    Start_y = np.max([int(np.floor((Geo_in[3] - latlim[1])/ -Geo_in[5])),0])
    End_y = np.min([int(np.ceil(((latlim[0]) - Geo_in[3])/Geo_in[5])), int(dest_in.RasterYSize)])

    #Create new GeoTransform
    Geo_in[0] = Geo_in[0] + Start_x * Geo_in[1]
    Geo_in[3] = Geo_in[3] + Start_y * Geo_in[5]
    Geo_out = tuple(Geo_in)

    if (End_y < Start_y) or (End_x < Start_x):
        print('%s contains not the area lat: %d - %d, lon: %d - %d' %(input_file, latlim[0], latlim[1], lonlim[0], lonlim[1]))
        data = []
    else:
        data = np.zeros([End_y - Start_y, End_x - Start_x])
        data = data_in[Start_y:End_y,Start_x:End_x]
    dest_in = None

    return(data, Geo_out)

def Save_as_MEM(data='', geo='', projection=''):
    """
    This function save the array as a memory file

    Keyword arguments:
    data -- [array], dataset of the geotiff
    geo -- [minimum lon, pixelsize, rotation, maximum lat, rotation,
            pixelsize], (geospatial dataset)
    projection -- interger, the EPSG code
    """
    # save as a geotiff
    driver = gdal.GetDriverByName("MEM")
    dst_ds = driver.Create('', int(data.shape[1]), int(data.shape[0]), 1,
                           gdal.GDT_Float32, ['COMPRESS=LZW'])
    srse = osr.SpatialReference()
    if projection == '':
        srse.SetWellKnownGeogCS("WGS84")
    else:
        srse.SetWellKnownGeogCS(projection)
    dst_ds.SetProjection(srse.ExportToWkt())
    dst_ds.GetRasterBand(1).SetNoDataValue(-9999)
    dst_ds.SetGeoTransform(geo)
    dst_ds.GetRasterBand(1).WriteArray(data)
    return(dst_ds)

def gap_filling(data,NoDataValue, method = 1):
    """
    This function fills the no data gaps in a numpy array

    Keyword arguments:
    dataset -- 'C:/'  path to the source data (dataset that must be filled)
    NoDataValue -- Value that must be filled
    """

    # fill the no data values
    if NoDataValue is np.nan:
        mask = ~(np.isnan(data))
    else:
        mask = ~(data==NoDataValue)
    xx, yy = np.meshgrid(np.arange(data.shape[1]), np.arange(data.shape[0]))
    xym = np.vstack( (np.ravel(xx[mask]), np.ravel(yy[mask])) ).T
    data0 = np.ravel( data[:,:][mask] )

    if method == 1:
        interp0 = interpolate.NearestNDInterpolator( xym, data0 )
        data_end = interp0(np.ravel(xx), np.ravel(yy)).reshape( xx.shape )

    if method == 2:
        interp0 = interpolate.LinearNDInterpolator( xym, data0 )
        data_end = interp0(np.ravel(xx), np.ravel(yy)).reshape( xx.shape )


    EndProduct = data_end

    return(EndProduct)