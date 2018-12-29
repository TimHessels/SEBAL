# -*- coding: utf-8 -*-
"""
Created on Sun Jun 03 06:15:11 2018

@author: tih
"""
import numpy as np
import os
import gdal
from openpyxl import load_workbook
import osr
from datetime import datetime
import pandas as pd
from netCDF4 import Dataset
from pyproj import Proj, transform
import scipy.interpolate

def main(VegetationExcel, KARIMET_netcdf, LST_netcdf):

####################################################################################################################
############################################# CREATE INPUT FOR preSEBAL RUN ########################################
####################################################################################################################

####################################################################################################################
##################################################### PreSEBAL_3 ###################################################
####################################################################################################################

# PreSEBAL_2
# Part 1: Define input by user
# Part 2: Set parameters and output folder
# Part 3: Calculate NDVI and Albedo based on PROBA-V

####################################################################################################################
#################################### part 1: Define input by user ##################################################
####################################################################################################################
  
    import SEBAL.pySEBAL.pySEBAL_code as SEBAL
    
####################################################################################################################
################################# part 2: Set parameters and output folder #########################################
####################################################################################################################

    # Open Excel workbook used for Vegetation c and p factor conversions
    wb_veg = load_workbook(VegetationExcel, data_only=True)
    ws_veg = wb_veg['General_Input']

    # Input for preSEBAL.py
    start_date = "%s" %str(ws_veg['B2'].value)
    end_date = "%s" %str(ws_veg['B3'].value)
    # end_date = "2016-01-31"
    inputExcel= r"%s" %str(ws_veg['B4'].value)              # The excel with all the SEBAL input data
    output_folder = r"%s" %str(ws_veg['B7'].value)
    LU_data_FileName = r"%s" %str(ws_veg['B5'].value)       # Path to Land Use map

    ######################## Load Excels ##########################################
    # Open Excel workbook for SEBAL inputs
    wb = load_workbook(inputExcel, data_only=True)

    # Get length of EXCEL sheet
    ws = wb['General_Input']
    ws2 = wb['VIIRS_PROBAV_Input']
    endExcel=int(ws.max_row)

    # Create Dict
    SEBAL_RUNS = dict()

    for number in range(2,endExcel+1):
        input_folder_SEBAL = str(ws['B%d' % number].value)
        output_folder_SEBAL = str(ws['C%d' % number].value)
        Image_Type = int(ws['D%d' % number].value)
        PROBA_V_name  = str(ws2['D%d' % number].value)
        VIIRS_name  = str(ws2['B%d' % number].value)
        SEBAL_RUNS[number] = {'input_folder': input_folder_SEBAL, 'output_folder': output_folder_SEBAL, 'image_type': Image_Type,'PROBA_V_name': PROBA_V_name,'VIIRS_name': VIIRS_name}

    Kind_Of_Runs_Dict = {}
    for k, v in iter(SEBAL_RUNS.items()):
        Kind_Of_Runs_Dict.setdefault(v['image_type'], []).append(k)

    PROBA_V_dict = {}
    for k, v in iter(SEBAL_RUNS.items()):
        PROBA_V_dict.setdefault(v['PROBA_V_name'], []).append(k)

    # Define dates
    Dates = pd.date_range(start_date, end_date, freq = "D")

    ######################## Create output folders  ##########################################

    # output maps for main directories
    output_folder_PreSEBAL_sub = os.path.join(output_folder,'PreSEBAL_SEBAL_out')         # outputs before HANTS
    output_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_out')                     # End outputs of preSEBAL
    temp_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_temp')                      # Temp outputs of preSEBAL

    # output maps for the end products
    ALBEDO_outfolder_end = os.path.join(output_folder_PreSEBAL,'ALBEDO')
    NDVI_outfolder_end = os.path.join(output_folder_PreSEBAL,'NDVI')
    WaterMask_outfolder_end = os.path.join(output_folder_PreSEBAL,'Water_Mask')
    LAI_outfolder = os.path.join(output_folder_PreSEBAL,'LAI')
    TRANS_outfolder = os.path.join(output_folder_PreSEBAL,'Transmissivity')
    output_folder_HANTS_end_sharp = os.path.join(output_folder_PreSEBAL, 'LST_Sharpened')
    output_folder_HANTS_end_Veg = os.path.join(output_folder_PreSEBAL, 'Vegetation_Height')
    output_folder_p_factor =  os.path.join(output_folder_PreSEBAL, 'p_factor')
    output_folder_LUE =  os.path.join(output_folder_PreSEBAL, 'LUE')
    Water_mask_output_tiff = os.path.join(output_folder_PreSEBAL,'Water_Mask','Water_Mask.tif')
    
    # preSEBAL side outputs
    output_folder_HANTS_outliers_PROBAV_combined = os.path.join(output_folder_PreSEBAL_sub, 'Outliers_PROBAV_combined')
    output_folder_HANTS_outliers_PROBAV_combined_buffer = os.path.join(output_folder_PreSEBAL_sub, 'Outliers_PROBAV_combined_buffer')
    Surface_Temperature_outfolder = os.path.join(output_folder_PreSEBAL_sub,'Surface_Temperature')

    # preSEBAL temp outputs
    output_folder_HANTS_outliers_PROBAV_albedo = os.path.join(temp_folder_PreSEBAL, 'Outliers_PROBAV_albedo')
    output_folder_HANTS_outliers_PROBAV_ndvi = os.path.join(temp_folder_PreSEBAL, 'Outliers_PROBAV_ndvi')
    output_folder_HANTS_value_PROBAV_albedo = os.path.join(temp_folder_PreSEBAL, 'HANTS_Value_PROBAV_albedo')
    output_folder_HANTS_value_PROBAV_ndvi = os.path.join(temp_folder_PreSEBAL, 'HANTS_Value_PROBAV_ndvi')
    output_folder_tir_emis =  os.path.join(temp_folder_PreSEBAL, 'tir_emis')

    # file name of the example dataset based on DEM
    Example_fileName = os.path.join(temp_folder_PreSEBAL, "DEM_Example_Projection.tif")

    if not os.path.exists(LAI_outfolder):
        os.makedirs(LAI_outfolder)
    if not os.path.exists(WaterMask_outfolder_end):
        os.makedirs(WaterMask_outfolder_end)
    if not os.path.exists(Surface_Temperature_outfolder):
        os.makedirs(Surface_Temperature_outfolder)
    if not os.path.exists(TRANS_outfolder):
        os.makedirs(TRANS_outfolder)
    if not os.path.exists(output_folder_HANTS_end_sharp):
        os.mkdir(output_folder_HANTS_end_sharp)
    if not os.path.exists(output_folder_HANTS_end_Veg):
        os.mkdir(output_folder_HANTS_end_Veg)
    if not os.path.exists(output_folder_p_factor):
        os.mkdir(output_folder_p_factor)
    if not os.path.exists(output_folder_LUE):
        os.mkdir(output_folder_LUE)
    if not os.path.exists(output_folder_tir_emis):
        os.mkdir(output_folder_tir_emis)
    if not os.path.exists(output_folder_HANTS_outliers_PROBAV_albedo):
        os.mkdir(output_folder_HANTS_outliers_PROBAV_albedo)
    if not os.path.exists(output_folder_HANTS_outliers_PROBAV_ndvi):
        os.mkdir(output_folder_HANTS_outliers_PROBAV_ndvi)
    if not os.path.exists(output_folder_HANTS_outliers_PROBAV_albedo):
        os.mkdir(output_folder_HANTS_outliers_PROBAV_combined)
    if not os.path.exists(output_folder_HANTS_value_PROBAV_albedo):
        os.mkdir(output_folder_HANTS_value_PROBAV_albedo)
    if not os.path.exists(output_folder_HANTS_value_PROBAV_ndvi):
        os.mkdir(output_folder_HANTS_value_PROBAV_ndvi)
    if not os.path.exists(output_folder_HANTS_outliers_PROBAV_combined_buffer):
        os.mkdir(output_folder_HANTS_outliers_PROBAV_combined_buffer)

    ################################################### Open Land Use map #############################################

    if not str(LU_data_FileName) == 'None':
        # reproject LU map to example file
        dest_LU, ulx_LU, lry_LU, lrx_LU, uly_LU, epsg_to = SEBAL.reproject_dataset_example(
                       LU_data_FileName, Example_fileName)
        LU_data = dest_LU.GetRasterBand(1).ReadAsArray()
 
    else:
        dest_ex = gdal.Open(Example_fileName)
        Size_x = dest_ex.RasterXSize
        Size_y = dest_ex.RasterYSize
        LU_data = np.ones([Size_y, Size_x])

    ############################################### Apply thermal sharpening ##########################################

    print('---------------------------------------------------------')
    print('-------------------- Downscale VIIRS --------------------')
    print('---------------------------------------------------------')

    # Upscale VIIRS and PROBA-V to 400m
    pixel_spacing_upscale = 0.004

    # Reproject from Geog Coord Syst to UTM -
    # 1) DEM - Original DEM coordinates is Geographic: lat, lon
    proyDEM_fileName_100 = os.path.join(temp_folder_PreSEBAL, 'DEM_Example_Projection.tif')
    dest = gdal.Open(proyDEM_fileName_100)
    ncol = dest.RasterXSize        # Get the reprojected dem column size
    nrow = dest.RasterYSize        # Get the reprojected dem row size
    shape=[ncol, nrow]

    # Create upscaled DEM
    proyDEM_fileName_400 = os.path.join(temp_folder_PreSEBAL,'DEM_Example_Projection_400.tif')
    dest_400, ulx_dem_400, lry_dem_400, lrx_dem_400, uly_dem_400, epsg_to = reproject_dataset_to_example(proyDEM_fileName_100, pixel_spacing_upscale, 4326)

    # find spatial parameters array
    DEM_400 = dest_400.GetRasterBand(1).ReadAsArray()
    Y_raster_size_400 = dest_400.RasterYSize
    X_raster_size_400 = dest_400.RasterXSize
    shape_400=([X_raster_size_400, Y_raster_size_400])

    # Save DEM file with the 400 meter resolution
    SEBAL.save_GeoTiff_proy(dest_400, DEM_400, proyDEM_fileName_400, shape_400, nband=1)

    # Open water mask based on proba v time series
    fileName_Watermask = os.path.join(WaterMask_outfolder_end, "Water_Mask.tif")
    dest_waterMask, ulx_ex, lry_ex, lrx_ex, uly_ex, epsg_to = SEBAL.reproject_dataset_example(
                          fileName_Watermask, Example_fileName)
    water_mask_proba_V_based = dest_waterMask.GetRasterBand(1).ReadAsArray()

    # Watermask upscale to 400
    Water_mask_Upscale, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SEBAL.reproject_dataset_example(
                  fileName_Watermask, proyDEM_fileName_400)
    watermask_400= Water_mask_Upscale.GetRasterBand(1).ReadAsArray()

    # Open the lST map from HANTS
    fh = Dataset(KARIMET_netcdf)
    time = fh.variables["time"][:]

    fh_HANTS = Dataset(LST_netcdf)
    time_HANTS = fh_HANTS.variables["time"][:]

    for date in Dates:

        Var_Name_VIIRS = "%d%02d%02d" %(date.year, date.month, date.day)
        IDz = np.argwhere(time == date.toordinal())[0][0]
        IDz_HANTS = np.argwhere(time_HANTS == date.toordinal())[0][0]
        
        # Define output names
        surf_temp_fileName = os.path.join(output_folder_HANTS_end_sharp, 'LST_sharpened_%s.tif' %(Var_Name_VIIRS))
        LAI_FileName = os.path.join(LAI_outfolder,'LAI_%d%02d%02d.tif' %(date.year, date.month, date.day))

        if not (os.path.exists(LAI_FileName) and os.path.exists(surf_temp_fileName)):

            ################################ Thermal Sharpening #####################################################

            # Define filename
            file_NDVI_after_HANTS = os.path.join(NDVI_outfolder_end, 'NDVI_%d%02d%02d.tif' %(date.year, date.month, date.day))

            # Open NDVI/LST destination folder
            dest_NDVI, ulx_ex, lry_ex, lrx_ex, uly_ex, epsg_to = SEBAL.reproject_dataset_example(
                          file_NDVI_after_HANTS, Example_fileName)
            NDVI = dest_NDVI.GetRasterBand(1).ReadAsArray()

            # Open data netcdf LST before sharpening
            LST_before = fh.variables['reconst_lin'][IDz,:,:] + 273.15
            LST_HANTS = fh_HANTS.variables['hants_values'][IDz_HANTS,:,:]
            
            # Get geostationary dataset netcdf
            lat = fh.variables['latitude'][:]
            lon = fh.variables['longitude'][:]
            diff_lon = 0.00375
            diff_lat = -0.00375
            minimum_lon = np.min(lon) - 0.5 * diff_lon
            maximum_lat = np.max(lat) - 0.5 * diff_lat
            geo = tuple([minimum_lon, diff_lon, 0, maximum_lat, 0, diff_lat])

            # Create memory file for LST before sharpening
            dest_LST_before = Save_as_MEM(LST_before, geo, 'WGS84')

            # Create memory file for LST before sharpening
            dest_LST_HANTS = Save_as_MEM(LST_HANTS, geo, 'WGS84')


            # Open original LST array
            dest_LST, ulx_ex, lry_ex, lrx_ex, uly_ex, epsg_to = SEBAL.reproject_dataset_example(
                          dest_LST_before, Example_fileName)
            LST = dest_LST.GetRasterBand(1).ReadAsArray()

            # Open original LST array
            dest_LST_HANTS, ulx_ex, lry_ex, lrx_ex, uly_ex, epsg_to = SEBAL.reproject_dataset_example(
                          dest_LST_HANTS, Example_fileName)
            LST_HANTS = dest_LST_HANTS.GetRasterBand(1).ReadAsArray()

            # replace the zeroes
            LST = gap_filling(LST,0)
            LST_HANTS = gap_filling(LST_HANTS,0)
            
            # Upscale thermal band VIIRS from 100m to 400m
            dest_LST_400, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SEBAL.reproject_dataset_example(
                           dest_LST_before, proyDEM_fileName_400)
            data_Temp_Surf_400 = dest_LST_400.GetRasterBand(1).ReadAsArray()
            data_Temp_Surf_400[data_Temp_Surf_400<200] = 0

            # replace the zeroes
            data_Temp_Surf_400 = gap_filling(data_Temp_Surf_400,0)

            # Upscale PROBA-V NDVI from 100m to 400m
            dest_NDVI_400, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SEBAL.reproject_dataset_example(
                          file_NDVI_after_HANTS, proyDEM_fileName_400)
            data_NDVI_400 = dest_NDVI_400.GetRasterBand(1).ReadAsArray()

            # Define the width of the moving window box
            Box=29

            # Apply the surface temperature sharpening
            temp_surface_sharpened = SEBAL.Thermal_Sharpening_Linear(data_Temp_Surf_400, data_NDVI_400, NDVI, Box, dest_NDVI_400, temp_folder_PreSEBAL, proyDEM_fileName_100, shape, dest, watermask_400)

            # Divide temporal watermask in snow and water mask by using surface temperature
            Snow_Mask_PROBAV, water_mask, ts_moist_veg_min, NDVI_max, NDVI_std = SEBAL.CalculateSnowWaterMask(NDVI,shape,water_mask_proba_V_based,temp_surface_sharpened)

            # Add the long term watermask
            dest_water_mask_long_term = gdal.Open(Water_mask_output_tiff)
            water_mask_long_term = dest_water_mask_long_term.GetRasterBand(1).ReadAsArray()

            # Replace water values
            temp_surface_sharpened[water_mask==1] = LST[water_mask == 1]
            temp_surface_sharpened = np.where(np.isnan(temp_surface_sharpened), LST, temp_surface_sharpened)

            # Replace water values for the long term water mask
            temp_surface_sharpened[water_mask_long_term==1] = LST_HANTS[water_mask_long_term == 1]

            SEBAL.save_GeoTiff_proy(dest, temp_surface_sharpened, surf_temp_fileName, shape, nband=1)

    ################################################## Calculate LAI ##################################################

            # Calculate LAI
            FPAR, tir_emis, Nitrogen, vegt_cover, LAI, b10_emissivity = SEBAL.Calc_vegt_para(NDVI,water_mask, shape)
            SEBAL.save_GeoTiff_proy(dest, LAI, LAI_FileName, shape, nband=1)

        ################################ Calculate the Vegetation height ########################

        veg_height_FileName = os.path.join(output_folder_HANTS_end_Veg,'Vegetation_Height_%d%02d%02d.tif' %(date.year, date.month, date.day))

        if not os.path.exists(veg_height_FileName):

            # Open preprosessing excel the Vegetation_Height sheet
            ws_veg = wb_veg['Vegetation_Height']

            # Read out the excel file coefficient numbers
            Array = np.zeros([ws_veg.max_row-1,4])
            for j in ['A','C','D','E']:
                j_number={'A' : 0, 'C' : 1, 'D' : 2, 'E' : 3}
                for i in range(2,ws_veg.max_row+1):
        	        Value = (ws_veg['%s%s' %(j,i)].value)
        	        Array[i-2, j_number[j]] = Value

            # Create maps with the coefficient numbers for the right land cover
            coeff = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1]),3])
            for coeff_nmbr in range(0,3):
                for Class in range(0,len(Array)):
        	        coeff[LU_data==Array[Class,0],coeff_nmbr] = Array[Class,coeff_nmbr+1]

            # Calculate the vegetation height in the LU projection
            Veg_Height = coeff[:,:,0] * np.power(LAI,2) + coeff[:,:,1] * LAI + coeff[:,:,2]
            Veg_Height = np.clip(Veg_Height, 0, 600)
            Veg_Height[Veg_Height == 0] = 0.4

            # Save Vegetation Height in the end folder
            SEBAL.save_GeoTiff_proy(dest, Veg_Height, veg_height_FileName, shape, nband=1)


    ######################## calculate p-factor by using the Landuse map #########################
    ws_p = wb_veg['p-factor']

    Array_P = np.zeros([ws_p.max_row-1,2])
    for j in ['A','C']:
        j_number={'A' : 0, 'C' : 1}
        for i in range(2,ws_p.max_row+1):
            Value = (ws_p['%s%s' %(j,i)].value)
            Array_P[i-2, j_number[j]] = Value

    p_factor = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1])])
    for Class in range(0,len(Array_P)):
	    p_factor[LU_data==Array_P[Class,0]] = Array_P[Class,1]

    p_factor[p_factor == 0] = 0.5

    dst_pfactor_FileName = os.path.join(output_folder_p_factor,'p_factor.tif')
    SEBAL.save_GeoTiff_proy(dest, p_factor, dst_pfactor_FileName, shape, nband=1)

######################## calculate c-factor by using the Landuse map #########################

    ws_c = wb_veg['C-factor']

    Array_C = np.zeros([ws_c.max_row-1,2])
    for j in ['A','C']:
        j_number={'A' : 0, 'C' : 1}
        for i in range(2,ws_c.max_row+1):
            Value = (ws_c['%s%s' %(j,i)].value)
            Array_C[i-2, j_number[j]] = Value

    c_factor = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1])])
    for Class in range(0,len(Array_C)):
	    c_factor[LU_data==Array_C[Class,0]] = Array_C[Class,1]

    c_factor[np.logical_and(c_factor != 3.0, c_factor != 4.0)] = np.nan

    LUE_max = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1])])
    LUE_max[c_factor == 3] = 2.5
    LUE_max[c_factor == 4] = 4.5
    LUE_max[LUE_max == 0] = 2.5

    dst_LUEmax_FileName = os.path.join(output_folder_LUE,'LUE_max.tif')
    SEBAL.save_GeoTiff_proy(dest, LUE_max, dst_LUEmax_FileName, shape, nband=1)

####################################################################################################################
################################################ Write output part 6 ###############################################
####################################################################################################################

   ############################################# Fill in the additional input sheet #########################################
   # things to be filled in:
   # Transmissivity (optional)
   # NDVI (additional input)
   # Albedo (additional input)
   # LST (additional input)
   # Water Mask (additional input)
   # p-factor (soil input)
   # c-factor  (soil input)
   # Vegetation height (meteo input)

    # VIIRS parameter copy
    VIIRS_Dict = {}
    for k, v in iter(SEBAL_RUNS.items()):
        VIIRS_Dict.setdefault(v['output_folder'], []).append(k)

    # get the worksheets and workbook of the input SEBAL excel file
    xfile = load_workbook(inputExcel)
    sheet_additional = xfile.get_sheet_by_name('Additional_Input')
    sheet_meteo = xfile.get_sheet_by_name('Meteo_Input')
    sheet_soil = xfile.get_sheet_by_name('Soil_Input')
    sheet_VIIRS_input = xfile.get_sheet_by_name('VIIRS_PROBAV_Input')

    for Row_number in Kind_Of_Runs_Dict[2]:

        output_name_run = [key for key, value in iter(VIIRS_Dict.items()) if len(np.argwhere(np.array(VIIRS_Dict[key]) == Row_number)) == 1][0]

        # Get General parameters
        VIIRS_date = "_".join(output_name_run.split('_')[-3:])
        VIIRS_datetime= datetime.strptime(VIIRS_date, '%Y%m%d_H%H_M%M')
        date_run = '%d%02d%02d' %(VIIRS_datetime.year,VIIRS_datetime.month,VIIRS_datetime.day)
        date_run_VIIRS = date_run #

        # import LST
        file_name_LST = os.path.join(output_folder_HANTS_end_sharp, 'LST_sharpened_%s.tif' % date_run_VIIRS )
        sheet_additional['D%d'%(Row_number)] = str(file_name_LST)

        # import NDVI in sheets
        file_name_NDVI = os.path.join(NDVI_outfolder_end, 'NDVI_%s.tif' % date_run)
        sheet_additional['B%d' % Row_number] = str(file_name_NDVI)

        # import Albedo
        file_name_Albedo = os.path.join(ALBEDO_outfolder_end, 'Albedo_%s.tif' % date_run)
        sheet_additional['C%d' % Row_number] = str(file_name_Albedo)

        # import Water Mask
        sheet_additional['E%d'%(Row_number)] = str(fileName_Watermask)

        # import p-factor
        file_name_p_factor = os.path.join(output_folder_p_factor,'p_factor.tif')
        sheet_soil['H%d'%(Row_number)] = str(file_name_p_factor)

        # import p-factor
        file_name_c_factor = os.path.join(output_folder_LUE, 'LUE_max.tif')
        sheet_soil['I%d'%(Row_number)] = str(file_name_c_factor)

        # import vegetation height
        file_name_vegt_height = os.path.join(output_folder_HANTS_end_Veg,'Vegetation_Height_%s.tif' %date_run)
        sheet_meteo['O%d'%(Row_number)] = str(file_name_vegt_height)
        
        if sheet_VIIRS_input['B%d'%(Row_number)].value == None:
            sheet_VIIRS_input['B%d'%(Row_number)] = str('npp_viirs_i05_%s_%02d%02d00.fake' %(date_run, VIIRS_datetime.hour, VIIRS_datetime.minute))
        

    xfile.save(inputExcel)
    return()
 #------------------------------------------------------------------------------
def reproject_dataset_to_example(dataset, pixel_spacing, proj_in):
    """
    A sample function to reproject and resample a GDAL dataset from within
    Python. The idea here is to reproject from one system to another, as well
    as to change the pixel size. The procedure is slightly long-winded, but
    goes like this:

    1. Set up the two Spatial Reference systems.
    2. Open the original dataset, and get the geotransform
    3. Calculate bounds of new geotransform by projecting the UL corners
    4. Calculate the number of pixels with the new projection & spacing
    5. Create an in-memory raster dataset
    6. Perform the projection
    """

    # 1) Open the dataset
    g = gdal.Open(dataset)
    if g is None:
        print('input folder does not exist')

     # Define the EPSG code...
    EPSG_code = '%d' % proj_in
    epsg_to = int(EPSG_code)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    try:
        proj = g.GetProjection()
        Proj_in=proj.split('EPSG","')
        epsg_from=int((str(Proj_in[-1]).split(']')[0])[0:-1])
    except:
        epsg_from = int(4326)    # Get the Geotransform vector:
    geo_t = g.GetGeoTransform()

    # Vector components:
    # 0- The Upper Left easting coordinate (i.e., horizontal)
    # 1- The E-W pixel spacing
    # 2- The rotation (0 degrees if image is "North Up")
    # 3- The Upper left northing coordinate (i.e., vertical)
    # 4- The rotation (0 degrees)
    # 5- The N-S pixel spacing, negative as it is counted from the UL corner
    x_size = g.RasterXSize  # Raster xsize
    y_size = g.RasterYSize  # Raster ysize

    epsg_to = int(epsg_to)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    osng = osr.SpatialReference()
    osng.ImportFromEPSG(epsg_to)
    wgs84 = osr.SpatialReference()
    wgs84.ImportFromEPSG(epsg_from)

    inProj = Proj(init='epsg:%d' %epsg_from)
    outProj = Proj(init='epsg:%d' %epsg_to)

    # Up to here, all  the projection have been defined, as well as a
    # transformation from the from to the to
    ulx, uly = transform(inProj,outProj,geo_t[0], geo_t[3])
    lrx, lry = transform(inProj,outProj,geo_t[0] + geo_t[1] * x_size,
                                        geo_t[3] + geo_t[5] * y_size)

    # See how using 27700 and WGS84 introduces a z-value!
    # Now, we create an in-memory raster
    mem_drv = gdal.GetDriverByName('MEM')

    # The size of the raster is given the new projection and pixel spacing
    # Using the values we calculated above. Also, setting it to store one band
    # and to use Float32 data type.
    col = int((lrx - ulx)/pixel_spacing)
    rows = int((uly - lry)/pixel_spacing)

    # Re-define lr coordinates based on whole number or rows and columns
    (lrx, lry) = (ulx + col * pixel_spacing, uly -
                  rows * pixel_spacing)

    dest = mem_drv.Create('', col, rows, 1, gdal.GDT_Float32)

    if dest is None:
        print('input folder to large for memory, clip input map')

   # Calculate the new geotransform
    new_geo = (ulx, pixel_spacing, geo_t[2], uly,
               geo_t[4], - pixel_spacing)

    # Set the geotransform
    dest.SetGeoTransform(new_geo)
    dest.SetProjection(osng.ExportToWkt())

    # Perform the projection/resampling
    gdal.ReprojectImage(g, dest, wgs84.ExportToWkt(), osng.ExportToWkt(),gdal.GRA_Bilinear)

    return dest, ulx, lry, lrx, uly, epsg_to

#------------------------------------------------------------------------------

def gap_filling(data,NoDataValue, method = 1):
    """
    This function fills the no data gaps in a numpy array

    Keyword arguments:
    dataset -- 'C:/'  path to the source data (dataset that must be filled)
    NoDataValue -- Value that must be filled
    """

    # fill the no data values
    if NoDataValue is np.nan:
        mask = ~(np.isnan(data))
    else:
        mask = ~(data==NoDataValue)
    xx, yy = np.meshgrid(np.arange(data.shape[1]), np.arange(data.shape[0]))
    xym = np.vstack( (np.ravel(xx[mask]), np.ravel(yy[mask])) ).T
    data0 = np.ravel( data[:,:][mask] )

    if method == 1:
        interp0 = scipy.interpolate.NearestNDInterpolator( xym, data0 )
        data_end = interp0(np.ravel(xx), np.ravel(yy)).reshape( xx.shape )

    if method == 2:
        interp0 = scipy.interpolate.LinearNDInterpolator( xym, data0 )
        data_end = interp0(np.ravel(xx), np.ravel(yy)).reshape( xx.shape )


    EndProduct = data_end

    return (EndProduct)

def Save_as_MEM(data='', geo='', projection=''):
    """
    This function save the array as a memory file

    Keyword arguments:
    data -- [array], dataset of the geotiff
    geo -- [minimum lon, pixelsize, rotation, maximum lat, rotation,
            pixelsize], (geospatial dataset)
    projection -- interger, the EPSG code
    """
    # save as a geotiff
    driver = gdal.GetDriverByName("MEM")
    dst_ds = driver.Create('', int(data.shape[1]), int(data.shape[0]), 1,
                           gdal.GDT_Float32, ['COMPRESS=LZW'])
    srse = osr.SpatialReference()
    if projection == '':
        srse.SetWellKnownGeogCS("WGS84")
    else:
        srse.SetWellKnownGeogCS(projection)
    dst_ds.SetProjection(srse.ExportToWkt())
    dst_ds.GetRasterBand(1).SetNoDataValue(-9999)
    dst_ds.SetGeoTransform(geo)
    dst_ds.GetRasterBand(1).WriteArray(data)
    return(dst_ds)
