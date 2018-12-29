# -*- coding: utf-8 -*-
"""
Created on Tue Jun 05 06:29:15 2018

@author: tih
"""

import numpy as np
import os
import gdal
from openpyxl import load_workbook
import osr
from datetime import datetime
from pyproj import Proj, transform

def main(VegetationExcel):

####################################################################################################################
############################################# CREATE INPUT FOR preSEBAL RUN ########################################
####################################################################################################################

####################################################################################################################
##################################################### PreSEBAL_4 ###################################################
####################################################################################################################

# PreSEBAL_4
# Part 1: Define input by user
# Part 2: Set parameters and output folder
# Part 3: 

####################################################################################################################
#################################### part 1: Define input by user ##################################################
####################################################################################################################

    import SEBAL.pySEBAL.pySEBAL_code as SEBAL
    
####################################################################################################################
################################# part 2: Set parameters and output folder #########################################
####################################################################################################################

    # Open Excel workbook used for Vegetation c and p factor conversions
    wb_veg = load_workbook(VegetationExcel, data_only=True)
    ws_veg = wb_veg['General_Input']

    # Input for preSEBAL.py
    inputExcel= r"%s" %str(ws_veg['B4'].value)              # The excel with all the SEBAL input data
    output_folder = r"%s" %str(ws_veg['B7'].value)

    # define temporary output folder
    temp_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_temp')                      # Temp outputs of preSEBAL

    temp_folder_PreSEBAL_cold_pixels = os.path.join(temp_folder_PreSEBAL,'Cold_pixels')                      # Temp outputs of preSEBAL
    temp_folder_PreSEBAL_hot_pixels = os.path.join(temp_folder_PreSEBAL,'Hot_pixels')                      # Temp outputs of preSEBAL

    if not os.path.exists(temp_folder_PreSEBAL_cold_pixels):
        os.makedirs(temp_folder_PreSEBAL_cold_pixels)
    if not os.path.exists(temp_folder_PreSEBAL_hot_pixels):
        os.makedirs(temp_folder_PreSEBAL_hot_pixels)

    ######################## Load Excels ##########################################
    # Open Excel workbook for SEBAL inputs
    wb = load_workbook(inputExcel, data_only=True)

    # Get length of EXCEL sheet
    ws = wb['General_Input']
    ws2 = wb['VIIRS_PROBAV_Input']
    ws3 = wb['Meteo_Input']
    ws4 = wb['Additional_Input']
    ws5 = wb['MODIS_Input']
    sheet_additional = wb.get_sheet_by_name('Additional_Input')
    endExcel=int(ws.max_row)

    # Create Dict
    SEBAL_RUNS = dict()

    for number in range(2,endExcel+1):
        input_folder_SEBAL = str(ws['B%d' % number].value)
        output_folder_SEBAL = str(ws['C%d' % number].value)
        Image_Type = int(ws['D%d' % number].value)
        PROBA_V_name  = str(ws2['D%d' % number].value)
        VIIRS_name  = str(ws2['B%d' % number].value)
        MODIS_name = str(ws5['B%d' % number].value)
        SEBAL_RUNS[number] = {'input_folder': input_folder_SEBAL, 'output_folder': output_folder_SEBAL, 'image_type': Image_Type,'PROBA_V_name': PROBA_V_name,'VIIRS_name': VIIRS_name,'MODIS_name': MODIS_name}

    Kind_Of_Runs_Dict = {}
    Kind_Of_Runs_Dict[2] = []
    Kind_Of_Runs_Dict[3] = []
    for k, v in SEBAL_RUNS.iteritems():
        Kind_Of_Runs_Dict.setdefault(v['image_type'], []).append(k)


    PROBA_V_dict = {}
    for k, v in SEBAL_RUNS.iteritems():
        PROBA_V_dict.setdefault(v['PROBA_V_name'], []).append(k)

    ######################## Calculate hot/cold pixel value  ##########################################

    for i in np.append(Kind_Of_Runs_Dict[2],Kind_Of_Runs_Dict[3]):

        NDVI_filename = '%s' %str(ws4['B%d' % i].value)
        WM_filename = '%s' %str(ws4['C%d' % i].value)
        Albedo_filename = '%s' %str(ws4['D%d' % i].value)
        LST_filename = '%s' %str(ws4['E%d' % i].value)
        DEM_filename = '%s' %str(ws['E%d' % i].value)
        Temp_filename_inst = '%s' %str(ws3['B%d' % i].value)
        Temp_filename_24 = '%s' %str(ws3['C%d' % i].value)
        Rs_in_filename_inst = '%s' %str(ws3['J%d' %i].value)

        image_type_now = SEBAL_RUNS[i]['image_type']

        if image_type_now == 2:
            Cold_Pixel_Constant = float(ws2['F%d' % i].value)
            Hot_Pixel_Constant = float(ws2['E%d' % i].value)
            UTM_number = float(ws2['G%d' % i].value)
            reproject_size = 100
            UTM_Zone = int("326%02d" %UTM_number)

        if image_type_now == 3:
            Cold_Pixel_Constant = float(ws5['F%d' % i].value)
            Hot_Pixel_Constant = float(ws5['E%d' % i].value)
            UTM_number = float(ws5['G%d' % i].value)
            reproject_size = 250
            UTM_Zone = int("326%02d" %UTM_number)

        if os.path.exists(LST_filename):

            # Create Example file
            dest_DEM, ulx, lry, lrx, uly, epsg_to = reproject_dataset_to_example(DEM_filename, reproject_size, UTM_Zone)
            ncol = dest_DEM.RasterXSize        # Get the reprojected dem column size
            nrow = dest_DEM.RasterYSize        # Get the reprojected dem row size
            shape = [ncol, nrow]

            # Create reprojected files
            dest_NDVI, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(NDVI_filename, dest_DEM)
            dest_WM, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(WM_filename, dest_DEM)
            dest_Albedo, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(Albedo_filename, dest_DEM)
            dest_LST, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(LST_filename, dest_DEM)
            dest_Temp_inst, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(Temp_filename_inst, dest_DEM)
            dest_Temp_24, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(Temp_filename_24, dest_DEM)
            dest_Rs_in_inst, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(Rs_in_filename_inst, dest_DEM)


            # Open reprojected files
            NDVI = dest_NDVI.GetRasterBand(1).ReadAsArray()
            WM = dest_WM.GetRasterBand(1).ReadAsArray()
            LST = dest_LST.GetRasterBand(1).ReadAsArray()
            DEM = dest_DEM.GetRasterBand(1).ReadAsArray()
            Temp_inst = dest_Temp_inst.GetRasterBand(1).ReadAsArray()
            Rs_in_inst = dest_Rs_in_inst.GetRasterBand(1).ReadAsArray()

            # Get time
            output_name_run =  '%s' %str(ws['C%d' % i].value)
            VIIRS_date = "_".join(output_name_run.split('_')[-3:])
            VIIRS_datetime= datetime.strptime(VIIRS_date, '%Y%m%d_H%H_M%M')
            year = VIIRS_datetime.year
            month = VIIRS_datetime.month
            day = VIIRS_datetime.day
            hour = VIIRS_datetime.hour
            minute = VIIRS_datetime.minute
            DOY = int(VIIRS_datetime.strftime("%j"))

            # Calculate radiation
            lat, lon, lat_fileName, lon_fileName = SEBAL.DEM_lat_lon(DEM_filename, temp_folder_PreSEBAL)

            dest_lat_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_to_example(lat_fileName,  reproject_size, UTM_Zone)
            dest_lon_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_to_example(lon_fileName,  reproject_size, UTM_Zone)

            # Get the reprojected latitude/longitude data
            lat_proy = dest_lat_rep.GetRasterBand(1).ReadAsArray()
            lon_proy = dest_lon_rep.GetRasterBand(1).ReadAsArray()

            # Calculate slope and aspect from the reprojected DEM
            deg2rad, rad2deg, slope, aspect = SEBAL.Calc_Gradient(DEM, reproject_size)

            # Calculation of extraterrestrial solar radiation for slope and aspect
            Ra_mountain_24, Ra_inst, cos_zn, dr, phi, delta = SEBAL.Calc_Ra_Mountain(lon, DOY, hour, minute, lon_proy, lat_proy, slope, aspect)
            Sun_elevation = 90 - (np.nanmean(cos_zn) * 180/np.pi)

            Transm_corr=Rs_in_inst/Ra_inst
            Transm_corr[Transm_corr<0.001]=0.1
            Transm_corr[Transm_corr>1]=1

            Temp_lapse_rate = 0.0065  # Temperature lapse rate (°K/m)
            Pair = 101.3 * np.power((293 - Temp_lapse_rate * DEM) / 293, 5.26)
            QC_Map = np.zeros(DEM.shape)
            ts_dem,air_dens,Temp_corr=SEBAL.Correct_Surface_Temp(LST,Temp_lapse_rate,DEM,Pair,dr,Transm_corr,cos_zn,Sun_elevation,deg2rad,QC_Map)

            ts_dem[DEM==0]=np.nan

            NDVI_std=np.nanstd(NDVI)
            NDVI_max=np.nanmax(NDVI)
            NDVIhot_low = 0.03               # Lower NDVI treshold for hot pixels
            NDVIhot_high = 0.20              # Higher NDVI treshold for hot pixels

            ts_dem_cold_veg = SEBAL.Calc_Cold_Pixels_Veg(NDVI,NDVI_max, NDVI_std, QC_Map,ts_dem, image_type_now, Cold_Pixel_Constant)

            # Cold pixels water
            ts_dem_cold,cold_pixels,ts_dem_cold_mean = Calc_Cold_Pixels(ts_dem, WM, QC_Map, ts_dem_cold_veg, Cold_Pixel_Constant)
            if np.isnan(ts_dem_cold) == True:
                ts_dem_cold = Temp_inst


            # Hot pixels
            ts_dem_hot,hot_pixels = Calc_Hot_Pixels(ts_dem,QC_Map, WM, NDVI, NDVIhot_low, NDVIhot_high, Hot_Pixel_Constant, ts_dem_cold)

            hot_pixels_fileName = os.path.join(temp_folder_PreSEBAL_hot_pixels,"Hot_Pixels_%s.tif" %VIIRS_datetime)
            cold_pixels_fileName = os.path.join(temp_folder_PreSEBAL_cold_pixels,"Cold_Pixels_%s.tif" %VIIRS_datetime)
            SEBAL.save_GeoTiff_proy(dest_DEM, cold_pixels, cold_pixels_fileName, shape, nband=1)
            SEBAL.save_GeoTiff_proy(dest_DEM, hot_pixels, hot_pixels_fileName, shape, nband=1)

            # store the value within the excel file
            sheet_additional['G%d'%(i)] = float(ts_dem_cold - 273.15)
            sheet_additional['H%d'%(i)] = float(ts_dem_hot - 273.15)

    wb.save(inputExcel)
    return()


def Calc_Hot_Pixels(ts_dem,QC_Map, water_mask, NDVI,NDVIhot_low,NDVIhot_high,Hot_Pixel_Constant, ts_dem_cold):
    """
    Function to calculates the hot pixels based on the surface temperature and NDVI
    """
    for_hot = np.copy(ts_dem)
    for_hot[NDVI <= NDVIhot_low] = 0.0
    for_hot[NDVI >= NDVIhot_high] = 0.0
    for_hot[np.logical_or(water_mask != 0.0, QC_Map != 0.0)] = 0.0
    hot_pixels = np.copy(for_hot)
    hot_pixels[for_hot < ts_dem_cold] = np.nan
    ts_dem_hot_max = np.nanmax(hot_pixels)    # Max
    ts_dem_hot_mean = np.nanmean(hot_pixels)  # Mean
    ts_dem_hot_std = np.nanstd(hot_pixels)    # Standard deviation
    #ts_dem_hot = ts_dem_hot_max - 0.25 * ts_dem_hot_std
    #ts_dem_hot = (ts_dem_hot_max + ts_dem_hot_mean)/2


    ts_dem_hot=ts_dem_hot_mean + Hot_Pixel_Constant * ts_dem_hot_std
    ts_dem_hot=np.nanpercentile(hot_pixels, 30)


    print('hot : max= %0.3f (Kelvin)' % ts_dem_hot_max, ', sd= %0.3f (Kelvin)' % ts_dem_hot_std, \
           ', mean= %0.3f (Kelvin)' % ts_dem_hot_mean, ', value= %0.3f (Kelvin)' % ts_dem_hot)
    return(ts_dem_hot,hot_pixels)




def Calc_Cold_Pixels(ts_dem,water_mask,QC_Map,ts_dem_cold_veg,Cold_Pixel_Constant):
    """
    Function to calculates the the cold pixels based on the surface temperature
    """
    for_cold = np.copy(ts_dem)
    for_cold[water_mask != 1.0] = 0.0
    for_cold[QC_Map != 0] = 0.0
    cold_pixels = np.copy(for_cold)
    cold_pixels[for_cold < 278.0] = np.nan
    cold_pixels[for_cold > 320.0] = np.nan
    # cold_pixels[for_cold < 285.0] = 285.0
    ts_dem_cold_std = np.nanstd(cold_pixels)     # Standard deviation
    ts_dem_cold_min = np.nanmin(cold_pixels)     # Min
    ts_dem_cold_mean = np.nanmean(cold_pixels)   # Mean

    # If average temperature is below zero or nan than use the vegetation cold pixel
    if ts_dem_cold_mean <= 0.0:
        ts_dem_cold = ts_dem_cold_veg + Cold_Pixel_Constant * ts_dem_cold_std
    if np.isnan(ts_dem_cold_mean) == True:
        ts_dem_cold = ts_dem_cold_veg + Cold_Pixel_Constant * ts_dem_cold_std
    else:
        ts_dem_cold = ts_dem_cold_mean + Cold_Pixel_Constant * ts_dem_cold_std

    if ts_dem_cold > ts_dem_cold_veg:
        ts_dem_cold = ts_dem_cold_veg


    ts_dem_cold = np.nanpercentile(cold_pixels, 95)


    print('cold water: min=%0.3f (Kelvin)' %ts_dem_cold_min , ', sd= %0.3f (Kelvin)' % ts_dem_cold_std, \
           ', mean= %0.3f (Kelvin)' % ts_dem_cold_mean, ', value= %0.3f (Kelvin)' % ts_dem_cold)
    return(ts_dem_cold,cold_pixels,ts_dem_cold_mean)

 #------------------------------------------------------------------------------
def reproject_dataset_to_example(dataset, pixel_spacing, proj_in):
    """
    A sample function to reproject and resample a GDAL dataset from within
    Python. The idea here is to reproject from one system to another, as well
    as to change the pixel size. The procedure is slightly long-winded, but
    goes like this:

    1. Set up the two Spatial Reference systems.
    2. Open the original dataset, and get the geotransform
    3. Calculate bounds of new geotransform by projecting the UL corners
    4. Calculate the number of pixels with the new projection & spacing
    5. Create an in-memory raster dataset
    6. Perform the projection
    """

    # 1) Open the dataset
    g = gdal.Open(dataset)
    if g is None:
        print('input folder does not exist')

     # Define the EPSG code...
    EPSG_code = '%d' % proj_in
    epsg_to = int(EPSG_code)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    try:
        proj = g.GetProjection()
        Proj_in=proj.split('EPSG","')
        epsg_from=int((str(Proj_in[-1]).split(']')[0])[0:-1])
    except:
        epsg_from = int(4326)    # Get the Geotransform vector:
    geo_t = g.GetGeoTransform()

    # Vector components:
    # 0- The Upper Left easting coordinate (i.e., horizontal)
    # 1- The E-W pixel spacing
    # 2- The rotation (0 degrees if image is "North Up")
    # 3- The Upper left northing coordinate (i.e., vertical)
    # 4- The rotation (0 degrees)
    # 5- The N-S pixel spacing, negative as it is counted from the UL corner
    x_size = g.RasterXSize  # Raster xsize
    y_size = g.RasterYSize  # Raster ysize

    epsg_to = int(epsg_to)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    osng = osr.SpatialReference()
    osng.ImportFromEPSG(epsg_to)
    wgs84 = osr.SpatialReference()
    wgs84.ImportFromEPSG(epsg_from)

    inProj = Proj(init='epsg:%d' %epsg_from)
    outProj = Proj(init='epsg:%d' %epsg_to)

    # Up to here, all  the projection have been defined, as well as a
    # transformation from the from to the to
    ulx, uly = transform(inProj,outProj,geo_t[0], geo_t[3])
    lrx, lry = transform(inProj,outProj,geo_t[0] + geo_t[1] * x_size,
                                        geo_t[3] + geo_t[5] * y_size)

    # See how using 27700 and WGS84 introduces a z-value!
    # Now, we create an in-memory raster
    mem_drv = gdal.GetDriverByName('MEM')

    # The size of the raster is given the new projection and pixel spacing
    # Using the values we calculated above. Also, setting it to store one band
    # and to use Float32 data type.
    col = int((lrx - ulx)/pixel_spacing)
    rows = int((uly - lry)/pixel_spacing)

    # Re-define lr coordinates based on whole number or rows and columns
    (lrx, lry) = (ulx + col * pixel_spacing, uly -
                  rows * pixel_spacing)

    dest = mem_drv.Create('', col, rows, 1, gdal.GDT_Float32)

    if dest is None:
        print('input folder to large for memory, clip input map')

   # Calculate the new geotransform
    new_geo = (ulx, pixel_spacing, geo_t[2], uly,
               geo_t[4], - pixel_spacing)

    # Set the geotransform
    dest.SetGeoTransform(new_geo)
    dest.SetProjection(osng.ExportToWkt())

    # Perform the projection/resampling
    gdal.ReprojectImage(g, dest, wgs84.ExportToWkt(), osng.ExportToWkt(),gdal.GRA_Bilinear)

    return dest, ulx, lry, lrx, uly, epsg_to


if __name__ == '__main__':
    main()



