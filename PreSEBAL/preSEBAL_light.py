# -*- coding: utf-8 -*-
"""
Created on Thu Nov 23 11:23:03 2017

@author: tih
"""
import numpy as np
import os
import scipy.interpolate
import gdal
from openpyxl import load_workbook
import osr
from datetime import datetime, timedelta
import pandas as pd
import shutil
import glob
from netCDF4 import Dataset
import warnings

import SEBAL

def main():
       
####################################################################################################################   
############################################# CREATE INPUT FOR SEBAL RUN ###########################################		
####################################################################################################################

####################################################################################################################   
##################################################### PreHANTS ####################################################		
####################################################################################################################

# PreHANTS
# Part 1: Define input by user 
# Part 2: Set parameters and output folder 
# Part 3: RUN SEBAL
# Part 4: HANTS
# Part 5: post HANTS
# Part 6: Write output

####################################################################################################################   
################################################# PreHANTS part 1 ##################################################		
####################################################################################################################

    VegetationExcel =r"X:\Excel_in_test_2\Vegetation height model.xlsx"  # This excel defines the p and c factor and vegetation height.

####################################################################################################################   
################################################# PreHANTS part 2 ##################################################		
####################################################################################################################

    # Open Excel workbook used for Vegetation c and p factor conversions				
    wb_veg = load_workbook(VegetationExcel, data_only=True)		
    ws_veg = wb_veg['General_Input']	
    
    # Input for preSEBAL.py
    start_date = "%s" %str(ws_veg['B2'].value)  
    end_date = "%s" %str(ws_veg['B3'].value)
    inputExcel= r"%s" %str(ws_veg['B4'].value)               # The excel with all the SEBAL input data
    LU_data_FileName = r"%s" %str(ws_veg['B5'].value)       # Path to Land Use map
    output_folder = r"%s" %str(ws_veg['B7'].value)  
    
    # optional paramater 
    DSSF_Folder= r"%s" %str(ws_veg['B6'].value)
    
    ######################## Load Excels ##########################################	
    # Open Excel workbook for SEBAL inputs
    wb = load_workbook(inputExcel)
				
    # Get length of EXCEL sheet
    ws = wb['General_Input']	
    ws2 = wb['VIIRS_PROBAV_Input']	
    endExcel=int(ws.max_row)
    
    # Create Dict
    SEBAL_RUNS = dict()
    
    for number in range(2,endExcel+1):
        input_folder_SEBAL = str(ws['B%d' % number].value)  
        output_folder_SEBAL = str(ws['C%d' % number].value)                    
        Image_Type = int(ws['D%d' % number].value)
        PROBA_V_name  = str(ws2['D%d' % number].value)
        VIIRS_name  = str(ws2['B%d' % number].value) 
        SEBAL_RUNS[number] = {'input_folder': input_folder_SEBAL, 'output_folder': output_folder_SEBAL, 'image_type': Image_Type,'PROBA_V_name': PROBA_V_name,'VIIRS_name': VIIRS_name}
    
    Kind_Of_Runs_Dict = {}
    for k, v in SEBAL_RUNS.iteritems():
        Kind_Of_Runs_Dict.setdefault(v['image_type'], []).append(k)

    ######################## Create output folders  ##########################################	
  
    output_folder_PreSEBAL_SEBAL = os.path.join(output_folder,'PreSEBAL_SEBAL_out') 
    input_folder_HANTS = os.path.join(output_folder,'HANTS_in')   
    output_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_out') 
    temp_folder_PreSEBAL = os.path.join(output_folder,'PreSEBAL_temp') 
    temp_folder_PreSEBAL_LST = os.path.join(temp_folder_PreSEBAL,'LST')         
    NDVI_outfolder = os.path.join(output_folder_PreSEBAL_SEBAL,'NDVI') 				
    Albedo_outfolder = os.path.join(output_folder_PreSEBAL_SEBAL,'Albedo') 								
    WaterMask_outfolder = os.path.join(output_folder_PreSEBAL_SEBAL,'WaterMask') 								
    LAI_outfolder = os.path.join(output_folder_PreSEBAL_SEBAL,'LAI') 
    TRANS_outfolder = os.path.join(output_folder_PreSEBAL,'Transmissivity') 
    Surface_Temperature_outfolder = os.path.join(output_folder_PreSEBAL_SEBAL,'Surface_Temperature')  
    output_folder_HANTS_end_sharp = os.path.join(output_folder_PreSEBAL, 'LST_Sharpened')
    output_folder_HANTS_end_LAI = os.path.join(output_folder_PreSEBAL, 'LAI')    
    output_folder_HANTS_end_Veg = os.path.join(output_folder_PreSEBAL, 'Vegetation_Height')
    output_folder_p_factor =  os.path.join(output_folder_PreSEBAL, 'p_factor')
    output_folder_LUE =  os.path.join(output_folder_PreSEBAL, 'LUE')
    
    if not os.path.exists(output_folder_PreSEBAL_SEBAL):
        os.makedirs(output_folder_PreSEBAL_SEBAL)
    if not os.path.exists(output_folder_PreSEBAL):
        os.mkdir(output_folder_PreSEBAL)
    if not os.path.exists(temp_folder_PreSEBAL):
        os.mkdir(temp_folder_PreSEBAL)        
    if not os.path.exists(NDVI_outfolder):
        os.makedirs(NDVI_outfolder)				
    if not os.path.exists(Albedo_outfolder):
        os.makedirs(Albedo_outfolder)
    if not os.path.exists(WaterMask_outfolder):
        os.makedirs(WaterMask_outfolder)        
    if not os.path.exists(LAI_outfolder):
        os.makedirs(LAI_outfolder)
    if not os.path.exists(temp_folder_PreSEBAL_LST):
        os.makedirs(temp_folder_PreSEBAL_LST)        
    if not os.path.exists(Surface_Temperature_outfolder):
        os.makedirs(Surface_Temperature_outfolder)
    if not os.path.exists(TRANS_outfolder):
        os.makedirs(TRANS_outfolder)	
    if not os.path.exists(output_folder_HANTS_end_sharp):
        os.mkdir(output_folder_HANTS_end_sharp)
    if not os.path.exists(output_folder_HANTS_end_LAI):
        os.mkdir(output_folder_HANTS_end_LAI)     
    if not os.path.exists(output_folder_HANTS_end_Veg):
        os.mkdir(output_folder_HANTS_end_Veg)           
    if not os.path.exists(output_folder_p_factor):
        os.mkdir(output_folder_p_factor)     
    if not os.path.exists(output_folder_LUE):
        os.mkdir(output_folder_LUE)       
        
    # Do not show warnings
    warnings.filterwarnings('ignore')  

####################################################################################################################   
################################################### RUN SEBAL part 3 ###############################################	
####################################################################################################################    
    
    ############################## Define General info ############################ 
    for number in Kind_Of_Runs_Dict[2]: # Number defines the column of the inputExcel
        print number
        if not (SEBAL_RUNS[number]['PROBA_V_name'] == 'None' and SEBAL_RUNS[number]['VIIRS_name'] == 'None'):
            Rp = 0.91                        # Path radiance in the 10.4-12.5 µm band (W/m2/sr/µm)
            tau_sky = 0.866                  # Narrow band transmissivity of air, range: [10.4-12.5 µm]
            surf_temp_offset = 3             # Surface temperature offset for water 
            
        ######################## Open General info from SEBAL Excel ################### 
    
            # Open the General_Input sheet			
            ws = wb['General_Input']
        
            # Extract the input and output folder, and Image type from the excel file			
            input_folder = str(ws['B%d' % number].value)                             
            Image_Type = int(2)                              # Type of Image (1=Landsat & 2 = VIIRS & GLOBA-V)     
      
            # Extract the Path to the DEM map from the excel file
            DEM_fileName = '%s' %str(ws['E%d' % number].value) #'DEM_HydroShed_m'  
    
            # Open DEM and create Latitude and longitude files
            lat,lon,lat_fileName,lon_fileName=SEBAL.DEM_lat_lon(DEM_fileName, temp_folder_PreSEBAL)
    
        ######################## Extract general data for Landsat ##########################################	
            if Image_Type == 1:	
           
                # Open the Landsat_Input sheet				
                ws = wb['Landsat_Input']		
          
                # Extract Landsat name, number and amount of thermal bands from excel file 
                Name_Landsat_Image = str(ws['B%d' % number].value)    # From glovis.usgs.gov
                Landsat_nr = int(ws['C%d' % number].value)            # Type of Landsat (LS) image used (LS5, LS7, or LS8)
                Bands_thermal = int(ws['D%d' %number].value)         # Number of LS bands to use to retrieve land surface 
     
                # Pixel size of the model
                pixel_spacing=int(30) 
                
                # the path to the MTL file of landsat				
                Landsat_meta_fileName = os.path.join(input_folder, '%s_MTL.txt' % Name_Landsat_Image)
                
                # read out the general info out of the MTL file in Greenwich Time
                year, DOY, hour, minutes, UTM_Zone, Sun_elevation = SEBAL.info_general_metadata(Landsat_meta_fileName) # call definition info_general_metadata
                date=datetime.strptime('%s %s'%(year,DOY), '%Y %j')
                month = date.month
                day = date.day
                
                # define the kind of sensor and resolution of the sensor
                sensor1 = 'L%d' % Landsat_nr
                sensor2 = 'L%d' % Landsat_nr
                sensor3 = 'L%d' % Landsat_nr
                res1 = '30m'
                res2 = '%sm' %int(pixel_spacing)
                res3 = '30m'
                
                # Set the start parameter for determining transmissivity at 0
                Determine_transmissivity = 0
                
        ######################## Extract general data for VIIRS-PROBAV ##########################################	
            if Image_Type == 2:	
                
                # Open the VIIRS_PROBAV_Input sheet					
                ws = wb['VIIRS_PROBAV_Input']
                
                # Extract the name of the thermal and quality VIIRS image from the excel file	
                Name_VIIRS_Image_TB = '%s' %str(ws['B%d' % number].value)
                
                # Extract the name to the PROBA-V image from the excel file	
                Name_PROBAV_Image = '%s' %str(ws['D%d' % number].value)    # Must be a tiff file 
                     
                # Pixel size of the model
                pixel_spacing=int(100) 	
     		
                # UTM Zone of the end results					
                UTM_Zone = float(ws['G%d' % number].value)
    
                if not Name_VIIRS_Image_TB == 'None':
    
                    #Get time from the VIIRS dataset name (IMPORTANT TO KEEP THE TEMPLATE OF THE VIIRS NAME CORRECT example: npp_viirs_i05_20150701_124752_wgs84_fit.tif)
                    Total_Day_VIIRS = Name_VIIRS_Image_TB.split('_')[3]
                    Total_Time_VIIRS = Name_VIIRS_Image_TB.split('_')[4]
        				
                    # Get the information out of the VIIRS name in GMT (Greenwich time)
                    year = int(Total_Day_VIIRS[0:4])
                    month = int(Total_Day_VIIRS[4:6])
                    day = int(Total_Day_VIIRS[6:8])
                    Startdate = '%d-%02d-%02d' % (year,month,day)
                    DOY=datetime.strptime(Startdate,'%Y-%m-%d').timetuple().tm_yday
                    hour = int(Total_Time_VIIRS[0:2])
                    minutes = int(Total_Time_VIIRS[2:4])
                    
                    # If this is runned correctly, we can determine transmissivity
                    ws = wb['Meteo_Input']
                    Field_Radiation_24 = '%s' %str(ws['J%d' % number].value)  
                    Field_Trans_24 = '%s' %str(ws['K%d' % number].value)  
    
                    Determine_transmissivity = 1
    
                    
                # else use PROBA-V day but than no transmissivity can be determined for now
                else:
                   
                    # Get the day and time from the PROBA-V                   
                    Band_PROBAVhdf_fileName = os.path.join(input_folder, '%s.HDF5' % (Name_PROBAV_Image))   
                    g=gdal.Open(Band_PROBAVhdf_fileName, gdal.GA_ReadOnly)
                        
                    Meta_data = g.GetMetadata()
                    Date_PROBAV = str(Meta_data['LEVEL3_RADIOMETRY_BLUE_OBSERVATION_START_DATE'])
                    year = int(Date_PROBAV.split("-")[0])
                    month = int(Date_PROBAV.split("-")[1])
                    day = int(Date_PROBAV.split("-")[2])                        
                    Var_name = '%d%02d%02d' %(year, month, day)  
                    DOY=datetime.strptime(Var_name,'%Y%m%d').timetuple().tm_yday
    
                    # We cannot determine transmissivity                                     
                    Determine_transmissivity = 0
                  
                # Determine the transmissivity if possible (Determine_transmissivity = 1)    
                if Determine_transmissivity == 1:                
                    
                    # Rounded difference of the local time from Greenwich (GMT) (hours):
                    delta_GTM = round(np.sign(lon[int(np.shape(lon)[0]/2), int(np.shape(lon)[1]/2)]) * lon[int(np.shape(lon)[0]/2), int(np.shape(lon)[1]/2)] * 24 / 360)
                    if np.isnan(delta_GTM) == True:
                        delta_GTM = round(np.nanmean(lon) * np.nanmean(lon)  * 24 / 360)
        
                    # Calculate local time 
                    hour += delta_GTM
                    if hour < 0.0:
                        day -= 1
                        hour += 24
                    if hour >= 24:
                        day += 1
                        hour -= 24         
                
                    # define the kind of sensor and resolution of the sensor	
                    sensor1 = 'PROBAV'
                    sensor2 = 'VIIRS'
                    res1 = '375m'
                    res2 = '%sm' %int(pixel_spacing)
                    res3 = '30m'
     
        ######################## Extract general data from DEM file and create Slope map ##########################################	
                
            # Variable date name
            Var_name = '%d%02d%02d' %(year, month, day)
        								
            # Reproject from Geog Coord Syst to UTM -
            # 1) DEM - Original DEM coordinates is Geographic: lat, lon
            dest, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SEBAL.reproject_dataset(
                           DEM_fileName, pixel_spacing, UTM_Zone=UTM_Zone)
            band = dest.GetRasterBand(1)   # Get the reprojected dem band
            ncol = dest.RasterXSize        # Get the reprojected dem column size
            nrow = dest.RasterYSize        # Get the reprojected dem row size
            shape=[ncol, nrow]
           
            # Read out the DEM band and print the DEM properties
            data_DEM = band.ReadAsArray(0, 0, ncol, nrow)
    
            # 2) Latitude file - reprojection    
            # reproject latitude to the landsat projection and save as tiff file																																
            lat_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SEBAL.reproject_dataset(
                            lat_fileName, pixel_spacing, UTM_Zone=UTM_Zone)
                        
            # Get the reprojected latitude data															
            lat_proy = lat_rep.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)
         
            # 3) Longitude file - reprojection
            # reproject longitude to the landsat projection	 and save as tiff file	
            lon_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = SEBAL.reproject_dataset(lon_fileName, pixel_spacing, UTM_Zone=UTM_Zone)
    
            # Get the reprojected longitude data	
            lon_proy = lon_rep.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)
            
            lon_fileName = os.path.join(temp_folder_PreSEBAL,'lon_resh.tif')
            SEBAL.save_GeoTiff_proy(dest, lon_proy, lon_fileName, shape, nband=1)	
    				
            # Calculate slope and aspect from the reprojected DEM
            deg2rad,rad2deg,slope,aspect=SEBAL.Calc_Gradient(data_DEM, pixel_spacing)
 
    
            if Determine_transmissivity == 1:
                
                # calculate the coz zenith angle
                Ra_mountain_24, Ra_inst, cos_zn_resh, dr, phi, delta = SEBAL.Calc_Ra_Mountain(lon,DOY,hour,minutes,lon_proy,lat_proy,slope,aspect) 
                cos_zn_fileName = os.path.join(temp_folder_PreSEBAL,'cos_zn.tif')
                SEBAL.save_GeoTiff_proy(dest, cos_zn_resh, cos_zn_fileName, shape, nband=1)
            
                # Save the Ra
                Ra_inst_fileName = os.path.join(temp_folder_PreSEBAL,'Ra_inst.tif')
                SEBAL.save_GeoTiff_proy(dest, Ra_inst, Ra_inst_fileName, shape, nband=1)	
                Ra_mountain_24_fileName = os.path.join(temp_folder_PreSEBAL,'Ra_mountain_24.tif')
                SEBAL.save_GeoTiff_proy(dest, Ra_mountain_24, Ra_mountain_24_fileName, shape, nband=1)	
                
                #################### Calculate Transmissivity ##########################################	           
                # Open the General_Input sheet			
                ws = wb['Meteo_Input']
        
                # Extract the method radiation value	
                Value_Method_Radiation_inst = '%s' %str(ws['L%d' % number].value)
        
                # Values to check if data is created
                Check_Trans_inst = 0
                Check_Trans_24 = 0
                
                '''  This is now turned off, so you need to fill in the instantanious transmissivity or Radiation
                # Extract the data to the method of radiation
                if int(Value_Method_Radiation_inst) == 2:
                    Field_Radiation_inst = '%s' %str(ws['N%d' % number].value)        
                        
                    if Field_Radiation_inst == 'None':
        
                        # Instantanious Transmissivity files must be created
                        Check_Trans_inst = 1
                        
                        # Calculate Transmissivity
                        quarters_hours = np.ceil(minutes/30.) * 30
                        hours_GMT = hour - delta_GTM
                        if quarters_hours >= 60:
                            hours_GMT += 1
                            quarters_hours = 0
                        
                        # Define the instantanious LANDSAF file
                        name_Landsaf_inst = 'HDF5_LSASAF_MSG_DSSF_MSG-Disk_%d%02d%02d%02d%02d.tif' %(year, month,day, hours_GMT, quarters_hours)
                        file_Landsaf_inst = os.path.join(DSSF_Folder,name_Landsaf_inst)  
                                     
                        # Reproject the Ra_inst data to match the LANDSAF data
                        Ra_inst_3Km_dest, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(Ra_inst_fileName, file_Landsaf_inst, method = 1)   
                        Ra_inst_3Km = Ra_inst_3Km_dest.GetRasterBand(1).ReadAsArray()
                        Ra_inst_3Km[Ra_inst_3Km==0] = np.nan
                        
                        # Open the Rs LANDSAF data           
                        dest_Rs_inst_3Km = gdal.Open(file_Landsaf_inst)
                        Rs_inst_3Km = dest_Rs_inst_3Km.GetRasterBand(1).ReadAsArray()
                        Rs_inst_3Km = np.float_(Rs_inst_3Km)/10
                        Rs_inst_3Km[Rs_inst_3Km<0]=np.nan
                        
                        # Get shape LANDSAF data
                        shape_trans=[dest_Rs_inst_3Km.RasterXSize , dest_Rs_inst_3Km.RasterYSize ]
                        
                        # Calculate Transmissivity 3Km
                        Transmissivity_3Km = Rs_inst_3Km/Ra_inst_3Km
                        Transmissivity_3Km_fileName = os.path.join(output_folder_temp,'Transmissivity_3Km.tif')
                        SEBAL.save_GeoTiff_proy(Ra_inst_3Km_dest, Transmissivity_3Km, Transmissivity_3Km_fileName, shape_trans, nband=1)	
                        
                        # Reproject Transmissivity to match DEM (now this is done by using the nearest neighbour method)
                        Transmissivity_inst_dest, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(Transmissivity_3Km_fileName, cos_zn_fileName, method = 3)  
                        Transmissivity_inst = Transmissivity_inst_dest.GetRasterBand(1).ReadAsArray()
                        Transmissivity_inst[Transmissivity_inst>0.98] = 0.98                
                        Transmissivity_inst_fileName = os.path.join(TRANS_outfolder,'Transmissivity_inst_%s.tif' %Var_name)                
                        SEBAL.save_GeoTiff_proy(Transmissivity_inst_dest, Transmissivity_inst, Transmissivity_inst_fileName, shape, nband=1)	
    
                '''                    
                # Extract the method radiation value	
                Value_Method_Radiation_24 = '%s' %str(ws['I%d' % number].value)
        
                # Extract the data to the method of radiation
                if int(Value_Method_Radiation_24) == 2:
                    Field_Radiation_24 = '%s' %str(ws['K%d' % number].value)        
                        
                    if Field_Radiation_24 == 'None':
                        
                        # Daily Transmissivity files must be created
                        Check_Trans_24 = 1
        
                        # Create times that are needed to calculate daily Rs (LANDSAF)
                        Starttime_GMT = datetime.strptime(Startdate,'%Y-%m-%d') + timedelta(hours=-delta_GTM)
                        Endtime_GMT = Starttime_GMT + timedelta(days=1)
                        Times = pd.date_range(Starttime_GMT, Endtime_GMT,freq = '30min')
         
                        for Time in Times[:-1]:
                            year_LANDSAF = Time.year
                            month_LANDSAF = Time.month
                            day_LANDSAF = Time.day
                            hour_LANDSAF = Time.hour
                            min_LANDSAF = Time.minute
         
                            # Define the instantanious LANDSAF file
                            #re = glob.glob('')
                            name_Landsaf_inst = 'HDF5_LSASAF_MSG_DSSF_MSG-Disk_%d%02d%02d%02d%02d.tif' %(year_LANDSAF, month_LANDSAF,day_LANDSAF, hour_LANDSAF, min_LANDSAF)
                            file_Landsaf_inst = os.path.join(DSSF_Folder,name_Landsaf_inst)  
        
                            # Open the Rs LANDSAF data           
                            dest_Rs_inst_3Km = gdal.Open(file_Landsaf_inst)
                            Rs_one_3Km = dest_Rs_inst_3Km.GetRasterBand(1).ReadAsArray()
                            Rs_one_3Km = np.float_(Rs_one_3Km)/10
                            Rs_one_3Km[Rs_one_3Km < 0]=np.nan
                          
                            if Time == Times[0]:
                                Rs_24_3Km_tot = Rs_one_3Km
                            else:
                                Rs_24_3Km_tot += Rs_one_3Km
                        
                        Rs_24_3Km = Rs_24_3Km_tot / len(Times[:-1])  
                         
                        # Reproject the Ra_inst data to match the LANDSAF data
                        Ra_24_3Km_dest, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(Ra_mountain_24_fileName, file_Landsaf_inst, method = 3)   
                        Ra_24_3Km = Ra_24_3Km_dest.GetRasterBand(1).ReadAsArray()
                        Ra_24_3Km[Ra_24_3Km==0] = np.nan  
                         
                        # Do gapfilling
                        Ra_24_3Km = gap_filling(Ra_24_3Km,np.nan)
                                 
                        # Get shape LANDSAF data
                        shape_trans=[dest_Rs_inst_3Km.RasterXSize , dest_Rs_inst_3Km.RasterYSize ]
                        
                        # Calculate Transmissivity 3Km
                        Transmissivity_24_3Km = Rs_24_3Km/Ra_24_3Km
                    
                        Transmissivity_24_3Km_fileName = os.path.join(temp_folder_PreSEBAL,'Transmissivity_24_3Km.tif')
                        SEBAL.save_GeoTiff_proy(Ra_24_3Km_dest, Transmissivity_24_3Km, Transmissivity_24_3Km_fileName, shape_trans, nband=1)	
                        
                        # Reproject Transmissivity to match DEM (now this is done by using the nearest neighbour method)
                        Transmissivity_24_dest, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(Transmissivity_24_3Km_fileName, lon_fileName, method = 3)  
                        Transmissivity_24 = Transmissivity_24_dest.GetRasterBand(1).ReadAsArray()
                        Transmissivity_24[Transmissivity_24>0.98] = 0.98
                        Transmissivity_24_fileName = os.path.join(TRANS_outfolder,'Transmissivity_24_%s.tif' %Var_name)                
                        SEBAL.save_GeoTiff_proy(Transmissivity_24_dest, Transmissivity_24, Transmissivity_24_fileName, shape, nband=1)	
    
        #################### Calculate NDVI for LANDSAT ##########################################	
    
            if Image_Type == 1:	
    
                # Define bands used for each Landsat number
                if Landsat_nr == 5 or Landsat_nr == 7:
                    Bands = np.array([1, 2, 3, 4, 5, 7, 6])
                elif Landsat_nr == 8:
                    Bands = np.array([2, 3, 4, 5, 6, 7, 10, 11])  
                else:
                    print 'Landsat image not supported, use Landsat 7 or 8'
    
                # Open MTL landsat and get the correction parameters   
                Landsat_meta_fileName = os.path.join(input_folder, '%s_MTL.txt' % Name_Landsat_Image)
                Lmin, Lmax, k1_c, k2_c = SEBAL.info_band_metadata(Landsat_meta_fileName, Bands)
               
                # Mean solar exo-atmospheric irradiance for each band (W/m2/microm)
                # for the different Landsat images (L5, L7, or L8)
                ESUN_L5 = np.array([1983, 1796, 1536, 1031, 220, 83.44])
                ESUN_L7 = np.array([1997, 1812, 1533, 1039, 230.8, 84.9])
                ESUN_L8 = np.array([1973.28, 1842.68, 1565.17, 963.69, 245, 82.106])
        
                # Open one band - To get the metadata of the landsat images only once (to get the extend)
                src_FileName = os.path.join(input_folder, '%s_B2.TIF' % Name_Landsat_Image)  # before 10!
                ls,band_data,ulx,uly,lrx,lry,x_size_ls,y_size_ls = SEBAL.Get_Extend_Landsat(src_FileName)
             
                # Crop the Landsat images to the DEM extent -
                dst_FileName = os.path.join(temp_folder_PreSEBAL,'cropped_LS_b2.tif')  # Before 10 !!
         							
                # Clip the landsat image to match the DEM map											
                lsc, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(src_FileName, lon_fileName)	
                data_LS = lsc.GetRasterBand(1).ReadAsArray()
                SEBAL.save_GeoTiff_proy(dest, data_LS, dst_FileName, shape, nband=1)	
        
                # Get the extend of the remaining landsat file after clipping based on the DEM file	
                lsc,band_data,ulx,uly,lrx,lry,x_size_lsc,y_size_lsc = SEBAL.Get_Extend_Landsat(dst_FileName)
    
                # Create the corrected signals of Landsat in 1 array
                Reflect = SEBAL.Landsat_Reflect(Bands,input_folder,Name_Landsat_Image,output_folder,shape,Lmax,Lmin,ESUN_L5,ESUN_L7,ESUN_L8,cos_zn_resh,dr,Landsat_nr, cos_zn_fileName)
    
                # Calculate temporal water mask
                water_mask_temp=SEBAL.Water_Mask(shape,Reflect)       
    
                # Calculate NDVI
                NDVI = SEBAL.Calc_NDVI(Reflect)
    
                # Calculate albedo
                albedo = SEBAL.Calc_albedo(Reflect)
    
                # Save NDVI
                NDVI_FileName = os.path.join(NDVI_outfolder,'NDVI_LS_%s.tif'%Var_name)
                SEBAL.save_GeoTiff_proy(dest, NDVI, NDVI_FileName, shape, nband=1)
    
                # Save albedo								
                albedo_FileName = os.path.join(Albedo_outfolder,'Albedo_LS_%s.tif'%Var_name)
                SEBAL.save_GeoTiff_proy(dest, albedo, albedo_FileName, shape, nband=1)
    
        ################### Extract Meteo data for Landsat days from SEBAL Excel ##################
    
            # Open the Meteo_Input sheet	
            ws = wb['Meteo_Input']	
            # ---------------------------- Instantaneous Air Temperature ------------
            # Open meteo data, first try to open as value, otherwise as string (path)	  
            try:
               Temp_inst = float(ws['B%d' %number].value)                # Instantaneous Air Temperature (°C)
    
            # if the data is not a value, than open as a string	
            except:
                Temp_inst_name = '%s' %str(ws['B%d' %number].value) 
                Temp_inst_fileName = os.path.join(output_folder, 'Temp', 'Temp_inst_input.tif')
                Temp_inst = SEBAL.Reshape_Reproject_Input_data(Temp_inst_name, Temp_inst_fileName, lon_fileName)
    
            try:
                RH_inst = float(ws['D%d' %number].value)                # Instantaneous Relative humidity (%)
     
            # if the data is not a value, than open as a string							
            except:
                RH_inst_name = '%s' %str(ws['D%d' %number].value) 
                RH_inst_fileName = os.path.join(output_folder, 'Temp', 'RH_inst_input.tif')
                RH_inst = SEBAL.Reshape_Reproject_Input_data(RH_inst_name, RH_inst_fileName, lon_fileName)			
             
            esat_inst = 0.6108 * np.exp(17.27 * Temp_inst / (Temp_inst + 237.3)) 													
            eact_inst = RH_inst * esat_inst / 100
            
         #################### Calculate NDVI for VIIRS-PROBAV ##########################################	
         
            if Image_Type == 2:	
                
                if Name_PROBAV_Image == 'None':
                    
                    offset_all = [-1, 1, -2, 2, -3, 3,-4, 4,-5 ,5 ,-6 , 6, -7, 7, -8, 8]
                    found_Name_PROBAV_Image = 0
                    for offset in offset_all:
                        
                        if found_Name_PROBAV_Image == 1:
                            continue
                        else:
                            try:
                                Name_PROBAV_Image = SEBAL_RUNS[number + offset]['PROBA_V_name']
                                if not Name_PROBAV_Image == 'None':
                                    found_Name_PROBAV_Image = 1  
                            except:
                                pass
 
                    # Get the day and time from the PROBA-V                   
                    Band_PROBAVhdf_fileName = os.path.join(input_folder, '%s.HDF5' % (Name_PROBAV_Image))   
                    g=gdal.Open(Band_PROBAVhdf_fileName, gdal.GA_ReadOnly)
                        
                    Meta_data = g.GetMetadata()
                    Date_PROBAV = str(Meta_data['LEVEL3_RADIOMETRY_BLUE_OBSERVATION_START_DATE'])
                    year = int(Date_PROBAV.split("-")[0])
                    month = int(Date_PROBAV.split("-")[1])
                    day = int(Date_PROBAV.split("-")[2])                        
                    Var_name_2 = '%d%02d%02d' %(year, month, day)  
                     
                    # Define the output name
                    NDVI_FileName = os.path.join(NDVI_outfolder,'NDVI_PROBAV_%s.tif' %Var_name_2)              
                    Albedo_FileName = os.path.join(Albedo_outfolder, 'Albedo_PROBAV_%s.tif' %Var_name_2) 	
                    water_mask_temp_FileName = os.path.join(WaterMask_outfolder, 'Water_Mask_PROBAV_%s.tif' %Var_name_2) 	
    
                else: 
                    NDVI_FileName = os.path.join(NDVI_outfolder,'NDVI_PROBAV_%s.tif' %Var_name)
                    Albedo_FileName = os.path.join(Albedo_outfolder, 'Albedo_PROBAV_%s.tif' %Var_name) 	
                    water_mask_temp_FileName = os.path.join(WaterMask_outfolder, 'Water_Mask_PROBAV_%s.tif' %Var_name) 	
    
                # vegetation maps that will be generated
               

                if not os.path.exists(NDVI_FileName):
             
                    # Define the bands that will be used
                    bands=['SM', 'B1', 'B2', 'B3', 'B4']  #'SM', 'BLUE', 'RED', 'NIR', 'SWIR'
        
                    # Set the index number at 0
                    index=0
        							
                    # create a zero array with the shape of the reprojected DEM file						
                    data_PROBAV=np.zeros((shape[1], shape[0]))
                    spectral_reflectance_PROBAV=np.zeros([shape[1], shape[0], 5])
               
                    # constants
                    n188_float=248       # Now it is 248, but we do not exactly know what this really means and if this is for constant for all images.
           
                    # write the data one by one to the spectral_reflectance_PROBAV       
                    for bandnmr in bands:
        
                        # Translate the PROBA-V names to the Landsat band names                
                        Band_number = {'SM':7,'B1':8,'B2':10,'B3':9,'B4':11}
        
                        # Open the dataset 
                        Band_PROBAVhdf_fileName = os.path.join(input_folder, '%s.HDF5' % (Name_PROBAV_Image))   
                        g=gdal.Open(Band_PROBAVhdf_fileName, gdal.GA_ReadOnly)
                        
                        # define data if it is not there yet
                        if not 'Var_name' in locals():
                            Meta_data = g.GetMetadata()
                            Date_PROBAV = str(Meta_data['LEVEL3_RADIOMETRY_BLUE_OBSERVATION_START_DATE'])
                            year = int(Date_PROBAV.split("-")[0])
                            month = int(Date_PROBAV.split("-")[0])
                            day = int(Date_PROBAV.split("-")[0])                        
                            Var_name = '%d%02d%02d' %(year, month, day)
        
                        
                        # Open the .hdf file              
                        name_out = os.path.join(input_folder, '%s_test.tif' % (Name_PROBAV_Image))   
                        name_in = g.GetSubDatasets()[Band_number[bandnmr]][0]
                        
                        # Get environmental variable
                        SEBAL_env_paths = os.environ["SEBAL"].split(';')
                        GDAL_env_path = SEBAL_env_paths[0]
                        GDAL_TRANSLATE = os.path.join(GDAL_env_path, 'gdal_translate.exe')
         
                        # run gdal translate command               
                        FullCmd = '%s -of GTiff %s %s' %(GDAL_TRANSLATE, name_in, name_out)            
                        SEBAL.Run_command_window(FullCmd)
                        
                        # Open data
                        dest_PV = gdal.Open(name_out)
                        Data = dest_PV.GetRasterBand(1).ReadAsArray()               
                        dest_PV = None
                        
                        # Remove temporary file
                        os.remove(name_out)
         
                        # Define the x and y spacing
                        Meta_data = g.GetMetadata()
                        Lat_Bottom = float(Meta_data['LEVEL3_GEOMETRY_BOTTOM_LEFT_LATITUDE'])
                        Lat_Top = float(Meta_data['LEVEL3_GEOMETRY_TOP_RIGHT_LATITUDE'])
                        Lon_Left = float(Meta_data['LEVEL3_GEOMETRY_BOTTOM_LEFT_LONGITUDE'])
                        Lon_Right = float(Meta_data['LEVEL3_GEOMETRY_TOP_RIGHT_LONGITUDE'])           
                        Pixel_size = float((Meta_data['LEVEL3_GEOMETRY_VNIR_VAA_MAPPING']).split(' ')[-3])
                        
                        # Define the georeference of the PROBA-V data
                        geo_PROBAV=[Lon_Left-0.5*Pixel_size, Pixel_size, 0, Lat_Top+0.5*Pixel_size, 0, -Pixel_size] #0.000992063492063
        		 									
                        # Define the name of the output file	
                        PROBAV_data_name=os.path.join(input_folder, '%s_%s.tif' % (Name_PROBAV_Image,bandnmr)) 	 																		
                        dst_fileName=os.path.join(input_folder, PROBAV_data_name)
                        
                        # create gtiff output with the PROBA-V band
                        fmt = 'GTiff'
                        driver = gdal.GetDriverByName(fmt)
        
                        dst_dataset = driver.Create(dst_fileName, int(Data.shape[1]), int(Data.shape[0]), 1,gdal.GDT_Float32)
                        dst_dataset.SetGeoTransform(geo_PROBAV)
                    
                        # set the reference info
                        srs = osr.SpatialReference()
                        srs.SetWellKnownGeogCS("WGS84")
                        dst_dataset.SetProjection(srs.ExportToWkt())
                   
                        # write the array in the geotiff band
                        dst_dataset.GetRasterBand(1).WriteArray(Data)
                        dst_dataset = None
         
                        # Open the PROBA-V band in SEBAL											
                        g=gdal.Open(PROBAV_data_name.replace("\\","/")) 
        	 										
                        # If the data cannot be opened, change the extension											
                        if g is None:
                            PROBAV_data_name=os.path.join(input_folder, '%s_%s.tiff' % (Name_PROBAV_Image,bandnmr))  
          
                        # Reproject the PROBA-V band  to match DEM's resolution          
                        PROBAV, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(
                                          PROBAV_data_name, lon_fileName)
        
                        # Open the reprojected PROBA-V band data                         
                        data_PROBAV_DN = PROBAV.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)
        	 										
                        # Define the filename to store the cropped Landsat image
                        dst_FileName = os.path.join(output_folder, 'Output_PROBAV','proy_PROBAV_%s.tif' % bandnmr)
        		 									
                        # close the PROBA-V 											
                        g=None
                                           
                        # If the band data is not SM change the DN values into PROBA-V values and write into the spectral_reflectance_PROBAV                      
                        if bandnmr is not 'SM':  
                            data_PROBAV[:, :]=data_PROBAV_DN/2000                           
                            spectral_reflectance_PROBAV[:, :, index]=data_PROBAV[:, :]
                        
                        # If the band data is the SM band than write the data into the spectral_reflectance_PROBAV  and create cloud mask            
                        else:
                            data_PROBAV[:, :]=data_PROBAV_DN
                            Cloud_Mask_PROBAV=np.zeros((shape[1], shape[0]))
                            Cloud_Mask_PROBAV[data_PROBAV[:,:]!=n188_float]=1
                            spectral_reflectance_PROBAV[:, :, index]=Cloud_Mask_PROBAV
          
                        # Change the spectral reflectance to meet certain limits                               
                        spectral_reflectance_PROBAV[:, :, index]=np.where(spectral_reflectance_PROBAV[:, :, index]<=0,np.nan,spectral_reflectance_PROBAV[:, :, index])   
                        spectral_reflectance_PROBAV[:, :, index]=np.where(spectral_reflectance_PROBAV[:, :, index]>=150,np.nan,spectral_reflectance_PROBAV[:, :, index])   
          										
                        # Go to the next index 									
                        index=index+1
         
                    # Bands in PROBAV spectral reflectance
                    # 0 = MS
                    # 1 = BLUE
                    # 2 = NIR
                    # 3 = RED
                    # 4 = SWIR
            
                    # Calculate surface albedo based on PROBA-V
                    Surface_Albedo_PROBAV = 0.219 * spectral_reflectance_PROBAV[:, :, 1] + 0.361 * spectral_reflectance_PROBAV[:, :, 2] + 0.379 * spectral_reflectance_PROBAV[:, :, 3] + 0.041 * spectral_reflectance_PROBAV[:, :, 4]
                        
                    # Create Water mask based on PROBA-V             
                    water_mask_temp = np.zeros((shape[1], shape[0])) 
                    water_mask_temp[np.logical_and(spectral_reflectance_PROBAV[:, :, 2] >= spectral_reflectance_PROBAV[:, :, 3],data_DEM>0)]=1
       
                    # Calculate the NDVI based on PROBA-V     
                    n218_memory = spectral_reflectance_PROBAV[:, :, 2] + spectral_reflectance_PROBAV[:, :, 3]
                    NDVI = np.zeros((shape[1], shape[0]))
                    NDVI[n218_memory != 0] =  ( spectral_reflectance_PROBAV[:, :, 3][n218_memory != 0] - spectral_reflectance_PROBAV[:, :, 2][n218_memory != 0] )/ ( spectral_reflectance_PROBAV[:, :, 2][n218_memory != 0] + spectral_reflectance_PROBAV[:, :, 3][n218_memory != 0] )
    
                    # Save Albedo for PROBA-V		
                    SEBAL.save_GeoTiff_proy(dest, Surface_Albedo_PROBAV, Albedo_FileName, shape, nband=1)	  
        
                    # Save NDVI for PROBA-V
                    SEBAL.save_GeoTiff_proy(dest, NDVI, NDVI_FileName, shape, nband=1)														

                    # Save Water Mask for PROBA-V	        
                    SEBAL.save_GeoTiff_proy(dest, water_mask_temp, water_mask_temp_FileName, shape, nband=1)	  
  
                else:
                    
                    dest_NDVI = gdal.Open(NDVI_FileName)
                    dest_water_mask_temp = gdal.Open(water_mask_temp_FileName)
                    NDVI = dest_NDVI.GetRasterBand(1).ReadAsArray()
                    water_mask_temp = dest_water_mask_temp.GetRasterBand(1).ReadAsArray()

             ############################ Calculate LAI ##########################################	
        
            # Calculate the LAI
            FPAR,tir_emis,Nitrogen,vegt_cover,LAI,b10_emissivity = SEBAL.Calc_vegt_para(NDVI,water_mask_temp,shape)
    
            # Create LAI name
            if Image_Type == 1:
                LAI_FileName = os.path.join(LAI_outfolder,'LAI_LS_%s.tif' %Var_name)
                SEBAL.save_GeoTiff_proy(dest, LAI, LAI_FileName, shape, nband=1)	
            
         #################### Calculate thermal for Landsat ##########################################	
    
            if Image_Type == 1:
    				
                # Calculate thermal				
                therm_data = SEBAL.Landsat_therm_data(Bands,input_folder,Name_Landsat_Image,output_folder,ulx_dem,lry_dem,lrx_dem,uly_dem,shape)          
    
                # Calculate surface temperature
                Surface_temp=SEBAL.Calc_surface_water_temp(Temp_inst,Landsat_nr,Lmax,Lmin,therm_data,b10_emissivity,k1_c,k2_c,eact_inst,shape,water_mask_temp,Bands_thermal,Rp,tau_sky,surf_temp_offset,Image_Type)
    
                # Save surface temperature
                therm_data_FileName = os.path.join(Surface_Temperature_outfolder,'Surface_Temperature_LS_%s.tif' %Var_name)
                SEBAL.save_GeoTiff_proy(dest, Surface_temp, therm_data_FileName, shape, nband=1)	
    
    
        ################################## Calculate VIIRS surface temperature ########################
    
            if Image_Type == 2:
     
               # If there is VIIRS data
               if not Name_VIIRS_Image_TB == 'None':
    			
                    # Define the VIIRS thermal data name
                    VIIRS_data_name=os.path.join(input_folder, '%s' % (Name_VIIRS_Image_TB))
    			
                    # Reproject VIIRS thermal data								
                    VIIRS, ulx, lry, lrx, uly, epsg_to = SEBAL.reproject_dataset_example(VIIRS_data_name, lon_fileName)
        																
                    # Open VIIRS thermal data																		
                    data_VIIRS = VIIRS.GetRasterBand(1).ReadAsArray()    
        				
                    # Define the thermal VIIRS output name
                    proyVIIRS_fileName = os.path.join(temp_folder_PreSEBAL, 'Surface_Temp_VIIRS_%s.tif' %Var_name)
        	 											
                    # Save the thermal VIIRS data 												
                    SEBAL.save_GeoTiff_proy(dest, data_VIIRS, proyVIIRS_fileName, shape, nband=1)	
        
                    # Set the conditions for the brightness temperature (100m)
                    brightness_temp=np.where(data_VIIRS>=250, data_VIIRS, np.nan)
         
                    # Constants
                    k1=606.399172
                    k2=1258.78
                    L_lambda_b10_100=((2*6.63e-34*(3.0e8)**2)/((11.45e-6)**5*(np.exp((6.63e-34*3e8)/(1.38e-23*(11.45e-6)*brightness_temp))-1)))*1e-6
                 
                    # Get Temperature for 100 and 375m resolution
                    Temp_TOA_100 = SEBAL.Get_Thermal(L_lambda_b10_100,Rp,Temp_inst,tau_sky,tir_emis,k1,k2) 
               
                    # Conditions for surface temperature (100m)
                    n120_surface_temp=Temp_TOA_100.clip(250, 450)
        
                    # Save the surface temperature of the VIIRS in 100m resolution
                    temp_surface_100_fileName_beforeTS = os.path.join(Surface_Temperature_outfolder,'Surface_Temperature_VIIRS_%s.tif' %Var_name)
                    SEBAL.save_GeoTiff_proy(dest, n120_surface_temp, temp_surface_100_fileName_beforeTS, shape, nband=1)     

    ###################### Create NC files for HANTS parameter testing ####################################

    # All HANTS variables
    VARS = ['NDVI', 'Albedo', 'Surface_Temperature']

    # loop over the HANTS parameters    
    for VAR in VARS:
            
        # Define paths for NDVI
        input_folder_HANTS_VAR = os.path.join(output_folder_PreSEBAL_SEBAL, VAR)
        if VAR == 'Surface_Temperature':
            name_format = 'Surface_Temperature_VIIRS_{0}.tif'
        else:    
            name_format = '%s_PROBAV_{0}.tif' %VAR

        # define NC output for the 3D array   
        nc_path = os.path.join(input_folder_HANTS_VAR,'%s_NC.nc' %VAR)

        # Set the working directory to the Tiff file of the HANTS parameter             
        os.chdir(input_folder_HANTS_VAR)
        
        # Rename the variable in the name_format
        name_format_var = name_format.replace('{0}','*')
        
        # Find all tiff files equal to format
        re_var = glob.glob(name_format_var)

        # Open the first tiff file
        path_to_first_variable = os.path.join(input_folder_HANTS_VAR,re_var[0])
        dest_first_var = gdal.Open(path_to_first_variable)
        
        # Read out the geo transforma paramters
        geo_first_var = dest_first_var.GetGeoTransform()
        x_size = int(dest_first_var.RasterXSize)
        y_size = int(dest_first_var.RasterYSize)
        cellsize = int(geo_first_var[1])
        latlim = [geo_first_var[3]+geo_first_var[5]*y_size,geo_first_var[3]]
        lonlim = [geo_first_var[0],geo_first_var[0]+geo_first_var[1]*x_size]

        # Create netcdf    
        import hants.wa_gdal as hants_gdal
        hants_gdal.create_netcdf(input_folder_HANTS_VAR, name_format, start_date, end_date,
                      latlim, lonlim, cellsize, nc_path)


# Functions
#################################################################################   
def Create_Buffer(Data_In):
    
   '''
   This function creates a 3D array which is used to apply the moving window
   '''
   Buffer_area = 7 # A block of 2 times Buffer_area + 1 will be 1 if there is the pixel in the middle is 1
   Data_Out=np.empty((len(Data_In),len(Data_In[1])))   
   Data_Out[:,:] = Data_In
   for ypixel in range(0,Buffer_area + 1):
                   
        for xpixel in range(1,Buffer_area + 1):

           if ypixel==0:
                for xpixel in range(1,Buffer_area + 1):
                    Data_Out[:,0:-xpixel] += Data_In[:,xpixel:]
                    Data_Out[:,xpixel:] += Data_In[:,:-xpixel]              
                
                for ypixel in range(1,Buffer_area + 1):
                
                    Data_Out[ypixel:,:] += Data_In[:-ypixel,:]
                    Data_Out[0:-ypixel,:] += Data_In[ypixel:,:]

           else:
               Data_Out[0:-xpixel,ypixel:] += Data_In[xpixel:,:-ypixel]
               Data_Out[xpixel:,ypixel:] += Data_In[:-xpixel,:-ypixel]
               Data_Out[0:-xpixel,0:-ypixel] += Data_In[xpixel:,ypixel:]
               Data_Out[xpixel:,0:-ypixel] += Data_In[:-xpixel,ypixel:]

   Data_Out[Data_Out>0.1] = 1
   Data_Out[Data_Out<=0.1] = 0  
           
   return(Data_Out)

#------------------------------------------------------------------------------
def Get_epsg(g):				
			
    try:
        # Get info of the dataset that is used for transforming     
        gland_proj = g.GetProjection()
        Projection=gland_proj.split('EPSG","')
        epsg_to=int((str(Projection[-1]).split(']')[0])[0:-1])				      
    except:
        epsg_to=4326	
        print 'Was not able to get the projection, so WGS84 is assumed'							
    return(epsg_to)	

#------------------------------------------------------------------------------
def gap_filling(data,NoDataValue):
    """
    This function fills the no data gaps in a numpy array
				
    Keyword arguments:
    dataset -- Array
    NoDataValue -- Value that must be filled
    """
	        
    # fill the no data values
    if NoDataValue is np.nan:
        mask = ~(np.isnan(data))
    else:
        mask = ~(data==NoDataValue)
    xx, yy = np.meshgrid(np.arange(data.shape[1]), np.arange(data.shape[0]))
    xym = np.vstack( (np.ravel(xx[mask]), np.ravel(yy[mask])) ).T
    data0 = np.ravel( data[:,:][mask] )
    interp0 = scipy.interpolate.NearestNDInterpolator( xym, data0 )
    data_end = interp0(np.ravel(xx), np.ravel(yy)).reshape( xx.shape )
		
    return (data_end)

#------------------------------------------------------------------------------				
if __name__ == '__main__':
    main()