# -*- coding: utf-8 -*-
"""
Created on Fri Nov 17 09:25:28 2017

@author: tih
"""
import os
from openpyxl import load_workbook
import pandas as pd

from hants.wa_gdal import *

# Create netcdf file
# Define parameters for the NDVI
VAR = 'NDVI'
VegetationExcel =r'G:\SEBAL_Tadla\Excel\Vegetation height model.xlsx'  # This excel defines the p and c factor and vegetation height.
point = [3542500.60, 681596.50]
cellsize = 100.0

# HANTS parameters to check
nf = 1  # number of frequencies to be considered above the zero frequency
low = 0.0 # valid range minimum
high = 1 # valid range maximum
HiLo = 'Lo' # 2-character string indicating rejection of high or low outliers
fet = 0.1  # fit error tolerance (point eviating more than fet from curve fit are rejected)
delta = 0.1 # small positive number e.g. 0.1 to supress high amplitudes
dod = 1 # degree of overdeterminedness ( iteration stops if number of points reaches the minimum requred for cure itting, plus dod). This is a safety measure


############################## select import paths ##########################################
# Open Excel workbook used for Vegetation c and p factor conversions				
wb_veg = load_workbook(VegetationExcel, data_only=True)		
ws_veg = wb_veg['General_Input']	

# Input for preSEBAL.py
start_date = "%s" %str(ws_veg['B2'].value)  
end_date = "%s" %str(ws_veg['B3'].value)
inputExcel= r"%s" %str(ws_veg['B4'].value)               # The excel with all the SEBAL input data

######################## Load Excels ##########################################	
# Open Excel workbook for SEBAL inputs
wb = load_workbook(inputExcel)
				
# Get length of EXCEL sheet
ws = wb['General_Input']	
ws2 = wb['VIIRS_PROBAV_Input']	
endExcel=int(ws.max_row)

# Create Dict
SEBAL_RUNS = dict()

for number in range(2,endExcel+1):
    input_folder = str(ws['B%d' % number].value)  
    output_folder = str(ws['C%d' % number].value)                    
    Image_Type = int(ws['D%d' % number].value)
    PROBA_V_name  = str(ws2['D%d' % number].value)
    VIIRS_name  = str(ws2['B%d' % number].value) 
    SEBAL_RUNS[number] = {'input_folder': input_folder, 'output_folder': output_folder, 'image_type': Image_Type,'PROBA_V_name': PROBA_V_name,'VIIRS_name': VIIRS_name}

Kind_Of_Runs_Dict = {}
for k, v in SEBAL_RUNS.iteritems():
    Kind_Of_Runs_Dict.setdefault(v['image_type'], []).append(k)
    
# Extract the input and output folder, and Image type from the excel file			
input_folder = str(ws['B%d' % number].value)   
# Create input folder for input Var
input_folder_HANTS = os.path.join(input_folder, 'HANTS_input')
if not os.path.exists(input_folder_HANTS):
    os.mkdir(input_folder_HANTS)

# Define paths for NDVI
input_folder_HANTS_VAR = os.path.join(input_folder_HANTS, VAR)
name_format = '%s_PROBAV_{0}.tif' %VAR
nc_path_ndvi = os.path.join(input_folder_HANTS_VAR,'%s_NC.nc' %VAR)

# Create output folder if not exists
output_folder_HANTS = os.path.join(input_folder, 'HANTS_output')
if not os.path.exists(output_folder_HANTS):
    os.mkdir(output_folder_HANTS)
    
# Create Output folder
rasters_path_out = os.path.join(output_folder_HANTS, VAR)
if not os.path.exists(rasters_path_out):
    os.mkdir(rasters_path_out)

# Run HANTS for a single point
Dates = pd.date_range(start_date, end_date, freq = 'D')
nb = int(len(Dates)) # nr of images

# Define paths for NDVI
input_folder_HANTS_VAR = os.path.join(input_folder_HANTS, VAR)
name_format = '%s_PROBAV_{0}.tif' %VAR
nc_path_ndvi = os.path.join(input_folder_HANTS_VAR,'%s_NC.nc' %VAR)

df = HANTS_singlepoint(nc_path_ndvi, point, nb, nf, HiLo, low, high, fet,
                       dod, delta)
print df

