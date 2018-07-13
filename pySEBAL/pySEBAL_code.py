# -*- coding: utf-8 -*-

"""
pySEBAL_3.4.0

@author: Tim Hessels, Jonna van Opstal, Patricia Trambauer, Wim Bastiaanssen, Mohamed Faouzi Smiej, Yasir Mohamed, and Ahmed Er-Raji
         UNESCO-IHE
         June 2018
"""
import sys
import os
import shutil
import numpy as np
import osr
import gdal
from math import sin, cos, pi, tan
import scipy.misc as misc
import subprocess
import numpy.polynomial.polynomial as poly
from openpyxl import load_workbook
from pyproj import Proj, transform
import warnings

def main(number, inputExcel):

    import SEBAL.pySEBAL.pySEBAL_input_LANDSAT as input_LS
    import SEBAL.pySEBAL.pySEBAL_input_PROBAV_VIIRS as input_PROBAV_VIIRS
    import SEBAL.pySEBAL.pySEBAL_input_MODIS as input_MODIS

    # Do not show warnings
    warnings.filterwarnings('ignore')

    # Open Excel workbook
    wb = load_workbook(inputExcel)

    # Open the General_Input sheet
    ws = wb['General_Input']

    # Extract the input and output folder, and Image type from the excel file
    input_folder = r"%s" %str(ws['B%d' %number].value)
    output_folder = r"%s" %str(ws['C%d' %number].value)
    Image_Type = int(ws['D%d' %number].value)           # Type of Image (1=Landsat & 2 = VIIRS & PROBA-V)

    # Create or empty output folder
    if os.path.isdir(output_folder):
        shutil.rmtree(output_folder)
    os.makedirs(output_folder)

    # Start log file
    filename_logfile = os.path.join(output_folder, 'log.txt')
    sys.stdout = open(filename_logfile, 'w')

    # Print data used from sheet General_Input
    print '.................................................................. '
    print '......................SEBAL Model running ........................ '
    print '.................................................................. '
    print 'pySEBAL version 3.4.0 Github'
    print 'General Input:'
    print 'input_folder = %s' %str(input_folder)
    print 'output_folder = %s' %str(output_folder)
    print 'Image_Type = %s' %int(Image_Type)

    print '.................................................................. '
    print '...........................Parameters ............................ '
    print '.................................................................. '

    # ------------------------------------------------------------------------
    # General constants that could be changed by the user:
    print ' '
    print '...................... General Constants ......................... '
    print ' '

    # Data for Module 1 - Open DEM and reproject
    print 'General Constants: Open DEM and reproject (Part 1)'
    print ' '

    # Data for Module 2 - Radiation
    print 'General Constants: Radiation (Part 2)'
    print ' '

    # Data for Module 3 - Read Soil and Meteo Input
    print 'General Constants: Read Soil and Meteo input (Part 3)'
    print ' '

    # Data for Module 4 - Calc meteo
    Temp_lapse_rate = 0.0065  # Temperature lapse rate (°K/m)
    Gsc = 1367        # Solar constant (W / m2)
    SB_const = 5.6703E-8  # Stefan-Bolzmann constant (watt/m2/°K4)
    print 'General Constants: Calc Meteo (Part 4)'
    print 'Lapse Rate Temperature = %s Kelvin/m' %Temp_lapse_rate
    print 'Solar Constant =  %s W/m2' %Gsc
    print 'Stefan Bolzmann Constant =  %s watt/m2/°K4' %SB_const
    print ' '

    # Data for Module 5 - Open VIS
    Apparent_atmosf_transm = 0.89    # This value is used for atmospheric correction of broad band albedo. This value is used for now, would be better to use tsw.
    path_radiance = 0.03             # Recommended, Range: [0.025 - 0.04], based on Bastiaanssen (2000).
    print 'General Constants: Open VIS (Part 5)'
    print 'Atmospheric correction of broad band albedo = %s' %Apparent_atmosf_transm
    print 'Path Radiance = %s' %path_radiance
    print ' '

    # Data for Module 6 - Open Thermal
    Thermal_Sharpening_not_needed = 0# (1 == off 0 == on)
    Rp = 0.91                        # Path radiance in the 10.4-12.5 µm band (W/m2/sr/µm)
    tau_sky = 0.866                  # Narrow band transmissivity of air, range: [10.4-12.5 µm]
    surf_temp_offset = 3             # Surface temperature offset for water
    Temperature_offset_shadow = -1   # Temperature offset for detecting shadow
    Maximum_shadow_albedo = 0.1      # Minimum albedo value for shadow
    Temperature_offset_clouds = -3   # Temperature offset for detecting clouds
    Minimum_cloud_albedo = 0.4       # Minimum albedo value for clouds
    print 'General Constants: Open Thermal (Part 6)'
    print 'Thermal Sharpening 0:on/1:off = %s'  %Thermal_Sharpening_not_needed
    print 'Path Radiance in the 10.4-12.5 band = %s (W/m2/sr/µm)'  %Rp
    print 'Narrow band transmissivity of air = %s' %tau_sky
    print 'Surface temperature offset for water = %s (Kelvin)' %surf_temp_offset
    print 'Temperature offset for detecting shadow = %s (Kelvin)' %Temperature_offset_shadow
    print 'Maximum albedo value for shadow = %s' %Maximum_shadow_albedo
    print 'Temperature offset for detecting clouds = %s (Kelvin)' %Temperature_offset_clouds
    print 'Minimum albedo value for clouds = %s' %Minimum_cloud_albedo
    print ' '

    # Data for Module 7 - Apply Thermal Sharpening
    print 'Apply Thermal Sharpening (Part 7)'
    print ' '

    # Data for Module 8 - Create Masks and Quality Layers
    print 'Create Masks and Quality Layers (Part 8)'
    print ' '

    # Data for Module 9 - Calc meteo and radiation
    print 'General Constants: Calc meteo and radiation (Part 9)'
    print ' '

    # Data for Module 10 - Calc Hot/Cold Pixel
    NDVIhot_low = 0.03               # Lower NDVI treshold for hot pixels
    NDVIhot_high = 0.25              # Higher NDVI treshold for hot pixels
    print 'General Constants: Calc Hot/Cold Pixel (Part 10)'
    print 'Lower NDVI treshold for hot pixels = %s' %NDVIhot_low
    print 'Higher NDVI treshold for hot pixels = %s' %NDVIhot_high
    print ' '

    # Data for Module 11 - Sensible Heat Flux
    surf_roughness_equation_used = 2 # NDVI model = 1, Raupach model = 2
    print 'General Constants: Sensible Heat Flux (Part 11)'
    print 'NDVI model(1), Raupach model(2) = %s' %surf_roughness_equation_used
    print ' '

    # Data for Module 12 - Evapotranspiration
    print 'General Constants: Evapotranspiration (Part 12)'
    print ' '

    # Data for Module 13 - Soil Moisture
    print 'General Constants: Soil Moisture (Part 13)'
    print ' '

    # Data for Module 14 - Biomass
    Th = 35.0                        # Upper limit of stomatal activity
    Kt = 23.0                        # Optimum conductance temperature (°C), range: [17 - 19]
    Tl = 0.0                         # Lower limit of stomatal activity
    rl = 130                         # Bulk stomatal resistance of the well-illuminated leaf (s/m)
    Light_use_extinction_factor = 0.5    # Light use extinction factor for Bear's Law
    print 'General Constants: Biomass (Part 14)'
    print 'Upper limit of stomatal activity = %s' %Th
    print 'Optimum conductance temperature = %s (Celcius Degrees)' %Kt
    print 'Lower limit of stomatal activity= %s' %Tl
    print 'Bulk stomatal resistance of the well-illuminated leaf = %s (s/m)' %rl
    print 'Light use extinction factor for Bears Law = %s' %(Light_use_extinction_factor)
    print ' '
    print '.................... Input Satellite ........................ '
    # ------------------------------------------------------------------------
    # ------------------------------------------------------------------------
    # ---   Extract general info from Landsat or VIIRS metadata: DOY, hour, minutes

    if Image_Type is 1:

        year, DOY, hour, minutes, UTM_Zone, Sun_elevation, Landsat_nr = input_LS.Get_Time_Info(wb, number)

        # define the kind of sensor and resolution of the sensor
        pixel_spacing = int(30)
        sensor1 = 'LS%d' %Landsat_nr
        sensor2 = 'LS%s' %Landsat_nr
        res1 = '30m'
        res2 = '30m'
        res3 = '30m'

        # Print data used from sheet General_Input
        print 'LANDSAT model Input:'
        print 'Landsat number = %s' %str(Landsat_nr)
        print 'UTM Zone = %s' %(UTM_Zone)
        print 'Pixel size model = %s (Meters)' %(pixel_spacing)

        # Open the Landsat_Input sheet
        ws = wb['Landsat_Input']

    if Image_Type is 2:

        year, DOY, hour, minutes, UTM_Zone = input_PROBAV_VIIRS.Get_Time_Info(wb, number)

        # define the kind of sensor and resolution of the sensor
        pixel_spacing = int(100)
        sensor1 = 'PROBAV'
        sensor2 = 'VIIRS'
        res1 = '375m'
        res2 = '100m'
        res3 = '30m'

        # Print data used from sheet General_Input
        print 'PROBA-V VIIRS model Input:'
        print 'UTM Zone = %s' %(UTM_Zone)
        print 'Pixel size model = %s (Meters)' %(pixel_spacing)

        # Open the VIIRS_PROBAV_Input sheet
        ws = wb['VIIRS_PROBAV_Input']

    if Image_Type is 3:

        year, DOY, UTM_Zone = input_MODIS.Get_Time_Info(wb, number)

        # define the kind of sensor and resolution of the sensor
        pixel_spacing = int(250)
        sensor1 = 'MODIS'
        sensor2 = 'MODIS'
        res1 = '1000m'
        res2 = '250m'
        res3 = '500m'

        # Print data used from sheet General_Input
        print 'MODIS model Input:'
        print 'UTM Zone = %s' %(UTM_Zone)
        print 'Pixel size model = %s (Meters)' %(pixel_spacing)

        # Open the MODIS_Input sheet
        ws = wb['MODIS_Input']

    # Calibartion constants Hot Pixels extracted from the excel file
    Hot_Pixel_Constant = float(ws['E%d' %number].value)          # Hot Pixel Value = Mean_Hot_Pixel + Hot_Pixel_Constant * Std_Hot_Pixel (only for VIIRS images)

    # Calibartion constants Cold Pixels from the excel file
    Cold_Pixel_Constant = float(ws['F%d' %number].value)         # Cold Pixel Value = Mean_Cold_Pixel + Cold_Pixel_Constant * Std_Cold_Pixel (only for VIIRS images)

    # ------------------------------------------------------------------------
    # Define the output maps names

    # output radiation balance
    proyDEM_fileName = os.path.join(output_folder, 'Output_radiation_balance', 'proy_DEM_%s.tif' %res2)
    slope_fileName = os.path.join(output_folder, 'Output_radiation_balance', 'slope_%s.tif' %res2)
    aspect_fileName = os.path.join(output_folder, 'Output_radiation_balance', 'aspect_%s.tif' %res2)
    radiation_inst_fileName = os.path.join(output_folder, 'Output_radiation_balance', 'Ra_inst_%s_%s_%s.tif' %(res2, year, DOY))
    phi_fileName = os.path.join(output_folder, 'Output_radiation_balance', 'phi_%s_%s_%s.tif' %(res2, year, DOY))
    radiation_fileName = os.path.join(output_folder, 'Output_radiation_balance', 'Ra24_mountain_%s_%s_%s.tif' %(res2, year, DOY))
    cos_zn_fileName = os.path.join(output_folder, 'Output_radiation_balance', 'cos_zn_%s_%s_%s.tif' %(res2, year, DOY))
    lon_fileName_rep = os.path.join(output_folder, 'Output_radiation_balance', 'longitude_proj_%s_%s_%s.tif' %(res1, year, DOY))
    lat_fileName_rep = os.path.join(output_folder, 'Output_radiation_balance', 'latitude_proj_%s_%s_%s.tif' %(res1, year, DOY))

    # output meteo
    Atmos_pressure_fileName = os.path.join(output_folder, 'Output_meteo', 'atmos_pressure_%s_%s_%s.tif' %(res2, year, DOY))
    Psychro_c_fileName = os.path.join(output_folder, 'Output_meteo', 'psychro_%s_%s_%s.tif' %(res2, year, DOY))

    # output soil moisture
    water_mask_temp_fileName = os.path.join(output_folder, 'Output_soil_moisture', '%s_Water_mask_temporary_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    snow_mask_fileName = os.path.join(output_folder, 'Output_soil_moisture', '%s_snow_mask_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    water_mask_fileName = os.path.join(output_folder, 'Output_soil_moisture', '%s_water_mask_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    total_soil_moisture_fileName = os.path.join(output_folder, 'Output_soil_moisture', '%s_%s_Total_soil_moisture_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    top_soil_moisture_fileName = os.path.join(output_folder, 'Output_soil_moisture', '%s_%s_Top_soil_moisture_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    RZ_SM_fileName = os.path.join(output_folder, 'Output_soil_moisture', '%s_%s_Root_zone_moisture_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    SM_stress_trigger_fileName = os.path.join(output_folder, 'Output_soil_moisture', '%s_%s_Moisture_stress_trigger_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    irrigation_needs_fileName = os.path.join(output_folder, 'Output_soil_moisture', '%s_%s_irrigation_needs_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))

    # output vegetation
    veg_cover_fileName = os.path.join(output_folder, 'Output_vegetation', '%s_vegt_cover_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    lai_fileName = os.path.join(output_folder, 'Output_vegetation', '%s_lai_average_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    nitrogen_fileName = os.path.join(output_folder, 'Output_vegetation', '%s_nitrogen_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    tir_emissivity_fileName = os.path.join(output_folder, 'Output_vegetation', '%s_tir_emissivity_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    fpar_fileName = os.path.join(output_folder, 'Output_vegetation', '%s_fpar_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    b10_emissivity_fileName = os.path.join(output_folder, 'Output_vegetation', '%s_b10_emissivity_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    surf_temp_fileName = os.path.join(output_folder, 'Output_vegetation', '%s_%s_surface_temp_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    temp_surface_sharpened_fileName =  os.path.join(output_folder, 'Output_vegetation', '%s_%s_surface_temp_sharpened_%s_%s_%s.tif' %(sensor1, sensor2, res1, year, DOY))
    surf_rough_fileName = os.path.join(output_folder, 'Output_vegetation', '%s_%s_surface_roughness_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    surface_albedo_fileName = os.path.join(output_folder, 'Output_vegetation','%s_surface_albedo_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    ndvi_fileName = os.path.join(output_folder, 'Output_vegetation','%s_ndvi_%s_%s_%s.tif' %(sensor1, res2, year, DOY))

    # output cloud mask
    cloud_mask_fileName = os.path.join(output_folder, 'Output_cloud_masked', '%s_cloud_mask_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    shadow_mask_fileName = os.path.join(output_folder, 'Output_cloud_masked', '%s_shadow_mask_%s_%s_%s.tif' %(sensor1, res2, year, DOY))
    QC_Map_fileName = os.path.join(output_folder, 'Output_cloud_masked', '%s_quality_mask_%s_%s_%s.tif.tif' %(sensor1, res2, year, DOY))

    # output energy balance
    Rn_24_fileName = os.path.join(output_folder, 'Output_energy_balance', '%s_%s_Rn_24_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    rn_inst_fileName = os.path.join(output_folder, 'Output_energy_balance', '%s_%s_Rn_inst_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    g_inst_fileName = os.path.join(output_folder, 'Output_energy_balance', '%s_%s_G_inst_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    h_inst_fileName = os.path.join(output_folder, 'Output_energy_balance', '%s_%s_h_inst_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    EF_inst_fileName = os.path.join(output_folder, 'Output_energy_balance', '%s_%s_EFinst_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    LE_inst_fileName = os.path.join(output_folder, 'Output_energy_balance', '%s_%s_LEinst_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))

    # output temporary
    temp_corr_fileName = os.path.join(output_folder, 'Output_temporary', '%s_%s_temp_corr_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    ts_dem_fileName = os.path.join(output_folder, 'Output_temporary', '%s_%s_ts_dem_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    hot_pixels_fileName = os.path.join(output_folder, 'Output_temporary', '%s_%s_hot_pixels_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    cold_pixels_fileName = os.path.join(output_folder, 'Output_temporary', '%s_%s_cold_pixels_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    QC_Map_after_VIS = os.path.join(output_folder, 'Output_temporary', '%s_QC_MAP_After_VIS_%s_%s_%s.tif' %(sensor1, res1, year, DOY))
    proyDEM_fileName_up = os.path.join(output_folder, 'Output_temporary', 'proy_DEM_up.tif')

    # output evapotranspiration
    min_bulk_surf_res_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_%s_min_bulk_surf_resis_24_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    ETref_24_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_ETref_24_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    ETA_24_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_ETact_24_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    ETP_24_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_ETpot_24_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    ET_24_deficit_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_ET_24_deficit_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    AF_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_Advection_Factor_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    kc_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_kc_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    kc_max_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_kc_max_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    bulk_surf_res_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_bulk_surf_resis_24_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    Tact24_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_Tact_24_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    Eact24_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_Eact_24_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    Tpot24_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_Tpot_24_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    T24_deficit_fileName = os.path.join(output_folder, 'Output_evapotranspiration', '%s_%s_T_24_deficit_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))

    # output biomass production
    moisture_stress_biomass_fileName = os.path.join(output_folder, 'Output_biomass_production', '%s_%s_Moisture_stress_biomass_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    LUE_fileName = os.path.join(output_folder, 'Output_biomass_production', '%s_%s_LUE_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    Biomass_prod_fileName = os.path.join(output_folder, 'Output_biomass_production', '%s_%s_Biomass_production_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    Biomass_wp_fileName = os.path.join(output_folder, 'Output_biomass_production', '%s_%s_Biomass_wp_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
    Biomass_deficit_fileName = os.path.join(output_folder, 'Output_biomass_production', '%s_%s_Biomass_deficit_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))

    print '---------------------------------------------------------'
    print '------------------ General info -------------------------'
    print '---------------------------------------------------------'
    print 'General info: '
    print '  DOY: ', DOY

    if not Image_Type == 3:
        print '  Hour: ', hour
        print '  Minutes: ', '%0.3f' % minutes

    print '  UTM_Zone: ', UTM_Zone

    print '---------------------------------------------------------'
    print '---------- Open DEM and reproject (Part 1) --------------'
    print '---------------------------------------------------------'

    ws = wb['General_Input']

    # Extract the Path to the DEM map from the excel file
    DEM_fileName = r"%s" %str(ws['E%d' %number].value) #'DEM_HydroShed_m'
    print 'Path to DEM file = %s' %str(DEM_fileName)

    # Open DEM and create Latitude and longitude files
    lat, lon, lat_fileName, lon_fileName = DEM_lat_lon(DEM_fileName, output_folder)

    # Reproject from Geog Coord Syst to UTM -
    # 1) DEM - Original DEM coordinates is Geographic: lat, lon
    lsc, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset(
                DEM_fileName, pixel_spacing, UTM_Zone = UTM_Zone)
    band = lsc.GetRasterBand(1)   # Get the reprojected dem band
    ncol = lsc.RasterXSize        # Get the reprojected dem column size
    nrow = lsc.RasterYSize        # Get the reprojected dem row size
    shape_lsc = [ncol, nrow]

    # Read out the DEM band and print the DEM properties
    DEM_resh = band.ReadAsArray(0, 0, ncol, nrow)
    #DEM_resh[DEM_resh<0] = 1

    print 'Projected DEM - '
    print '   Size: ', ncol, nrow
    print '   Upper Left corner x, y: ', ulx_dem, ',', uly_dem
    print '   Lower right corner x, y: ', lrx_dem, ',', lry_dem

    # 2) Latitude File - reprojection
    # reproject latitude to the landsat projection and save as tiff file
    lat_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset(
               lat_fileName, pixel_spacing, UTM_Zone=UTM_Zone)

    # Get the reprojected latitude data
    lat_proy = lat_rep.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)

    # 3) Longitude file - reprojection
    # reproject longitude to the landsat projection	 and save as tiff file
    lon_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset(lon_fileName, pixel_spacing, UTM_Zone)

    # Get the reprojected longitude data
    lon_proy = lon_rep.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)

    # Calculate slope and aspect from the reprojected DEM
    deg2rad, rad2deg, slope, aspect = Calc_Gradient(DEM_resh, pixel_spacing)

    # Saving the reprojected maps
    save_GeoTiff_proy(lsc, DEM_resh, proyDEM_fileName, shape_lsc, nband = 1)
    save_GeoTiff_proy(lsc, slope, slope_fileName, shape_lsc, nband = 1)
    save_GeoTiff_proy(lsc, aspect, aspect_fileName, shape_lsc, nband = 1)
    save_GeoTiff_proy(lon_rep, lon_proy, lon_fileName_rep, shape_lsc, nband = 1)
    save_GeoTiff_proy(lat_rep, lat_proy, lat_fileName_rep, shape_lsc, nband = 1)

    print '---------------------------------------------------------'
    print '---------------- Radiation (Part 2) ---------------------'
    print '---------------------------------------------------------'

    # now we can also get the time for a MODIS run
    if Image_Type == 3:

           hour, minutes = input_MODIS.Modis_Time(wb, epsg_to, number, proyDEM_fileName)
           hour = np.nanmean(hour)
           minutes = np.nanmean(minutes)


    # Calculation of extraterrestrial solar radiation for slope and aspect
    Ra_mountain_24, Ra_inst, cos_zn, dr, phi, delta = Calc_Ra_Mountain(lon, DOY, hour, minutes, lon_proy, lat_proy, slope, aspect)

    if Image_Type == 2 or Image_Type == 3:
        Sun_elevation = 90 - (np.nanmean(cos_zn) * 180/np.pi)

    # Save files created in module 1
    save_GeoTiff_proy(lsc, cos_zn, cos_zn_fileName, shape_lsc, nband = 1)
    save_GeoTiff_proy(lsc, Ra_mountain_24, radiation_fileName, shape_lsc, nband = 1)
    save_GeoTiff_proy(lsc, Ra_inst, radiation_inst_fileName, shape_lsc, nband = 1 )
    save_GeoTiff_proy(lsc, phi, phi_fileName, shape_lsc, nband = 1 )

    print '---------------------------------------------------------'
    print '------- Read Meteo and Soil inputs (Part 3) -------------'
    print '---------------------------------------------------------'

    # Open the Meteo_Input sheet
    ws = wb['Meteo_Input']

    # 6a) Instantanious Temperature
    Output_filename_temp_inst = os.path.join(output_folder, 'Output_radiation_balance', 'Temp_24_input.tif')
    Temp_inst, Temp_inst_source = Open_constant_or_spatial_map(ws, "B%d" %number, Output_filename_temp_inst, proyDEM_fileName)
    print '_____________________Instantanious Temperature______________________'
    print 'Source of instantanious temperature = %s' %str(Temp_inst_source)
    print 'Average instantanious temperature = %s Kelvin\n' %float(np.nanmean(Temp_inst))

    # 6b) Daily Temperature
    Output_filename_temp_24 = os.path.join(output_folder, 'Output_radiation_balance', 'Temp_24_input.tif')
    Temp_24, Temp_24_source = Open_constant_or_spatial_map(ws, "C%d" %number, Output_filename_temp_24, proyDEM_fileName)
    print '__________________________Daily Temperature_________________________'
    print 'Source of daily temperature = %s' %str(Temp_24_source)
    print 'Average daily temperature = %s Kelvin\n' %float(np.nanmean(Temp_24))

    # 6c) Instantanious Relative Humidity
    Output_filename_RH_inst = os.path.join(output_folder, 'Output_radiation_balance', 'RH_inst_input.tif')
    RH_inst, RH_inst_source = Open_constant_or_spatial_map(ws, "D%d" %number, Output_filename_RH_inst, proyDEM_fileName)
    print '________________Instantanious Relative Humidity_____________________'
    print 'Source of instantanious relative humidity = %s' %str(RH_inst_source)
    print 'Average instantanious relative humidity = %s Procent\n' %float(np.nanmean(RH_inst))

    # 6d) Daily Relative Humidity
    Output_filename_RH_24 = os.path.join(output_folder, 'Output_radiation_balance', 'RH_24_input.tif')
    RH_24, RH_24_source = Open_constant_or_spatial_map(ws, "E%d" %number, Output_filename_RH_24, proyDEM_fileName)
    print '____________________Daily Relative Humidity_________________________'
    print 'Source of daily relative humidity = %s' %str(RH_24_source)
    print 'Average daily relative humidity = %s Procent\n' %float(np.nanmean(RH_24))

    # 6) Wind speed measurement height
    zx = float(ws['F%d' %number].value)
    print '___________________Measurement Height Wind Speed____________________'
    print 'Height at which wind speed is measured = %s (m)\n' %(zx)

    # 6e) Instantanious wind speed
    Output_filename_wind_inst = os.path.join(output_folder, 'Output_radiation_balance', 'Wind_inst_input.tif')
    Wind_inst, Wind_inst_source = Open_constant_or_spatial_map(ws, "G%d" %number, Output_filename_wind_inst, proyDEM_fileName)
    print '_____________________Instantanious Wind Speed_______________________'
    print 'Source of instantanious wind speed = %s' %str(Wind_inst_source)
    print 'Average instantanious wind speed = %s m/s\n' %float(np.nanmean(Wind_inst))

    # 6f) Daily wind speed
    Output_filename_wind_24 = os.path.join(output_folder, 'Output_radiation_balance', 'Wind_24_input.tif')
    Wind_24, Wind_24_source = Open_constant_or_spatial_map(ws, "H%d" %number, Output_filename_wind_24, proyDEM_fileName)
    print '__________________________Daily Wind Speed__________________________'
    print 'Source of daily wind speed = %s' %str(Wind_24_source)
    print 'Average daily wind speed = %s m/s\n' %float(np.nanmean(Wind_24))

    # 6g) instantanious radiation or transmissivity

    # Define the method of radiation (1 or 2)
    Method_Radiation_inst=int(ws['I%d' %number].value)     # 1=Transm_inst will be calculated Rs_inst must be given
                                                           # 2=Rs_inst will be determined Transm_inst must be given
    print '________________________Instantanious Solar_________________________'
    print 'Method for instantanious radiation (1=Rs_inst, 2=Transm_inst) = %s\n' %(Method_Radiation_inst)

    if Method_Radiation_inst == 1:
        Output_filename_radiation_inst = os.path.join(output_folder, 'Output_radiation_balance', 'Rs_inst_input.tif')
        Rs_inst, Rs_inst_source = Open_constant_or_spatial_map(ws, "J%d" %number, Output_filename_radiation_inst, proyDEM_fileName)
        print '____________________Instantanious Radiation_________________________'
        print 'Source of instantanious solar radiation = %s' %str(Rs_inst_source)
        print 'Average instantanious solar radiation = %s W/m2\n' %float(np.nanmean(Rs_inst))

    if Method_Radiation_inst == 2:
        Output_filename_transm_inst = os.path.join(output_folder, 'Output_radiation_balance', 'Transm_inst_input.tif')
        Transm_inst, Transm_inst_source = Open_constant_or_spatial_map(ws, "K%d" %number, Output_filename_transm_inst, proyDEM_fileName)
        print '___________________Instantanious Transmissivity_____________________'
        print 'Source of instantanious transmissivity = %s' %str(Transm_inst_source)
        print 'Average instantanious transmissivity = %s\n' %float(np.nanmean(Transm_inst))

    # 6h) daily radiation or transmissivity

    # Define the method of radiation (1 or 2)
    Method_Radiation_24=int(ws['L%d' %number].value)     # 1=Transm_inst will be calculated Rs_24 must be given
                                                           # 2=Rs_inst will be determined Transm_24 must be given
    print '____________________________Daily Solar_____________________________'
    print 'Method for daily radiation (1=Rs_24, 2=Transm_24) = %s\n' %(Method_Radiation_24)

    if Method_Radiation_24 == 1:
        Output_filename_radiation_24 = os.path.join(output_folder, 'Output_radiation_balance', 'Rs_24_input.tif')
        Rs_24, Rs_24_source = Open_constant_or_spatial_map(ws, "M%d" %number, Output_filename_radiation_24, proyDEM_fileName)
        print '____________________________Daily Radiation_________________________'
        print 'Source of daily solar radiation = %s' %str(Rs_24_source)
        print 'Average daily solar radiation = %s W/m2\n' %float(np.nanmean(Rs_24))

    if Method_Radiation_24 == 2:
        Output_filename_transm_24 = os.path.join(output_folder, 'Output_radiation_balance', 'Transm_24_input.tif')
        Transm_24, Transm_24_source = Open_constant_or_spatial_map(ws, "N%d" %number, Output_filename_transm_24, proyDEM_fileName)
        print '___________________________Daily Transmissivity_____________________'
        print 'Source of daily transmissivity = %s' %str(Transm_24_source)
        print 'Average daily transmissivity = %s\n' %float(np.nanmean(Transm_24))

    # 6i) Obstacle height
    Output_filename_h_obst = os.path.join(output_folder, 'Output_soil_moisture', 'Obst_h_input.tif')
    h_obst, h_obst_source = Open_constant_or_spatial_map(ws, "O%d" %number, Output_filename_h_obst, proyDEM_fileName)
    print '___________________________Obstacle Height__________________________'
    print 'Source of obstacle height = %s' %str(h_obst_source)
    print 'Average obstacle height = %s meter\n' %float(np.nanmean(h_obst))

    # Open the Meteo_Input sheet
    ws = wb['Soil_Input']

    # 6j) Saturated Soil Moisture Content topsoil
    Output_filename_Theta_sat_top = os.path.join(output_folder, 'Output_soil_moisture', 'Theta_sat_top_input.tif')
    Theta_sat_top, Theta_sat_top_source = Open_constant_or_spatial_map(ws, "B%d" %number, Output_filename_Theta_sat_top, proyDEM_fileName)
    print '________________Saturated Soil Moisture Content Topsoil_____________'
    print 'Source of the saturated soil moisture content topsoil = %s' %str(Theta_sat_top_source)
    print 'Average saturated soil moisture content topsoil = %s\n' %float(np.nanmean(Theta_sat_top))

    # 6k) Saturated Soil Moisture Content subsoil
    Output_filename_Theta_sat_sub = os.path.join(output_folder, 'Output_soil_moisture', 'Theta_sat_sub_input.tif')
    Theta_sat_sub, Theta_sat_sub_source = Open_constant_or_spatial_map(ws, "C%d" %number, Output_filename_Theta_sat_sub, proyDEM_fileName)
    print '________________Saturated Soil Moisture Content Subsoil_____________'
    print 'Source of the saturated soil moisture content subsoil = %s' %str(Theta_sat_sub_source)
    print 'Average saturated soil moisture content subsoil = %s\n' %float(np.nanmean(Theta_sat_sub))

    # 6l) Residual Soil Moisture Content topsoil
    Output_filename_Theta_res_top = os.path.join(output_folder, 'Output_soil_moisture', 'Theta_res_top_input.tif')
    Theta_res_top, Theta_res_top_source = Open_constant_or_spatial_map(ws, "D%d" %number, Output_filename_Theta_res_top, proyDEM_fileName)
    print '_________________Residual Soil Moisture Content Topsoil_____________'
    print 'Source of the residual soil moisture content topsoil = %s' %str(Theta_res_top_source)
    print 'Average residual soil moisture content topsoil = %s\n' %float(np.nanmean(Theta_res_top))

    # 6m) Residual Soil Moisture Content subsoil
    Output_filename_Theta_res_sub = os.path.join(output_folder, 'Output_soil_moisture', 'Theta_res_sub_input.tif')
    Theta_res_sub, Theta_res_sub_source = Open_constant_or_spatial_map(ws, "E%d" %number, Output_filename_Theta_res_sub, proyDEM_fileName)
    print '_________________Residual Soil Moisture Content Subsoil_____________'
    print 'Source of the residual soil moisture content subsoil = %s' %str(Theta_res_sub_source)
    print 'Average residual soil moisture content subsoil = %s\n' %float(np.nanmean(Theta_res_sub))

    # 6n) Soil Moisture Wilting point
    Output_filename_soil_wilting_point = os.path.join(output_folder, 'Output_soil_moisture', 'Soil_moisture_wilting_point_input.tif')
    Soil_moisture_wilting_point, Soil_moisture_wilting_point_source = Open_constant_or_spatial_map(ws, "G%d" %number, Output_filename_soil_wilting_point, proyDEM_fileName)
    print '_______________________Soil Moisture Wilting point__________________'
    print 'Source of the soil moisture wilting point = %s' %str(Soil_moisture_wilting_point_source)
    print 'Average soil moisture wilting point = %s\n' %float(np.nanmean(Soil_moisture_wilting_point))

    # 6o) Fraction Field Capacity
    Output_filename_Field_Capacity = os.path.join(output_folder, 'Output_soil_moisture', 'Fraction_Field_Capacity_input.tif')
    Field_Capacity, Field_Capacity_source = Open_constant_or_spatial_map(ws, "F%d" %number, Output_filename_Field_Capacity, proyDEM_fileName)
    print '_________________________Fraction Field Capacity____________________'
    print 'Source of the fraction field capacity = %s' %str(Field_Capacity_source)
    print 'Average fraction field capacity = %s\n' %float(np.nanmean(Field_Capacity))

    # 6p) Light Use Efficiency
    Output_filename_LUEmax = os.path.join(output_folder, 'Output_soil_moisture', 'LUEmax_input.tif')
    LUEmax, LUEmax_source = Open_constant_or_spatial_map(ws, "I%d" %number, Output_filename_LUEmax, proyDEM_fileName)
    print '______________________Maximum Light Use Efficiency__________________'
    print 'Source of the Maximum Light Use Efficiency = %s' %str(LUEmax_source)
    print 'Average Maximum Light Use Efficiency = %s\n' %float(np.nanmean(LUEmax))

    # 6p) Depletion Factor
    Output_filename_depl_factor = os.path.join(output_folder, 'Output_soil_moisture', 'depl_factor_input.tif')
    depl_factor, depl_factor_source = Open_constant_or_spatial_map(ws, "H%d" %number, Output_filename_depl_factor, proyDEM_fileName)
    print '______________________________Depletion Factor______________________'
    print 'Source of the Depletion Factor = %s' %str(depl_factor_source)
    print 'Average Depletion Factor = %s\n' %float(np.nanmean(depl_factor))

    print '---------------------------------------------------------'
    print '---------------- Calc Meteo (Part 4) --------------------'
    print '---------------------------------------------------------'

    # Atmospheric pressure for altitude:
    Pair = 101.3 * np.power((293 - Temp_lapse_rate * DEM_resh) / 293, 5.26)

    # Psychrometric constant (kPa / °C), FAO 56, eq 8.:
    Psychro_c = 0.665E-3 * Pair

    # Saturation Vapor Pressure at the air temperature (kPa):
    esat_inst = 0.6108 * np.exp(17.27 * Temp_inst / (Temp_inst + 237.3))
    esat_24 = 0.6108 * np.exp(17.27 * Temp_24 / (Temp_24 + 237.3))

    # Actual vapour pressure (kPa), FAO 56, eq 19.:
    eact_inst = RH_inst * esat_inst / 100
    eact_24 = RH_24 * esat_24 / 100
    print 'Instantaneous Saturation Vapor Pressure = ', '%0.3f (kPa)' % np.nanmean(esat_inst)
    print 'Instantaneous Actual vapour pressure =  ', '%0.3f (kPa)' % np.nanmean(eact_inst)
    print 'Daily Saturation Vapor Pressure = ', '%0.3f (kPa)' % np.nanmean(esat_24)
    print 'Daily Actual vapour pressure =  ', '%0.3f (kPa)' % np.nanmean(eact_24)

    print '---------------------------------------------------------'
    print '------------ Open VIS Parameters (Part 5) ---------------'
    print '---------------------------------------------------------'

    if Image_Type == 1:
        Surf_albedo, NDVI, LAI, vegt_cover, FPAR, Nitrogen, tir_emis, b10_emissivity, water_mask_temp, QC_Map = input_LS.Get_LS_Para_Veg(wb, number, proyDEM_fileName, year, DOY, path_radiance, Apparent_atmosf_transm, cos_zn, dr)

    if Image_Type == 2:
        Surf_albedo, NDVI, LAI, vegt_cover, FPAR, Nitrogen, tir_emis, b10_emissivity, water_mask_temp, QC_Map = input_PROBAV_VIIRS.Get_PROBAV_Para_Veg(wb, number, proyDEM_fileName, year, DOY, path_radiance, Apparent_atmosf_transm, cos_zn, dr, DEM_resh)

    if Image_Type == 3:
        Surf_albedo, NDVI, LAI, vegt_cover, FPAR, Nitrogen, tir_emis, b10_emissivity, water_mask_temp, QC_Map = input_MODIS.Get_MODIS_Para_Veg(wb, number, proyDEM_fileName, year, DOY, path_radiance, Apparent_atmosf_transm, cos_zn, dr, DEM_resh, epsg_to)

    # Save output maps
    save_GeoTiff_proy(lsc, water_mask_temp, water_mask_temp_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, FPAR, fpar_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, tir_emis, tir_emissivity_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, Nitrogen, nitrogen_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, vegt_cover, veg_cover_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, LAI, lai_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, b10_emissivity, b10_emissivity_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, NDVI, ndvi_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, Surf_albedo, surface_albedo_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, QC_Map, QC_Map_after_VIS, shape_lsc, nband=1)

    print '---------------------------------------------------------'
    print '--------- Open Thermal Parameters (Part 6) --------------'
    print '---------------------------------------------------------'

    if Image_Type == 1:
        Surface_temp, cloud_mask_temp, Thermal_Sharpening_not_needed = input_LS.Get_LS_Para_Thermal(wb, number, proyDEM_fileName, year, DOY,  water_mask_temp, b10_emissivity, Temp_inst, Rp, tau_sky, surf_temp_offset, Thermal_Sharpening_not_needed, DEM_fileName, UTM_Zone, eact_inst, QC_Map)

    if Image_Type == 2:
        Surface_temp, cloud_mask_temp , Thermal_Sharpening_not_needed = input_PROBAV_VIIRS.Get_VIIRS_Para_Thermal(wb, number, proyDEM_fileName, year, DOY, water_mask_temp, b10_emissivity, Temp_inst,  Rp, tau_sky, surf_temp_offset, Thermal_Sharpening_not_needed)

    if Image_Type == 3:
        Surface_temp, cloud_mask_temp, Thermal_Sharpening_not_needed = input_MODIS.Get_MODIS_Para_Thermal(wb, number, proyDEM_fileName, year, DOY, water_mask_temp, b10_emissivity, Temp_inst,  Rp, tau_sky, surf_temp_offset, Thermal_Sharpening_not_needed, epsg_to)

    # Save output maps
    save_GeoTiff_proy(lsc, Surface_temp, surf_temp_fileName, shape_lsc, nband=1)

    print '---------------------------------------------------------'
    print '------ Apply Thermal Sharpening (Part 7) ----------------'
    print '---------------------------------------------------------'

    # Perform Thermal sharpening for the thermal band
    if Thermal_Sharpening_not_needed is 1:
        temp_surface_sharpened = Surface_temp

    if Thermal_Sharpening_not_needed is 0:

        # Create mask for thermal sharpening
        Total_mask_thermal = QC_Map + cloud_mask_temp + water_mask_temp
        Total_mask_thermal[Total_mask_thermal > 0] = 1

    	  # Upscale DEM
        if Image_Type == 1:
            pixel_spacing_upscale=90
            Box = 7
        if Image_Type == 2:
            pixel_spacing_upscale=400
            Box = 9
        if Image_Type == 3:
            pixel_spacing_upscale=1000
            Box = 9

        dest_up, ulx_dem_up, lry_dem_up, lrx_dem_up, uly_dem_up, epsg_to = reproject_dataset(
            DEM_fileName, pixel_spacing_upscale, UTM_Zone = UTM_Zone)

        DEM_up = dest_up.GetRasterBand(1).ReadAsArray()
        Y_raster_size_up = dest_up.RasterYSize
        X_raster_size_up = dest_up.RasterXSize
        shape_up=([X_raster_size_up, Y_raster_size_up])

        save_GeoTiff_proy(dest_up, DEM_up, proyDEM_fileName_up, shape_up, nband=1)

        # save landsat surface temperature
        surf_temp_fileName = os.path.join(output_folder, 'Output_vegetation','%s_%s_surface_temp_%s_%s_%s.tif' %(sensor1, sensor2, res2, year, DOY))
        save_GeoTiff_proy(lsc, Surface_temp, surf_temp_fileName, shape_lsc, nband=1)

        # Upscale NDVI data
        dest_up, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
                                   ndvi_fileName, proyDEM_fileName_up)

        NDVI_Landsat_up = dest_up.GetRasterBand(1).ReadAsArray()

        # upscale the mask to coarser resolution
        Total_mask_thermal_up = resize_array_example(Total_mask_thermal, NDVI_Landsat_up, method=2)
        Total_mask_thermal_up[Total_mask_thermal_up>0]=1

        # Upscale Thermal data
        dest_up, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
                                   surf_temp_fileName, proyDEM_fileName_up)
        surface_temp_up = dest_up.GetRasterBand(1).ReadAsArray()

        # Remove wrong values
        surface_temp_up[surface_temp_up==0] = np.nan
        NDVI_Landsat_up[NDVI_Landsat_up==0] = np.nan
        surface_temp_up[surface_temp_up==1] = np.nan
        NDVI_Landsat_up[Total_mask_thermal_up==1] = np.nan
        NDVI[Total_mask_thermal==1] = np.nan

        # Apply thermal sharpening
        temp_surface_sharpened = Thermal_Sharpening(surface_temp_up, NDVI_Landsat_up, NDVI, Box, dest_up, output_folder, ndvi_fileName, shape_lsc, lsc)

        # Replace water values to original thermal  values
        temp_surface_sharpened[water_mask_temp == 1] = Surface_temp[water_mask_temp == 1]
        temp_surface_sharpened[np.isnan(temp_surface_sharpened)] = Surface_temp[np.isnan(temp_surface_sharpened)]

    # remove low temperature values
    temp_surface_sharpened[temp_surface_sharpened <= 253.0]=np.nan

    # Calculate the tempearture of the water
    Temperature_water_std=np.nanstd(temp_surface_sharpened[water_mask_temp != 0])
    Temperature_water_mean=np.nanmean(temp_surface_sharpened[water_mask_temp != 0])
    print 'Mean water Temperature = %0.3f (K)' % Temperature_water_mean
    print 'Standard deviation water temperature = %0.3f (K)' % Temperature_water_std

	 # save landsat surface temperature
    save_GeoTiff_proy(lsc, temp_surface_sharpened, temp_surface_sharpened_fileName, shape_lsc, nband=1)

    print '---------------------------------------------------------'
    print '------- Create Masks and Quality Layers (Part 8) --------'
    print '---------------------------------------------------------'

    # Check Quality
    try:
        ws = wb['Additional_Input']
        if (ws['F%d' % number].value) is not None:
            # Output folder QC defined by the user
            QC_Map_fileName = os.path.join(output_folder, 'Output_cloud_masked', 'User_quality_mask_%s_%s_%s.tif' %(res2, year, DOY))

            # Reproject and reshape users NDVI
            QC_Map = Reshape_Reproject_Input_data(r'%s' %str(ws['F%d' % number].value), QC_Map_fileName, proyDEM_fileName)

        else:

            snow_mask, water_mask, ts_moist_veg_min, NDVI_max, NDVI_std = CalculateSnowWaterMask(NDVI,shape_lsc,water_mask_temp,Surface_temp)
            Temperature_water_mean=np.nanmean(temp_surface_sharpened[water_mask != 0])

            if np.isnan(Temperature_water_mean) == True or Temperature_water_mean < 0.0:
                ts_cold_land=ts_moist_veg_min
            else:
                ts_cold_land=Temperature_water_mean

            # Make shadow mask
            shadow_mask=np.zeros((shape_lsc[1], shape_lsc[0]))
            shadow_mask[np.logical_and.reduce((temp_surface_sharpened < (ts_cold_land+Temperature_offset_shadow),Surf_albedo < Maximum_shadow_albedo,water_mask!=1))]=1
            shadow_mask = Create_Buffer(shadow_mask)

            # Improve cloud mask for Landsat
            if Image_Type == 1:

                # open worksheet
                ws = wb['Landsat_Input']

                # Extract Landsat name
                Name_Landsat_Image = str(ws['B%d' %number].value)

                if os.path.exists(os.path.join(input_folder, '%s_BQA.TIF' %Name_Landsat_Image)):
                    cloud_mask_temp[np.logical_and.reduce((Surface_temp < (ts_cold_land+Temperature_offset_clouds),Surf_albedo > Minimum_cloud_albedo,NDVI<0.7,snow_mask!=1))]=1
                    cloud_mask = Create_Buffer(cloud_mask_temp)            # if there are no cold water pixels than use cold vegetation pixels
                else:
                    cloud_mask = cloud_mask_temp
            else:
                cloud_mask_temp[np.logical_and.reduce((Surface_temp < (ts_cold_land+Temperature_offset_clouds),Surf_albedo > Minimum_cloud_albedo,NDVI<0.7,snow_mask!=1))]=1
                cloud_mask = Create_Buffer(cloud_mask_temp)

            # Total Quality Mask
            Tot_Masks = cloud_mask + snow_mask + shadow_mask + QC_Map
            QC_Map[Tot_Masks>0] = 1

            # Output folder QC defined by the user
            QC_Map_fileName = os.path.join(output_folder, 'Output_cloud_masked', '%s_quality_mask_%s_%s_%s.tif.tif' %(sensor1, res2, year, DOY))

            # Save output maps
            save_GeoTiff_proy(lsc, cloud_mask, cloud_mask_fileName, shape_lsc, nband=1)
            save_GeoTiff_proy(lsc, snow_mask, snow_mask_fileName, shape_lsc, nband=1)
            save_GeoTiff_proy(lsc, shadow_mask, shadow_mask_fileName, shape_lsc, nband=1)
            save_GeoTiff_proy(lsc, QC_Map, QC_Map_fileName, shape_lsc, nband=1)

    except:
        assert "Please check the quality path"

    # Check Water Mask and replace the temporay
    try:
        ws = wb['Additional_Input']
        if (ws['E%d' % number].value) is not None:

            # Overwrite the Water mask and change the output name
            water_mask_fileName = os.path.join(output_folder, 'Output_soil_moisture', 'User_Water_mask_temporary_%s_%s_%s.tif' %(res2, year, DOY))
            water_mask = Reshape_Reproject_Input_data(r'%s' %str(ws['E%d' % number].value), water_mask_temp_fileName, proyDEM_fileName)

    except:
        assert "Please check the Water Mask input path"

    if not "water_mask" in locals():
        water_mask = water_mask_temp

    # Save output maps
    save_GeoTiff_proy(lsc, water_mask, water_mask_fileName, shape_lsc, nband=1)

    print '---------------------------------------------------------'
    print '------- Meteo and Radiation Continue (Part 9) -----------'
    print '---------------------------------------------------------'

    # Slope of satur vapour pressure curve at air temp (kPa / °C)
    sl_es_24 = 4098 * esat_24 / np.power(Temp_24 + 237.3, 2)

    # Daily 24 hr radiation - For flat terrain only !
    ws_angle = np.arccos(-np.tan(phi)*tan(delta))   # Sunset hour angle ws

    # Extraterrestrial daily radiation, Ra (W/m2):
    Ra24_flat = (Gsc/np.pi * dr * (ws_angle * np.sin(phi[nrow/2, ncol/2]) * np.sin(delta) +
                    np.cos(phi[nrow/2, ncol/2]) * np.cos(delta) * np.sin(ws_angle)))

    # calculate the daily radiation or daily transmissivity or daily surface radiation based on the method defined by the user
    if Method_Radiation_24==1:
       Transm_24 = Rs_24/Ra_mountain_24

    if Method_Radiation_24==2:
       Rs_24 = Ra_mountain_24 * Transm_24

    # Solar radiation from extraterrestrial radiation
    Rs_24_flat = Ra24_flat * Transm_24
    print 'Mean Daily Transmissivity = %0.3f (-)' % np.nanmean(Transm_24)
    print 'Mean Daily incoming net Radiation = %0.3f (W/m2)' % np.nanmean(Rs_24)
    print 'Mean Daily incoming net Radiation Flat Terrain = %0.3f (W/m2)' % np.nanmean(Rs_24_flat)

    # If method of instantaneous radiation 1 is used than calculate the Transmissivity
    if Method_Radiation_inst==1:
        Transm_corr=Rs_inst/Ra_inst

    # If method of instantaneous radiation 2 is used than calculate the instantaneous incomming Radiation
    if Method_Radiation_inst==2:
        # calculate the transmissivity index for direct beam radiation
        Transm_corr = Transm_inst + 2e-5 * DEM_resh
        # Instantaneous incoming short wave radiation (W/m2):
        Rs_inst = Ra_inst * Transm_corr

    # Atmospheric emissivity, by Bastiaanssen (1995):
    Transm_corr[Transm_corr<0.001]=0.1
    Transm_corr[Transm_corr>1]=1
    atmos_emis = 0.85 * np.power(-np.log(Transm_corr), 0.09)

    # Instantaneous incoming longwave radiation:
    lw_in_inst = atmos_emis * SB_const * np.power(Temp_inst + 273.15, 4)
    print 'Instantaneous longwave incoming radiation = %0.3f (W/m2)' % np.nanmean(lw_in_inst)
    print 'Atmospheric emissivity = %0.3f' % np.nanmean(atmos_emis)

    # calculates the ground heat flux and the solar radiation
    Rn_24,rn_inst,g_inst,Rnl_24_FAO = Calc_Meteo(Rs_24,eact_24,Temp_24,Surf_albedo,dr,tir_emis,temp_surface_sharpened,water_mask,NDVI,Transm_24,SB_const,lw_in_inst,Rs_inst)

    print 'Mean Daily Net Radiation (FAO) = %0.3f (W/m2)' % np.nanmean(Rnl_24_FAO)
    print 'Mean Daily Net Radiation = %0.3f (W/m2)' % np.nanmean(Rn_24)
    print 'Mean instantaneous Net Radiation = %0.3f (W/m2)' % np.nanmean(rn_inst)
    print 'Mean instantaneous Ground Heat Flux = %0.3f (W/m2)' % np.nanmean(g_inst)

    # Save output maps
    save_GeoTiff_proy(lsc, Rn_24, Rn_24_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, rn_inst, rn_inst_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, g_inst, g_inst_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, Pair, Atmos_pressure_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, Psychro_c, Psychro_c_fileName, shape_lsc, nband=1)

    print '---------------------------------------------------------'
    print '---------------- Hot/Cold Pixels (Part 10) --------------'
    print '---------------------------------------------------------'

    # Temperature at sea level corrected for elevation: ??
    ts_dem,air_dens,Temp_corr=Correct_Surface_Temp(temp_surface_sharpened,Temp_lapse_rate,DEM_resh,Pair,dr,Transm_corr,cos_zn,Sun_elevation,deg2rad,QC_Map)

    # Selection of hot and cold pixels

    # Open Additional_Input sheet in the excel
    ws = wb['Additional_Input']
    if (ws['G%d' % number].value) is not None:
        ts_dem_cold = float(ws['G%d' % number].value) + 273.15
        print 'cold pixel defined by the user: value=%0.3f (Kelvin)' %ts_dem_cold

    else:
        if not "NDVI_max" in locals():
            NDVI_max = np.nanmax(NDVI)
            NDVI_std = np.nanstd(NDVI)

        # Cold pixels vegetation
        ts_dem_cold_veg = Calc_Cold_Pixels_Veg(NDVI,NDVI_max,NDVI_std, QC_Map,ts_dem,Image_Type, Cold_Pixel_Constant)

        # Cold pixels water
        ts_dem_cold,cold_pixels,ts_dem_cold_mean = Calc_Cold_Pixels(ts_dem,water_mask,QC_Map,ts_dem_cold_veg,Cold_Pixel_Constant)
        if np.isnan(ts_dem_cold) == True:
            ts_dem_cold = Temp_inst
        save_GeoTiff_proy(lsc, cold_pixels, cold_pixels_fileName, shape_lsc, nband=1)

    if (ws['H%d' % number].value) is not None:
        ts_dem_hot = float(ws['H%d' % number].value) + 273.15
        print 'hot pixel defined by the user: value=%0.3f (Kelvin)' %ts_dem_hot
        for_hot = np.copy(ts_dem)
        for_hot[NDVI <= NDVIhot_low] = 0.0
        for_hot[NDVI >= NDVIhot_high] = 0.0
        for_hot[np.logical_or(water_mask != 0.0, QC_Map != 0.0)] = 0.0
        hot_pixels = np.copy(for_hot)
        hot_pixels[for_hot < ts_dem_cold] = np.nan

    else:
        # Hot pixels
        ts_dem_hot,hot_pixels = Calc_Hot_Pixels(ts_dem,QC_Map, water_mask,NDVI,NDVIhot_low,NDVIhot_high, Hot_Pixel_Constant, ts_dem_cold)
        save_GeoTiff_proy(lsc, hot_pixels, hot_pixels_fileName, shape_lsc, nband=1)

    # Save files
    save_GeoTiff_proy(lsc, Temp_corr, temp_corr_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, ts_dem, ts_dem_fileName, shape_lsc, nband=1)

    print '---------------------------------------------------------'
    print '------------ Sensible heat flux (Part 11) ---------------'
    print '---------------------------------------------------------'

    # Change the minimum windspeed to prevent high values in further calculations
    Wind_inst = np.where(Wind_inst<1.5, 1.5, Wind_inst)
    Wind_24 = np.where(Wind_24<1.5, 1.5, Wind_24)

    # calculate windspeed at the blending height and the friction velocity by using the Raupach model or NDVI
    Surf_roughness,u_200,ustar_1=Calc_Wind_Speed_Friction(h_obst,Wind_inst,zx,LAI,NDVI,Surf_albedo,water_mask,surf_roughness_equation_used)
    save_GeoTiff_proy(lsc, Surf_roughness, surf_rough_fileName, shape_lsc, nband=1)

    # Computation of surface roughness for momentum transport
    k_vk = 0.41      # Von Karman constant

    # Sensible heat 1 (Step 5)
    # Corrected value for the aerodynamic resistance (eq 41 with psi2 = psi1):
    rah1 = np.log(2.0/0.01) / (k_vk * ustar_1)
    i=0
    L, psi_m200_stable, psi, psi_m200,h_inst,dT, slope_dt, offset_dt = sensible_heat(
            rah1, ustar_1, rn_inst, g_inst, ts_dem, ts_dem_hot, ts_dem_cold,
            air_dens, temp_surface_sharpened, k_vk,QC_Map, hot_pixels, slope)

    # do the calculation iteratively 10 times
    for i in range(1,10):
        L,psi,psi_m200,psi_m200_stable,h_inst,ustar_corr,rah_corr,dT, slope_dt, offset_dt = Iterate_Friction_Velocity(k_vk,u_200,Surf_roughness,g_inst,rn_inst, ts_dem, ts_dem_hot, ts_dem_cold,air_dens, temp_surface_sharpened,L,psi,psi_m200,psi_m200_stable,QC_Map, hot_pixels, slope)

    # Save files
    save_GeoTiff_proy(lsc, h_inst, h_inst_fileName, shape_lsc, nband=1)

    print '---------------------------------------------------------'
    print '-------------- Evaporation (Part 12) --------------------'
    print '---------------------------------------------------------'

    # calculate reference net radiation
    Rn_ref, Refl_rad_water, rah_grass=Calc_Rn_Ref(shape_lsc,water_mask,Rn_24,Ra_mountain_24,Transm_24,Rnl_24_FAO,Wind_24)

    # Calculate rah of PM for the ET act (dT after iteration) and ETpot (4 degrees)
    rah_pm_act=((np.log((2.0-0.0)/(Surf_roughness*0.1))*np.log((2.0-0.0)/(Surf_roughness)))/(k_vk*1.5**2))*((1-5*(-9.82*dT*(2.0-0.0))/((273.15+Temp_inst)*1.5**2))**(-0.75))
    rah_pm_act[rah_pm_act<25]=25

    rah_pm_pot=((np.log((2.0-0.0)/(Surf_roughness*0.1))*np.log((2.0-0.0)/(Surf_roughness)))/(k_vk*1.5**2))*((1-5*(-9.82*4.0*(2.0-0.0))/((273.15+Temp_inst)*1.5**2))**(-0.75))
    rah_pm_pot[rah_pm_pot<25]=25

    # calculate reference potential evaporation.
    ETpot_24,ETref_24,Lhv,rs_min=Calc_Ref_Pot_ET(LAI,temp_surface_sharpened,sl_es_24,Rn_ref,air_dens,esat_24,eact_24,rah_grass,Psychro_c,Rn_24,Refl_rad_water,rah_pm_pot,rl)

    # Instantaneous evapotranspiration
    LE_inst = rn_inst - g_inst - h_inst

    # Evaporative fraction
    EF_inst=Calc_instantaneous_ET_fraction(LE_inst,rn_inst,g_inst)

    # Daily Evaporation and advection factor
    ETA_24, AF=Calc_ETact(esat_24,eact_24,EF_inst,Rn_24,Refl_rad_water,Lhv, Image_Type)

    # Bulk surface resistance (s/m):
    bulk_surf_resis_24=Calc_Bulk_surface_resistance(sl_es_24,Rn_24,Refl_rad_water,air_dens,esat_24,eact_24,rah_pm_act,ETA_24,Lhv,Psychro_c)

    # crop factor
    kc = ETA_24 / ETref_24  # Crop factor
    ETP_24 = np.where(ETpot_24 < ETA_24, ETA_24, ETpot_24)
    ET_24_deficit = ETP_24 - ETA_24
    kc_max = ETP_24 / ETref_24

    # Save files
    save_GeoTiff_proy(lsc, rs_min, min_bulk_surf_res_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, EF_inst, EF_inst_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, LE_inst, LE_inst_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, ETref_24, ETref_24_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, ETA_24, ETA_24_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, ETP_24, ETP_24_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, ET_24_deficit, ET_24_deficit_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, AF, AF_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, kc, kc_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, kc_max, kc_max_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, bulk_surf_resis_24, bulk_surf_res_fileName, shape_lsc, nband=1)

    print '---------------------------------------------------------'
    print '--------------- Soil Moisture (Part 13) -----------------'
    print '---------------------------------------------------------'

    #  Calculate soil properties
    #SM_stress_trigger, total_soil_moisture, RZ_SM,moisture_stress_biomass,irrigation_needs,top_soil_moisture=Calc_Soil_Moisture(ETA_24,accum_prec_14d,accum_ETo_14d,EF_inst,water_mask,vegt_cover,Theta_sat,Theta_res)
    SM_stress_trigger, total_soil_moisture, root_zone_moisture_first, moisture_stress_biomass_first,top_soil_moisture,RZ_SM_NAN = Calc_Soil_Moisture(ETA_24,EF_inst,QC_Map,water_mask,vegt_cover,Theta_sat_top,Theta_sat_sub, Theta_res_top,Theta_res_sub, depl_factor,Field_Capacity,FPAR, Soil_moisture_wilting_point)

    # seperation of E and T
    Eact_24,Tpot_24,Tact_24,moisture_stress_biomass,T24_deficit,beneficial_fraction,root_zone_moisture_final,top_zone_moisture_final=Separate_E_T(Light_use_extinction_factor,LAI,ETP_24,Theta_res_top, Theta_res_sub,Theta_sat_top,Theta_sat_sub,top_soil_moisture,sl_es_24, Psychro_c,moisture_stress_biomass_first,vegt_cover,ETA_24,SM_stress_trigger,root_zone_moisture_first,total_soil_moisture)

    # Irrigation:
    irrigation_needs = Classify_Irrigation(moisture_stress_biomass, vegt_cover)

    # Save files
    save_GeoTiff_proy(lsc, Tact_24, Tact24_fileName,shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, Eact_24, Eact24_fileName,shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, Tpot_24, Tpot24_fileName,shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, T24_deficit, T24_deficit_fileName,shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, total_soil_moisture, total_soil_moisture_fileName,shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, top_zone_moisture_final, top_soil_moisture_fileName,shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, root_zone_moisture_final, RZ_SM_fileName, shape_lsc,nband=1)
    save_GeoTiff_proy(lsc, SM_stress_trigger, SM_stress_trigger_fileName,shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, moisture_stress_biomass, moisture_stress_biomass_fileName,shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, irrigation_needs, irrigation_needs_fileName,shape_lsc, nband=1)

    print '---------------------------------------------------------'
    print '------------------ Biomass (Part 14)---------------------'
    print '---------------------------------------------------------'

    # calculate biomass production
    LUE,Biomass_prod,Biomass_wp,Biomass_deficit = Calc_Biomass_production(LAI,ETP_24,moisture_stress_biomass,ETA_24,Ra_mountain_24,Transm_24,FPAR,esat_24,eact_24,Th,Kt,Tl,Temp_24,LUEmax)

    # Save files
    save_GeoTiff_proy(lsc, LUE, LUE_fileName,shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, Biomass_prod, Biomass_prod_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, Biomass_wp, Biomass_wp_fileName, shape_lsc, nband=1)
    save_GeoTiff_proy(lsc, Biomass_deficit, Biomass_deficit_fileName,shape_lsc, nband=1)
    lsc=None

    print '...................................................................'
    print '............................DONE!..................................'
    print '...................................................................'

    # ------------------------------------------------------------------------
    # ------------------------------------------------------------------------
    # FUNCTIONS
    #-------------------------------------------------------------------------

def Create_Buffer(Data_In):

   '''
   This function creates a 3D array which is used to apply the moving window
   '''
   Buffer_area = 2 # A block of 2 times Buffer_area + 1 will be 1 if there is the pixel in the middle is 1
   Data_Out=np.empty((len(Data_In),len(Data_In[1])))
   Data_Out[:,:] = Data_In
   for ypixel in range(0,Buffer_area + 1):

        for xpixel in range(1,Buffer_area + 1):

           if ypixel==0:
                for xpixel in range(1,Buffer_area + 1):
                    Data_Out[:,0:-xpixel] += Data_In[:,xpixel:]
                    Data_Out[:,xpixel:] += Data_In[:,:-xpixel]

                for ypixel in range(1,Buffer_area + 1):

                    Data_Out[ypixel:,:] += Data_In[:-ypixel,:]
                    Data_Out[0:-ypixel,:] += Data_In[ypixel:,:]

           else:
               Data_Out[0:-xpixel,ypixel:] += Data_In[xpixel:,:-ypixel]
               Data_Out[xpixel:,ypixel:] += Data_In[:-xpixel,:-ypixel]
               Data_Out[0:-xpixel,0:-ypixel] += Data_In[xpixel:,ypixel:]
               Data_Out[xpixel:,0:-ypixel] += Data_In[:-xpixel,ypixel:]

   Data_Out[Data_Out>0.1] = 1
   Data_Out[Data_Out<=0.1] = 0

   return(Data_Out)

def Calc_Biomass_production(LAI,ETP_24,moisture_stress_biomass,ETA_24,Ra_mountain_24,Transm_24,FPAR,esat_24,eact_24,Th,Kt,Tl,Temp_24,LUEmax):
    """
    Function to calculate the biomass production and water productivity
    """

    Ksolar = Ra_mountain_24 * Transm_24

    # Incident Photosynthetically active radiation (PAR, MJ/m2) per time period
    PAR = 0.48 * Ksolar

    # Aborbed Photosynthetical Active Radiation (APAR) by the vegetation:
    APAR = FPAR * PAR

    vapor_stress = 0.88 - 0.183 * np.log(esat_24 - eact_24)
    vapor_stress_biomass = vapor_stress.clip(0.0, 1.0)
    Jarvis_coeff = (Th - Kt) / (Kt - Tl)
    heat_stress_biomass = ((Temp_24 - Tl) * np.power(Th - Temp_24, Jarvis_coeff) /
                           ((Kt - Tl) * np.power(Th - Kt, Jarvis_coeff)))
    print 'vapor stress biomass =', '%0.3f' % np.nanmean(vapor_stress_biomass)
    print 'heat stress biomass =', '%0.3f' % np.nanmean(heat_stress_biomass)

    # Light use efficiency, reduced below its potential value by low
    # temperature or water shortage:
    LUE = (LUEmax * heat_stress_biomass * vapor_stress_biomass * moisture_stress_biomass)

    # Dry matter production (kg/ha/d):
    Biomass_prod = APAR * LUE * 0.864             # C3 vegetation

    # Water productivity
    Biomass_wp = Biomass_prod/ (ETA_24 * 10)  # C3 vegetation
    Biomass_wp[ETA_24 == 0.0] = 0.0

    # Water deficit
    Biomass_deficit = (Biomass_prod / moisture_stress_biomass -
                          Biomass_prod)


    return(LUE,Biomass_prod,Biomass_wp,Biomass_deficit)

#------------------------------------------------------------------------------
def Classify_Irrigation(moisture_stress_biomass, vegt_cover):
   '''
   This function makes a classification with 4 categories which show the irrigation needs
   '''
   for_irrigation = np.copy(moisture_stress_biomass)

   # make a discreed irrigation needs map with the following categories
      # Irrigation needs:
       # 0: No need for irrigation
       # 1: Perhaps irrigate
       # 2: Irrigate
       # 3: Irrigate immediately
   irrigation_needs = np.copy(for_irrigation)
   irrigation_needs[np.where(irrigation_needs >= 1.0)] == 0.0
   irrigation_needs[np.logical_and(irrigation_needs >= 0.9, irrigation_needs < 1.0)] = 1.0
   irrigation_needs[np.where((irrigation_needs >= 0.8) & (irrigation_needs < 0.9))] = 2.0
   irrigation_needs[np.where(irrigation_needs < 0.8)] = 3.0
   irrigation_needs[vegt_cover <= 0.3] = 0.0
   return(irrigation_needs)

#------------------------------------------------------------------------------
def Separate_E_T(Light_use_extinction_factor,LAI,ETP_24,Theta_res_top,Theta_res_sub, Theta_sat_top, Theta_sat_sub, top_soil_moisture,sl_es_24, Psychro_c,moisture_stress_biomass_first,vegt_cover,ETA_24,SM_stress_trigger,root_zone_moisture_first,total_soil_moisture):
   '''
   Separate the Evapotranspiration into evaporation and Transpiration
   '''
   # constants

   Tpot_24_estimate=(1-np.exp(-Light_use_extinction_factor*LAI))*ETP_24
   SE_top = (top_soil_moisture-Theta_res_top)/(Theta_sat_top-Theta_res_top)
   Eact_24_estimate=np.minimum(1,1 / np.power(SE_top + 0.1,-2.0))*(ETP_24-Tpot_24_estimate)
   #RS_soil = RS_soil_min * np.power(SE_top,-2.0)
   #Eact_24_estimate=(sl_es_24+Psychro_c*(1+RS_soil_min/Rah_PM))/(sl_es_24+Psychro_c*(1+RS_soil/Rah_PM))*(ETP_24-Tpot_24_estimate)
   n66_memory = moisture_stress_biomass_first * Tpot_24_estimate

   # calulate the first estimation of actual daily tranpiration
   Tact_24_estimate = np.copy(n66_memory)
   Tact_24_estimate[n66_memory > 0.99*ETA_24]=ETA_24[n66_memory > 0.99*ETA_24]
   Tact_24_estimate[vegt_cover == 0.0] = 0.0

   # calculate the second estimation and end estimation of the actual daily tranpiration
   Tact_24 = np.abs((Tact_24_estimate/(Tact_24_estimate + Eact_24_estimate))*ETA_24)

   # calculate the actual daily potential transpiration
   Tpot_24 = np.copy(Tpot_24_estimate)
   Tpot_24[Tpot_24_estimate < Tact_24] = Tact_24[Tpot_24_estimate < Tact_24]

   # calculate moisture stress biomass
   moisture_stress_biomass = Tact_24 / Tpot_24

   # Calculate root zone moisture final
   Se_Poly=2.23*np.power(moisture_stress_biomass,3)-3.35*np.power(moisture_stress_biomass,2)+1.98*moisture_stress_biomass+0.07
   root_zone_moisture1=Se_Poly*(SM_stress_trigger+0.02-Theta_res_sub)+Theta_res_sub
   root_zone_moisture_final=np.where(root_zone_moisture1>root_zone_moisture_first,root_zone_moisture1,root_zone_moisture_first)

   # Calculate top zone moisture final
   top_zone_moisture1=(total_soil_moisture-root_zone_moisture_final*vegt_cover)/(1-vegt_cover)
   top_zone_moisture_final=top_zone_moisture1.clip(Theta_res_top,Theta_sat_top)

   # calculate the actual daily evaporation
   Eact_24 = ETA_24 - Tact_24

   # calculate the Transpiration deficit
   T24_deficit = Tpot_24 - Tact_24

   # calculate the beneficial fraction
   beneficial_fraction=Tact_24 / ETA_24
   beneficial_fraction[ETA_24 == 0.0] = 0.0

   return(Eact_24,Tpot_24,Tact_24,moisture_stress_biomass,T24_deficit,beneficial_fraction,root_zone_moisture_final,top_zone_moisture_final)

#------------------------------------------------------------------------------
def Calc_Soil_Moisture(ETA_24,EF_inst,QC_Map, water_mask,vegt_cover,Theta_sat_top, Theta_sat_sub,Theta_res_top, Theta_res_sub,depl_factor,Field_Capacity,FPAR, Soil_moisture_wilting_point):
    """
    Function to calculate soil characteristics
    """
    # constants:
    Veg_Cover_Threshold_RZ = 0.9 # Threshold vegetation cover for root zone moisture

    # Average fraction of TAW that can be depleted from the root zone
    # before stress:
    p_factor = depl_factor + 0.04 * (5.0 - ETA_24)  # page 163 of FAO 56
    # The factor p differs from one crop to another. It normally varies from
    # 0.30 for shallow rooted plants at high rates of ETc (> 8 mm d-1)
    # to 0.70 for deep rooted plants at low rates of ETc (< 3 mm d-1)

    # Critical value under which plants get stressed:
    SM_stress_trigger = Field_Capacity - p_factor * (Field_Capacity - Soil_moisture_wilting_point)
    EF_inst[EF_inst >= 1.0] = 0.999

    # Total soil water content (cm3/cm3):
    total_soil_moisture = Theta_sat_sub * np.exp((EF_inst - 1.0) / 0.421)   # asce paper Scott et al. 2003
    total_soil_moisture[np.logical_or(water_mask == 1.0,QC_Map == 1.0)] = 1.0  # In water and snow is 1
    total_soil_moisture[QC_Map == 1.0] = np.nan  # Where clouds no data

    # Root zone soil moisture:
    RZ_SM = np.copy(total_soil_moisture)
    RZ_SM[vegt_cover <= Veg_Cover_Threshold_RZ] = np.nan
    if np.isnan(np.nanmean(RZ_SM)) == True:
        Veg_Cover_Threshold_RZ = np.nanpercentile(vegt_cover, 80)
        RZ_SM = np.copy(total_soil_moisture)
        RZ_SM[vegt_cover <= Veg_Cover_Threshold_RZ] = np.nan
        print 'No RZ_SM so the vegetation Threshold for RZ is adjusted from 0,9 to =', '%0.3f' % Veg_Cover_Threshold_RZ

    #RZ_SM = RZ_SM.clip(Theta_res, (0.85 * Theta_sat))
    #RZ_SM[np.logical_or(water_mask == 1.0, water_mask == 2.0)] = 1.0
    RZ_SM_NAN = np.copy(RZ_SM)
    RZ_SM_NAN[RZ_SM==0] = np.nan
    RZ_SM_min = np.nanmin(RZ_SM_NAN)
    RZ_SM_max = np.nanmax(RZ_SM_NAN)
    RZ_SM_mean = np.nanmean(RZ_SM_NAN)
    print 'Root Zone Soil moisture mean =', '%0.3f (cm3/cm3)' % RZ_SM_mean
    print 'Root Zone Soil moisture min =', '%0.3f (cm3/cm3)' % RZ_SM_min
    print 'Root Zone Soil moisture max =', '%0.3f (cm3/cm3)' % RZ_SM_max

    Max_moisture_RZ = vegt_cover * (RZ_SM_max - RZ_SM_mean) + RZ_SM_mean

    # Soil moisture in the top (temporary)
    top_soil_moisture_temp = np.copy(total_soil_moisture)
    top_soil_moisture_temp[np.logical_or(vegt_cover <= 0.02, vegt_cover >= 0.1)] = 0
    top_soil_moisture_temp[top_soil_moisture_temp == 0] = np.nan
    top_soil_moisture_std = np.nanstd(top_soil_moisture_temp)
    top_soil_moisture_mean = np.nanmean(top_soil_moisture_temp)
    print 'Top Soil moisture mean =', '%0.3f (cm3/cm3)' % top_soil_moisture_mean
    print 'Top Soil moisture Standard Deviation', '%0.3f (cm3/cm3)' % top_soil_moisture_std

    # calculate root zone moisture
    root_zone_moisture_temp = (total_soil_moisture - (top_soil_moisture_mean + top_soil_moisture_std) * (1-vegt_cover))/vegt_cover # total soil moisture = soil moisture no vegtatation *(1-vegt_cover)+soil moisture root zone * vegt_cover
    try:
        root_zone_moisture_temp[root_zone_moisture_temp <= Theta_res_sub] = Theta_res_sub[root_zone_moisture_temp <= Theta_res_sub]
    except:
        root_zone_moisture_temp[root_zone_moisture_temp <= Theta_res_sub] = Theta_res_sub


    root_zone_moisture_temp[root_zone_moisture_temp >= Max_moisture_RZ] = Max_moisture_RZ[root_zone_moisture_temp >= Max_moisture_RZ]

    root_zone_moisture_first = np.copy(root_zone_moisture_temp)
    root_zone_moisture_first[np.logical_or(QC_Map ==1.0 ,np.logical_or(water_mask == 1.0, vegt_cover < 0.0))] = 0

    # Normalized stress trigger:
    norm_trigger = (root_zone_moisture_first - Soil_moisture_wilting_point)/ (SM_stress_trigger + 0.02 - Soil_moisture_wilting_point)
    norm_trigger[norm_trigger > 1.0] = 1.0

    # moisture stress biomass:
    moisture_stress_biomass_first = norm_trigger - (np.sin(2 * np.pi * norm_trigger)) / (2 * np.pi)
    moisture_stress_biomass_first=np.where(moisture_stress_biomass_first<0.5*FPAR,0.5*FPAR,moisture_stress_biomass_first)
    moisture_stress_biomass_first[moisture_stress_biomass_first <= 0.0] = 0
    moisture_stress_biomass_first[moisture_stress_biomass_first > 1.0] = 1.0

     # Soil moisture in the top layer - Recalculated ??
    top_soil_moisture = ((total_soil_moisture - root_zone_moisture_first * vegt_cover) / (1.0 - vegt_cover))

    try:
        top_soil_moisture[top_soil_moisture > Theta_sat_top] = Theta_sat_top [top_soil_moisture > Theta_sat_top]
    except:
        top_soil_moisture[top_soil_moisture > Theta_sat_top] = Theta_sat_top

    top_soil_moisture[np.logical_or(water_mask == 1.0, QC_Map == 1.0)] = 1.0

    return(SM_stress_trigger, total_soil_moisture, root_zone_moisture_first, moisture_stress_biomass_first,top_soil_moisture,RZ_SM_NAN)

#------------------------------------------------------------------------------
def Calc_Bulk_surface_resistance(sl_es_24,Rn_24,Refl_rad_water,air_dens,esat_24,eact_24,rah_pm_act,ETA_24,Lhv,Psychro_c):
    """
    Function to calculate the bulk surface resistance
    """
    # Bulk surface resistance (s/m):
    bulk_surf_resis_24 = ((((sl_es_24 * (Rn_24 - Refl_rad_water) + air_dens *
                          1004 * (esat_24 - eact_24) / rah_pm_act) / (ETA_24 * Lhv / 86400) -
                          sl_es_24) / Psychro_c - 1.0) * rah_pm_act)
    bulk_surf_resis_24[ETA_24 <= 0.0] = 100000.0
    bulk_surf_resis_24 = bulk_surf_resis_24.clip(0.0, 100000.0)
    return(bulk_surf_resis_24)

#------------------------------------------------------------------------------
def Calc_ETact(esat_24, eact_24, EF_inst, Rn_24, Refl_rad_water, Lhv, Image_Type):
    """
    Function to calculate the daily evaporation
    """
    # Advection factor
    if Image_Type == 2:
        AF = np.ones(Rn_24.shape)
    else:
        AF = 1 + 0.985 * (np.exp((esat_24 - eact_24) * 0.08) - 1.0) * EF_inst
    # Daily evapotranspiration:
    ETA_24 = EF_inst * AF * (Rn_24 - Refl_rad_water) / (Lhv * 1000) * 86400000
    ETA_24=ETA_24.clip(0,15.0)
    return(ETA_24, AF)

#------------------------------------------------------------------------------
def Calc_instantaneous_ET_fraction(LE_inst,rn_inst,g_inst):
    """
    Function to calculate the evaporative fraction
    """
    EF_inst = LE_inst / (rn_inst - g_inst)    # Evaporative fraction
    EF_inst = EF_inst.clip(0.0, 1.8)
    EF_inst[LE_inst<0] = 0

    return(EF_inst)

#------------------------------------------------------------------------------
def Calc_Ref_Pot_ET(LAI,Surface_temp,sl_es_24,Rn_ref,air_dens,esat_24,eact_24,rah_grass,Psychro_c,Rn_24,Refl_rad_water,rah_pm_pot,rl):
    """
    Function to calculate the reference potential evapotransporation and potential evaporation
    """

    # Effective leaf area index involved, see Allen et al. (2006):
    LAI_eff = LAI / (0.3 * LAI + 1.2)
    rs_min = rl / LAI_eff  # Min (Bulk) surface resistance (s/m)
    # Latent heat of vaporization (J/kg):
    Lhv = (2.501 - 2.361e-3 * (Surface_temp - 273.15)) * 1E6

    # Reference evapotranspiration- grass
    # Penman-Monteith of the combination equation (eq 3 FAO 56) (J/s/m2)
    LET_ref_24 = ((sl_es_24 * Rn_ref + air_dens * 1004 * (esat_24 - eact_24) /
                  rah_grass) / (sl_es_24 + Psychro_c * (1 + 70.0/rah_grass)))
    # Reference evaportranspiration (mm/d):
    ETref_24 = LET_ref_24 / (Lhv * 1000) * 86400000

    # Potential evapotranspiration
    # Penman-Monteith of the combination equation (eq 3 FAO 56) (J/s/m2)
    LETpot_24 = ((sl_es_24 * (Rn_24 - Refl_rad_water) + air_dens * 1004 *
               (esat_24 - eact_24)/rah_pm_pot) / (sl_es_24 + Psychro_c * (1 + rs_min/rah_pm_pot)))
    # Potential evaportranspiration (mm/d)
    ETpot_24 = LETpot_24 / (Lhv * 1000) * 86400000
    ETpot_24[ETpot_24 > 15.0] = 15.0
    return(ETpot_24,ETref_24,Lhv,rs_min)

#------------------------------------------------------------------------------
def Calc_Rn_Ref(shape_lsc,water_mask,Rn_24,Ra_mountain_24,Transm_24,Rnl_24_FAO,Wind_24):
    """
    Function to calculate the net solar radiation
    """
    # constants:
    G24_water = 0.1  # G24 ratio for water - reflectivity?

    # Reflected radiation at water surface: ??
    Refl_rad_water = np.zeros((shape_lsc[1], shape_lsc[0]))
    Refl_rad_water = np.where(water_mask != 0.0, G24_water * Rn_24, 0.0)

    # Aerodynamic resistance (s/m) for grass surface:
    rah_grass = 208.0 / Wind_24
    print  'rah_grass=', '%0.3f (s/m)' % np.nanmean(rah_grass)
    # Net radiation for grass Rn_ref, eq 40, FAO56:
    Rn_ref = Ra_mountain_24 * Transm_24 * (1 - 0.23) - Rnl_24_FAO  # Rnl avg(fao-slob)?
    return(Rn_ref, Refl_rad_water,rah_grass)


#------------------------------------------------------------------------------
def Iterate_Friction_Velocity(k_vk,u_200,Surf_roughness,g_inst,rn_inst, ts_dem, ts_dem_hot, ts_dem_cold,air_dens, Surface_temp,L,psi,psi_m200,psi_m200_stable,QC_Map, hot_pixels, slope):
    """
    Function to correct the windspeed and aerodynamic resistance for the iterative process the output can be used as the new input for this model
    """
    # Sensible heat 2 (Step 6)
    # Corrected value for the friction velocity, unstable
    ustar_corr_unstable = (k_vk * u_200 / (np.log(200.0 / Surf_roughness) -
                        psi_m200))
    # Corrected value for the friction velocity, stable
    ustar_corr_stable = (k_vk * u_200 / (np.log(200.0 / Surf_roughness) -
                      psi_m200_stable))
    ustar_corr = np.where(L > 0.0, ustar_corr_stable, ustar_corr_unstable)
    ustar_corr[ustar_corr < 0.02] = 0.02

    rah_corr_unstable = (np.log(2.0/0.01) - psi) / (k_vk * ustar_corr)     # unstable
    rah_corr_stable = (np.log(2.0/0.01) - 0.0) / (k_vk * ustar_corr)       # stable
    rah_corr = np.where(L > 0.0, rah_corr_stable, rah_corr_unstable)
    L_corr, psi_m200_corr_stable, psi_corr, psi_m200_corr,h,dT, slope_dt, offset_dt = sensible_heat(
            rah_corr, ustar_corr, rn_inst, g_inst, ts_dem, ts_dem_hot, ts_dem_cold,
            air_dens, Surface_temp, k_vk,QC_Map, hot_pixels, slope)
    return(L_corr,psi_corr,psi_m200_corr,psi_m200_corr_stable,h,ustar_corr,rah_corr,dT,slope_dt, offset_dt)

#------------------------------------------------------------------------------
def Calc_Wind_Speed_Friction(h_obst,Wind_inst,zx,LAI,NDVI,Surf_albedo,water_mask,surf_roughness_equation_used):
    """
    Function to calculate the windspeed and friction by using the Raupach or NDVI model
    """

    # constants
    k_vk = 0.41      # Von Karman constant
    h_grass = 0.12   # Grass height (m)
    cd = 53          # Free parameter for displacement height, default = 20.6
    # 1) Raupach model
    zom_Raupach=Raupach_Model(h_obst,cd,LAI)

    # 2) NDVI model
    zom_NDVI=NDVI_Model(NDVI,Surf_albedo,water_mask)

    if surf_roughness_equation_used == 1:
        Surf_roughness = zom_NDVI
    else:
        Surf_roughness = zom_Raupach

    zom_grass = 0.123 * h_grass
    # Friction velocity for grass (m/s):
    ustar_grass = k_vk * Wind_inst / np.log(zx / zom_grass)
    print 'u*_grass = ', '%0.3f (m/s)' % np.mean(ustar_grass)
    # Wind speed (m/s) at the "blending height" (200m):
    u_200 = ustar_grass * np.log(200 / zom_grass) / k_vk
    print 'Wind speed at the blending height, u200 =', '%0.3f (m/s)' % np.mean(u_200)
    # Friction velocity (m/s):
    ustar_1 = k_vk * u_200 / np.log(200 / Surf_roughness)
    return(Surf_roughness,u_200,ustar_1)

#------------------------------------------------------------------------------
def Raupach_Model(h_obst,cd,LAI):
    """
    Function for the Raupach model to calculate the surface roughness (based on Raupach 1994)
    """
    # constants
    cw = 2.0
    LAIshelter = 2.5

    # calculate psi
    psi = np.log(cw) - 1 + np.power(2.0, -1)  # Vegetation influence function

    # Calculate Ustar divided by U
    ustar_u = np.power((0.003+0.3*LAI/2), 0.5)
    ustar_u[LAI<LAIshelter] = 0.3

    # calculate: 1 - d/hv
    inv_d_hv =(1-np.exp(-1*np.power((cd*LAI),0.5)))/np.power((cd * LAI),0.5)

    # Calculate: surface roughness/hv
    zom_hv = inv_d_hv * np.exp(-0.41/ustar_u-psi)

    # Calculate: surface roughness
    zom_Raupach = zom_hv * h_obst

    return(zom_Raupach)

#------------------------------------------------------------------------------
def NDVI_Model(NDVI,Surf_albedo,water_mask):
    """
    Function for the NDVI model to calculate the surface roughness
    """
    zom_NDVI = np.exp(1.096 * NDVI / Surf_albedo - 5.307)
    zom_NDVI[water_mask == 1.0] = 0.001
    zom_NDVI[zom_NDVI > 10.0] = 10.0
    return(zom_NDVI)

#------------------------------------------------------------------------------
def Correct_Surface_Temp(Surface_temp,Temp_lapse_rate,DEM_resh,Pair,dr,Transm_corr,cos_zn,Sun_elevation,deg2rad,ClipLandsat):
    """
    Function to correct the surface temperature based on the DEM map
    """
    #constants:
    Gsc = 1367        # Solar constant (W / m2)

    cos_zenith_flat = np.cos((90 - Sun_elevation) * deg2rad)
    Temp_corr = Surface_temp + Temp_lapse_rate * DEM_resh  # rescale everything to sea level
    Temp_corr[Surface_temp == 350.0] = 0.0
    air_dens = 1000 * Pair / (1.01 * Surface_temp * 287)
    #
    ts_dem = (Temp_corr + (Gsc * dr * Transm_corr * cos_zn -
              Gsc * dr * Transm_corr * cos_zenith_flat) / (air_dens * 1004 * 0.050))
    #(Temp_corr - (Gsc * dr * Transm_corr * cos_zn -
    #          Gsc * dr * Transm_corr * cos_zenith_flat) / (air_dens * 1004 * 0.050))
    ts_dem[ClipLandsat==1]=np.nan
    ts_dem[ts_dem==0]=np.nan
    ts_dem[ts_dem<273]=np.nan
    ts_dem[ts_dem>350]=np.nan

    return(ts_dem,air_dens,Temp_corr)

#------------------------------------------------------------------------------
def Calc_Hot_Pixels(ts_dem,QC_Map, water_mask, NDVI,NDVIhot_low,NDVIhot_high,Hot_Pixel_Constant, ts_dem_cold):
    """
    Function to calculates the hot pixels based on the surface temperature and NDVI
    """
    for_hot = np.copy(ts_dem)
    for_hot[NDVI <= NDVIhot_low] = 0.0
    for_hot[NDVI >= NDVIhot_high] = 0.0
    for_hot[np.logical_or(water_mask != 0.0, QC_Map != 0.0)] = 0.0
    hot_pixels = np.copy(for_hot)
    hot_pixels[for_hot < ts_dem_cold] = np.nan
    ts_dem_hot_max = np.nanmax(hot_pixels)    # Max
    ts_dem_hot_mean = np.nanmean(hot_pixels)  # Mean
    ts_dem_hot_std = np.nanstd(hot_pixels)    # Standard deviation
    #ts_dem_hot = ts_dem_hot_max - 0.25 * ts_dem_hot_std
    #ts_dem_hot = (ts_dem_hot_max + ts_dem_hot_mean)/2


    ts_dem_hot=ts_dem_hot_mean + Hot_Pixel_Constant * ts_dem_hot_std


    print 'hot : max= %0.3f (Kelvin)' % ts_dem_hot_max, ', sd= %0.3f (Kelvin)' % ts_dem_hot_std, \
           ', mean= %0.3f (Kelvin)' % ts_dem_hot_mean, ', value= %0.3f (Kelvin)' % ts_dem_hot
    return(ts_dem_hot,hot_pixels)

#------------------------------------------------------------------------------
def Calc_Cold_Pixels(ts_dem,water_mask,QC_Map,ts_dem_cold_veg,Cold_Pixel_Constant):
    """
    Function to calculates the the cold pixels based on the surface temperature
    """
    for_cold = np.copy(ts_dem)
    for_cold[water_mask != 1.0] = 0.0
    for_cold[QC_Map != 0] = 0.0
    cold_pixels = np.copy(for_cold)
    cold_pixels[for_cold < 278.0] = np.nan
    cold_pixels[for_cold > 320.0] = np.nan
    # cold_pixels[for_cold < 285.0] = 285.0
    ts_dem_cold_std = np.nanstd(cold_pixels)     # Standard deviation
    ts_dem_cold_min = np.nanmin(cold_pixels)     # Min
    ts_dem_cold_mean = np.nanmean(cold_pixels)   # Mean

    # If average temperature is below zero or nan than use the vegetation cold pixel
    if ts_dem_cold_mean <= 0.0:
        ts_dem_cold = ts_dem_cold_veg + Cold_Pixel_Constant * ts_dem_cold_std
    if np.isnan(ts_dem_cold_mean) == True:
        ts_dem_cold = ts_dem_cold_veg + Cold_Pixel_Constant * ts_dem_cold_std
    else:
        ts_dem_cold = ts_dem_cold_mean + Cold_Pixel_Constant * ts_dem_cold_std

    if ts_dem_cold > ts_dem_cold_veg:
        ts_dem_cold = ts_dem_cold_veg
    if np.isnan(ts_dem_cold):
        ts_dem_cold = ts_dem_cold_veg

    print 'cold water: min=%0.3f (Kelvin)' %ts_dem_cold_min , ', sd= %0.3f (Kelvin)' % ts_dem_cold_std, \
           ', mean= %0.3f (Kelvin)' % ts_dem_cold_mean, ', value= %0.3f (Kelvin)' % ts_dem_cold
    return(ts_dem_cold,cold_pixels,ts_dem_cold_mean)

#------------------------------------------------------------------------------
def Calc_Cold_Pixels_Veg(NDVI,NDVI_max,NDVI_std,QC_Map,ts_dem,Image_Type, Cold_Pixel_Constant):
    """
    Function to calculates the the cold pixels based on vegetation
    """
    cold_pixels_vegetation = np.copy(ts_dem)
    cold_pixels_vegetation[np.logical_or(NDVI <= (NDVI_max-0.1*NDVI_std),QC_Map != 0.0)] = 0.0
    cold_pixels_vegetation[cold_pixels_vegetation==0.0] = np.nan
    ts_dem_cold_std_veg = np.nanstd(cold_pixels_vegetation)
    ts_dem_cold_min_veg = np.nanmin(cold_pixels_vegetation)
    ts_dem_cold_mean_veg = np.nanmean(cold_pixels_vegetation)

    if Image_Type == 1:
            ts_dem_cold_veg = ts_dem_cold_mean_veg + Cold_Pixel_Constant * ts_dem_cold_std_veg
    if Image_Type == 2:
            ts_dem_cold_veg = ts_dem_cold_mean_veg + Cold_Pixel_Constant * ts_dem_cold_std_veg
    if Image_Type == 3:
            ts_dem_cold_veg = ts_dem_cold_mean_veg + Cold_Pixel_Constant * ts_dem_cold_std_veg

    print 'cold vegetation: min=%0.3f (Kelvin)' %ts_dem_cold_min_veg , ', sd= %0.3f (Kelvin)' % ts_dem_cold_std_veg, \
				', mean= %0.3f (Kelvin)' % ts_dem_cold_mean_veg, ', value= %0.3f (Kelvin)' % ts_dem_cold_veg
    return(ts_dem_cold_veg)

#------------------------------------------------------------------------------
def Calc_Meteo(Rs_24,eact_24,Temp_24,Surf_albedo,dr,tir_emis,Surface_temp,water_mask,NDVI,Transm_24,SB_const,lw_in_inst,Rs_inst):
    """
    Calculates the instantaneous Ground heat flux and solar radiation.
    """

    # Net shortwave radiation (W/m2):
    Rns_24 = Rs_24 * (1 - Surf_albedo)


    # Net outgoing longwave radiation (W/m2):
    Rnl_24_FAO = (SB_const * np.power(Temp_24 + 273.15, 4) * (0.34-0.14 *
                  np.power(eact_24, 0.5)) * (1.35 * Transm_24 / 0.8 - 0.35))

    Rnl_24_Slob = 110 * Transm_24

    print 'Mean Daily Net longwave Radiation (Slob) = %0.3f (W/m2)' % np.nanmean(Rnl_24_Slob)
    print 'Mean Daily Net longwave Radiation (FAO) = %0.3f (W/m2)' % np.nanmean(Rnl_24_FAO)

    # Net 24 hrs radiation (W/m2):
    Rn_24_FAO = Rns_24 - Rnl_24_FAO          # FAO equation
    Rn_24_Slob = Rns_24 - Rnl_24_Slob       # Slob equation
    Rn_24 = (Rn_24_FAO + Rn_24_Slob) / 2  # Average

    print 'Mean Daily Net Radiation (Slob) = %0.3f (W/m2)' % np.nanmean(Rn_24_Slob)
    print 'Mean Daily Net Radiation (FAO) = %0.3f (W/m2)' % np.nanmean(Rn_24_FAO)

    # Instantaneous outgoing longwave radiation:
    lw_out_inst = tir_emis * SB_const * np.power(Surface_temp, 4)

    # Instantaneous net radiation
    rn_inst = (Rs_inst * (1 - Surf_albedo) + lw_in_inst - lw_out_inst -
               (1 - tir_emis) * lw_in_inst)
    # Instantaneous Soil heat flux
    g_inst = np.where(water_mask != 0.0, 0.4 * rn_inst,
                      ((Surface_temp - 273.15) * (0.0038 + 0.0074 * Surf_albedo) *
                       (1 - 0.978 * np.power(NDVI, 4))) * rn_inst)
    return(Rn_24,rn_inst,g_inst,Rnl_24_FAO)

#------------------------------------------------------------------------------
def Calc_surface_water_temp(Temp_inst,Landsat_nr,Lmax,Lmin,therm_data,b10_emissivity,k1_c,k2_c,eact,shape_lsc,water_mask_temp,Bands_thermal,Rp,tau_sky,surf_temp_offset,Image_Type):
    """
    Calculates the surface temperature and create a water mask
    """

    # Spectral radiance for termal
    if Landsat_nr == 8:
        if Bands_thermal == 1:
            k1 = k1_c[0]
            k2 = k2_c[0]
            L_lambda_b10 = (Lmax[-1] - Lmin[-1]) / (65535-1) * therm_data[:, :, 0] + Lmin[-1]

            # Get Temperature
            Temp_TOA = Get_Thermal(L_lambda_b10,Rp,Temp_inst,tau_sky,b10_emissivity,k1,k2)

        elif Bands_thermal == 2:
            L_lambda_b10 = (Lmax[-2] - Lmin[-2]) / (65535-1) * therm_data[:, :, 0] + Lmin[-2]
            L_lambda_b11 = (Lmax[-1] - Lmin[-1]) / (65535-1) * therm_data[:, :, 1] + Lmin[-1]

            # Brightness temperature
            # From Band 10:
            Temp_TOA_10 = (k2_c[0] / np.log(k1_c[0] / L_lambda_b10 + 1.0))
            # From Band 11:
            Temp_TOA_11 = (k2_c[1] / np.log(k1_c[1] / L_lambda_b11 + 1.0))
            # Combined:
            Temp_TOA = (Temp_TOA_10 + 1.378 * (Temp_TOA_10 - Temp_TOA_11) +
                           0.183 * np.power(Temp_TOA_10 - Temp_TOA_11, 2) - 0.268 +
                           (54.30 - 2.238 * eact) * (1 - b10_emissivity))

    elif Landsat_nr == 7:
        k1=666.09
        k2=1282.71
        L_lambda_b6 = (Lmax[-1] - Lmin[-1]) / (256-1) * therm_data[:, :, 0] + Lmin[-1]

        # Brightness temperature - From Band 6:
        Temp_TOA = Get_Thermal(L_lambda_b6,Rp,Temp_inst,tau_sky,b10_emissivity,k1,k2)

    elif Landsat_nr == 5:
        k1=607.76
        k2=1260.56
        L_lambda_b6 = ((Lmax[-1] - Lmin[-1]) / (256-1) * therm_data[:, :, 0] +
                       Lmin[-1])

       # Brightness temperature - From Band 6:
        Temp_TOA = Get_Thermal(L_lambda_b6,Rp,Temp_inst,tau_sky,b10_emissivity,k1,k2)

    # Surface temperature
    Surface_temp = Temp_TOA
    Surface_temp = Surface_temp.clip(230.0, 360.0)

    # Cloud mask:
    temp_water = np.zeros((shape_lsc[1], shape_lsc[0]))
    temp_water = np.copy(Surface_temp)
    temp_water[water_mask_temp == 0.0] = np.nan
    temp_water_sd = np.nanstd(temp_water)     # Standard deviation
    temp_water_mean = np.nanmean(temp_water)  # Mean
    print 'Mean water temperature = ', '%0.3f (Kelvin)' % temp_water_mean
    print 'SD water temperature = ', '%0.3f (Kelvin)' % temp_water_sd
    cloud_mask = np.zeros((shape_lsc[1], shape_lsc[0]))
    cloud_mask[Surface_temp < np.minimum((temp_water_mean - 1.0 * temp_water_sd -
               surf_temp_offset),290)] = 1.0

    return(Surface_temp, cloud_mask)


#------------------------------------------------------------------------------
def Get_Thermal(lambda_b10,Rp,Temp_inst,tau_sky,TIR_Emissivity,k1,k2):

    # Narrow band downward thermal radiation from clear sky, rsky (W/m2/sr/µm)
    rsky = (1.807E-10 * np.power(Temp_inst + 273.15, 4) * (1 - 0.26 *
            np.exp(-7.77E-4 * np.power((-Temp_inst), -2))))
    print 'Rsky = ', '%0.3f (W/m2/sr/µm)' % np.nanmean(rsky)

    # Corrected thermal radiance from the surface, Wukelikc et al. (1989):
    correc_lambda_b10 = ((lambda_b10 - Rp) / tau_sky -
                               (1.0 - TIR_Emissivity) * rsky)
    # Brightness temperature - From Band 10:
    Temp_TOA = (k2 / np.log(TIR_Emissivity * k1 /
                       correc_lambda_b10 + 1.0))

    return(Temp_TOA)
#------------------------------------------------------------------------------
def Calc_vegt_para(NDVI,water_mask_temp,shape_lsc):
    """
    Calculates the Fraction of PAR, Thermal infrared emissivity, Nitrogen, Vegetation Cover, LAI, b10_emissivity
    """
    # Fraction of PAR absorbed by the vegetation canopy (FPAR):
    FPAR = -0.161 + 1.257 * NDVI
    FPAR[NDVI < 0.125] = 0.0

    # Termal infrared emissivity
    tir_emis = 1.009 + 0.047 * np.log(NDVI)
    tir_emis[np.logical_or(water_mask_temp == 1.0, water_mask_temp == 2.0)] = 1.0
    tir_emis[np.logical_and(NDVI < 0.125, water_mask_temp == 0.0)] = 0.92

    # Vegetation Index - Regression model from Bagheri et al. (2013)
    VI = 38.764 * np.square(NDVI) - 24.605 * NDVI + 5.8103

    # Nitrogen computation
    Nitrogen = np.copy(VI)
    Nitrogen[VI <= 0.0] = 0.0
    Nitrogen[NDVI <= 0.0] = 0.0

    # Vegetation cover:
    vegt_cover = 1 - np.power((0.8 - NDVI)/(0.8 - 0.125), 0.7)
    vegt_cover[NDVI < 0.125] = 0.0
    vegt_cover[NDVI > 0.8] = 0.99

    # Leaf Area Index (LAI)
    LAI_1 = np.log(-(vegt_cover - 1)) / -0.45
    LAI_1[LAI_1 > 8] = 8.0
    LAI_2 = (9.519 * np.power(NDVI, 3) + 0.104 * np.power(NDVI, 2) +
             1.236 * NDVI - 0.257)

    LAI = (LAI_1 + LAI_2) / 2.0  # Average LAI
    LAI[LAI < 0.001] = 0.001

    b10_emissivity = np.zeros((shape_lsc[1], shape_lsc[0]))
    b10_emissivity = np.where(LAI <= 3.0, 0.95 + 0.01 * LAI, 0.98)
    b10_emissivity[water_mask_temp != 0.0] = 1.0

    return(FPAR,tir_emis,Nitrogen,vegt_cover,LAI,b10_emissivity)

#------------------------------------------------------------------------------
def Water_Mask(shape_lsc,Reflect):
    """
    Calculates the water and cloud mask
    """
    mask = np.zeros((shape_lsc[1], shape_lsc[0]))
    mask[np.logical_and(Reflect[:, :, 3] < Reflect[:, :, 2],
                        Reflect[:, :, 4] < Reflect[:, :, 1])] = 1.0
    water_mask_temp = np.copy(mask)

    return(water_mask_temp)

#------------------------------------------------------------------------------
def Calc_albedo(Reflect,path_radiance,Apparent_atmosf_transm):
    """
    This function calculates and returns the Surface albedo, NDVI by using the refectance from the landsat image.
    """
    # Surface albedo:
    Surf_albedo = (0.254 * Reflect[:, :, 0] + 0.149 * Reflect[:, :, 1] +
                   0.147 * Reflect[:, :, 2] + 0.311 * Reflect[:, :, 3] +
                   0.103 * Reflect[:, :, 4] + 0.036 * Reflect[:, :, 5] -
                   path_radiance) / np.power(Apparent_atmosf_transm, 2)

    # Better tsw instead of Apparent_atmosf_transm ??
    Surf_albedo = Surf_albedo.clip(0.0, 0.6)

    return(Surf_albedo)
#------------------------------------------------------------------------------
def Calc_NDVI(Reflect):
    """
    This function calculates and returns the Surface albedo, NDVI by using the refectance from the landsat image.
    """
    # Computation of Normalized Difference Vegetation Index (NDVI)
    NDVI = ((Reflect[:, :, 3] - Reflect[:, :, 2]) /
            (Reflect[:, :, 3] + Reflect[:, :, 2]))

    return(NDVI)

#------------------------------------------------------------------------------
def CalculateSnowWaterMask(NDVI,shape_lsc,water_mask_temp,Surface_temp):
   '''
   Devides the temporaly water mask into a snow and water mask by using the surface temperature
   '''
   NDVI_nan=np.copy(NDVI)
   NDVI_nan[NDVI==0]=np.nan
   NDVI_nan=np.float32(NDVI_nan)
   NDVI_std=np.nanstd(NDVI_nan)
   NDVI_max=np.nanmax(NDVI_nan)
   NDVI_treshold_cold_pixels=NDVI_max-0.1*NDVI_std
   print 'NDVI treshold for cold pixels = ', '%0.3f' % NDVI_treshold_cold_pixels
   ts_moist_veg_min=np.nanmin(Surface_temp[NDVI>NDVI_treshold_cold_pixels])

   # calculate new water mask
   mask=np.zeros((shape_lsc[1], shape_lsc[0]))
   mask[np.logical_and(np.logical_and(water_mask_temp==1, Surface_temp <= 275),NDVI>=0.3)]=1
   snow_mask=np.copy(mask)

   # calculate new water mask
   mask=np.zeros((shape_lsc[1], shape_lsc[0]))
   mask[np.logical_and(water_mask_temp==1, Surface_temp > 273)]=1
   water_mask=np.copy(mask)

   return(snow_mask,water_mask,ts_moist_veg_min, NDVI_max, NDVI_std)

#------------------------------------------------------------------------------
def Calc_Ra_Mountain(lon,DOY,hour,minutes,lon_proy,lat_proy,slope,aspect):
    """
    Calculates the extraterrestiral solar radiation by using the date, slope and aspect.
    """

    # Constants
    deg2rad = np.pi / 180.0  # Factor to transform from degree to rad
    Min_cos_zn = 0.1  # Min value for cos zenith angle
    Max_cos_zn = 1.0  # Max value for cos zenith angle
    Gsc = 1367        # Solar constant (W / m2)
    try:
        Loc_time = float(hour) + float(minutes)/60  # Local time (hours)
    except:
        Loc_time = np.float_(hour) + np.float_(minutes)/60  # Local time (hours)
    # Rounded difference of the local time from Greenwich (GMT) (hours):
    offset_GTM = round(np.sign(lon[int(lon.shape[0])/2, int(lon.shape[1])/2]) * lon[int(lon.shape[0])/2,int(lon.shape[1])/2] * 24 / 360)

    print '  Local time: ', '%0.3f' % np.nanmean(Loc_time)
    print '  Difference of local time (LT) from Greenwich (GMT): ', offset_GTM

    # 1. Calculation of extraterrestrial solar radiation for slope and aspect
    # Computation of Hour Angle (HRA = w)
    B = 360./365 * (DOY-81)           # (degrees)
    # Computation of cos(theta), where theta is the solar incidence angle
    # relative to the normal to the land surface
    delta=np.arcsin(np.sin(23.45*deg2rad)*np.sin(np.deg2rad(B))) # Declination angle (radians)
    phi = lat_proy * deg2rad                                     # latitude of the pixel (radians)
    s = slope * deg2rad                                          # Surface slope (radians)
    gamma = (aspect-180) * deg2rad                               # Surface aspect angle (radians)
    w=w_time(Loc_time, lon_proy, DOY)                            # Hour angle (radians)
    a,b,c = Constants(delta,s,gamma,phi)
    cos_zn= AngleSlope(a,b,c,w)
    cos_zn = cos_zn.clip(Min_cos_zn, Max_cos_zn)

    print 'Average Cos Zenith Angle: ', '%0.3f (Radians)' % np.nanmean(cos_zn)

    dr = 1 + 0.033 * cos(DOY*2*pi/365)  # Inverse relative distance Earth-Sun
    # Instant. extraterrestrial solar radiation (W/m2), Allen et al.(2006):
    Ra_inst = Gsc * cos_zn * dr

    # 24-hours extraterrestrial radiation
    # 1.) determine if there are one or two periods of sun
    # 2.) calculate the 24-hours extraterrestrial radiation if there are two periods of sun
    # 3.) calculate the 24-hours extraterrestrial radiation if there is one period of sun

    #1.) determine amount of sun periods
    Ra_24 = np.zeros(np.shape(lat_proy))*np.nan
    constant=Gsc*dr/(2*np.pi)
    TwoPeriod= TwoPeriods(delta,s,phi)  # all input in radians

    #2.) calculate the 24-hours extraterrestrial radiation (2 periods)
    ID = np.where(np.ravel(TwoPeriod==True))
    Ra_24.flat[ID]=TwoPeriodSun(constant,delta,s.flat[ID],gamma.flat[ID],phi.flat[ID])

    #3.) calculate the 24-hours extraterrestrial radiation (1 period)
    ID = np.where(np.ravel(TwoPeriod==False))
    Ra_24.flat[ID]=OnePeriodSun(constant,delta,s.flat[ID],gamma.flat[ID],phi.flat[ID])

    # Horizontal surface
    ws = np.arccos(-np.tan(delta) * np.tan(phi))  # Sunrise/sunset time angle

    # Extraterrestial radiation for a horizontal surface for 24-h period:
    Ra_hor_24 = (Gsc * dr / np.pi * (np.sin(delta) * np.sin(phi) * ws + np.cos(delta) * np.cos(phi) * np.sin(ws)))
    # cos_theta_flat = (np.sin(delta) * np.sin(phi) + np.cos(delta) * np.cos(phi) * np.cos(w))

    # Mountain radiation
    Ra_mountain_24 = np.where(Ra_24 > Min_cos_zn * Ra_hor_24, Ra_24 / np.cos(s),
                           Ra_hor_24)
    Ra_mountain_24[Ra_mountain_24 > 600.0] = 600.0

    return(Ra_mountain_24,Ra_inst,cos_zn,dr,phi,delta)

#------------------------------------------------------------------------------
def OnePeriodSun(constant,delta,s,gamma,phi):
    '''
    Based on Richard G. Allen 2006
    Calculate the 24-hours extraterrestrial radiation when there is one sun period
    '''
    sunrise,sunset = SunHours(delta,s,gamma,phi)
    Vals=IntegrateSlope(constant,sunrise,sunset,delta,s,gamma,phi)

    return(Vals)

#------------------------------------------------------------------------------
def TwoPeriodSun(constant,delta,s,gamma,phi):
    '''
    Based on Richard G. Allen 2006
    Calculate the 24-hours extraterrestrial radiation when there are two sun period
    '''
    A1, A2 = SunHours(delta,s,gamma,phi)
    a,b,c = Constants(delta,s,gamma,phi)
    riseSlope, setSlope = BoundsSlope(a,b,c)
    B1 = np.maximum(riseSlope,setSlope)
    B2 = np.minimum(riseSlope,setSlope)
    Angle_B1 = AngleSlope(a,b,c,B1)
    Angle_B2 = AngleSlope(a,b,c,B2)

    B1[abs(Angle_B1) > 0.001] = np.pi - B1[abs(Angle_B1) > 0.001]
    B2[abs(Angle_B2) > 0.001] = -np.pi - B2[abs(Angle_B2) > 0.001]

    # Check if two periods really exist
    ID = np.ravel_multi_index(np.where(np.logical_and(B2 >= A1, B1 >= A2) == True),a.shape)
    Val = IntegrateSlope(constant,B2.flat[ID],B1.flat[ID],delta,s.flat[ID],gamma.flat[ID],phi.flat[ID])
    ID = ID[Val < 0]

    # Finally calculate resulting values
    Vals = np.zeros(B1.shape)

    Vals.flat[ID] = (IntegrateSlope(constant,A1.flat[ID],B2.flat[ID],delta,s.flat[ID],gamma.flat[ID],phi.flat[ID])  +
                   IntegrateSlope(constant,B1.flat[ID],A2.flat[ID],delta,s.flat[ID],gamma.flat[ID],phi.flat[ID]))
    ID = np.ravel_multi_index(np.where(Vals == 0),a.shape)
    Vals.flat[ID] = IntegrateSlope(constant,A1.flat[ID],A2.flat[ID],delta,s.flat[ID],gamma.flat[ID],phi.flat[ID])

    return(Vals)

#------------------------------------------------------------------------------
def IntegrateSlope(constant,sunrise,sunset,delta,s,gamma,phi):
    '''
    Based on Richard G. Allen 2006 equation 5
    Calculate the 24 hours extraterrestrial radiation
    '''
    # correct the sunset and sunrise angels for days that have no sunset or no sunrise
    SunOrNoSun = np.logical_or(((np.abs(delta + phi)) > (np.pi/2)),((np.abs(delta - phi)) > (np.pi/2)))
    integral=np.zeros(s.shape)
    ID = np.where(np.ravel(SunOrNoSun==True))

    # No sunset
    if abs(delta+phi.flat[ID])>(np.pi/2):
        sunset1=np.pi
        sunrise1=-np.pi
        integral.flat[ID] = constant * (np.sin(delta)*np.sin(phi)*np.cos(s)*(sunset1-sunrise1)
            - np.sin(delta)*np.cos(phi)*np.sin(s)*np.cos(gamma)*(sunset1-sunrise1)
            + np.cos(delta)*np.cos(phi)*np.cos(s)*(np.sin(sunset1)-np.sin(sunrise1))
            + np.cos(delta)*np.sin(phi)*np.sin(s)*np.cos(gamma)*(np.sin(sunset1)-np.sin(sunrise1))
            - np.cos(delta)*np.sin(s)*np.sin(gamma)*(np.cos(sunset1)-np.cos(sunrise1)))

    # No sunrise
    elif np.abs(delta-phi.flat[ID])>(np.pi/2):
        integral.flat[ID]=constant * (np.sin(delta)*np.sin(phi)*np.cos(s)*(0)
            - np.sin(delta)*np.cos(phi)*np.sin(s)*np.cos(gamma)*(0)
            + np.cos(delta)*np.cos(phi)*np.cos(s)*(np.sin(0)-np.sin(0))
            + np.cos(delta)*np.sin(phi)*np.sin(s)*np.cos(gamma)*(np.sin(0)-np.sin(0))
            - np.cos(delta)*np.sin(s)*np.sin(gamma)*(np.cos(0)-np.cos(0)))

    ID = np.where(np.ravel(SunOrNoSun==False))
    integral.flat[ID] = constant * (np.sin(delta)*np.sin(phi)*np.cos(s)*(sunset-sunrise)
            - np.sin(delta)*np.cos(phi)*np.sin(s)*np.cos(gamma)*(sunset-sunrise)
            + np.cos(delta)*np.cos(phi)*np.cos(s)*(np.sin(sunset)-np.sin(sunrise))
            + np.cos(delta)*np.sin(phi)*np.sin(s)*np.cos(gamma)*(np.sin(sunset)-np.sin(sunrise))
            - np.cos(delta)*np.sin(s)*np.sin(gamma)*(np.cos(sunset)-np.cos(sunrise)))

    return(integral)

#------------------------------------------------------------------------------
def TwoPeriods(delta,s,phi):
    '''
    Based on Richard G. Allen 2006
    Create a boolean map with True values for places with two sunsets
    '''
    TwoPeriods = (np.sin(s) > np.ones(s.shape)*np.sin(phi)*np.sin(delta)+np.cos(phi)*np.cos(delta))

    return(TwoPeriods)

#------------------------------------------------------------------------------
def SunHours(delta,slope,slopedir,lat):
    # Define sun hours in case of one sunlight period

    a,b,c = Constants(delta,slope,slopedir,lat)
    riseSlope, setSlope = BoundsSlope(a,b,c)
    bound = BoundsHorizontal(delta,lat)

    Calculated = np.zeros(slope.shape, dtype = bool)
    RiseFinal = np.zeros(slope.shape)
    SetFinal = np.zeros(slope.shape)

    # First check sunrise is not nan
    # This means that their is either no sunrise (whole day night) or no sunset (whole day light)
    # For whole day light, use the horizontal sunrise and whole day night a zero..
    Angle4 = AngleSlope(a,b,c,-bound)
    RiseFinal[np.logical_and(np.isnan(riseSlope),Angle4 >= 0)] = -bound[np.logical_and(np.isnan(riseSlope),Angle4 >= 0)]
    Calculated[np.isnan(riseSlope)] = True

    # Step 1 > 4
    Angle1 = AngleSlope(a,b,c,riseSlope)
    Angle2 = AngleSlope(a,b,c,-bound)

    ID = np.ravel_multi_index(np.where(np.logical_and(np.logical_and(Angle2 < Angle1+0.001 ,Angle1 < 0.001),Calculated == False) == True),a.shape)
    RiseFinal.flat[ID] = riseSlope.flat[ID]
    Calculated.flat[ID] = True
    # step 5 > 7
    Angle3 = AngleSlope(a,b,c,-np.pi - riseSlope)

    ID = np.ravel_multi_index(np.where(np.logical_and(np.logical_and(-bound<(-np.pi-riseSlope),Angle3 <= 0.001),Calculated == False) == True),a.shape)
    RiseFinal.flat[ID] = -np.pi -riseSlope.flat[ID]
    Calculated.flat[ID] = True

    # For all other values we use the horizontal sunset if it is positive, otherwise keep a zero
    RiseFinal[Calculated == False] = -bound[Calculated == False]

    # Then check sunset is not nan or < 0
    Calculated = np.zeros(slope.shape, dtype = bool)

    Angle4 = AngleSlope(a,b,c,bound)
    SetFinal[np.logical_and(np.isnan(setSlope),Angle4 >= 0)] = bound[np.logical_and(np.isnan(setSlope),Angle4 >= 0)]
    Calculated[np.isnan(setSlope)] = True

    # Step 1 > 4
    Angle1 = AngleSlope(a,b,c,setSlope)
    Angle2 = AngleSlope(a,b,c,bound)

    ID = np.ravel_multi_index(np.where(np.logical_and(np.logical_and(Angle2 < Angle1+0.001,Angle1 < 0.001),Calculated == False) == True),a.shape)
    SetFinal.flat[ID] = setSlope.flat[ID]
    Calculated.flat[ID] = True
    # step 5 > 7
    Angle3 = AngleSlope(a,b,c,np.pi - setSlope)

    ID = np.ravel_multi_index(np.where(np.logical_and(np.logical_and(bound>(np.pi-setSlope),Angle3 <= 0.001),Calculated == False) == True),a.shape)
    SetFinal.flat[ID] = np.pi - setSlope.flat[ID]
    Calculated.flat[ID] = True

    # For all other values we use the horizontal sunset if it is positive, otherwise keep a zero
    SetFinal[Calculated == False] = bound[Calculated == False]

    #    Angle4 = AngleSlope(a,b,c,bound)
    #    SetFinal[np.logical_and(Calculated == False,Angle4 >= 0)] = bound[np.logical_and(Calculated == False,Angle4 >= 0)]

    # If Sunrise is after Sunset there is no sunlight during the day
    SetFinal[SetFinal <= RiseFinal] = 0
    RiseFinal[SetFinal <= RiseFinal] = 0

    return(RiseFinal,SetFinal)

#------------------------------------------------------------------------------
def Constants(delta,s,gamma,phi):
    '''
    Based on Richard G. Allen 2006 equation 11
    determines constants for calculating the exterrestial solar radiation
    '''
    a = np.sin(delta)*np.cos(phi)*np.sin(s)*np.cos(gamma) - np.sin(delta)*np.sin(phi)*np.cos(s)
    b = np.cos(delta)*np.cos(phi)*np.cos(s) + np.cos(delta)*np.sin(phi)*np.sin(s)*np.cos(gamma)
    c = np.cos(delta)*np.sin(s)*np.sin(gamma)

    return(a,b,c)

#------------------------------------------------------------------------------
def BoundsSlope(a,b,c):
    '''
    Based on Richard G. Allen 2006 equation 13
    This function calculates candidate values for sunrise and sunset hour angles
    '''
    Div = (b**2+c**2)
    Div[Div <= 0] = 0.00001
    sinB = (a*c + b*np.sqrt(b**2+c**2-a**2)) / Div
    sinA = (a*c - b*np.sqrt(b**2+c**2-a**2)) / Div

    sinB[sinB < -1] = -1; sinB[sinB > 1] = 1    # Limits see appendix A.2.i
    sinA[sinA < -1] = -1; sinA[sinA > 1] = 1    # Limits see appendix A.2.i

    sunrise = np.arcsin(sinA)
    sunset = np.arcsin(sinB)

    return(sunrise,sunset)

#------------------------------------------------------------------------------
def BoundsHorizontal(delta,phi):
    ''''
    Based on Richard G. Allen 2006
    This function calculates sunrise hours based on earth inclination and latitude
    If there is no sunset or sunrise hours the values are either set to 0 (polar night) or pi (polar day)
    '''
    bound = np.arccos(-np.tan(delta)*np.tan(phi))
    bound[abs(delta+phi) > np.pi/2] = np.pi
    bound[abs(delta-phi) > np.pi/2] = 0

    return(bound)

#------------------------------------------------------------------------------
def AngleSlope(a,b,c,w):
    '''
    Based on Richard G. Allen 2006
    Calculate the cos zenith angle by using the hour angle and constants
    '''
    angle = -a + b*np.cos(w) + c*np.sin(w)

    return(angle)

#------------------------------------------------------------------------------
def Calc_Gradient(dataset,pixel_spacing):
    """
    This function calculates the slope and aspect of a DEM map.
    """
    # constants
    deg2rad = np.pi / 180.0  # Factor to transform from degree to rad
    rad2deg = 180.0 / np.pi  # Factor to transform from rad to degree

    # Calculate slope
    x, y = np.gradient(dataset)
    slope = np.arctan(np.sqrt(np.square(x/pixel_spacing) + np.square(y/pixel_spacing))) * rad2deg

    # calculate aspect
    aspect = np.arctan2(y/pixel_spacing, -x/pixel_spacing) * rad2deg
    aspect = 180 + aspect

    return(deg2rad,rad2deg,slope,aspect)

#------------------------------------------------------------------------------
def DEM_lat_lon(DEM_fileName,output_folder):
    """
    This function retrieves information about the latitude and longitude of the
    DEM map.

    """
    # name for output
    lat_fileName = os.path.join(output_folder, 'Output_radiation_balance','latitude.tif')
    lon_fileName = os.path.join(output_folder, 'Output_radiation_balance','longitude.tif')

    g = gdal.Open(DEM_fileName)     # Open DEM
    geo_t = g.GetGeoTransform()     # Get the Geotransform vector:
    x_size = g.RasterXSize          # Raster xsize - Columns
    y_size = g.RasterYSize          # Raster ysize - Rows

    # create a longitude and a latitude array
    lon = np.zeros((y_size, x_size))
    lat = np.zeros((y_size, x_size))
    for col in np.arange(x_size):
        lon[:, col] = geo_t[0] + col * geo_t[1] + geo_t[1]/2
        # ULx + col*(E-W pixel spacing) + E-W pixel spacing
    for row in np.arange(y_size):
        lat[row, :] = geo_t[3] + row * geo_t[5] + geo_t[5]/2
        # ULy + row*(N-S pixel spacing) + N-S pixel spacing,
        # negative as we will be counting from the UL corner

    # Define shape of the raster
    shape = [x_size, y_size]

    # Save lat and lon files in geo- coordinates
    save_GeoTiff_proy(g, lat, lat_fileName, shape, nband=1)
    save_GeoTiff_proy(g, lon, lon_fileName, shape, nband=1)

    return(lat,lon,lat_fileName,lon_fileName)

#------------------------------------------------------------------------------
def reproject_dataset(dataset, pixel_spacing, UTM_Zone):
    """
    A sample function to reproject and resample a GDAL dataset from within
    Python. The idea here is to reproject from one system to another, as well
    as to change the pixel size. The procedure is slightly long-winded, but
    goes like this:

    1. Set up the two Spatial Reference systems.
    2. Open the original dataset, and get the geotransform
    3. Calculate bounds of new geotransform by projecting the UL corners
    4. Calculate the number of pixels with the new projection & spacing
    5. Create an in-memory raster dataset
    6. Perform the projection
    """

    # 1) Open the dataset
    g = gdal.Open(dataset)
    if g is None:
        print 'input folder does not exist'

     # Define the EPSG code...
    EPSG_code = '326%02d' % UTM_Zone
    epsg_to = int(EPSG_code)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    try:
        proj = g.GetProjection()
        Proj_in=proj.split('EPSG","')
        epsg_from=int((str(Proj_in[-1]).split(']')[0])[0:-1])
    except:
        epsg_from = int(4326)    # Get the Geotransform vector:
    geo_t = g.GetGeoTransform()

    # Vector components:
    # 0- The Upper Left easting coordinate (i.e., horizontal)
    # 1- The E-W pixel spacing
    # 2- The rotation (0 degrees if image is "North Up")
    # 3- The Upper left northing coordinate (i.e., vertical)
    # 4- The rotation (0 degrees)
    # 5- The N-S pixel spacing, negative as it is counted from the UL corner
    x_size = g.RasterXSize  # Raster xsize
    y_size = g.RasterYSize  # Raster ysize

    epsg_to = int(epsg_to)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    osng = osr.SpatialReference()
    osng.ImportFromEPSG(epsg_to)
    wgs84 = osr.SpatialReference()
    wgs84.ImportFromEPSG(epsg_from)

    inProj = Proj(init='epsg:%d' %epsg_from)
    outProj = Proj(init='epsg:%d' %epsg_to)


    nrow_skip = round((0.06*y_size)/2)
    ncol_skip = round((0.06*x_size)/2)

    # Up to here, all  the projection have been defined, as well as a
    # transformation from the from to the to
    ulx, uly = transform(inProj,outProj,geo_t[0] + nrow_skip * geo_t[1], geo_t[3] + nrow_skip * geo_t[5])
    lrx, lry = transform(inProj,outProj,geo_t[0] + geo_t[1] * (x_size-ncol_skip),
                                        geo_t[3] + geo_t[5] * (y_size-nrow_skip))

    # See how using 27700 and WGS84 introduces a z-value!
    # Now, we create an in-memory raster
    mem_drv = gdal.GetDriverByName('MEM')

    # The size of the raster is given the new projection and pixel spacing
    # Using the values we calculated above. Also, setting it to store one band
    # and to use Float32 data type.
    col = int((lrx - ulx)/pixel_spacing)
    rows = int((uly - lry)/pixel_spacing)

    # Re-define lr coordinates based on whole number or rows and columns
    (lrx, lry) = (ulx + col * pixel_spacing, uly -
                  rows * pixel_spacing)

    dest = mem_drv.Create('', col, rows, 1, gdal.GDT_Float32)

    if dest is None:
        print 'input folder to large for memory, clip input map'

   # Calculate the new geotransform
    new_geo = (ulx, pixel_spacing, geo_t[2], uly,
               geo_t[4], - pixel_spacing)

    # Set the geotransform
    dest.SetGeoTransform(new_geo)
    dest.SetProjection(osng.ExportToWkt())

    # Perform the projection/resampling
    gdal.ReprojectImage(g, dest, wgs84.ExportToWkt(), osng.ExportToWkt(),gdal.GRA_Bilinear)

    return dest, ulx, lry, lrx, uly, epsg_to
#------------------------------------------------------------------------------
def reproject_dataset_example(dataset, dataset_example, method = 1):

    try:
        if (os.path.splitext(dataset)[-1] == '.tif' or os.path.splitext(dataset)[-1] == '.TIF'):
            g_in = gdal.Open(dataset)
        else:
            g_in = dataset
    except:
            g_in = dataset
    epsg_from = Get_epsg(g_in)

    #exceptions
    if epsg_from == 9001:
        epsg_from = 5070

    # open dataset that is used for transforming the dataset
    try:
        if (os.path.splitext(dataset_example)[-1] == '.tif' or os.path.splitext(dataset_example)[-1] == '.TIF'):
            g_ex = gdal.Open(dataset_example)
        else:
            g_ex = dataset_example

    except:
            g_ex = dataset_example
    epsg_to = Get_epsg(g_ex)

    Y_raster_size = g_ex.RasterYSize
    X_raster_size = g_ex.RasterXSize

    Geo = g_ex.GetGeoTransform()
    ulx = Geo[0]
    uly = Geo[3]
    lrx = ulx + X_raster_size * Geo[1]
    lry = uly + Y_raster_size * Geo[5]

    # Set the EPSG codes
    osng = osr.SpatialReference()
    osng.ImportFromEPSG(epsg_to)
    wgs84 = osr.SpatialReference()
    wgs84.ImportFromEPSG(epsg_from)

    # Create new raster
    mem_drv = gdal.GetDriverByName('MEM')
    dest1 = mem_drv.Create('', X_raster_size, Y_raster_size, 1, gdal.GDT_Float32)
    dest1.SetGeoTransform(Geo)
    dest1.SetProjection(osng.ExportToWkt())

    # Perform the projection/resampling
    if method == 1:
        gdal.ReprojectImage(g_in, dest1, wgs84.ExportToWkt(), osng.ExportToWkt(), gdal.GRA_NearestNeighbour)
    if method == 2:
        gdal.ReprojectImage(g_in, dest1, wgs84.ExportToWkt(), osng.ExportToWkt(), gdal.GRA_Average)
    if method == 3:
        gdal.ReprojectImage(g_in, dest1, wgs84.ExportToWkt(), osng.ExportToWkt(), gdal.GRA_Cubic)

    return(dest1, ulx, lry, lrx, uly, epsg_to)

#------------------------------------------------------------------------------
def save_GeoTiff_proy(src_dataset, dst_dataset_array, dst_fileName, shape_lsc, nband):
    """
    This function saves an array dataset in GeoTiff, using the parameters
    from the source dataset, in projected coordinates

    """
    dst_dataset_array	= np.float_(dst_dataset_array)
    dst_dataset_array[dst_dataset_array<-9999] = np.nan
    geotransform = src_dataset.GetGeoTransform()
    spatialreference = src_dataset.GetProjection()

    # create dataset for output
    fmt = 'GTiff'
    driver = gdal.GetDriverByName(fmt)
    dir_name = os.path.dirname(dst_fileName)

    # If the directory does not exist, make it.
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)
    dst_dataset = driver.Create(dst_fileName, shape_lsc[0], shape_lsc[1], nband,gdal.GDT_Float32)
    dst_dataset.SetGeoTransform(geotransform)
    dst_dataset.SetProjection(spatialreference)
    dst_dataset.GetRasterBand(1).SetNoDataValue(-9999)
    dst_dataset.GetRasterBand(1).WriteArray(dst_dataset_array)
    dst_dataset = None

#------------------------------------------------------------------------------
def w_time(LT,lon_proy, DOY):
    """
    This function computes the hour angle (radians) of an image given the
    local time, longitude, and day of the year.

    """
    nrow, ncol = lon_proy.shape

    # Difference of the local time (LT) from Greenwich Mean Time (GMT) (hours):
    delta_GTM = np.sign(lon_proy[nrow/2, ncol/2]) * lon_proy[nrow/2, ncol/2] * 24 / 360
    if np.isnan(delta_GTM) == True:
         delta_GTM = np.nanmean(lon_proy) * np.nanmean(lon_proy)  * 24 / 360


    # Local Standard Time Meridian (degrees):
    LSTM = 15 * delta_GTM

    # Ecuation of time (EoT, minutes):
    B = 360./365 * (DOY-81)  # (degrees)
    EoT = 9.87*sin(np.deg2rad(2*B))-7.53*cos(np.deg2rad(B))-1.5*sin(np.deg2rad(B))

    # Net Time Correction Factor (minutes) at the center of the image:
    TC = 4 * (lon_proy - LSTM) + EoT     # Difference in time over the longitude
    LST = LT + delta_GTM + TC/60         # Local solar time (hours)
    HRA = 15 * (LST-12)                  # Hour angle HRA (degrees)
    deg2rad = np.pi / 180.0              # Factor to transform from degree to rad
    w = HRA * deg2rad                    # Hour angle HRA (radians)
    return w



#------------------------------------------------------------------------------
def sensible_heat(rah, ustar, rn_inst, g_inst, ts_dem, ts_dem_hot, ts_dem_cold,
                  air_dens, Surf_temp, k_vk, QC_Map, hot_pixels, slope):
    """
    This function computes the instantaneous sensible heat given the
    instantaneous net radiation, ground heat flux, and other parameters.

    """
    # Near surface temperature difference (dT):
    dT_ini = (rn_inst - g_inst) * rah / (air_dens * 1004)
    dT_hot = np.copy(dT_ini)

    #dT_hot_fileName = os.path.join(output_folder, 'Output_cloud_masked','test.tif')
    #save_GeoTiff_proy(dest, dT_hot, dT_hot_fileName,shape, nband=1)

    # dT for hot pixels - hot, (dry) agricultural fields with no green veget.:
    dT_hot[ts_dem <= (ts_dem_hot - 0.5)] = np.nan
    dT_hot[QC_Map == 1] = np.nan
    dT_hot[dT_hot == 0] = np.nan
    if np.all(np.isnan(dT_hot)) == True:
       dT_hot = np.copy(dT_ini)
       ts_dem_hot = np.nanpercentile(hot_pixels, 99.5)
       dT_hot[ts_dem <= (ts_dem_hot - 0.5)] = np.nan
       dT_hot[dT_hot == 0] = np.nan

    dT_hot=np.float32(dT_hot)
    dT_hot[slope > 10]=np.nan

    dT_hot_mean = np.nanmean(dT_hot)

    # Compute slope and offset of linear relationship dT = b + a * Ts
    slope_dt = (dT_hot_mean - 0.0) / (ts_dem_hot - ts_dem_cold)  # EThot = 0.0
    offset_dt = dT_hot_mean - slope_dt * ts_dem_hot

    dT = offset_dt + slope_dt * ts_dem

    # Sensible heat flux:
    h = air_dens * 1004 * dT / rah
    h[QC_Map == 1] = np.nan
    h[h==0]=np.nan
    h[QC_Map != 0] = np.nan

    # Monin-Obukhov length (m):
    L_MO = ((-1.0 * air_dens * 1004 * np.power(ustar, 3) * Surf_temp) /
            (k_vk * 9.81 * h))
    L_MO[L_MO < -1000] = -1000

    # Stability correction for momentum, stable conditions (L_MO >= 0):
    psi_200_stable = -0.05 * 200 / L_MO

    # Stability correction for momentum and heat transport, unstable
    # conditions (L_MO < 0):
    x2 = np.power((1.0 - 16.0 * (2.0/L_MO)), 0.25)  # x at 2m
    x200 = np.power(1.0 - 16.0 * (200/L_MO), 0.25)  # x at 200m
    psi_h = 2 * np.log((1 + np.power(x2, 2))/2)
    psi_m200 = (2 * np.log((1 + x200) / 2) + np.log((1 + np.power(x200, 2)) /
                2) - 2 * np.arctan(x200) + 0.5*np.pi)
    print 'Sensible Heat ', np.nanmean(h)
    print 'dT' , np.nanmean(dT)

    return L_MO, psi_200_stable, psi_h, psi_m200, h, dT, slope_dt, offset_dt

#------------------------------------------------------------------------------

def Reshape_Reproject_Input_data(input_File_Name, output_File_Name, Example_extend_fileName):

   # Reproject the dataset based on the example
   data_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
       input_File_Name, Example_extend_fileName)

   # Get the array information from the new created map
   band_data = data_rep.GetRasterBand(1) # Get the reprojected dem band
   ncol_data = data_rep.RasterXSize
   nrow_data = data_rep.RasterYSize
   shape_data=[ncol_data, nrow_data]

   # Save new dataset
   #stats = band.GetStatistics(0, 1)
   data = band_data.ReadAsArray(0, 0, ncol_data, nrow_data)
   save_GeoTiff_proy(data_rep, data, output_File_Name, shape_data, nband=1)

   return(data)

#------------------------------------------------------------------------------
def Thermal_Sharpening(surface_temp_up, NDVI_up, NDVI, Box, dest_up, output_folder, ndvi_fileName, shape_down, dest_down):

    # Creating arrays to store the coefficients
    CoefA=np.zeros((len(surface_temp_up),len(surface_temp_up[1])))
    CoefB=np.zeros((len(surface_temp_up),len(surface_temp_up[1])))
    CoefC=np.zeros((len(surface_temp_up),len(surface_temp_up[1])))

    # Fit a second polynominal fit to the NDVI and Thermal data and save the coefficients for each pixel
    # NOW USING FOR LOOPS PROBABLY NOT THE FASTEST METHOD
    for i in range(0,len(surface_temp_up)):
        for j in range(0,len(surface_temp_up[1])):
            if np.isnan(np.sum(surface_temp_up[i,j]))==False and np.isnan(np.sum(NDVI_up[i,j]))==False:
                x_data=NDVI_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)][np.logical_and(np.logical_not(np.isnan(NDVI_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)])),np.logical_not(np.isnan(surface_temp_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)])))]
                y_data=surface_temp_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)][np.logical_and(np.logical_not(np.isnan(NDVI_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)])),np.logical_not(np.isnan(surface_temp_up[np.maximum(0,i-(Box-1)/2):np.minimum(len(surface_temp_up),i+(Box-1)/2+1),np.maximum(0,j-(Box-1)/2):np.minimum(len(surface_temp_up[1]),j+(Box-1)/2+1)])))]
                x_data[~np.isnan(x_data)]
                y_data[~np.isnan(y_data)]
                if len(x_data)>6:
                    coefs = poly.polyfit(x_data, y_data, 2)
                    CoefA[i,j] = coefs[2]
                    CoefB[i,j] = coefs[1]
                    CoefC[i,j] = coefs[0]
                else:
                    CoefA[i,j] = np.nan
                    CoefB[i,j] = np.nan
                    CoefC[i,j] = np.nan
            else:
                CoefA[i,j] = np.nan
                CoefB[i,j] = np.nan
                CoefC[i,j] = np.nan

    # Define the shape of the surface temperature with the resolution of 400m
    shape_up=[len(surface_temp_up[1]),len(surface_temp_up)]

    # Save the coefficients
    CoefA_fileName_Optie2 = os.path.join(output_folder, 'Output_temporary','coef_A.tif')
    save_GeoTiff_proy(dest_up,CoefA, CoefA_fileName_Optie2,shape_up, nband=1)

    CoefB_fileName_Optie2 = os.path.join(output_folder, 'Output_temporary','coef_B.tif')
    save_GeoTiff_proy(dest_up,CoefB, CoefB_fileName_Optie2,shape_up, nband=1)

    CoefC_fileName_Optie2 = os.path.join(output_folder, 'Output_temporary','coef_C.tif')
    save_GeoTiff_proy(dest_up,CoefC, CoefC_fileName_Optie2,shape_up, nband=1)

    # Downscale the fitted coefficients
    CoefA_Downscale, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
                                  CoefA_fileName_Optie2, ndvi_fileName)
    CoefA = CoefA_Downscale.GetRasterBand(1).ReadAsArray()

    CoefB_Downscale, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
                                  CoefB_fileName_Optie2, ndvi_fileName)
    CoefB = CoefB_Downscale.GetRasterBand(1).ReadAsArray()

    CoefC_downscale, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
                                CoefC_fileName_Optie2, ndvi_fileName)
    CoefC = CoefC_downscale.GetRasterBand(1).ReadAsArray()

    # Calculate the surface temperature based on the fitted coefficents and NDVI
    temp_surface_sharpened=CoefA*NDVI**2+CoefB*NDVI+CoefC
    temp_surface_sharpened[temp_surface_sharpened < 250] = np.nan
    temp_surface_sharpened[temp_surface_sharpened > 400] = np.nan

    return(temp_surface_sharpened)

#------------------------------------------------------------------------------

def Run_command_window(argument):
    """
    This function runs the argument in the command window without showing cmd window

    Keyword Arguments:
    argument -- string, name of the adf file
    """
    if os.name == 'posix':
        argument = argument.replace(".exe","")
        os.system(argument)

    else:
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW

        process = subprocess.Popen(argument, startupinfo=startupinfo, stderr=subprocess.PIPE, stdout=subprocess.PIPE)
        process.wait()

    return()

#------------------------------------------------------------------------------
def Get_epsg(g, extension = 'tiff'):
    """
    This function reads the projection of a GEOGCS file or tiff file

    Keyword arguments:
    g -- string
        Filename to the file that must be read
    extension -- tiff or GEOGCS
        Define the extension of the dataset (default is tiff)
    """
    try:
        if extension == 'tiff':
            # Get info of the dataset that is used for transforming
            g_proj = g.GetProjection()
            Projection=g_proj.split('EPSG","')
        if extension == 'GEOGCS':
            Projection = g
        epsg_to=int((str(Projection[-1]).split(']')[0])[0:-1])
    except:
       epsg_to=4326
       #print 'Was not able to get the projection, so WGS84 is assumed'
    return(epsg_to)

#------------------------------------------------------------------------------
def Open_constant_or_spatial_map(worksheet, CellID, Output_filename, Example_file):

    # Open data, first try to open as value, otherwise as string (path)
    try:
        Constant_or_Map = float(worksheet['%s' %CellID].value)
        Map_file_name = "Constant value of: " + str(Constant_or_Map)

    # if the data is not a value, than open as a string
    except:
        Map_file_name = '%s' %str(worksheet['%s' %CellID].value)

        try:
            Constant_or_Map = Reshape_Reproject_Input_data(Map_file_name, Output_filename, Example_file)
        except:
            print 'ERROR: One of the INPUTS is NOT CORRECT'

    return(Constant_or_Map, Map_file_name)

#------------------------------------------------------------------------------
def resize_array_example(Array_in, Array_example, method=1):
    """
    This function resizes an array so it has the same size as an example array
    The extend of the array must be the same

    Keyword arguments:
    Array_in -- []
        Array: 2D or 3D array
    Array_example -- []
        Array: 2D or 3D array
    method: -- 1 ... 5
        int: Resampling method
    """

    # Create old raster
    Array_out_shape = np.int_(Array_in.shape)
    Array_out_shape[-1] = Array_example.shape[-1]
    Array_out_shape[-2] = Array_example.shape[-2]

    if method == 1:
        interpolation_method='nearest'
    if method == 2:
        interpolation_method='bicubic'
    if method == 3:
        interpolation_method='bilinear'
    if method == 4:
        interpolation_method='cubic'
    if method == 5:
        interpolation_method='lanczos'

    if len(Array_out_shape) == 3:
        Array_out = np.zeros(Array_out_shape)

        for i in range(0, Array_out_shape[0]):
            Array_in_slice = Array_in[i,:,:]
            size=tuple(Array_out_shape[1:])

            Array_out_slice= misc.imresize(np.float_(Array_in_slice), size, interp=interpolation_method, mode='F')
            Array_out[i,:,:] = Array_out_slice

    elif len(Array_out_shape) == 2:

        size=tuple(Array_out_shape)
        Array_out= misc.imresize(np.float_(Array_in), size, interp=interpolation_method, mode='F')

    else:
        print('only 2D or 3D dimensions are supported')

    return(Array_out)